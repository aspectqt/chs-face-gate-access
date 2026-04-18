from flask import Flask, render_template, request, redirect, url_for, session, jsonify, Response, stream_with_context, has_request_context
from bson.objectid import ObjectId
from pymongo.errors import DuplicateKeyError
from pymongo import ASCENDING, DESCENDING
from datetime import datetime, timedelta, time as dtime
from decimal import Decimal, InvalidOperation
import atexit
import os
import cv2
import face_recognition
import numpy as np
from flask import send_file
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image as RLImage,
    LongTable,
    KeepTogether,
    HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from dotenv import load_dotenv, dotenv_values
from face_matching import (
    build_face_registration_metadata,
    build_student_face_index,
    detect_face_registration_conflict,
    face_distance,
    match_face_probe,
    sanitize_face_encodings,
)
from config import (
    DB_NAME,
    client,
    ensure_indexes,
    get_student_enrollment_collection,
    list_student_enrollment_collection_names,
    student_enrollment_collection_name,
    students,
    attendance_logs,
    attendance_logs_archive,
    sms_logs,
    sms_logs_archive,
    otp_requests,
    users,
    alerts,
    alerts_archive,
    login_history,
    failed_scans,
    sections,
    school_years,
    student_enrollments,
    audit_logs,
    login_attempts,
    attendance_corrections,
    attendance_corrections_archive,
    scheduled_reports,
    scheduled_report_runs,
    anomaly_rules,
    anomaly_events,
    system_settings,
    calendar_events,
    calendar_events_archive,
    early_timeout_requests,
    early_timeout_requests_archive,
)
import json
from PIL import Image
import base64
import threading
import time
import sys
import unicodedata
from functools import wraps
from urllib.parse import urlencode, urlparse
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
import re
import uuid
import requests
import traceback
import secrets
import hashlib
import smtplib
import socket
import shutil
import subprocess
from xml.sax.saxutils import escape as xml_escape
from email.message import EmailMessage
from services.sms_provider import SmsProvider, create_sms_provider_from_env
from services.otp_service import generate_otp_code, hash_otp_code, verify_otp_code
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None
    Alignment = None
    Border = None
    Font = None
    Side = None
    get_column_letter = None

# =====================================
# ENVIRONMENT
# =====================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_FILE_PATH = os.path.join(BASE_DIR, ".env")
load_dotenv(ENV_FILE_PATH)
db = client[DB_NAME]


def coerce_env_int(raw, default, minimum=None, maximum=None, name="value"):
    try:
        value = int(raw)
    except (TypeError, ValueError):
        print(f"[WARNING] Invalid integer for {name}: {raw!r}. Using default={default}.")
        value = int(default)

    if minimum is not None and value < minimum:
        value = minimum
    if maximum is not None and value > maximum:
        value = maximum
    return value


def env_int(name, default, minimum=None, maximum=None):
    raw = os.getenv(name, str(default))
    return coerce_env_int(raw, default, minimum=minimum, maximum=maximum, name=name)


def env_float(name, default, minimum=None, maximum=None):
    raw = os.getenv(name, str(default))
    try:
        value = float(raw)
    except (TypeError, ValueError):
        print(f"[WARNING] Invalid float for {name}: {raw!r}. Using default={default}.")
        value = float(default)

    if minimum is not None and value < minimum:
        value = minimum
    if maximum is not None and value > maximum:
        value = maximum
    return value


def coerce_env_bool(raw, default=False):
    if raw is None:
        raw = str(int(bool(default)))
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}


def env_bool(name, default=False):
    raw = os.getenv(name, str(int(bool(default))))
    return coerce_env_bool(raw, default)


def runtime_setting_value(*names, default=""):
    env_file_values = {}
    if ENV_FILE_PATH and os.path.exists(ENV_FILE_PATH):
        try:
            env_file_values = dotenv_values(ENV_FILE_PATH) or {}
        except Exception:
            env_file_values = {}

    for name in names:
        file_value = env_file_values.get(name)
        if file_value is not None and str(file_value).strip():
            return str(file_value)
        env_value = os.getenv(name, "")
        if str(env_value).strip():
            return str(env_value)
    return str(default)

HTTPS_ENABLED = env_bool("HTTPS_ENABLED", False)
FORCE_HTTPS = env_bool("FORCE_HTTPS", HTTPS_ENABLED)
TRUST_PROXY_HEADERS = env_bool("TRUST_PROXY_HEADERS", False)
FLASK_DEBUG_MODE = env_bool("FLASK_DEBUG", True)
DEV_AUTO_RELOAD = env_bool("DEV_AUTO_RELOAD", False)
SSL_CERT_FILE = os.getenv("SSL_CERT_FILE", "").strip()
SSL_KEY_FILE = os.getenv("SSL_KEY_FILE", "").strip()
FLASK_HOST = os.getenv("FLASK_HOST", "0.0.0.0").strip() or "0.0.0.0"
FLASK_PORT = env_int("FLASK_PORT", 5000, minimum=1, maximum=65535)
DEV_RELOAD_POLL_INTERVAL_MS = env_int("DEV_RELOAD_POLL_INTERVAL_MS", 1200, minimum=500, maximum=10000)
PREFERRED_URL_SCHEME = (
    os.getenv("PREFERRED_URL_SCHEME", "https" if HTTPS_ENABLED else "http").strip().lower()
    or ("https" if HTTPS_ENABLED else "http")
)
REPORT_PREPARED_BY_TITLE = os.getenv("REPORT_PREPARED_BY_TITLE", "Administrator").strip() or "Administrator"
REPORT_PRINCIPAL_NAME = os.getenv("REPORT_PRINCIPAL_NAME", "").strip()
REPORT_PRINCIPAL_TITLE = os.getenv("REPORT_PRINCIPAL_TITLE", "Principal").strip() or "Principal"
PDF_EXPORT_HEADER_REPUBLIC = "Republic of the Philippines"
PDF_EXPORT_HEADER_DEPARTMENT = "Department of Education"
PDF_EXPORT_HEADER_REGION = "REGION VII - CENTRAL VISAYAS"
PDF_EXPORT_HEADER_DIVISION = "SCHOOLS DIVISION OF NEGROS ORIENTAL"
PDF_EXPORT_HEADER_DISTRICT = "Sta. Catalina District 2"
PDF_EXPORT_HEADER_SCHOOL = "CAWITAN HIGH SCHOOL"
PDF_EXPORT_HEADER_ADDRESS = "Cawitan, Santa Catalina, Negros Oriental"
PDF_EXPORT_OLD_ENGLISH_FONT_NAME = "PdfExportOldEnglish"
PDF_EXPORT_OLD_ENGLISH_FONT_PATHS = (
    r"C:\Windows\Fonts\OLDENGL.TTF",
    r"C:\Windows\Fonts\GOTHIC.TTF",
)

# =====================================
# FLASK SETUP
# =====================================
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "super_secret_key_change_this")
if app.secret_key == "super_secret_key_change_this":
    print("[WARNING] FLASK_SECRET_KEY is not set. Using insecure default key.")
app.permanent_session_lifetime = timedelta(days=14)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = env_bool("SESSION_COOKIE_SECURE", HTTPS_ENABLED)
app.config["PREFERRED_URL_SCHEME"] = PREFERRED_URL_SCHEME
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024
AVATAR_UPLOAD_DIR = os.path.join(app.root_path, "static", "avatars")
os.makedirs(AVATAR_UPLOAD_DIR, exist_ok=True)

if TRUST_PROXY_HEADERS:
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# =====================================
# CONFIGURATION
# =====================================
SCAN_COOLDOWN_SECONDS = env_int("SCAN_COOLDOWN_SECONDS", 8, minimum=5, maximum=30)
SCAN_FRAME_WIDTH = env_int("SCAN_FRAME_WIDTH", 640, minimum=320, maximum=1920)
SCAN_FRAME_HEIGHT = env_int("SCAN_FRAME_HEIGHT", 480, minimum=240, maximum=1080)
SCAN_TARGET_FPS = env_int("SCAN_TARGET_FPS", 20, minimum=5, maximum=60)
SCAN_PROCESS_EVERY_N_FRAMES = env_int("SCAN_PROCESS_EVERY_N_FRAMES", 2, minimum=1, maximum=8)
SCAN_RECOGNITION_INTERVAL_MS = env_int("SCAN_RECOGNITION_INTERVAL_MS", 120, minimum=50, maximum=1000)
SCAN_RECOGNITION_SCALE_PERCENT = env_int("SCAN_RECOGNITION_SCALE_PERCENT", 50, minimum=25, maximum=100)
SCAN_MULTI_FACE_RETRY_SCALE_PERCENT = env_int(
    "SCAN_MULTI_FACE_RETRY_SCALE_PERCENT",
    max(int(SCAN_RECOGNITION_SCALE_PERCENT) + 15, 55),
    minimum=40,
    maximum=85,
)
SCAN_MULTI_FACE_RETRY_MAX_FACE_AREA_RATIO = env_float(
    "SCAN_MULTI_FACE_RETRY_MAX_FACE_AREA_RATIO",
    0.16,
    minimum=0.02,
    maximum=0.8,
)
SCAN_JPEG_QUALITY = env_int("SCAN_JPEG_QUALITY", 80, minimum=40, maximum=95)
SCAN_CAPTURE_FLUSH_GRABS = env_int("SCAN_CAPTURE_FLUSH_GRABS", 2, minimum=0, maximum=10)
SCAN_FORCE_RESIZE = env_bool("SCAN_FORCE_RESIZE", True)
SCAN_FACE_INDEX_ALLOW_LEGACY_IMAGE_FALLBACK = env_bool("SCAN_FACE_INDEX_ALLOW_LEGACY_IMAGE_FALLBACK", False)
SCAN_LIVE_IN_COOLDOWN_SECONDS = env_int("SCAN_LIVE_IN_COOLDOWN_SECONDS", SCAN_COOLDOWN_SECONDS, minimum=5, maximum=300)
SCAN_LIVE_OUT_COOLDOWN_MINUTES = env_int("SCAN_LIVE_OUT_COOLDOWN_MINUTES", 5, minimum=5, maximum=60)
SCAN_OUT_MINUTES = env_int("SCAN_OUT_MINUTES", 30, minimum=1, maximum=240)
SCAN_REPEAT_SUPPRESSION_SECONDS = env_int("SCAN_REPEAT_SUPPRESSION_SECONDS", 20, minimum=5, maximum=120)
SCAN_DUPLICATE_EVENT_WINDOW_SECONDS = env_int("SCAN_DUPLICATE_EVENT_WINDOW_SECONDS", 90, minimum=15, maximum=600)
SCAN_FACE_PRESENCE_RESET_SECONDS = env_int("SCAN_FACE_PRESENCE_RESET_SECONDS", 1, minimum=1, maximum=5)
UNKNOWN_ALERT_COOLDOWN_SECONDS = 30
UNREGISTERED_EVENT_COOLDOWN_SECONDS = 2
FACE_INDEX_MAX_ENCODINGS_PER_STUDENT = env_int("FACE_INDEX_MAX_ENCODINGS_PER_STUDENT", 20, minimum=5, maximum=60)
FACE_CAPTURE_DUPLICATE_DISTANCE = env_float("FACE_CAPTURE_DUPLICATE_DISTANCE", 0.018, minimum=0.0, maximum=0.1)
FACE_REGISTRATION_CONFLICT_DISTANCE = env_float("FACE_REGISTRATION_CONFLICT_DISTANCE", 0.30, minimum=0.0, maximum=0.6)
FACE_REGISTRATION_CONFLICT_MEAN_DISTANCE = env_float("FACE_REGISTRATION_CONFLICT_MEAN_DISTANCE", 0.36, minimum=0.0, maximum=0.8)
FACE_REGISTRATION_CONFLICT_SUPPORT_DISTANCE = env_float("FACE_REGISTRATION_CONFLICT_SUPPORT_DISTANCE", 0.40, minimum=0.0, maximum=0.8)
FACE_REGISTRATION_CONFLICT_SUPPORT_COUNT = env_int("FACE_REGISTRATION_CONFLICT_SUPPORT_COUNT", 3, minimum=1, maximum=10)
FACE_REGISTRATION_BRIGHTNESS_MIN = env_float("FACE_REGISTRATION_BRIGHTNESS_MIN", 12.0, minimum=0.0, maximum=255.0)
FACE_REGISTRATION_BRIGHTNESS_MAX = env_float("FACE_REGISTRATION_BRIGHTNESS_MAX", 250.0, minimum=0.0, maximum=255.0)
FACE_REGISTRATION_CONTRAST_MIN = env_float("FACE_REGISTRATION_CONTRAST_MIN", 4.5, minimum=0.0, maximum=100.0)
FACE_REGISTRATION_SHARPNESS_MIN = env_float("FACE_REGISTRATION_SHARPNESS_MIN", 3.8, minimum=0.0, maximum=200.0)
FACE_REGISTRATION_LOW_LIGHT_BRIGHTNESS = env_float("FACE_REGISTRATION_LOW_LIGHT_BRIGHTNESS", 80.0, minimum=0.0, maximum=255.0)
FACE_REGISTRATION_LOW_LIGHT_SHARPNESS_MIN = env_float("FACE_REGISTRATION_LOW_LIGHT_SHARPNESS_MIN", 2.2, minimum=0.0, maximum=200.0)
FACE_REGISTRATION_ENCODING_JITTERS = env_int("FACE_REGISTRATION_ENCODING_JITTERS", 1, minimum=1, maximum=4)
FACE_REGISTRATION_MIN_ENCODING_SPREAD = env_float("FACE_REGISTRATION_MIN_ENCODING_SPREAD", 0.03, minimum=0.0, maximum=0.5)
FACE_REGISTRATION_MIN_POSE_BUCKETS_STANDARD = env_int("FACE_REGISTRATION_MIN_POSE_BUCKETS_STANDARD", 4, minimum=0, maximum=10)
FACE_REGISTRATION_MIN_POSE_BUCKETS_SIMILAR = env_int("FACE_REGISTRATION_MIN_POSE_BUCKETS_SIMILAR", 6, minimum=0, maximum=20)
RECOGNITION_TOLERANCE = env_float("RECOGNITION_TOLERANCE", 0.45, minimum=0.2, maximum=0.75)
RECOGNITION_SCORE_TOLERANCE = env_float("RECOGNITION_SCORE_TOLERANCE", 0.47, minimum=0.2, maximum=0.8)
RECOGNITION_SUPPORT_TOLERANCE = env_float("RECOGNITION_SUPPORT_TOLERANCE", 0.50, minimum=0.2, maximum=0.8)
RECOGNITION_STRONG_MATCH_DISTANCE = env_float("RECOGNITION_STRONG_MATCH_DISTANCE", 0.36, minimum=0.2, maximum=0.75)
RECOGNITION_MARGIN_THRESHOLD = env_float("RECOGNITION_MARGIN_THRESHOLD", 0.035, minimum=0.0, maximum=0.2)
RECOGNITION_SHORTLIST_SIZE = env_int("RECOGNITION_SHORTLIST_SIZE", 8, minimum=1, maximum=24)
RECOGNITION_MIN_SUPPORT_COUNT = env_int("RECOGNITION_MIN_SUPPORT_COUNT", 2, minimum=1, maximum=6)
MIN_RECOGNITION_CONFIDENCE = env_float("MIN_RECOGNITION_CONFIDENCE", 55.0, minimum=0.0, maximum=100.0)
LIVE_RECOGNITION_DISTANCE_TOLERANCE = env_float("LIVE_RECOGNITION_DISTANCE_TOLERANCE", 0.41, minimum=0.2, maximum=0.75)
LIVE_RECOGNITION_MARGIN_THRESHOLD = env_float("LIVE_RECOGNITION_MARGIN_THRESHOLD", 0.055, minimum=0.0, maximum=0.2)
LIVE_RECOGNITION_DISTANCE_MARGIN_THRESHOLD = env_float("LIVE_RECOGNITION_DISTANCE_MARGIN_THRESHOLD", 0.028, minimum=0.0, maximum=0.2)
LIVE_RECOGNITION_MIN_CONFIDENCE = env_float("LIVE_RECOGNITION_MIN_CONFIDENCE", 60.0, minimum=0.0, maximum=100.0)
LIVE_RECOGNITION_CONFIRMATION_FRAMES = env_int("LIVE_RECOGNITION_CONFIRMATION_FRAMES", 2, minimum=1, maximum=6)
LIVE_RECOGNITION_CONFIRMATION_WINDOW_MS = env_int("LIVE_RECOGNITION_CONFIRMATION_WINDOW_MS", 900, minimum=300, maximum=3000)
LIVE_RECOGNITION_STRONG_MATCH_CONFIRMATION_FRAMES = env_int("LIVE_RECOGNITION_STRONG_MATCH_CONFIRMATION_FRAMES", 1, minimum=1, maximum=6)
LIVE_RECOGNITION_MIN_SHARPNESS = env_float("LIVE_RECOGNITION_MIN_SHARPNESS", 6.0, minimum=0.0, maximum=200.0)
LIVE_RECOGNITION_LOW_LIGHT_BRIGHTNESS = env_float("LIVE_RECOGNITION_LOW_LIGHT_BRIGHTNESS", 90.0, minimum=0.0, maximum=255.0)
LIVE_RECOGNITION_LOW_LIGHT_MIN_SHARPNESS = env_float("LIVE_RECOGNITION_LOW_LIGHT_MIN_SHARPNESS", 3.5, minimum=0.0, maximum=200.0)
LIVE_RECOGNITION_DOMINANT_FACE_RATIO = env_float("LIVE_RECOGNITION_DOMINANT_FACE_RATIO", 0.45, minimum=0.0, maximum=1.0)
LIVE_RECOGNITION_IMMEDIATE_MATCH_DISTANCE = env_float("LIVE_RECOGNITION_IMMEDIATE_MATCH_DISTANCE", 0.34, minimum=0.2, maximum=0.75)
LIVE_RECOGNITION_IMMEDIATE_CONFIDENCE = env_float("LIVE_RECOGNITION_IMMEDIATE_CONFIDENCE", 66.0, minimum=0.0, maximum=100.0)
LIVE_RECOGNITION_ENCODING_MAX_WIDTH = env_int("LIVE_RECOGNITION_ENCODING_MAX_WIDTH", 960, minimum=480, maximum=1920)
LIVE_RECOGNITION_ENCODING_JITTERS = env_int("LIVE_RECOGNITION_ENCODING_JITTERS", 1, minimum=1, maximum=4)
LIVE_RECOGNITION_FACE_QUALITY_CACHE_SECONDS = env_float("LIVE_RECOGNITION_FACE_QUALITY_CACHE_SECONDS", 0.18, minimum=0.05, maximum=1.0)
LIVE_RECOGNITION_ENCODING_CACHE_SECONDS = env_float("LIVE_RECOGNITION_ENCODING_CACHE_SECONDS", 0.24, minimum=0.05, maximum=1.0)
LIVE_RECOGNITION_CACHE_MAX_MOTION_COMPONENT = env_float("LIVE_RECOGNITION_CACHE_MAX_MOTION_COMPONENT", 0.085, minimum=0.0, maximum=1.0)
LIVE_RECOGNITION_CACHE_MAX_AREA_COMPONENT = env_float("LIVE_RECOGNITION_CACHE_MAX_AREA_COMPONENT", 0.18, minimum=0.0, maximum=1.5)
LIVE_RECOGNITION_CACHE_MAX_POSE_COMPONENT = env_float("LIVE_RECOGNITION_CACHE_MAX_POSE_COMPONENT", 0.055, minimum=0.0, maximum=1.0)
LIVE_RECOGNITION_MAX_TRACKED_FACES = env_int("LIVE_RECOGNITION_MAX_TRACKED_FACES", 12, minimum=2, maximum=40)
LIVE_RECOGNITION_MAX_RECOGNITIONS_PER_FRAME = env_int("LIVE_RECOGNITION_MAX_RECOGNITIONS_PER_FRAME", 5, minimum=1, maximum=20)
LIVE_RECOGNITION_MULTI_FACE_MAX_CANDIDATES = env_int("LIVE_RECOGNITION_MULTI_FACE_MAX_CANDIDATES", 8, minimum=1, maximum=24)
LIVE_RECOGNITION_MAX_NEW_ENCODINGS_PER_FRAME = env_int(
    "LIVE_RECOGNITION_MAX_NEW_ENCODINGS_PER_FRAME",
    LIVE_RECOGNITION_MAX_RECOGNITIONS_PER_FRAME,
    minimum=1,
    maximum=20,
)
LIVE_RECOGNITION_TRACK_TIMEOUT_SECONDS = env_float("LIVE_RECOGNITION_TRACK_TIMEOUT_SECONDS", 1.4, minimum=0.5, maximum=5.0)
LIVE_RECOGNITION_TRACK_STABILITY_FRAMES = env_int("LIVE_RECOGNITION_TRACK_STABILITY_FRAMES", 2, minimum=1, maximum=8)
LIVE_RECOGNITION_TRACK_RETRY_SECONDS = env_float("LIVE_RECOGNITION_TRACK_RETRY_SECONDS", 0.45, minimum=0.1, maximum=3.0)
LIVE_RECOGNITION_TRACK_UNKNOWN_RETRY_SECONDS = env_float("LIVE_RECOGNITION_TRACK_UNKNOWN_RETRY_SECONDS", 0.75, minimum=0.1, maximum=5.0)
LIVE_RECOGNITION_CONFIRMATION_RETRY_SECONDS = env_float("LIVE_RECOGNITION_CONFIRMATION_RETRY_SECONDS", 0.18, minimum=0.05, maximum=1.0)
LIVE_RECOGNITION_LIVENESS_RETRY_SECONDS = env_float("LIVE_RECOGNITION_LIVENESS_RETRY_SECONDS", 0.12, minimum=0.05, maximum=1.0)
LIVE_RECOGNITION_TRACK_MATCH_DISTANCE_RATIO = env_float("LIVE_RECOGNITION_TRACK_MATCH_DISTANCE_RATIO", 0.32, minimum=0.05, maximum=1.0)
LIVE_RECOGNITION_LOW_LIGHT_ENCODING_RETRY_ENABLED = env_bool("LIVE_RECOGNITION_LOW_LIGHT_ENCODING_RETRY_ENABLED", True)
LIVE_RECOGNITION_LOW_LIGHT_ENCODING_BRIGHTNESS = env_float("LIVE_RECOGNITION_LOW_LIGHT_ENCODING_BRIGHTNESS", 82.0, minimum=0.0, maximum=255.0)
LIVE_LIVENESS_ENABLED = env_bool("LIVE_LIVENESS_ENABLED", True)
# Anti-spoofing is intentionally bypassed in the live gate workflow to keep
# recognition responsive without changing the rest of the scan pipeline.
LIVE_SCAN_LIVENESS_ENABLED = False
LIVE_LIVENESS_MIN_TRACK_FRAMES = env_int("LIVE_LIVENESS_MIN_TRACK_FRAMES", 3, minimum=1, maximum=12)
LIVE_LIVENESS_MIN_MOTION_SCORE = env_float("LIVE_LIVENESS_MIN_MOTION_SCORE", 0.16, minimum=0.0, maximum=3.0)
LIVE_LIVENESS_MIN_POSE_SCORE = env_float("LIVE_LIVENESS_MIN_POSE_SCORE", 0.02, minimum=0.0, maximum=1.0)
LIVE_LIVENESS_POSE_GRACE_SECONDS = env_float("LIVE_LIVENESS_POSE_GRACE_SECONDS", 1.6, minimum=0.2, maximum=8.0)
LIVE_LIVENESS_SCORE_DECAY = env_float("LIVE_LIVENESS_SCORE_DECAY", 0.88, minimum=0.5, maximum=0.99)
LIVE_LIVENESS_TEXTURE_CHECK_ENABLED = env_bool("LIVE_LIVENESS_TEXTURE_CHECK_ENABLED", True)
LIVE_LIVENESS_MIN_TEXTURE = env_float("LIVE_LIVENESS_MIN_TEXTURE", 7.5, minimum=0.0, maximum=80.0)
LIVE_LIVENESS_MIN_CONTRAST = env_float("LIVE_LIVENESS_MIN_CONTRAST", 16.0, minimum=0.0, maximum=100.0)
LIVE_LIVENESS_MAX_HIGHLIGHT_RATIO = env_float("LIVE_LIVENESS_MAX_HIGHLIGHT_RATIO", 0.28, minimum=0.0, maximum=1.0)
LIVE_LIVENESS_PATCH_PARALLAX_ENABLED = env_bool("LIVE_LIVENESS_PATCH_PARALLAX_ENABLED", True)
LIVE_LIVENESS_PATCH_SIZE = env_int("LIVE_LIVENESS_PATCH_SIZE", 24, minimum=12, maximum=48)
LIVE_LIVENESS_MIN_PARALLAX_SCORE = env_float("LIVE_LIVENESS_MIN_PARALLAX_SCORE", 0.045, minimum=0.0, maximum=2.0)
LIVE_LIVENESS_PLANAR_STREAK_FRAMES = env_int("LIVE_LIVENESS_PLANAR_STREAK_FRAMES", 4, minimum=2, maximum=16)
LIVE_LIVENESS_MOTION_COMPONENT_THRESHOLD = env_float("LIVE_LIVENESS_MOTION_COMPONENT_THRESHOLD", 0.09, minimum=0.0, maximum=2.0)
LIVE_LIVENESS_MAX_PLANAR_POSE_COMPONENT = env_float("LIVE_LIVENESS_MAX_PLANAR_POSE_COMPONENT", 0.03, minimum=0.0, maximum=1.0)
LIVE_LIVENESS_MAX_PLANAR_PATCH_DELTA = env_float("LIVE_LIVENESS_MAX_PLANAR_PATCH_DELTA", 0.022, minimum=0.0, maximum=0.2)
LIVE_LIVENESS_MAX_PLANAR_GRAD_DELTA = env_float("LIVE_LIVENESS_MAX_PLANAR_GRAD_DELTA", 0.03, minimum=0.0, maximum=0.2)
LIVE_LIVENESS_BLINK_ENABLED = env_bool("LIVE_LIVENESS_BLINK_ENABLED", True)
LIVE_LIVENESS_BLINK_REQUIRED = env_bool("LIVE_LIVENESS_BLINK_REQUIRED", True)
LIVE_LIVENESS_BLINK_EAR_CLOSED = env_float("LIVE_LIVENESS_BLINK_EAR_CLOSED", 0.20, minimum=0.08, maximum=0.4)
LIVE_LIVENESS_BLINK_EAR_OPEN = env_float("LIVE_LIVENESS_BLINK_EAR_OPEN", 0.245, minimum=0.1, maximum=0.5)
LIVE_LIVENESS_BLINK_MIN_CLOSED_FRAMES = env_int("LIVE_LIVENESS_BLINK_MIN_CLOSED_FRAMES", 1, minimum=1, maximum=8)
LIVE_LIVENESS_BLINK_INTERVAL_SECONDS = env_float("LIVE_LIVENESS_BLINK_INTERVAL_SECONDS", 0.22, minimum=0.05, maximum=2.0)
LIVE_LIVENESS_BLINK_RECENT_SECONDS = env_float("LIVE_LIVENESS_BLINK_RECENT_SECONDS", 5.0, minimum=0.5, maximum=20.0)
LIVE_LIVENESS_LANDMARK_POSE_ENABLED = env_bool("LIVE_LIVENESS_LANDMARK_POSE_ENABLED", True)
LIVE_LIVENESS_POSE_SAMPLE_INTERVAL_SECONDS = env_float("LIVE_LIVENESS_POSE_SAMPLE_INTERVAL_SECONDS", 0.14, minimum=0.05, maximum=1.0)
LIVE_LIVENESS_POSE_WINDOW_SECONDS = env_float("LIVE_LIVENESS_POSE_WINDOW_SECONDS", 3.5, minimum=1.0, maximum=10.0)
LIVE_LIVENESS_MIN_LANDMARK_POSE_SPAN = env_float("LIVE_LIVENESS_MIN_LANDMARK_POSE_SPAN", 0.075, minimum=0.0, maximum=1.0)
LIVE_LIVENESS_MIN_LANDMARK_POSE_SAMPLES = env_int("LIVE_LIVENESS_MIN_LANDMARK_POSE_SAMPLES", 2, minimum=1, maximum=8)
LIVE_LIVENESS_MIN_PATCH_DIVERSITY = env_float("LIVE_LIVENESS_MIN_PATCH_DIVERSITY", 0.12, minimum=0.0, maximum=2.0)
LIVE_LIVENESS_DISPLAY_CHECK_ENABLED = env_bool("LIVE_LIVENESS_DISPLAY_CHECK_ENABLED", True)
LIVE_LIVENESS_DISPLAY_MIN_STREAK = env_int("LIVE_LIVENESS_DISPLAY_MIN_STREAK", 2, minimum=1, maximum=6)
LIVE_LIVENESS_DISPLAY_MIN_RECT_AREA_RATIO = env_float("LIVE_LIVENESS_DISPLAY_MIN_RECT_AREA_RATIO", 1.45, minimum=1.0, maximum=20.0)
LIVE_LIVENESS_DISPLAY_MAX_RECT_AREA_RATIO = env_float("LIVE_LIVENESS_DISPLAY_MAX_RECT_AREA_RATIO", 6.8, minimum=1.0, maximum=30.0)
LIVE_LIVENESS_AI_MODEL_ENABLED = env_bool("LIVE_LIVENESS_AI_MODEL_ENABLED", False)
LIVE_LIVENESS_AI_MODEL_PATH = os.getenv("LIVE_LIVENESS_AI_MODEL_PATH", "").strip()
LIVE_LIVENESS_AI_MIN_LIVE_SCORE = env_float("LIVE_LIVENESS_AI_MIN_LIVE_SCORE", 0.72, minimum=0.0, maximum=1.0)
LIVE_LIVENESS_PROFILE = (os.getenv("LIVE_LIVENESS_PROFILE", "strict").strip().lower() or "strict")
if LIVE_LIVENESS_PROFILE not in {"strict", "balanced", "lenient"}:
    LIVE_LIVENESS_PROFILE = "strict"

LIVE_LIVENESS_BASE_MIN_TRACK_FRAMES = int(LIVE_LIVENESS_MIN_TRACK_FRAMES)
LIVE_LIVENESS_BASE_MIN_MOTION_SCORE = float(LIVE_LIVENESS_MIN_MOTION_SCORE)
LIVE_LIVENESS_BASE_MIN_POSE_SCORE = float(LIVE_LIVENESS_MIN_POSE_SCORE)
LIVE_LIVENESS_BASE_POSE_GRACE_SECONDS = float(LIVE_LIVENESS_POSE_GRACE_SECONDS)
LIVE_LIVENESS_BASE_MIN_TEXTURE = float(LIVE_LIVENESS_MIN_TEXTURE)
LIVE_LIVENESS_BASE_MIN_CONTRAST = float(LIVE_LIVENESS_MIN_CONTRAST)
LIVE_LIVENESS_BASE_MAX_HIGHLIGHT_RATIO = float(LIVE_LIVENESS_MAX_HIGHLIGHT_RATIO)
LIVE_LIVENESS_BASE_MIN_PARALLAX_SCORE = float(LIVE_LIVENESS_MIN_PARALLAX_SCORE)
LIVE_LIVENESS_BASE_PLANAR_STREAK_FRAMES = int(LIVE_LIVENESS_PLANAR_STREAK_FRAMES)
LIVE_LIVENESS_BASE_MAX_PLANAR_POSE_COMPONENT = float(LIVE_LIVENESS_MAX_PLANAR_POSE_COMPONENT)
LIVE_LIVENESS_BASE_MAX_PLANAR_PATCH_DELTA = float(LIVE_LIVENESS_MAX_PLANAR_PATCH_DELTA)
LIVE_LIVENESS_BASE_MAX_PLANAR_GRAD_DELTA = float(LIVE_LIVENESS_MAX_PLANAR_GRAD_DELTA)
LIVE_LIVENESS_BASE_BLINK_REQUIRED = bool(LIVE_LIVENESS_BLINK_REQUIRED)
LIVE_LIVENESS_LENIENT_BLINK_REQUIRED = bool(env_bool("LIVE_LIVENESS_BLINK_REQUIRED", False))
LIVE_LIVENESS_BASE_BLINK_INTERVAL_SECONDS = float(LIVE_LIVENESS_BLINK_INTERVAL_SECONDS)
LIVE_LIVENESS_BASE_BLINK_RECENT_SECONDS = float(LIVE_LIVENESS_BLINK_RECENT_SECONDS)
LIVE_LIVENESS_BASE_POSE_SAMPLE_INTERVAL_SECONDS = float(LIVE_LIVENESS_POSE_SAMPLE_INTERVAL_SECONDS)
LIVE_LIVENESS_BASE_POSE_WINDOW_SECONDS = float(LIVE_LIVENESS_POSE_WINDOW_SECONDS)
LIVE_LIVENESS_BASE_MIN_LANDMARK_POSE_SPAN = float(LIVE_LIVENESS_MIN_LANDMARK_POSE_SPAN)
LIVE_LIVENESS_BASE_MIN_LANDMARK_POSE_SAMPLES = int(LIVE_LIVENESS_MIN_LANDMARK_POSE_SAMPLES)
LIVE_LIVENESS_BASE_MIN_PATCH_DIVERSITY = float(LIVE_LIVENESS_MIN_PATCH_DIVERSITY)

if LIVE_LIVENESS_PROFILE == "strict":
    LIVE_LIVENESS_MIN_TRACK_FRAMES = max(int(LIVE_LIVENESS_MIN_TRACK_FRAMES), 4)
    LIVE_LIVENESS_MIN_MOTION_SCORE = max(float(LIVE_LIVENESS_MIN_MOTION_SCORE), 0.24)
    LIVE_LIVENESS_MIN_POSE_SCORE = max(float(LIVE_LIVENESS_MIN_POSE_SCORE), 0.032)
    LIVE_LIVENESS_POSE_GRACE_SECONDS = min(float(LIVE_LIVENESS_POSE_GRACE_SECONDS), 1.25)
    LIVE_LIVENESS_MIN_TEXTURE = max(float(LIVE_LIVENESS_MIN_TEXTURE), 9.2)
    LIVE_LIVENESS_MIN_CONTRAST = max(float(LIVE_LIVENESS_MIN_CONTRAST), 18.0)
    LIVE_LIVENESS_MAX_HIGHLIGHT_RATIO = min(float(LIVE_LIVENESS_MAX_HIGHLIGHT_RATIO), 0.22)
    LIVE_LIVENESS_MIN_PARALLAX_SCORE = max(float(LIVE_LIVENESS_MIN_PARALLAX_SCORE), 0.058)
    LIVE_LIVENESS_PLANAR_STREAK_FRAMES = min(int(LIVE_LIVENESS_PLANAR_STREAK_FRAMES), 3)
    LIVE_LIVENESS_MAX_PLANAR_PATCH_DELTA = min(float(LIVE_LIVENESS_MAX_PLANAR_PATCH_DELTA), 0.018)
    LIVE_LIVENESS_MAX_PLANAR_GRAD_DELTA = min(float(LIVE_LIVENESS_MAX_PLANAR_GRAD_DELTA), 0.024)
    LIVE_LIVENESS_MAX_PLANAR_POSE_COMPONENT = min(float(LIVE_LIVENESS_MAX_PLANAR_POSE_COMPONENT), 0.026)
    LIVE_LIVENESS_BLINK_RECENT_SECONDS = min(float(LIVE_LIVENESS_BLINK_RECENT_SECONDS), 4.0)
    LIVE_LIVENESS_BLINK_INTERVAL_SECONDS = min(float(LIVE_LIVENESS_BLINK_INTERVAL_SECONDS), 0.16)
    LIVE_LIVENESS_POSE_SAMPLE_INTERVAL_SECONDS = min(float(LIVE_LIVENESS_POSE_SAMPLE_INTERVAL_SECONDS), 0.18)
    LIVE_LIVENESS_POSE_WINDOW_SECONDS = min(float(LIVE_LIVENESS_POSE_WINDOW_SECONDS), 4.0)
    LIVE_LIVENESS_MIN_LANDMARK_POSE_SPAN = max(float(LIVE_LIVENESS_MIN_LANDMARK_POSE_SPAN), 0.082)
    LIVE_LIVENESS_MIN_LANDMARK_POSE_SAMPLES = max(int(LIVE_LIVENESS_MIN_LANDMARK_POSE_SAMPLES), 2)
    LIVE_LIVENESS_MIN_PATCH_DIVERSITY = max(float(LIVE_LIVENESS_MIN_PATCH_DIVERSITY), 0.115)
elif LIVE_LIVENESS_PROFILE == "lenient":
    LIVE_LIVENESS_MIN_TRACK_FRAMES = min(int(LIVE_LIVENESS_MIN_TRACK_FRAMES), 2)
    LIVE_LIVENESS_MIN_MOTION_SCORE = min(float(LIVE_LIVENESS_MIN_MOTION_SCORE), 0.11)
    LIVE_LIVENESS_MIN_POSE_SCORE = min(float(LIVE_LIVENESS_MIN_POSE_SCORE), 0.012)
    LIVE_LIVENESS_POSE_GRACE_SECONDS = max(float(LIVE_LIVENESS_POSE_GRACE_SECONDS), 2.2)
    LIVE_LIVENESS_MIN_TEXTURE = min(float(LIVE_LIVENESS_MIN_TEXTURE), 6.0)
    LIVE_LIVENESS_MIN_CONTRAST = min(float(LIVE_LIVENESS_MIN_CONTRAST), 13.0)
    LIVE_LIVENESS_MAX_HIGHLIGHT_RATIO = max(float(LIVE_LIVENESS_MAX_HIGHLIGHT_RATIO), 0.34)
    LIVE_LIVENESS_MIN_PARALLAX_SCORE = min(float(LIVE_LIVENESS_MIN_PARALLAX_SCORE), 0.032)
    LIVE_LIVENESS_PLANAR_STREAK_FRAMES = max(int(LIVE_LIVENESS_PLANAR_STREAK_FRAMES), 5)
    LIVE_LIVENESS_MAX_PLANAR_PATCH_DELTA = max(float(LIVE_LIVENESS_MAX_PLANAR_PATCH_DELTA), 0.028)
    LIVE_LIVENESS_MAX_PLANAR_GRAD_DELTA = max(float(LIVE_LIVENESS_MAX_PLANAR_GRAD_DELTA), 0.038)
    LIVE_LIVENESS_MAX_PLANAR_POSE_COMPONENT = max(float(LIVE_LIVENESS_MAX_PLANAR_POSE_COMPONENT), 0.04)
    LIVE_LIVENESS_BLINK_REQUIRED = bool(LIVE_LIVENESS_LENIENT_BLINK_REQUIRED)
    LIVE_LIVENESS_BLINK_RECENT_SECONDS = max(float(LIVE_LIVENESS_BLINK_RECENT_SECONDS), 6.5)
    LIVE_LIVENESS_BLINK_INTERVAL_SECONDS = max(float(LIVE_LIVENESS_BLINK_INTERVAL_SECONDS), 0.25)
    LIVE_LIVENESS_POSE_SAMPLE_INTERVAL_SECONDS = max(float(LIVE_LIVENESS_POSE_SAMPLE_INTERVAL_SECONDS), 0.2)
    LIVE_LIVENESS_POSE_WINDOW_SECONDS = max(float(LIVE_LIVENESS_POSE_WINDOW_SECONDS), 4.5)
    LIVE_LIVENESS_MIN_LANDMARK_POSE_SPAN = min(float(LIVE_LIVENESS_MIN_LANDMARK_POSE_SPAN), 0.06)
    LIVE_LIVENESS_MIN_LANDMARK_POSE_SAMPLES = min(int(LIVE_LIVENESS_MIN_LANDMARK_POSE_SAMPLES), 2)
    LIVE_LIVENESS_MIN_PATCH_DIVERSITY = min(float(LIVE_LIVENESS_MIN_PATCH_DIVERSITY), 0.095)

PASSWORD_HASH_METHOD = "pbkdf2:sha256:600000"
ALLOWED_AVATAR_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}
MAX_AVATAR_SIZE_BYTES = 5 * 1024 * 1024
MIN_PASSWORD_LENGTH = 8
MAX_PASSWORD_LENGTH = 128
PASSWORD_RESET_TOKEN_TTL_MINUTES = env_int("PASSWORD_RESET_TOKEN_TTL_MINUTES", 30, minimum=5, maximum=120)
PASSWORD_RESET_REQUEST_WINDOW_MINUTES = env_int("PASSWORD_RESET_REQUEST_WINDOW_MINUTES", 15, minimum=1, maximum=240)
PASSWORD_RESET_MAX_REQUESTS_PER_WINDOW = env_int("PASSWORD_RESET_MAX_REQUESTS_PER_WINDOW", 3, minimum=1, maximum=20)
PASSWORD_RESET_RATE_LIMIT_ENABLED = env_bool("PASSWORD_RESET_RATE_LIMIT_ENABLED", True)
PASSWORD_RESET_DEV_LINK_FALLBACK = env_bool("PASSWORD_RESET_DEV_LINK_FALLBACK", False)
LOGIN_ATTEMPT_WINDOW_MINUTES = env_int("LOGIN_ATTEMPT_WINDOW_MINUTES", 15, minimum=1, maximum=240)
LOGIN_MAX_ATTEMPTS = env_int("LOGIN_MAX_ATTEMPTS", 5, minimum=3, maximum=20)
LOGIN_LOCKOUT_MINUTES = env_int("LOGIN_LOCKOUT_MINUTES", 15, minimum=1, maximum=240)
SMS_BALANCE_CACHE_TTL_SECONDS = env_int("SMS_BALANCE_CACHE_TTL_SECONDS", 60, minimum=10, maximum=600)
SMS_BALANCE_LOW_THRESHOLD = env_int("SMS_BALANCE_LOW_THRESHOLD", 50, minimum=0, maximum=1000000)
SMS_TEMPLATE_MAX_LENGTH = env_int("SMS_TEMPLATE_MAX_LENGTH", 480, minimum=80, maximum=2000)
ATTENDANCE_SMS_TEMPLATE_DEFAULT = (
    os.getenv("ATTENDANCE_SMS_TEMPLATE", "").strip()
    or "CHS Gate Access: {student_name} {movement_text} the gate ({status}) at {time} on {date}."
)
ATTENDANCE_SMS_TEMPLATE_DOC_ID = "attendance_gate_scan"
ATTENDANCE_SMS_NOTIFICATION_SETTINGS_KEY = "attendance_sms_notifications"
ATTENDANCE_SMS_TEMPLATE_VARIABLES = (
    "student_name",
    "student_id",
    "movement_text",
    "gate_action",
    "status",
    "session",
    "time",
    "date",
)
CORRECTION_ALLOWED_STATUSES = {"Present", "Late", "Absent"}
MORNING_START = dtime(hour=5, minute=0)
NOON_START = dtime(hour=12, minute=0)
AFTERNOON_START = dtime(hour=13, minute=0)
AFTERNOON_END_START = dtime(hour=17, minute=0)
MORNING_LATE_THRESHOLD = dtime(hour=8, minute=15)
AFTERNOON_LATE_THRESHOLD = dtime(hour=13, minute=15)
GRADE_LEVEL_OPTIONS = ["Grade 7", "Grade 8", "Grade 9", "Grade 10", "Grade 11", "Grade 12"]
SCHOOL_YEAR_START_MONTH = env_int("SCHOOL_YEAR_START_MONTH", 6, minimum=1, maximum=12)
SCHOOL_YEAR_SESSION_KEY = "selected_school_year"
STUDENT_IMPORT_ALLOWED_EXTENSIONS = {"xlsx"}
STUDENT_IMPORT_MAX_ROWS = env_int("STUDENT_IMPORT_MAX_ROWS", 0, minimum=0, maximum=200000)
REQUIRED_STUDENT_IMPORT_FIELDS = {"lrn", "name", "gender"}
STUDENT_IMPORT_HEADER_ALIASES = {
    "lrn": {"lrn", "learnerreferencenumber", "learnerreference", "studentid", "studentnumber"},
    "name": {"name", "studentname", "fullname", "full_name"},
    "grade_level": {"gradelevel", "grade", "gradelevelsection", "yearlevel"},
    "gender": {"gender", "sex", "sexgender", "sexorgender"},
    "section": {"section", "advisory", "class", "homeroom"},
}
PREDEFINED_SECTIONS_BY_GRADE = {
    "Grade 7": ["AVILA", "CALINGACION", "GUIRON", "VILLASAN"],
    "Grade 8": ["ELNAR", "FERRATER", "FLORES", "SARNE", "TRACES"],
    "Grade 9": ["NUIQUE", "PALENCIA", "RUBIO"],
    "Grade 10": ["BORROMEO", "FEROLINO", "PONSICA", "SY"],
    "Grade 11": ["ABEJO", "CABILES", "DAGAMI", "ESTRELLA"],
    "Grade 12": ["BSINT"],
}
PREDEFINED_SECTION_LOOKUP = {
    section.lower(): {"grade_level": grade_level, "section": section}
    for grade_level, section_values in PREDEFINED_SECTIONS_BY_GRADE.items()
    for section in section_values
}
STUDENT_IMPORT_TEMPLATE_HEADERS = ["LRN", "", "NAME", "", "", "", "Sex / Gender", "Section", "Grade Level"]
STUDENT_IMPORT_TEMPLATE_SAMPLE_ROWS = [
    ["120526180006", "ARADAN,LOUIS MIGUEL, SITOY", "M", "AVILA", "Grade 7"],
    ["120507180005", "AUJERO,IYAN, ARDIENTE", "M", "AVILA", "Grade 7"],
    ["120508130014", "BALIGASA,RICKY, AURILIO", "M", "AVILA", "Grade 7"],
    ["120507180026", "ALFONSO,CHADITH, GAUDIA", "F", "AVILA", "Grade 7"],
    ["120526180025", "BANDICO,REXCYN MAE, QUILAT", "F", "AVILA", "Grade 7"],
    ["120511180001", "SANTOS,JUAN, DELA CRUZ", "M", "ABEJO", "Grade 11"],
    ["120511180002", "REYES,MARIA, SANTOS", "F", "BSINT", "Grade 12"],
]
OTP_CODE_LENGTH = env_int("OTP_CODE_LENGTH", 6, minimum=4, maximum=10)
OTP_EXPIRES_MINUTES = env_int("OTP_EXPIRES_MINUTES", 5, minimum=1, maximum=30)
OTP_MAX_ATTEMPTS = env_int("OTP_MAX_ATTEMPTS", 5, minimum=1, maximum=10)
OTP_THROTTLE_SECONDS = env_int("OTP_THROTTLE_SECONDS", 60, minimum=0, maximum=3600)
OTP_MAX_PER_HOUR = env_int("OTP_MAX_PER_HOUR", 5, minimum=1, maximum=100)
ENABLE_SECURITY_HEADERS = env_bool("ENABLE_SECURITY_HEADERS", True)
CSP_ENFORCE = env_bool("CSP_ENFORCE", False)
VALID_SCAN_SESSION_MODES = {"auto", "manual_in", "manual_out"}
VALID_SCAN_LIVENESS_PROFILES = {"strict", "balanced", "lenient"}
CSRF_SESSION_KEY = "_csrf_token"
CSRF_HEADER_NAME = "X-CSRF-Token"
CSRF_ALLOWED_METHODS = {"POST", "PUT", "PATCH", "DELETE"}
SCHEDULED_REPORT_ALLOWED_FREQUENCIES = {"daily", "weekly", "monthly"}
SCHEDULED_REPORT_DEFAULT_SEND_TIME = "07:00"
SCHEDULED_REPORT_MAX_RECIPIENTS = 20
SCHEDULED_REPORT_MAX_RESULTS = 100
ANOMALY_ALLOWED_METRICS = {"late_count", "failed_sms_count", "unknown_scan_count", "pending_correction_count"}
ANOMALY_ALLOWED_OPERATORS = {"gt", "gte", "lt", "lte"}
ANOMALY_ALLOWED_SEVERITIES = {"info", "warn", "high"}
ANOMALY_DEFAULT_COOLDOWN_MINUTES = env_int("ANOMALY_DEFAULT_COOLDOWN_MINUTES", 60, minimum=5, maximum=1440)
BACKGROUND_JOB_INTERVAL_SECONDS = env_int("BACKGROUND_JOB_INTERVAL_SECONDS", 60, minimum=15, maximum=600)
ENABLE_BACKGROUND_JOBS = env_bool("ENABLE_BACKGROUND_JOBS", True)
ALERT_NOTIFICATION_CLEANUP_ENABLED = env_bool("ALERT_NOTIFICATION_CLEANUP_ENABLED", True)
ALERT_NOTIFICATION_RETENTION_MONTHS = env_int("ALERT_NOTIFICATION_RETENTION_MONTHS", 1, minimum=1, maximum=24)
CONTENT_SECURITY_POLICY = "; ".join([
    "default-src 'self'",
    "base-uri 'self'",
    "object-src 'none'",
    "frame-ancestors 'none'",
    "form-action 'self'",
    "img-src 'self' data: blob:",
    "script-src 'self' 'unsafe-inline' 'unsafe-eval' 'wasm-unsafe-eval' https://cdn.tailwindcss.com https://cdn.jsdelivr.net",
    "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net",
    "font-src 'self' data: https://cdn.jsdelivr.net",
    "connect-src 'self' https://cdn.jsdelivr.net",
    "media-src 'self' blob:",
    "worker-src 'self' blob:",
])

sms_provider = create_sms_provider_from_env()
sms_provider_startup_status = sms_provider.validate_configuration(raise_on_error=False)
print(f"[INFO] MongoDB database selected: {DB_NAME}")
if DB_NAME != "face_gate_db":
    print(f"[WARNING] MONGODB_DB_NAME is '{DB_NAME}'. Expected 'face_gate_db'.")
if sms_provider_startup_status.get("status") != "ok":
    print(f"[WARNING] SMS provider not ready at startup: {sms_provider_startup_status.get('message')}")

# =====================================
# GLOBAL STATE
# =====================================
APP_START_TS = time.time()
background_jobs_started = False
background_jobs_lock = threading.Lock()
SCAN_RECOGNITION_SCALE = SCAN_RECOGNITION_SCALE_PERCENT / 100.0
last_scanned = {}
scan_presence_locks = {}
dev_reload_lock = threading.Lock()
dev_reload_cache = {
    "checked_at": 0.0,
    "token": "",
}

scan_lock = threading.Lock()
scan_event_condition = threading.Condition(scan_lock)
frame_processing_lock = threading.Lock()
scan_state = {
    "active": False,
    "capture": None,
    "events": [],
    "event_counter": 0,
    "last_unknown_alert_ts": 0.0,
    "last_not_registered_ts": 0.0,
    "last_multi_face_ts": 0.0,
    "model_status": "idle",
    "known_encodings": [],
    "known_students": [],
    "face_index_loading": False,
    "session_mode": "auto",
    "liveness_profile": "strict",
    "pending_recognition": {},
    "face_tracks": {},
    "next_face_track_id": 1,
    "face_track_cursor": 0,
    "last_faces_payload": [],
}
liveness_ai_model = None
liveness_ai_model_loaded = False
liveness_ai_model_lock = threading.Lock()
liveness_ai_model_error = ""

alert_lock = threading.Lock()
alert_revision = 0
alert_cleanup_lock = threading.Lock()
alert_cleanup_state = {
    "last_checked_date": "",
}
data_change_lock = threading.Lock()
data_change_revision = 0
data_change_domains = {
    "students": 0,
    "sections": 0,
    "gate_logs": 0,
    "sms_logs": 0,
    "users": 0,
}

sms_balance_lock = threading.Lock()
sms_balance_cache = {
    "status": "idle",
    "units": None,
    "message": "",
    "provider": "PHILSMS",
    "http_status": None,
    "probe_path": "",
    "checked_at": "",
    "checked_ts": 0.0,
}

ROLE_FULL_ADMIN = "Full Admin"
ROLE_STAFF = "Staff"
LEGACY_LIMITED_ACCESS_ROLE = "Limited Access"

ROLE_PERMISSIONS = {
    ROLE_FULL_ADMIN: {"dashboard", "scan", "students_read", "students_write", "face_register", "logs", "analytics", "users_manage", "alerts_manage"},
    ROLE_STAFF: {"dashboard", "scan", "students_read", "face_register"},
}
password_reset_tokens = users.database["password_reset_tokens"]
sms_templates = users.database["sms_templates"]
try:
    password_reset_tokens.create_index([("token_hash", 1)], unique=True)
    password_reset_tokens.create_index([("expiresAt", 1)], expireAfterSeconds=0)
    password_reset_tokens.create_index([("email", 1), ("createdAt", -1)])
    password_reset_tokens.create_index([("requestIp", 1), ("createdAt", -1)])
except Exception as exc:
    print(f"[WARNING] Could not initialize password_reset_tokens indexes: {exc}")
try:
    sms_templates.create_index([("updatedAt", -1)])
except Exception as exc:
    print(f"[WARNING] Could not initialize sms_templates indexes: {exc}")


# =====================================
# HELPER FUNCTIONS
# =====================================
def login_required():
    return "admin" in session


def hash_password(password):
    return generate_password_hash(password, method=PASSWORD_HASH_METHOD)


def normalize_role_value(value, default=ROLE_STAFF):
    normalized = str(value or "").strip()
    if normalized == LEGACY_LIMITED_ACCESS_ROLE:
        normalized = ROLE_STAFF
    if normalized in ROLE_PERMISSIONS:
        return normalized
    return default


def normalize_account_role(value, username=""):
    username_text = str(username or "").strip().lower()
    default_role = ROLE_FULL_ADMIN if username_text == "admin" else ROLE_STAFF
    return normalize_role_value(value, default=default_role)


def post_login_redirect(role):
    # Role-based redirect map can be extended when distinct staff pages exist.
    role_routes = {
        ROLE_FULL_ADMIN: "dashboard",
        ROLE_STAFF: "dashboard",
    }
    return url_for(role_routes.get(normalize_role_value(role), "dashboard"))


def validate_email_format(value):
    if not value:
        return False
    return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value) is not None


def normalize_email_value(value):
    return str(value or "").strip().lower()


def build_internal_account_email(username):
    username_text = str(username or "").strip().lower()
    local_part = re.sub(r"[^a-z0-9._-]+", ".", username_text)
    local_part = re.sub(r"[._-]{2,}", ".", local_part).strip("._-")
    if not local_part:
        local_part = f"user-{uuid.uuid4().hex[:8]}"
    return normalize_email_value(f"{local_part}@chs.local")


def validate_password_reset_input(new_password, confirm_password):
    if not new_password:
        return "New password is required.", "newPassword"
    if not confirm_password:
        return "Please confirm your new password.", "confirmPassword"
    if len(new_password) < MIN_PASSWORD_LENGTH or len(new_password) > MAX_PASSWORD_LENGTH:
        return f"Password must be between {MIN_PASSWORD_LENGTH} and {MAX_PASSWORD_LENGTH} characters.", "newPassword"
    if new_password != confirm_password:
        return "Passwords do not match.", "confirmPassword"

    checks = [
        bool(re.search(r"[A-Z]", new_password)),
        bool(re.search(r"[a-z]", new_password)),
        bool(re.search(r"[0-9]", new_password)),
        bool(re.search(r"[^A-Za-z0-9]", new_password)),
    ]
    if sum(checks) < 3:
        return "Use at least 3 of: uppercase, lowercase, number, special character.", "newPassword"

    return "", ""


def hash_password_reset_token(token):
    return hashlib.sha256(str(token or "").encode("utf-8")).hexdigest()


def build_password_reset_link(token):
    base_url = os.getenv("APP_BASE_URL", "").strip().rstrip("/")
    path = url_for("reset_password", token=token)
    if base_url:
        return f"{base_url}{path}"
    return url_for("reset_password", token=token, _external=True)


def normalize_smtp_password(password, username="", sender=""):
    password_value = str(password or "")
    if not password_value:
        return ""

    trimmed_password = password_value.strip()
    account_candidates = (
        str(username or "").strip().lower(),
        str(sender or "").strip().lower(),
    )
    is_gmail_account = any(candidate.endswith("@gmail.com") for candidate in account_candidates if candidate)
    if not is_gmail_account:
        return trimmed_password

    compact_password = re.sub(r"[\s-]+", "", trimmed_password)
    # Gmail app passwords are often pasted in grouped blocks; normalize them before SMTP login.
    if re.fullmatch(r"[A-Za-z0-9\s-]+", trimmed_password) and len(compact_password) == 16:
        return compact_password
    return trimmed_password


def smtp_settings():
    smtp_host = (
        runtime_setting_value("SMTP_HOST", "MAIL_SERVER", "EMAIL_HOST").strip()
    )
    smtp_port = coerce_env_int(
        runtime_setting_value("SMTP_PORT", "MAIL_PORT", default="587"),
        587,
        minimum=1,
        maximum=65535,
        name="SMTP_PORT",
    )
    smtp_username = (
        runtime_setting_value(
            "SMTP_USERNAME",
            "MAIL_USERNAME",
            "EMAIL_HOST_USER",
            "EMAIL_ADDRESS",
            "GMAIL_ADDRESS",
        ).strip()
    )
    smtp_password = (
        runtime_setting_value(
            "SMTP_PASSWORD",
            "MAIL_PASSWORD",
            "EMAIL_HOST_PASSWORD",
            "EMAIL_APP_PASSWORD",
            "GMAIL_APP_PASSWORD",
        )
    )
    smtp_from = (
        runtime_setting_value("SMTP_FROM", "MAIL_DEFAULT_SENDER").strip()
        or smtp_username
    )
    smtp_password = normalize_smtp_password(smtp_password, smtp_username, smtp_from)
    smtp_security = runtime_setting_value("SMTP_SECURITY", default="").strip().lower()
    smtp_use_ssl = coerce_env_bool(runtime_setting_value("SMTP_USE_SSL", default="0"), False)
    smtp_use_tls = coerce_env_bool(runtime_setting_value("SMTP_USE_TLS", default="1"), True)

    if smtp_security == "ssl":
        smtp_use_ssl = True
        smtp_use_tls = False
    elif smtp_security in {"starttls", "tls"}:
        smtp_use_ssl = False
        smtp_use_tls = True
    elif smtp_security == "none":
        smtp_use_ssl = False
        smtp_use_tls = False
    elif smtp_port == 465:
        smtp_use_ssl = True
        smtp_use_tls = False

    if not smtp_host and smtp_username.lower().endswith("@gmail.com"):
        smtp_host = "smtp.gmail.com"
        if not runtime_setting_value("SMTP_PORT", "MAIL_PORT", default="").strip():
            smtp_port = 587
            smtp_use_ssl = False
            smtp_use_tls = True

    return {
        "host": smtp_host,
        "port": smtp_port,
        "username": smtp_username,
        "password": smtp_password,
        "sender": smtp_from,
        "use_ssl": smtp_use_ssl,
        "use_tls": smtp_use_tls,
    }


def smtp_configuration_error():
    settings = smtp_settings()
    if not settings["host"]:
        return "Email service is not configured. Set SMTP_HOST (or MAIL_SERVER)."
    if not settings["sender"]:
        return "Email sender is not configured. Set SMTP_FROM (or MAIL_DEFAULT_SENDER)."
    if settings["username"] and not settings["password"]:
        return "Email password is missing. Set SMTP_PASSWORD (or MAIL_PASSWORD)."
    return ""


def send_email_message(subject, body_text, recipients, from_name="CHS Gate Access"):
    config_error = smtp_configuration_error()
    if config_error:
        return False, config_error

    recipient_list = []
    if isinstance(recipients, str):
        recipient_list = [r.strip() for r in recipients.split(",")]
    elif isinstance(recipients, (list, tuple, set)):
        recipient_list = [str(r).strip() for r in recipients]
    recipient_list = [r for r in recipient_list if r]
    if not recipient_list:
        return False, "No valid recipients were provided."

    settings = smtp_settings()
    smtp_host = settings["host"]
    smtp_port = settings["port"]
    smtp_username = settings["username"]
    smtp_password = settings["password"]
    smtp_from = settings["sender"]
    smtp_use_ssl = settings["use_ssl"]
    smtp_use_tls = settings["use_tls"]

    message = EmailMessage()
    message["Subject"] = str(subject or "").strip() or "CHS Gate Access Notification"
    message["From"] = f"{from_name} <{smtp_from}>" if from_name and "<" not in smtp_from else smtp_from
    message["To"] = ", ".join(recipient_list)
    message.set_content(str(body_text or "").strip())

    try:
        if smtp_use_ssl:
            server_ctx = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=20)
        else:
            server_ctx = smtplib.SMTP(smtp_host, smtp_port, timeout=20)

        with server_ctx as server:
            server.ehlo()
            if smtp_use_tls and not smtp_use_ssl:
                server.starttls()
                server.ehlo()
            if smtp_username:
                server.login(smtp_username, smtp_password)
            server.send_message(message)
    except smtplib.SMTPAuthenticationError:
        return False, "SMTP authentication failed. Verify your Gmail address and App Password."
    except (smtplib.SMTPConnectError, smtplib.SMTPServerDisconnected, TimeoutError, socket.timeout):
        return False, "Unable to connect to the email server. Check SMTP host/port and network access."
    except Exception as exc:
        print(f"[ERROR] Failed to send email: {exc}")
        return False, "Failed to send email."
    return True, ""


def send_password_reset_email(to_email, reset_link):
    body_text = (
        "A password reset was requested for your CHS Gate Access account.\n\n"
        f"Open this link to reset your password (valid for {PASSWORD_RESET_TOKEN_TTL_MINUTES} minutes):\n"
        f"{reset_link}\n\n"
        "If you did not request this, you can safely ignore this message."
    )
    return send_email_message(
        subject="CHS Gate Access Password Reset",
        body_text=body_text,
        recipients=[to_email],
        from_name="CHS Gate Access",
    )


def find_user_by_email(email_value):
    if not email_value:
        return None
    return users.find_one({"email": {"$regex": f"^{re.escape(email_value)}$", "$options": "i"}})


def get_request_client_ip():
    xff = request.headers.get("X-Forwarded-For", "").strip()
    if xff:
        return xff.split(",")[0].strip()
    return (request.remote_addr or "").strip()


def password_reset_request_rate_limited(email_value, client_ip):
    window_after = datetime.utcnow() - timedelta(minutes=PASSWORD_RESET_REQUEST_WINDOW_MINUTES)
    email_hits = password_reset_tokens.count_documents({
        "email": email_value,
        "createdAt": {"$gte": window_after},
    })
    if email_hits >= PASSWORD_RESET_MAX_REQUESTS_PER_WINDOW:
        return True

    if client_ip:
        ip_hits = password_reset_tokens.count_documents({
            "requestIp": client_ip,
            "createdAt": {"$gte": window_after},
        })
        if ip_hits >= (PASSWORD_RESET_MAX_REQUESTS_PER_WINDOW * 3):
            return True

    return False


def get_password_reset_record(token):
    token_hash = hash_password_reset_token(token)
    now_utc = datetime.utcnow()
    return password_reset_tokens.find_one({
        "token_hash": token_hash,
        "used": False,
        "expiresAt": {"$gt": now_utc},
    })


def validate_phone_format(value):
    if not value:
        return True
    return re.match(r"^[0-9+\-\s()]{7,20}$", value) is not None


def normalize_parent_contact_value(value):
    raw = str(value or "").strip()
    if not raw:
        return ""

    compact = re.sub(r"[\s\-\(\)]", "", raw)
    if compact in {"+63", "63"}:
        return ""

    digits = ""
    if compact.startswith("+63"):
        digits = compact[3:]
    elif compact.startswith("63"):
        digits = compact[2:]
    elif compact.startswith("09"):
        digits = compact[1:]
    elif compact.startswith("9"):
        digits = compact
    else:
        raise ValueError("Parent contact must be a Philippine mobile number (+639XXXXXXXXX).")

    if not digits.isdigit():
        raise ValueError("Parent contact must contain numbers only.")

    normalized = f"+63{digits}"
    if re.match(r"^\+639\d{9}$", normalized) is None:
        raise ValueError("Parent contact must be in +639XXXXXXXXX format.")
    return normalized


def normalize_parent_contact_display(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        return normalize_parent_contact_value(raw)
    except ValueError:
        return raw


def normalize_section_value(value):
    cleaned = re.sub(r"\s+", " ", str(value or "")).strip()
    return cleaned[:64]


def normalize_student_import_header(value):
    normalized = normalize_text_value(value).lower()
    return re.sub(r"[^a-z0-9]+", "", normalized)


def map_student_import_columns(header_row):
    mapping = {}
    for idx, raw_header in enumerate(header_row or []):
        header_key = normalize_student_import_header(raw_header)
        if not header_key:
            continue
        for field_name, aliases in STUDENT_IMPORT_HEADER_ALIASES.items():
            if field_name in mapping:
                continue
            if header_key in aliases:
                mapping[field_name] = idx
                break
    return mapping


def is_student_import_summary_row(row_payload):
    if not isinstance(row_payload, dict):
        return False
    lrn_text = normalize_text_value(row_payload.get("lrn", "")).lower()
    name_text = normalize_text_value(row_payload.get("name", "")).lower()
    combined = f"{lrn_text} {name_text}".strip()
    if not combined:
        return False
    if "total male" in combined or "total female" in combined:
        return True
    if "<==" in combined and "total" in combined:
        return True
    return False


def parse_student_import_workbook(file_bytes):
    if load_workbook is None:
        raise ValueError("Excel import dependency is not installed on the server.")
    if not file_bytes:
        raise ValueError("Uploaded Excel file is empty.")

    workbook = None
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=False)
        sheet = None
        header_row_number = 0
        column_mapping = {}

        for candidate_sheet in workbook.worksheets:
            for row_index, row_values in enumerate(candidate_sheet.iter_rows(values_only=True), start=1):
                if all(normalize_text_value(value) == "" for value in row_values or []):
                    continue
                candidate_mapping = map_student_import_columns(row_values)
                if REQUIRED_STUDENT_IMPORT_FIELDS.issubset(set(candidate_mapping.keys())):
                    sheet = candidate_sheet
                    header_row_number = row_index
                    column_mapping = candidate_mapping
                    break
            if column_mapping:
                break

        if not column_mapping:
            expected = "LRN, NAME, Sex / Gender (required). GRADE LEVEL and SECTION are optional if provided as defaults."
            raise ValueError(f"Excel template is invalid. Required columns: {expected}.")

        max_rows_allowed = STUDENT_IMPORT_MAX_ROWS if STUDENT_IMPORT_MAX_ROWS and STUDENT_IMPORT_MAX_ROWS > 0 else 0
        parsed_rows = []
        for row_index, row_values in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            if all(normalize_text_value(value) == "" for value in row_values or []):
                continue

            if max_rows_allowed and len(parsed_rows) >= max_rows_allowed:
                raise ValueError(f"Import limit reached. Maximum allowed rows is {max_rows_allowed}.")

            row_payload = {"row_number": row_index}
            for field_name, column_index in column_mapping.items():
                row_payload[field_name] = row_values[column_index] if column_index < len(row_values) else ""
            parsed_rows.append(row_payload)

        if not parsed_rows:
            raise ValueError("No student rows found in the uploaded Excel file.")
        return parsed_rows
    except ValueError:
        raise
    except Exception:
        raise ValueError("Unable to read Excel file. Please upload a valid .xlsx template.")
    finally:
        if workbook is not None:
            workbook.close()


def build_student_import_template_bytes():
    if Workbook is None:
        raise ValueError("Excel import dependency is not installed on the server.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet.append(STUDENT_IMPORT_TEMPLATE_HEADERS)
    sheet.append([""] * len(STUDENT_IMPORT_TEMPLATE_HEADERS))
    for lrn, name, gender, section, grade_level in STUDENT_IMPORT_TEMPLATE_SAMPLE_ROWS:
        sheet.append([lrn, "", name, "", "", "", gender, section, grade_level])

    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    ) if Border and Side else None

    header_font = Font(bold=True) if Font else None
    center_alignment = Alignment(horizontal="center", vertical="center") if Alignment else None
    left_alignment = Alignment(horizontal="left", vertical="center") if Alignment else None

    total_rows = 2 + len(STUDENT_IMPORT_TEMPLATE_SAMPLE_ROWS)
    total_cols = len(STUDENT_IMPORT_TEMPLATE_HEADERS)
    for row_idx in range(1, total_rows + 1):
        for col_idx in range(1, total_cols + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if border:
                cell.border = border
            if row_idx == 1 and col_idx in (1, 3, 7, 8, 9):
                if header_font:
                    cell.font = header_font
                if center_alignment:
                    cell.alignment = center_alignment
            elif row_idx >= 3:
                if col_idx in (1, 3) and left_alignment:
                    cell.alignment = left_alignment
                elif col_idx in (7, 8, 9) and center_alignment:
                    cell.alignment = center_alignment

    sheet.merge_cells("A1:B2")
    sheet.merge_cells("C1:F2")
    sheet.merge_cells("G1:G2")
    sheet.merge_cells("H1:H2")
    sheet.merge_cells("I1:I2")
    for row_idx in range(3, total_rows + 1):
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        sheet.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)

    column_widths = {
        1: 13,
        2: 13,
        3: 13,
        4: 13,
        5: 13,
        6: 13,
        7: 11.42578125,
        8: 9.140625,
        9: 13.28515625,
    }
    for col_idx, width_value in column_widths.items():
        if get_column_letter:
            sheet.column_dimensions[get_column_letter(col_idx)].width = width_value

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.getvalue()


def normalize_lrn_value(value):
    if value is None:
        return ""

    if isinstance(value, Decimal):
        if value.is_nan() or value.is_infinite():
            return ""
        text = str(value.quantize(Decimal(1))) if value == value.to_integral_value() else format(value.normalize(), "f")
    elif isinstance(value, (int, np.integer)):
        text = str(int(value))
    elif isinstance(value, (float, np.floating)):
        if not np.isfinite(value):
            return ""
        if float(value).is_integer():
            text = str(int(value))
        else:
            try:
                dec = Decimal(str(value))
                text = format(dec.normalize(), "f")
            except InvalidOperation:
                text = str(value)
    else:
        text = str(value)

    text = unicodedata.normalize("NFKC", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) not in {"Cf", "Cc"})
    text = text.replace("\ufeff", "").replace("\u200b", "").replace("\xa0", " ")
    text = text.strip()

    # Unwrap formula-like values such as ="120526180006" before strict validation.
    if text.startswith("="):
        formula_text = text[1:].strip()
        quoted_match = re.fullmatch(r"""['"](.+)['"]""", formula_text)
        if quoted_match:
            text = quoted_match.group(1).strip()
        else:
            text = formula_text

    text = re.sub(r"^[`'\"\u2018\u2019\u201c\u201d]+", "", text).strip()
    text = re.sub(r"[`'\"\u2018\u2019\u201c\u201d]+$", "", text).strip()
    text = re.sub(r"[\u2010-\u2015\u2212]", "-", text)
    text = text.replace(",", "")
    normalized = re.sub(r"\s+", "", text)

    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?[eE][+-]?\d+", normalized):
        try:
            sci = Decimal(normalized)
            if sci == sci.to_integral_value():
                normalized = str(sci.quantize(Decimal(1)))
            else:
                normalized = format(sci.normalize(), "f").rstrip("0").rstrip(".")
        except InvalidOperation:
            pass

    if re.fullmatch(r"\d+\.0+", normalized):
        normalized = normalized.split(".", 1)[0]
    return normalized[:32]


def validate_lrn_value(value):
    lrn = normalize_lrn_value(value)
    if not lrn:
        return "", "LRN is required."
    if re.match(r"^[A-Za-z0-9_-]+$", lrn) is None:
        return "", "LRN may contain only letters, numbers, dashes, and underscores."
    return lrn, ""


def sanitize_profile_text(value, max_length, allow_newlines=False):
    raw = str(value or "")
    if allow_newlines:
        cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", raw)
        cleaned = re.sub(r"\r\n?", "\n", cleaned)
        cleaned = "\n".join(line.strip() for line in cleaned.split("\n"))
        cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
    else:
        cleaned = re.sub(r"[\x00-\x1f\x7f]", "", raw)
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:max_length]


def normalize_theme_value(value, default="light"):
    normalized = (value or "").strip().lower()
    if normalized in ("light", "dark"):
        return normalized
    return default


def normalize_profile_user_doc(user_doc):
    if not user_doc:
        return None

    username = (user_doc.get("username") or "").strip()
    email = (user_doc.get("email") or "").strip()
    if not email:
        email = build_internal_account_email(username) if username else ""

    full_name = (user_doc.get("fullName") or "").strip() or username
    avatar_url = (user_doc.get("avatarUrl") or "").strip()
    updated_at = (user_doc.get("updatedAt") or user_doc.get("updated_at") or user_doc.get("created_at") or "").strip()

    return {
        "username": username,
        "role": normalize_account_role(user_doc.get("role"), username),
        "fullName": full_name,
        "email": email,
        "phone": (user_doc.get("phone") or "").strip(),
        "address": (user_doc.get("address") or "").strip(),
        "bio": (user_doc.get("bio") or "").strip(),
        "avatarUrl": avatar_url,
        "twoFactorEnabled": bool(user_doc.get("twoFactorEnabled", False)),
        "updatedAt": updated_at,
        "theme": normalize_theme_value(user_doc.get("theme")),
    }


def current_user_profile():
    username = session.get("admin", "").strip()
    if not username:
        return None, None
    user_doc = users.find_one({"username": username})
    if not user_doc:
        return None, None
    return user_doc, normalize_profile_user_doc(user_doc)


def serialize_dashboard_identity_user(user_doc):
    profile = normalize_profile_user_doc(user_doc)
    if not profile:
        return None

    username = str(profile.get("username") or "").strip()
    display_name = str(profile.get("fullName") or "").strip() or username or "User"
    return {
        "_id": str(user_doc.get("_id") or ""),
        "username": username,
        "displayName": display_name,
        "role": normalize_account_role(profile.get("role"), username),
        "avatarUrl": str(profile.get("avatarUrl") or "").strip(),
    }


def build_dashboard_users_list():
    rows = []
    for user_doc in users.find({}, {"password": 0, "password_hash": 0}).sort("username", 1):
        username = str(user_doc.get("username") or "").strip()
        if username.lower() == "admin":
            continue
        serialized = serialize_dashboard_identity_user(user_doc)
        if serialized:
            rows.append(serialized)
    return rows


def build_dashboard_user_stats(users_list):
    stats = {
        "total": len(users_list or []),
        "full_admin": 0,
        "staff": 0,
    }
    for user in users_list or []:
        role_name = normalize_account_role(user.get("role"), user.get("username"))
        if role_name == ROLE_FULL_ADMIN:
            stats["full_admin"] += 1
        else:
            stats["staff"] += 1
    return stats


def generate_csrf_token():
    token = secrets.token_urlsafe(32)
    session[CSRF_SESSION_KEY] = token
    return token


def get_csrf_token():
    token = session.get(CSRF_SESSION_KEY, "").strip()
    if not token:
        token = generate_csrf_token()
    return token


def csrf_failure_response(message):
    wants_json = (
        request.path.startswith("/api/")
        or request.is_json
        or "application/json" in (request.headers.get("Accept", "") or "")
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )
    if wants_json:
        return jsonify({"status": "error", "message": message}), 400
    return Response(message, status=400, mimetype="text/plain")


def extract_csrf_token_from_request():
    token = (request.headers.get(CSRF_HEADER_NAME, "") or "").strip()
    if token:
        return token
    token = (request.form.get("csrf_token", "") or request.form.get("_csrf", "")).strip()
    if token:
        return token
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        token = str(payload.get("csrf_token") or payload.get("_csrf") or "").strip()
        if token:
            return token
    return ""


@app.context_processor
def inject_csrf_token():
    return {"csrf_token": get_csrf_token}


@app.context_processor
def inject_global_theme():
    theme = normalize_theme_value(session.get("theme"), default="")
    if theme:
        return {"current_theme": theme}

    if not session.get("admin"):
        return {"current_theme": ""}

    try:
        user_doc, profile = current_user_profile()
        if user_doc and profile:
            theme = normalize_theme_value(profile.get("theme"))
            session["theme"] = theme
            return {"current_theme": theme}
    except Exception:
        pass
    return {"current_theme": ""}


@app.context_processor
def inject_dev_runtime_flags():
    return {
        "dev_auto_reload": DEV_AUTO_RELOAD,
        "dev_reload_poll_interval_ms": DEV_RELOAD_POLL_INTERVAL_MS,
        "dev_reload_token": compute_dev_reload_token() if DEV_AUTO_RELOAD else str(int(time.time() * 1000)),
    }


def current_role():
    normalized = normalize_account_role(session.get("role"), session.get("admin"))
    if has_request_context() and session.get("admin") and session.get("role") != normalized:
        session["role"] = normalized
    return normalized


def has_permission(permission):
    perms = ROLE_PERMISSIONS.get(current_role(), set())
    return permission in perms or current_role() == ROLE_FULL_ADMIN


def require_permission(permission, api=False):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not login_required():
                if api:
                    return jsonify({"status": "error", "message": "Unauthorized"}), 401
                return redirect(url_for("login"))

            if permission and not has_permission(permission):
                create_alert(
                    "warning",
                    f"Unauthorized permission attempt by {session.get('admin', 'unknown')}: {permission}",
                    "security",
                    {"permission": permission},
                )
                log_audit_event(
                    action="auth.permission_denied",
                    outcome="blocked",
                    severity="warn",
                    target_type="permission",
                    target_id=permission,
                )
                if api:
                    return jsonify({"status": "error", "message": "Forbidden"}), 403
                return redirect(url_for("dashboard"))
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def is_https_request():
    if request.is_secure:
        return True
    proto_header = (request.headers.get("X-Forwarded-Proto") or "").split(",")[0].strip().lower()
    return proto_header == "https"


@app.before_request
def enforce_https_redirect():
    if not FORCE_HTTPS:
        return None
    if is_https_request():
        return None
    https_url = request.url.replace("http://", "https://", 1)
    return redirect(https_url, code=308)


def start_background_jobs_if_needed():
    global background_jobs_started
    if not ENABLE_BACKGROUND_JOBS:
        return
    with background_jobs_lock:
        if background_jobs_started:
            return
        worker = threading.Thread(target=background_jobs_worker_loop, name="phase2-background-worker", daemon=True)
        worker.start()
        background_jobs_started = True


@app.before_request
def csrf_and_background_guard():
    if request.endpoint == "static":
        return None

    get_csrf_token()
    cleanup_notification_alerts()
    start_background_jobs_if_needed()

    if request.method.upper() not in CSRF_ALLOWED_METHODS:
        return None

    # Exempt enhanced scanning endpoints from CSRF check
    exempt_endpoints = {"detect_faces", "recognize_face"}
    if request.endpoint in exempt_endpoints:
        return None

    expected = session.get(CSRF_SESSION_KEY, "").strip()
    provided = extract_csrf_token_from_request()
    if not expected or not provided or not secrets.compare_digest(expected, provided):
        return csrf_failure_response("Invalid or missing CSRF token.")
    return None


@app.after_request
def apply_security_headers(response):
    if not ENABLE_SECURITY_HEADERS:
        return response

    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("X-Permitted-Cross-Domain-Policies", "none")
    response.headers.setdefault("Permissions-Policy", "camera=(self), microphone=(), geolocation=(), payment=(), usb=()")
    response.headers.setdefault("Cross-Origin-Opener-Policy", "same-origin")
    response.headers.setdefault("Cross-Origin-Resource-Policy", "same-origin")

    csp_header_name = "Content-Security-Policy" if CSP_ENFORCE else "Content-Security-Policy-Report-Only"
    response.headers.setdefault(csp_header_name, CONTENT_SECURITY_POLICY)

    if is_https_request():
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")

    if login_required() and request.endpoint != "static":
        response.headers.setdefault("Cache-Control", "no-store")
        response.headers.setdefault("Pragma", "no-cache")
        response.headers.setdefault("Expires", "0")
    elif request.endpoint == "static" and (request.path or "").endswith("/students.js"):
        response.headers["Cache-Control"] = "no-store, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"

    return response


@app.route("/api/dev/reload-token", methods=["GET"])
def api_dev_reload_token():
    if not DEV_AUTO_RELOAD:
        return jsonify({"status": "error", "message": "Dev auto reload is disabled."}), 404

    response = jsonify({"status": "ok", "token": compute_dev_reload_token()})
    response.headers["Cache-Control"] = "no-store"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


def now_local():
    return datetime.now()


def now_iso():
    return now_local().isoformat(timespec="seconds")


def month_start_local(dt=None):
    value = dt or now_local()
    return value.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def month_start_months_back(months_back=0, dt=None):
    base = month_start_local(dt)
    safe_months_back = max(0, int(months_back or 0))
    absolute_month = (base.year * 12) + (base.month - 1) - safe_months_back
    year = absolute_month // 12
    month = (absolute_month % 12) + 1
    return base.replace(year=year, month=month)


def build_school_year_label(start_year):
    try:
        start = int(start_year)
    except (TypeError, ValueError):
        return ""
    return f"{start}-{start + 1}"


def normalize_school_year_value(value):
    text = str(value or "").strip().replace("–", "-").replace("/", "-")
    if not text:
        return ""
    match = re.fullmatch(r"(\d{4})\s*-\s*(\d{4})", text)
    if not match:
        return ""
    start_year = int(match.group(1))
    end_year = int(match.group(2))
    if end_year != start_year + 1:
        return ""
    return f"{start_year}-{end_year}"


def derive_default_school_year_label(today_date=None):
    date_value = today_date or now_local().date()
    start_year = date_value.year if date_value.month >= SCHOOL_YEAR_START_MONTH else date_value.year - 1
    return build_school_year_label(start_year)


def school_year_sort_key(label):
    normalized = normalize_school_year_value(label)
    if not normalized:
        return (1, str(label or ""))
    return (0, -int(normalized.split("-", 1)[0]))


def ensure_school_year_exists(label, set_current=False, created_by="system", allow_create=True):
    normalized = normalize_school_year_value(label)
    if not normalized:
        raise ValueError("Invalid school year format. Use YYYY-YYYY.")

    # Check if school year already exists
    existing = school_years.find_one({"label": normalized})
    if existing:
        # Update existing school year
        now_value = now_iso()
        school_years.update_one(
            {"label": normalized},
            {"$set": {"updated_at": now_value}}
        )
        if set_current and not existing.get("is_current"):
            school_years.update_many(
                {"label": {"$ne": normalized}, "is_current": True},
                {"$set": {"is_current": False, "updated_at": now_value}}
            )
            school_years.update_one(
                {"label": normalized},
                {"$set": {"is_current": True, "updated_at": now_value}}
            )
        return existing
    
    # Only create if allowed
    if not allow_create:
        return None

    start_year, end_year = [int(part) for part in normalized.split("-", 1)]
    now_value = now_iso()
    school_years.update_one(
        {"label": normalized},
        {
            "$set": {
                "label": normalized,
                "start_year": start_year,
                "end_year": end_year,
                "updated_at": now_value,
            },
            "$setOnInsert": {
                "created_at": now_value,
                "created_by": str(created_by or "").strip(),
                "is_current": False,
            },
        },
        upsert=True,
    )
    if set_current:
        school_years.update_many(
            {"label": {"$ne": normalized}, "is_current": True},
            {"$set": {"is_current": False, "updated_at": now_value}},
        )
        school_years.update_one(
            {"label": normalized},
            {"$set": {"is_current": True, "updated_at": now_value}},
        )
    get_student_enrollment_collection(normalized)
    return school_years.find_one({"label": normalized}) or {
        "label": normalized,
        "start_year": start_year,
        "end_year": end_year,
        "is_current": bool(set_current),
    }


def ensure_default_school_year(allow_create=True):
    current_doc = school_years.find_one({"is_current": True})
    if current_doc:
        return normalize_school_year_value(current_doc.get("label")) or derive_default_school_year_label()
    if not allow_create:
        return derive_default_school_year_label()
    label = derive_default_school_year_label()
    ensure_school_year_exists(label, set_current=True, created_by="system", allow_create=allow_create)
    return label


def list_school_year_docs():
    docs = list(school_years.find().sort("start_year", -1))
    # Don't auto-create default school year - return empty list if none exist
    normalized_docs = []
    for doc in docs:
        label = normalize_school_year_value(doc.get("label"))
        if not label:
            continue
        normalized_docs.append({
            "_id": str(doc.get("_id") or ""),
            "label": label,
            "start_year": int(doc.get("start_year") or int(label.split("-", 1)[0])),
            "end_year": int(doc.get("end_year") or int(label.split("-", 1)[1])),
            "is_current": bool(doc.get("is_current")),
            "created_at": str(doc.get("created_at") or ""),
            "updated_at": str(doc.get("updated_at") or ""),
        })
    normalized_docs.sort(key=lambda row: school_year_sort_key(row.get("label")))
    return normalized_docs


def get_current_school_year_doc():
    # Try to find existing current school year
    current_doc = school_years.find_one({"is_current": True})
    if current_doc:
        return current_doc
    
    # Don't auto-create - return None if no school year exists
    return None


def get_current_school_year_label():
    current_doc = school_years.find_one({"is_current": True})
    if current_doc:
        return normalize_school_year_value(current_doc.get("label"))
    
    # If no current school year exists, check if any school years exist at all
    any_year = school_years.find_one({}, sort=[("label", -1)])
    if any_year:
        return normalize_school_year_value(any_year.get("label"))
    
    # Fallback to current calendar year
    now = now_local()
    return f"{now.year}-{now.year + 1}"


def get_school_year_for_date(date_obj):
    """Derive school year label from a date object."""
    if isinstance(date_obj, str):
        from datetime import datetime
        try:
            date_obj = datetime.strptime(date_obj, "%Y-%m-%d")
        except ValueError:
            return ""
    
    year = date_obj.year
    # School years typically start in June (month 6)
    if date_obj.month >= 6:
        return f"{year}-{year + 1}"
    else:
        return f"{year - 1}-{year}"


def school_year_date_bounds(school_year=""):
    school_year_label = normalize_school_year_value(school_year)
    if not school_year_label:
        return None, None
    start_year, end_year = [int(part) for part in school_year_label.split("-", 1)]
    return datetime(start_year, 6, 1).date(), datetime(end_year, 5, 31).date()


def school_year_date_strings(school_year=""):
    start_date, end_date = school_year_date_bounds(school_year)
    if not start_date or not end_date:
        return "", ""
    return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")


def is_archived_school_year(label):
    normalized = normalize_school_year_value(label)
    return bool(normalized) and normalized != get_current_school_year_label()


def resolve_school_year_storage(active_collection, archive_collection, school_year=""):
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    return (
        archive_collection if is_archived_school_year(school_year_label) else active_collection,
        school_year_label,
        is_archived_school_year(school_year_label),
    )


def get_attendance_logs_storage(school_year=""):
    return resolve_school_year_storage(attendance_logs, attendance_logs_archive, school_year)


def get_sms_logs_storage(school_year=""):
    return resolve_school_year_storage(sms_logs, sms_logs_archive, school_year)


def get_alerts_storage(school_year=""):
    return resolve_school_year_storage(alerts, alerts_archive, school_year)


def get_attendance_corrections_storage(school_year=""):
    return resolve_school_year_storage(attendance_corrections, attendance_corrections_archive, school_year)


def get_early_timeout_requests_storage(school_year=""):
    return resolve_school_year_storage(early_timeout_requests, early_timeout_requests_archive, school_year)


def get_calendar_events_storage(school_year=""):
    return resolve_school_year_storage(calendar_events, calendar_events_archive, school_year)


def find_record_in_active_or_archive(active_collection, archive_collection, object_id):
    document = active_collection.find_one({"_id": object_id})
    if document:
        return document, active_collection, False
    document = archive_collection.find_one({"_id": object_id})
    if document:
        return document, archive_collection, True
    return None, None, False


def school_year_contains_date(school_year="", value=None):
    start_date, end_date = school_year_date_bounds(school_year)
    if not start_date or not end_date or value is None:
        return False
    if isinstance(value, datetime):
        target_date = value.date()
    else:
        target_date = value
    return start_date <= target_date <= end_date


def derive_school_year_label_from_value(value, fallback=""):
    normalized_fallback = normalize_school_year_value(fallback)
    if value in (None, ""):
        return normalized_fallback

    parsed_date = None
    if isinstance(value, datetime):
        parsed_date = value.date()
    else:
        raw = str(value).strip()
        if not raw:
            return normalized_fallback

        parsed_date = parse_date_or_none(raw[:10])
        if parsed_date is None:
            normalized = raw.replace("Z", "+00:00")
            try:
                parsed_date = datetime.fromisoformat(normalized).date()
            except Exception:
                for pattern in ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%Y %H:%M:%S"):
                    try:
                        parsed_date = datetime.strptime(raw[:19], pattern).date()
                        break
                    except Exception:
                        continue

    if parsed_date is None:
        return normalized_fallback

    start_year = parsed_date.year if parsed_date.month >= 6 else parsed_date.year - 1
    return build_school_year_label(start_year)


def infer_document_school_year(doc, date_keys=None, timestamp_keys=None, fallback=""):
    if not isinstance(doc, dict):
        return normalize_school_year_value(fallback)

    normalized = normalize_school_year_value(doc.get("school_year"))
    if normalized:
        return normalized

    meta = doc.get("meta") if isinstance(doc.get("meta"), dict) else {}
    normalized_meta = normalize_school_year_value(meta.get("school_year"))
    if normalized_meta:
        return normalized_meta

    detail_keys = list(date_keys or []) + list(timestamp_keys or [])
    for key in detail_keys:
        label = derive_school_year_label_from_value(doc.get(key), fallback)
        if label:
            return label

    return normalize_school_year_value(fallback)


def backfill_collection_school_year(collection, date_keys=None, timestamp_keys=None, patch_meta=False):
    updated = 0
    query = {"$or": [{"school_year": {"$exists": False}}, {"school_year": ""}, {"school_year": None}]}
    projection = {"school_year": 1, "meta": 1}
    for key in set(list(date_keys or []) + list(timestamp_keys or [])):
        projection[key] = 1

    for row in collection.find(query, projection):
        school_year_label = infer_document_school_year(row, date_keys=date_keys, timestamp_keys=timestamp_keys)
        if not school_year_label:
            continue

        ensure_school_year_exists(
            school_year_label,
            set_current=is_current_school_year(school_year_label),
            created_by="system",
            allow_create=False,
        )

        update_doc = {"school_year": school_year_label}
        if patch_meta:
            meta = row.get("meta") if isinstance(row.get("meta"), dict) else {}
            if not normalize_school_year_value(meta.get("school_year")):
                update_doc["meta.school_year"] = school_year_label

        result = collection.update_one({"_id": row["_id"]}, {"$set": update_doc})
        updated += int(result.modified_count or 0)

    return updated


def ensure_school_year_scope_defaults():
    attendance_updated = backfill_collection_school_year(
        attendance_logs,
        date_keys=["date"],
        timestamp_keys=["timestamp"],
    )
    attendance_archive_updated = backfill_collection_school_year(
        attendance_logs_archive,
        date_keys=["date"],
        timestamp_keys=["timestamp"],
    )
    sms_updated = backfill_collection_school_year(
        sms_logs,
        date_keys=["date"],
        timestamp_keys=["timestamp", "createdAt", "updatedAt"],
    )
    sms_archive_updated = backfill_collection_school_year(
        sms_logs_archive,
        date_keys=["date"],
        timestamp_keys=["timestamp", "createdAt", "updatedAt"],
    )
    alerts_updated = backfill_collection_school_year(
        alerts,
        timestamp_keys=["timestamp", "created_at"],
        patch_meta=True,
    )
    alerts_archive_updated = backfill_collection_school_year(
        alerts_archive,
        timestamp_keys=["timestamp", "created_at"],
        patch_meta=True,
    )
    corrections_updated = backfill_collection_school_year(
        attendance_corrections,
        timestamp_keys=["log_timestamp", "requested_at", "requestedAt", "reviewed_at", "reviewedAt"],
    )
    corrections_archive_updated = backfill_collection_school_year(
        attendance_corrections_archive,
        timestamp_keys=["log_timestamp", "requested_at", "requestedAt", "reviewed_at", "reviewedAt"],
    )
    total_updated = (
        attendance_updated
        + attendance_archive_updated
        + sms_updated
        + sms_archive_updated
        + alerts_updated
        + alerts_archive_updated
        + corrections_updated
        + corrections_archive_updated
    )
    if total_updated:
        print(
            "[INFO] Backfilled school_year fields. "
            f"Attendance active/archive: {attendance_updated}/{attendance_archive_updated}. "
            f"SMS active/archive: {sms_updated}/{sms_archive_updated}. "
            f"Alerts active/archive: {alerts_updated}/{alerts_archive_updated}. "
            f"Corrections active/archive: {corrections_updated}/{corrections_archive_updated}."
        )


def migrate_non_current_school_year_records(active_collection, archive_collection, collection_label):
    current_school_year = get_current_school_year_label()
    moved_count = 0
    cursor = active_collection.find({"school_year": {"$nin": ["", None, current_school_year]}})
    for row in cursor:
        try:
            archive_collection.replace_one({"_id": row["_id"]}, row, upsert=True)
            delete_result = active_collection.delete_one({"_id": row["_id"]})
            moved_count += int(delete_result.deleted_count or 0)
        except Exception as exc:
            print(f"[WARNING] Failed archiving {collection_label} record {row.get('_id')}: {exc}")
    if moved_count:
        print(
            f"[INFO] Archived {moved_count} {collection_label} record(s) "
            f"outside the current school year {current_school_year}."
        )
    return moved_count


def verify_legacy_enrollment_collection_mirrored(collection_name):
    if collection_name not in db.list_collection_names():
        return True
    collection = db[collection_name]
    if collection.count_documents({}, limit=1) == 0:
        return True

    for row in collection.find({}, {"school_year": 1, "student_id": 1, "lrn": 1}):
        school_year_label = normalize_school_year_value(row.get("school_year"))
        student_id = normalize_lrn_value(row.get("student_id") or row.get("lrn"))
        if not school_year_label or not student_id:
            return False
        if get_school_year_enrollment_collection(school_year_label).count_documents({"student_id": student_id}, limit=1) == 0:
            return False
    return True


def cleanup_obsolete_collections():
    dropped = []

    for collection_name in ("student_enrollments", "Attendance"):
        if collection_name not in db.list_collection_names():
            continue
        collection = db[collection_name]
        if collection.count_documents({}, limit=1) == 0:
            try:
                collection.drop()
                dropped.append(collection_name)
            except Exception as exc:
                print(f"[WARNING] Failed dropping empty obsolete collection {collection_name}: {exc}")

    legacy_backup_name = "student_enrollments_legacy"
    if verify_legacy_enrollment_collection_mirrored(legacy_backup_name):
        if legacy_backup_name in db.list_collection_names():
            try:
                db[legacy_backup_name].drop()
                dropped.append(legacy_backup_name)
            except Exception as exc:
                print(f"[WARNING] Failed dropping mirrored legacy collection {legacy_backup_name}: {exc}")

    if dropped:
        print(f"[INFO] Dropped obsolete MongoDB collection(s): {', '.join(sorted(dropped))}.")


def archive_historical_school_year_records():
    moved_total = 0
    moved_total += migrate_non_current_school_year_records(attendance_logs, attendance_logs_archive, "attendance_logs")
    moved_total += migrate_non_current_school_year_records(sms_logs, sms_logs_archive, "sms_logs")
    moved_total += migrate_non_current_school_year_records(alerts, alerts_archive, "alerts")
    moved_total += migrate_non_current_school_year_records(
        attendance_corrections,
        attendance_corrections_archive,
        "attendance_corrections",
    )
    moved_total += migrate_non_current_school_year_records(
        early_timeout_requests,
        early_timeout_requests_archive,
        "early_timeout_requests",
    )
    moved_total += migrate_non_current_school_year_records(
        calendar_events,
        calendar_events_archive,
        "calendar_events",
    )
    return moved_total


def get_school_year_enrollment_collection(school_year=""):
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    
    # Check if school year exists before creating/accessing collection
    if not school_years.find_one({"label": school_year_label}):
        # Return a reference to a non-existent collection that will be empty
        # This prevents errors while maintaining the expectation that deleted data stays deleted
        collection_name = student_enrollment_collection_name(school_year_label)
        return db[collection_name]
    
    # Don't auto-create school year when accessing collection
    ensure_school_year_exists(school_year_label, set_current=is_current_school_year(school_year_label), created_by="system", allow_create=False)
    return get_student_enrollment_collection(school_year_label)


def list_student_enrollment_school_year_labels(include_legacy=False):
    labels = set()
    for row in list_school_year_docs():
        label = normalize_school_year_value(row.get("label"))
        if label:
            labels.add(label)
    current_label = normalize_school_year_value(get_current_school_year_label())
    if current_label:
        labels.add(current_label)
    for collection_name in list_student_enrollment_collection_names():
        suffix = collection_name[len("student_"):] if collection_name.startswith("student_") else collection_name
        label = normalize_school_year_value(suffix)
        if not label:
            continue
        if school_years.find_one({"label": label}, {"_id": 1}):
            labels.add(label)
            continue
        try:
            if db[collection_name].count_documents({}, limit=1) > 0:
                labels.add(label)
        except Exception:
            continue
    if include_legacy and student_enrollments.count_documents({}, limit=1) > 0:
        for label in student_enrollments.distinct("school_year"):
            normalized = normalize_school_year_value(label)
            if normalized:
                labels.add(normalized)
    return sorted(labels, key=school_year_sort_key)


def update_student_base_fields_across_enrollments(student_id, update_fields):
    normalized_student_id = normalize_lrn_value(student_id)
    if not normalized_student_id or not isinstance(update_fields, dict) or not update_fields:
        return 0
    modified_count = 0
    for school_year_label in list_student_enrollment_school_year_labels():
        collection = get_school_year_enrollment_collection(school_year_label)
        result = collection.update_many({"student_id": normalized_student_id}, {"$set": dict(update_fields)})
        modified_count += int(result.modified_count or 0)
    return modified_count


def build_student_reference_cleanup_query(student_id="", student_ref_id=None):
    normalized_student_id = normalize_lrn_value(student_id)
    clauses = []
    if normalized_student_id:
        clauses.extend([
            {"student_id": normalized_student_id},
            {"lrn": normalized_student_id},
        ])

    student_ref_text = ""
    if isinstance(student_ref_id, ObjectId):
        student_ref_text = str(student_ref_id)
    else:
        student_ref_text = str(student_ref_id or "").strip()
        parsed_ref = parse_student_oid(student_ref_text)
        if parsed_ref:
            student_ref_text = str(parsed_ref)

    if student_ref_text:
        clauses.append({"student_ref_id": student_ref_text})

    if not clauses:
        return {}
    if len(clauses) == 1:
        return clauses[0]
    return {"$or": clauses}


def purge_student_runtime_face_references(student_id):
    normalized_student_id = normalize_lrn_value(student_id)
    if not normalized_student_id:
        return {"removed": False, "events_pruned": 0, "runtime_remaining": False}

    removed = False
    events_pruned = 0
    runtime_remaining = False

    with scan_lock:
        last_scanned.pop(normalized_student_id, None)
        scan_presence_locks.pop(normalized_student_id, None)

        existing_events = list(scan_state.get("events") or [])
        filtered_events = [
            row for row in existing_events
            if normalize_lrn_value(row.get("student_id") or (row.get("activity_entry") or {}).get("student_id")) != normalized_student_id
        ]
        events_pruned = max(len(existing_events) - len(filtered_events), 0)
        if events_pruned:
            scan_state["events"] = filtered_events

        known_students = list(scan_state.get("known_students") or [])
        keep_indices = []
        filtered_students = []
        for index, row in enumerate(known_students):
            current_student_id = normalize_lrn_value((row or {}).get("student_id") or (row or {}).get("lrn"))
            if current_student_id == normalized_student_id:
                removed = True
                continue
            keep_indices.append(index)
            filtered_students.append(row)

        if removed:
            scan_state["known_students"] = filtered_students
            known_encodings = scan_state.get("known_encodings", np.empty((0, 128), dtype=np.float64))
            if isinstance(known_encodings, np.ndarray) and known_encodings.ndim == 2 and known_encodings.shape[0] == len(known_students):
                scan_state["known_encodings"] = known_encodings[keep_indices] if keep_indices else np.empty((0, 128), dtype=np.float64)
            elif isinstance(known_encodings, list) and len(known_encodings) == len(known_students):
                scan_state["known_encodings"] = [known_encodings[index] for index in keep_indices]
            else:
                scan_state["known_encodings"] = np.empty((0, 128), dtype=np.float64)
                if scan_state.get("active"):
                    scan_state["model_status"] = "loading"

            if scan_state.get("active") and not filtered_students:
                scan_state["model_status"] = "no_registered_students"

        face_tracks = scan_state.get("face_tracks")
        if isinstance(face_tracks, dict) and face_tracks:
            stale_track_ids = [
                track_id
                for track_id, row in face_tracks.items()
                if normalize_lrn_value((row or {}).get("student_id")) == normalized_student_id
            ]
            for track_id in stale_track_ids:
                face_tracks.pop(track_id, None)

        runtime_remaining = (
            any(normalize_lrn_value((row or {}).get("student_id") or (row or {}).get("lrn")) == normalized_student_id for row in scan_state.get("known_students", []))
            or normalized_student_id in last_scanned
            or normalized_student_id in scan_presence_locks
            or any(normalize_lrn_value((row or {}).get("student_id")) == normalized_student_id for row in (scan_state.get("face_tracks") or {}).values())
        )

    return {
        "removed": removed,
        "events_pruned": events_pruned,
        "runtime_remaining": runtime_remaining,
    }


def refresh_loaded_face_processors():
    refresh_scan_face_index_if_active()
    enhanced_module = sys.modules.get("enhanced_face_processor")
    face_processor = getattr(enhanced_module, "face_processor", None) if enhanced_module else None
    if face_processor is None:
        return
    try:
        face_processor.load_known_faces()
    except Exception as exc:
        print(f"[WARNING] Could not refresh enhanced face processor after student cleanup: {exc}")


def delete_student_profile_and_related_data(student_doc=None, enrollment_doc=None):
    base_student = dict(student_doc or {})
    source_enrollment = dict(enrollment_doc or {})

    student_ref_id = parse_student_oid(base_student.get("_id") or source_enrollment.get("student_ref_id"))
    if not base_student and student_ref_id:
        base_student = students.find_one({"_id": student_ref_id}) or {}

    normalized_student_id = normalize_lrn_value(
        base_student.get("student_id")
        or base_student.get("lrn")
        or source_enrollment.get("student_id")
        or source_enrollment.get("lrn")
    )
    cleanup_query = build_student_reference_cleanup_query(normalized_student_id, student_ref_id)
    if not cleanup_query and student_ref_id is None:
        return {
            "ok": False,
            "message": "Student profile could not be resolved for deletion.",
            "counts": {},
            "validation": {"completed": False},
        }

    counts = {
        "student_profiles_deleted": 0,
        "enrollment_records_deleted": 0,
        "legacy_enrollment_records_deleted": 0,
        "attendance_logs_deleted": 0,
        "attendance_logs_archive_deleted": 0,
        "sms_logs_deleted": 0,
        "sms_logs_archive_deleted": 0,
        "failed_scans_deleted": 0,
        "attendance_corrections_deleted": 0,
        "attendance_corrections_archive_deleted": 0,
        "early_timeout_requests_deleted": 0,
        "early_timeout_requests_archive_deleted": 0,
        "alerts_deleted": 0,
        "alerts_archive_deleted": 0,
    }

    if student_ref_id:
        delete_result = students.delete_one({"_id": student_ref_id})
        counts["student_profiles_deleted"] += int(delete_result.deleted_count or 0)
    if normalized_student_id:
        duplicate_cleanup = students.delete_many(build_lrn_duplicate_query(normalized_student_id, exclude_oid=student_ref_id))
        counts["student_profiles_deleted"] += int(duplicate_cleanup.deleted_count or 0)

    school_year_labels = set(list_student_enrollment_school_year_labels(include_legacy=True))
    source_school_year = normalize_school_year_value(source_enrollment.get("school_year"))
    if source_school_year:
        school_year_labels.add(source_school_year)

    for school_year_label in school_year_labels:
        collection = get_school_year_enrollment_collection(school_year_label)
        if cleanup_query:
            result = collection.delete_many(cleanup_query)
            counts["enrollment_records_deleted"] += int(result.deleted_count or 0)

    if cleanup_query and student_enrollments.count_documents({}, limit=1) > 0:
        legacy_result = student_enrollments.delete_many(cleanup_query)
        counts["legacy_enrollment_records_deleted"] += int(legacy_result.deleted_count or 0)

    if normalized_student_id:
        counts["attendance_logs_deleted"] = int(attendance_logs.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["attendance_logs_archive_deleted"] = int(attendance_logs_archive.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["sms_logs_deleted"] = int(sms_logs.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["sms_logs_archive_deleted"] = int(sms_logs_archive.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["failed_scans_deleted"] = int(failed_scans.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["attendance_corrections_deleted"] = int(attendance_corrections.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["attendance_corrections_archive_deleted"] = int(attendance_corrections_archive.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["early_timeout_requests_deleted"] = int(early_timeout_requests.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["early_timeout_requests_archive_deleted"] = int(early_timeout_requests_archive.delete_many({"student_id": normalized_student_id}).deleted_count or 0)
        counts["alerts_deleted"] = int(alerts.delete_many({"meta.student_id": normalized_student_id}).deleted_count or 0)
        counts["alerts_archive_deleted"] = int(alerts_archive.delete_many({"meta.student_id": normalized_student_id}).deleted_count or 0)

    runtime_cleanup = purge_student_runtime_face_references(normalized_student_id)
    refresh_loaded_face_processors()

    remaining_profile_count = 0
    remaining_enrollment_count = 0
    if student_ref_id:
        remaining_profile_count += int(students.count_documents({"_id": student_ref_id}, limit=1))
    if normalized_student_id:
        remaining_profile_count += int(students.count_documents(build_lrn_duplicate_query(normalized_student_id)))
    if cleanup_query:
        for school_year_label in school_year_labels:
            collection = get_school_year_enrollment_collection(school_year_label)
            remaining_enrollment_count += int(collection.count_documents(cleanup_query))
        if student_enrollments.count_documents({}, limit=1) > 0:
            remaining_enrollment_count += int(student_enrollments.count_documents(cleanup_query))

    validation = {
        "completed": remaining_profile_count == 0 and remaining_enrollment_count == 0 and not runtime_cleanup.get("runtime_remaining"),
        "remaining_profiles": remaining_profile_count,
        "remaining_enrollment_records": remaining_enrollment_count,
        "runtime_remaining": bool(runtime_cleanup.get("runtime_remaining")),
        "events_pruned": int(runtime_cleanup.get("events_pruned") or 0),
    }

    return {
        "ok": validation["completed"],
        "message": "Student account and all related face data were deleted successfully." if validation["completed"] else "Student deletion completed with leftover references that need attention.",
        "counts": counts,
        "validation": validation,
        "student_id": normalized_student_id,
        "student_name": str(base_student.get("name") or source_enrollment.get("name") or "").strip(),
    }


def find_student_enrollment_record(enrollment_oid, school_year=""):
    if not enrollment_oid:
        return None, ""
    normalized_school_year = normalize_school_year_value(school_year)
    if normalized_school_year:
        collection = get_school_year_enrollment_collection(normalized_school_year)
        return collection.find_one({"_id": enrollment_oid}), normalized_school_year
    for school_year_label in list_student_enrollment_school_year_labels():
        collection = get_school_year_enrollment_collection(school_year_label)
        document = collection.find_one({"_id": enrollment_oid})
        if document:
            return document, school_year_label
    return None, ""


def count_school_year_enrollments(school_year, query=None):
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    collection = get_school_year_enrollment_collection(school_year_label)
    return int(collection.count_documents(query or {}))


def migrate_legacy_student_enrollments():
    if student_enrollments.count_documents({}, limit=1) == 0:
        return {"migrated": 0, "updated": 0, "skipped": 0, "archived": False}

    migrated = 0
    updated = 0
    skipped = 0
    for row in student_enrollments.find({}):
        school_year_label = normalize_school_year_value(row.get("school_year"))
        student_id = normalize_lrn_value(row.get("student_id") or row.get("lrn"))
        if not school_year_label or not student_id:
            skipped += 1
            continue

        ensure_school_year_exists(school_year_label, set_current=is_current_school_year(school_year_label), created_by="system", allow_create=False)
        collection = get_school_year_enrollment_collection(school_year_label)
        enrollment_doc = dict(row)
        enrollment_doc["school_year"] = school_year_label
        enrollment_doc["student_id"] = student_id
        enrollment_doc["lrn"] = normalize_lrn_value(enrollment_doc.get("lrn") or student_id) or student_id

        existing = collection.find_one({"student_id": student_id}, {"_id": 1})
        if existing:
            updated += 1
            continue

        collection.insert_one(enrollment_doc)
        migrated += 1

    unresolved = 0
    for row in student_enrollments.find({}, {"school_year": 1, "student_id": 1, "lrn": 1}):
        school_year_label = normalize_school_year_value(row.get("school_year"))
        student_id = normalize_lrn_value(row.get("student_id") or row.get("lrn"))
        if not school_year_label or not student_id:
            continue
        collection = get_school_year_enrollment_collection(school_year_label)
        if collection.count_documents({"student_id": student_id}, limit=1) == 0:
            unresolved += 1
            break

    archived = False
    if unresolved == 0:
        backup_collection_name = "student_enrollments_legacy"
        if backup_collection_name not in db.list_collection_names():
            try:
                student_enrollments.rename(backup_collection_name)
                archived = True
            except Exception as exc:
                print(f"[WARNING] Could not archive legacy student_enrollments collection: {exc}")

    if migrated or updated or skipped:
        print(
            "[INFO] Enrollment migration to per-school-year collections completed. "
            f"Migrated: {migrated}. Updated: {updated}. Skipped: {skipped}. Archived legacy: {archived}."
        )

    return {"migrated": migrated, "updated": updated, "skipped": skipped, "archived": archived}


def is_current_school_year(label):
    normalized = normalize_school_year_value(label)
    return bool(normalized) and normalized == get_current_school_year_label()


def remember_selected_school_year(label):
    normalized = normalize_school_year_value(label)
    if has_request_context() and normalized:
        session[SCHOOL_YEAR_SESSION_KEY] = normalized
    return normalized


def resolve_selected_school_year(explicit_value=""):
    requested = normalize_school_year_value(explicit_value)
    if not requested and has_request_context():
        requested = normalize_school_year_value(request.args.get("school_year", ""))
    if not requested and has_request_context():
        requested = normalize_school_year_value(request.form.get("school_year", ""))
    if not requested and has_request_context():
        payload = request.get_json(silent=True) if request.is_json else {}
        if isinstance(payload, dict):
            requested = normalize_school_year_value(payload.get("school_year", ""))
    if not requested and has_request_context():
        requested = normalize_school_year_value(session.get(SCHOOL_YEAR_SESSION_KEY, ""))
    if not requested:
        requested = get_current_school_year_label()
    ensure_school_year_exists(requested, set_current=is_current_school_year(requested), created_by="system", allow_create=False)
    return remember_selected_school_year(requested)


def compute_dev_reload_token():
    if not DEV_AUTO_RELOAD:
        return ""

    now_ts = time.time()
    with dev_reload_lock:
        last_checked = float(dev_reload_cache.get("checked_at", 0.0) or 0.0)
        if now_ts - last_checked < 0.75:
            return dev_reload_cache.get("token", "")

        latest_mtime_ns = 0
        file_count = 0
        watched_extensions = {".py", ".html", ".js", ".css"}
        ignored_dirs = {"venv", "__pycache__", ".git", "certs"}

        for root, dirs, files in os.walk(BASE_DIR):
            dirs[:] = [d for d in dirs if d not in ignored_dirs]
            for filename in files:
                if os.path.splitext(filename)[1].lower() not in watched_extensions:
                    continue
                file_path = os.path.join(root, filename)
                try:
                    stat = os.stat(file_path)
                except OSError:
                    continue
                file_count += 1
                latest_mtime_ns = max(
                    latest_mtime_ns,
                    int(getattr(stat, "st_mtime_ns", int(stat.st_mtime * 1_000_000_000))),
                )

        if latest_mtime_ns <= 0:
            latest_mtime_ns = int(APP_START_TS * 1_000_000_000)

        token = f"{latest_mtime_ns}:{file_count}"
        dev_reload_cache["checked_at"] = now_ts
        dev_reload_cache["token"] = token
        return token


def _now_utc():
    return datetime.utcnow()


def _safe_client_ip():
    try:
        return get_request_client_ip()
    except Exception:
        if has_request_context():
            return (request.remote_addr or "").strip()
        return ""


def log_audit_event(action, outcome="success", severity="info", target_type="", target_id="", details=None):
    try:
        actor_username = session.get("admin", "system")
        actor_role = session.get("role", "System")
    except Exception:
        actor_username = "system"
        actor_role = "System"

    payload = {
        "action": str(action or "").strip() or "unknown_action",
        "outcome": str(outcome or "").strip().lower() or "success",
        "severity": str(severity or "").strip().lower() or "info",
        "target_type": str(target_type or "").strip(),
        "target_id": str(target_id or "").strip(),
        "details": details if isinstance(details, dict) else {},
        "actor": {
            "username": str(actor_username or "").strip() or "system",
            "role": str(actor_role or "").strip() or "System",
        },
        "ip": _safe_client_ip(),
        "user_agent": (request.headers.get("User-Agent", "")[:300] if has_request_context() else ""),
        "created_at": now_iso(),
        "createdAt": _now_utc(),
    }
    try:
        audit_logs.insert_one(payload)
    except Exception as exc:
        print(f"[ERROR] Failed to write audit log: {exc}")


def login_attempt_key(username, ip_address):
    return {
        "username_lower": str(username or "").strip().lower(),
        "ip": str(ip_address or "").strip() or "unknown",
    }


def get_login_lockout_seconds(username, ip_address):
    key = login_attempt_key(username, ip_address)
    if not key["username_lower"]:
        return 0

    doc = login_attempts.find_one(key, {"lockout_until": 1})
    if not doc:
        return 0

    lockout_until = doc.get("lockout_until")
    if isinstance(lockout_until, datetime):
        delta = int((lockout_until - _now_utc()).total_seconds())
        if delta > 0:
            return delta

    login_attempts.update_one(
        key,
        {"$set": {"attempts": 0, "lockout_until": None}},
    )
    return 0


def register_failed_login_attempt(username, ip_address):
    key = login_attempt_key(username, ip_address)
    if not key["username_lower"]:
        return

    now_utc = _now_utc()
    window_start = now_utc - timedelta(minutes=LOGIN_ATTEMPT_WINDOW_MINUTES)
    previous = login_attempts.find_one(key, {"attempts": 1, "last_attempt_at": 1})

    attempts = 1
    if previous:
        last_attempt_at = previous.get("last_attempt_at")
        if isinstance(last_attempt_at, datetime) and last_attempt_at >= window_start:
            attempts = int(previous.get("attempts") or 0) + 1

    lockout_until = None
    if attempts >= LOGIN_MAX_ATTEMPTS:
        lockout_until = now_utc + timedelta(minutes=LOGIN_LOCKOUT_MINUTES)

    login_attempts.update_one(
        key,
        {
            "$set": {
                "attempts": attempts,
                "last_attempt_at": now_utc,
                "lockout_until": lockout_until,
                "updated_at": now_iso(),
            },
            "$setOnInsert": {"createdAt": now_utc},
        },
        upsert=True,
    )


def clear_login_attempts(username, ip_address):
    key = login_attempt_key(username, ip_address)
    if not key["username_lower"]:
        return
    login_attempts.delete_one(key)


def normalize_timestamp_value(value):
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if value is None:
        return ""
    return str(value)


def parse_date_or_none(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def contains_regex_filter(value):
    term = (value or "").strip()
    if not term:
        return None
    return {"$regex": re.escape(term), "$options": "i"}


def normalize_student_doc(student_doc):
    if not student_doc:
        return {}
    doc = dict(student_doc)
    doc["_id"] = str(doc.get("_id", ""))
    lrn_value = normalize_lrn_value(doc.get("lrn") or doc.get("student_id"))
    doc["lrn"] = lrn_value
    doc["student_id"] = lrn_value
    doc["status"] = doc.get("status", "Active") or "Active"
    doc["gender"] = normalize_gender_value(doc.get("gender") or doc.get("sex"))
    doc["grade_level"] = normalize_grade_level(doc.get("grade_level") or doc.get("grade"))
    doc["grade"] = doc["grade_level"]
    doc["sex"] = doc["gender"]
    doc["created_at"] = normalize_timestamp_value(doc.get("created_at"))
    doc["updated_at"] = normalize_timestamp_value(doc.get("updated_at"))
    created_at_text = doc.get("created_at", "")
    doc["created_date"] = created_at_text[:10] if created_at_text else ""
    faces = doc.get("face_data", doc.get("faces"))
    if not isinstance(faces, list):
        faces = []
    doc["faces"] = faces[:5]
    doc["face_data"] = doc["faces"]
    doc["profile_photo"] = doc.get("profile_photo") or (doc["faces"][0] if doc["faces"] else "")
    has_face_payload = bool(doc.get("faces")) or bool(doc.get("face_encodings")) or bool(doc.get("face_embeddings"))
    doc["face_registered"] = bool(doc.get("face_registered")) or has_face_payload
    doc["face_updated_at"] = normalize_timestamp_value(doc.get("face_updated_at"))
    return doc


def normalize_enrollment_doc(enrollment_doc):
    if not enrollment_doc:
        return {}
    doc = dict(enrollment_doc)
    doc["_id"] = str(doc.get("_id", ""))
    doc["student_ref_id"] = str(doc.get("student_ref_id") or "")
    lrn_value = normalize_lrn_value(doc.get("lrn") or doc.get("student_id"))
    doc["lrn"] = lrn_value
    doc["student_id"] = lrn_value
    doc["name"] = normalize_student_name_value(doc.get("name"))
    doc["status"] = doc.get("status", "Active") or "Active"
    doc["gender"] = normalize_gender_value(doc.get("gender") or doc.get("sex"))
    doc["sex"] = doc["gender"]
    doc["grade_level"] = normalize_grade_level(doc.get("grade_level") or doc.get("grade"))
    doc["grade"] = doc["grade_level"]
    doc["section"] = normalize_section_value(doc.get("section"))
    doc["school_year"] = normalize_school_year_value(doc.get("school_year"))
    doc["parent_contact"] = normalize_parent_contact_display(doc.get("parent_contact"))
    doc["created_at"] = normalize_timestamp_value(doc.get("created_at"))
    doc["updated_at"] = normalize_timestamp_value(doc.get("updated_at"))
    doc["profile_photo"] = doc.get("profile_photo") or ""
    doc["face_registered"] = bool(doc.get("face_registered"))
    return doc


def normalize_gender_value(value):
    v = normalize_text_value(value).lower()
    if v in {"male", "m"}:
        return "Male"
    if v in {"female", "f"}:
        return "Female"
    return ""


def normalize_grade_level(value):
    if value is None:
        return ""

    if isinstance(value, Decimal):
        if value.is_nan() or value.is_infinite():
            return ""
        if value == value.to_integral_value():
            value = str(value.quantize(Decimal(1)))
        else:
            value = format(value.normalize(), "f")
    elif isinstance(value, (int, np.integer)):
        value = str(int(value))
    elif isinstance(value, (float, np.floating)):
        if not np.isfinite(value):
            return ""
        if float(value).is_integer():
            value = str(int(value))
        else:
            try:
                value = format(Decimal(str(value)).normalize(), "f")
            except InvalidOperation:
                value = str(value)

    v = normalize_text_value(value)
    if not v:
        return ""
    lower = v.lower()
    if lower in {"nan", "none", "null", "n/a", "na", "-"}:
        return ""

    grade_number_match = re.fullmatch(r"(7|8|9|10|11|12)(?:\.0+)?", lower)
    if grade_number_match:
        return f"Grade {grade_number_match.group(1)}"

    explicit_grade_match = re.search(r"\bgrade\s*(7|8|9|10|11|12)\b", lower, re.IGNORECASE)
    if explicit_grade_match:
        return f"Grade {explicit_grade_match.group(1)}"

    shorthand_grade_match = re.search(r"\bg\s*(7|8|9|10|11|12)\b", lower, re.IGNORECASE)
    if shorthand_grade_match:
        return f"Grade {shorthand_grade_match.group(1)}"

    combined_grade_match = re.search(r"(?<!\d)(7|8|9|10|11|12)(?!\d)\s*[-/]\s*[A-Za-z]", v)
    if combined_grade_match:
        return f"Grade {combined_grade_match.group(1)}"

    for grade_label in GRADE_LEVEL_OPTIONS:
        if grade_label.lower() == lower:
            return grade_label

    return v


def normalize_text_value(value):
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = "".join(
        (" " if unicodedata.category(ch) == "Zs" else ch)
        for ch in text
        if unicodedata.category(ch) not in {"Cf", "Cc"}
    )
    text = text.replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def build_pagination_payload(page, per_page, total, filters_payload, endpoint):
    total_pages = max((total + per_page - 1) // per_page, 1)
    page = min(max(page, 1), total_pages)

    def build_page_link(page_number):
        params = {**filters_payload, "page": page_number}
        params = {k: v for k, v in params.items() if v not in ("", None)}
        return f"{url_for(endpoint)}?{urlencode(params)}" if params else url_for(endpoint)

    page_start = max(1, page - 2)
    page_end = min(total_pages, page + 2)
    page_numbers = list(range(page_start, page_end + 1))

    return {
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_url": build_page_link(page - 1) if page > 1 else "",
        "next_url": build_page_link(page + 1) if page < total_pages else "",
        "page_numbers": page_numbers,
        "page_links": {p: build_page_link(p) for p in page_numbers},
    }


def clamp_int_value(value, default, minimum=None, maximum=None):
    try:
        normalized = int(value)
    except (TypeError, ValueError):
        normalized = int(default)
    if minimum is not None:
        normalized = max(int(minimum), normalized)
    if maximum is not None:
        normalized = min(int(maximum), normalized)
    return normalized


def time_to_hhmm(value, fallback):
    parsed = parse_time_str(value) or parse_time_str(fallback)
    if parsed is None:
        parsed = dtime(hour=0, minute=0)
    return parsed.strftime("%H:%M")


def hhmm_to_minutes(value, fallback="00:00"):
    parsed = parse_time_str(value) or parse_time_str(fallback)
    if parsed is None:
        return 0
    return (int(parsed.hour) * 60) + int(parsed.minute)


def minutes_to_hhmm(total_minutes):
    normalized = max(0, int(total_minutes or 0)) % (24 * 60)
    hour, minute = divmod(normalized, 60)
    return f"{hour:02d}:{minute:02d}"


def derive_late_threshold_minutes(raw_schedule, morning_start, afternoon_start, default_minutes=15):
    explicit_minutes = raw_schedule.get("late_threshold_minutes")
    if explicit_minutes not in (None, ""):
        return clamp_int_value(explicit_minutes, default_minutes, minimum=1, maximum=180)

    morning_late = raw_schedule.get("morning_late")
    afternoon_late = raw_schedule.get("afternoon_late")
    derived_candidates = []
    if morning_late:
        derived_candidates.append(hhmm_to_minutes(morning_late) - hhmm_to_minutes(morning_start))
    if afternoon_late:
        derived_candidates.append(hhmm_to_minutes(afternoon_late) - hhmm_to_minutes(afternoon_start))
    for candidate in derived_candidates:
        if candidate and candidate > 0:
            return clamp_int_value(candidate, default_minutes, minimum=1, maximum=180)
    return clamp_int_value(default_minutes, default_minutes, minimum=1, maximum=180)


def normalize_attendance_schedule(raw_schedule=None):
    schedule = dict(raw_schedule or {})
    morning_start = time_to_hhmm(schedule.get("morning_start"), "05:00")
    noon_start = time_to_hhmm(schedule.get("noon_start"), "12:00")
    afternoon_start = time_to_hhmm(schedule.get("afternoon_start"), "13:00")
    afternoon_end = time_to_hhmm(schedule.get("afternoon_end"), "17:00")
    late_threshold_minutes = derive_late_threshold_minutes(
        schedule,
        morning_start,
        afternoon_start,
        default_minutes=15,
    )
    morning_late = time_to_hhmm(
        schedule.get("morning_late"),
        minutes_to_hhmm(hhmm_to_minutes(morning_start) + late_threshold_minutes),
    )
    afternoon_late = time_to_hhmm(
        schedule.get("afternoon_late"),
        minutes_to_hhmm(hhmm_to_minutes(afternoon_start) + late_threshold_minutes),
    )
    scan_cooldown_minutes = clamp_int_value(
        schedule.get("scan_cooldown_minutes"),
        30,
        minimum=1,
        maximum=240,
    )
    scan_in_cooldown_seconds = clamp_int_value(
        schedule.get("scan_in_cooldown_seconds"),
        SCAN_LIVE_IN_COOLDOWN_SECONDS,
        minimum=5,
        maximum=300,
    )
    return {
        "morning_start": morning_start,
        "morning_late": morning_late,
        "noon_start": noon_start,
        "afternoon_start": afternoon_start,
        "afternoon_late": afternoon_late,
        "afternoon_end": afternoon_end,
        "late_threshold_minutes": late_threshold_minutes,
        "scan_cooldown_minutes": scan_cooldown_minutes,
        "scan_in_cooldown_seconds": scan_in_cooldown_seconds,
    }


def get_default_schedule():
    settings = system_settings.find_one({"key": "default_schedule"}) or {}
    return normalize_attendance_schedule(settings.get("schedule") or {})


def get_active_schedule(date_obj):
    date_str = date_obj.strftime("%Y-%m-%d")
    school_year = get_school_year_for_date(date_obj)
    calendar_collection, _, _ = get_calendar_events_storage(school_year)
    base_schedule = get_default_schedule()
    event = calendar_collection.find_one({"date": date_str})
    if event:
        custom_schedule = event.get("custom_schedule") or {}
        if event.get("custom_schedule"):
            return {
                "type": event.get("type", "event"),
                "special_condition": event.get("special_condition"),
                "schedule": normalize_attendance_schedule({**base_schedule, **custom_schedule}),
            }
        return {
            "type": event.get("type", "event"),
            "special_condition": event.get("special_condition"),
            "schedule": base_schedule,
        }
    return {
        "type": "regular",
        "special_condition": "",
        "schedule": base_schedule,
    }


def parse_time_str(time_str):
    if not time_str: return None
    try:
        parts = str(time_str).split(":")
        return dtime(hour=int(parts[0]), minute=int(parts[1]))
    except:
        return None


def session_info_for_time(dt):
    active_sched = get_active_schedule(dt)
    sched = normalize_attendance_schedule(active_sched.get("schedule", {}))
    
    m_start = parse_time_str(sched.get("morning_start")) or MORNING_START
    m_late_thr = parse_time_str(sched.get("morning_late")) or MORNING_LATE_THRESHOLD
    n_start = parse_time_str(sched.get("noon_start")) or NOON_START
    a_start = parse_time_str(sched.get("afternoon_start")) or AFTERNOON_START
    a_late_thr = parse_time_str(sched.get("afternoon_late")) or AFTERNOON_LATE_THRESHOLD
    a_end = parse_time_str(sched.get("afternoon_end")) or AFTERNOON_END_START
    
    t = dt.time()
    
    is_holiday = active_sched.get("type") == "holiday"
    special_cond = active_sched.get("special_condition")
    
    def make_res(session, action, status, label):
        # Update label to "Welcome" / "Thank You" as requested
        if action == "IN":
            voice_msg = "Welcome"
            if label == "Verified In":
                label = "Welcome"
        else:
            voice_msg = "Thank you"
            if label == "Verified Out":
                label = "Thank You"

        display_msg = f"{label}"
        if special_cond:
            display_msg += f" ({special_cond})"
            
        if is_holiday:
            status = "Holiday"
            display_msg = f"{label} (Holiday)"
        elif status == "Late":
            display_msg += " - You are Late"
            
        return {
            "session": session,
            "gate_action": action,
            "verification_label": label,
            "status": status,
            "display_message": display_msg,
            "voice_message": voice_msg,
        }

    if m_start <= t < n_start:
        is_late = t >= m_late_thr
        return make_res("Morning In", "IN", "Late" if is_late else "Present", "Verified In")

    if n_start <= t < a_start:
        return make_res("Noon Out", "OUT", "Present", "Verified Out")

    if a_start <= t < a_end:
        is_late = t >= a_late_thr
        return make_res("Afternoon In", "IN", "Late" if is_late else "Present", "Verified In")

    if t >= a_end:
        return make_res("Afternoon Out", "OUT", "Present", "Verified Out")

    return make_res("Morning In", "IN", "Present", "Verified In")


def normalize_scan_session_mode(value, default="auto"):
    mode = str(value or "").strip().lower().replace("-", "_")
    if not mode:
        return default
    if mode not in VALID_SCAN_SESSION_MODES:
        raise ValueError("Invalid session mode. Allowed values: auto, manual_in, manual_out.")
    return mode


def scan_mode_forced_gate_action(mode=None):
    normalized_mode = normalize_scan_session_mode(mode or get_scan_session_mode(), default="auto")
    if normalized_mode == "manual_in":
        return "IN"
    if normalized_mode == "manual_out":
        return "OUT"
    return ""


def scan_session_mode_label(mode):
    normalized = normalize_scan_session_mode(mode)
    labels = {
        "auto": "Smart IN/OUT Tracking",
        "manual_in": "Manual IN",
        "manual_out": "Manual OUT",
    }
    return labels.get(normalized, "Smart IN/OUT Tracking")


def get_scan_session_mode():
    with scan_lock:
        return normalize_scan_session_mode(scan_state.get("session_mode", "auto"), default="auto")


def set_scan_session_mode(mode):
    normalized = normalize_scan_session_mode(mode)
    with scan_lock:
        scan_state["session_mode"] = normalized
    return normalized


def normalize_scan_liveness_profile(value, default=None):
    # Strict mode is enforced internally to keep anti-spoofing hardened.
    # Any external/requested profile value is normalized to strict.
    return "strict"


def scan_liveness_profile_label(profile):
    normalized = normalize_scan_liveness_profile(profile)
    labels = {
        "strict": "Strict Liveness",
        "balanced": "Balanced Liveness",
        "lenient": "Lenient Liveness",
    }
    return labels.get(normalized, "Strict Liveness")


def resolve_live_liveness_profile_thresholds(profile=None):
    normalized = normalize_scan_liveness_profile(profile if profile is not None else get_scan_liveness_profile())
    thresholds = {
        "profile": normalized,
        "min_track_frames": max(int(LIVE_LIVENESS_BASE_MIN_TRACK_FRAMES or 1), 1),
        "min_motion_score": max(float(LIVE_LIVENESS_BASE_MIN_MOTION_SCORE or 0.0), 0.0),
        "min_pose_score": max(float(LIVE_LIVENESS_BASE_MIN_POSE_SCORE or 0.0), 0.0),
        "pose_grace_seconds": max(float(LIVE_LIVENESS_BASE_POSE_GRACE_SECONDS or 0.0), 0.0),
        "min_texture": max(float(LIVE_LIVENESS_BASE_MIN_TEXTURE or 0.0), 0.0),
        "min_contrast": max(float(LIVE_LIVENESS_BASE_MIN_CONTRAST or 0.0), 0.0),
        "max_highlight_ratio": max(float(LIVE_LIVENESS_BASE_MAX_HIGHLIGHT_RATIO or 0.0), 0.0),
        "min_parallax_score": max(float(LIVE_LIVENESS_BASE_MIN_PARALLAX_SCORE or 0.0), 0.0),
        "planar_streak_frames": max(int(LIVE_LIVENESS_BASE_PLANAR_STREAK_FRAMES or 2), 2),
        "max_planar_pose_component": max(float(LIVE_LIVENESS_BASE_MAX_PLANAR_POSE_COMPONENT or 0.0), 0.0),
        "max_planar_patch_delta": max(float(LIVE_LIVENESS_BASE_MAX_PLANAR_PATCH_DELTA or 0.0), 0.0),
        "max_planar_grad_delta": max(float(LIVE_LIVENESS_BASE_MAX_PLANAR_GRAD_DELTA or 0.0), 0.0),
        "blink_required": bool(LIVE_LIVENESS_BASE_BLINK_REQUIRED),
        "blink_interval_seconds": max(float(LIVE_LIVENESS_BASE_BLINK_INTERVAL_SECONDS or 0.0), 0.05),
        "blink_recent_seconds": max(float(LIVE_LIVENESS_BASE_BLINK_RECENT_SECONDS or 0.0), 0.5),
        "pose_sample_interval_seconds": max(float(LIVE_LIVENESS_BASE_POSE_SAMPLE_INTERVAL_SECONDS or 0.0), 0.05),
        "pose_window_seconds": max(float(LIVE_LIVENESS_BASE_POSE_WINDOW_SECONDS or 0.0), 1.0),
        "min_landmark_pose_span": max(float(LIVE_LIVENESS_BASE_MIN_LANDMARK_POSE_SPAN or 0.0), 0.0),
        "min_landmark_pose_samples": max(int(LIVE_LIVENESS_BASE_MIN_LANDMARK_POSE_SAMPLES or 1), 1),
        "min_patch_diversity": max(float(LIVE_LIVENESS_BASE_MIN_PATCH_DIVERSITY or 0.0), 0.0),
    }

    if normalized == "strict":
        # Tuned strict defaults: still hardened against spoofing, but less likely
        # to reject real users under normal motion/lighting.
        thresholds["min_track_frames"] = max(int(thresholds["min_track_frames"]), 3)
        thresholds["min_motion_score"] = max(float(thresholds["min_motion_score"]), 0.135)
        thresholds["min_pose_score"] = max(float(thresholds["min_pose_score"]), 0.015)
        thresholds["pose_grace_seconds"] = min(float(thresholds["pose_grace_seconds"]), 2.15)
        thresholds["min_texture"] = max(float(thresholds["min_texture"]), 7.1)
        thresholds["min_contrast"] = max(float(thresholds["min_contrast"]), 14.0)
        thresholds["max_highlight_ratio"] = min(float(thresholds["max_highlight_ratio"]), 0.26)
        thresholds["min_parallax_score"] = max(float(thresholds["min_parallax_score"]), 0.031)
        thresholds["planar_streak_frames"] = max(int(thresholds["planar_streak_frames"]), 6)
        thresholds["max_planar_patch_delta"] = min(float(thresholds["max_planar_patch_delta"]), 0.022)
        thresholds["max_planar_grad_delta"] = min(float(thresholds["max_planar_grad_delta"]), 0.03)
        thresholds["max_planar_pose_component"] = min(float(thresholds["max_planar_pose_component"]), 0.034)
        thresholds["blink_required"] = True
        thresholds["blink_recent_seconds"] = min(float(thresholds["blink_recent_seconds"]), 6.2)
        thresholds["blink_interval_seconds"] = min(float(thresholds["blink_interval_seconds"]), 0.12)
        thresholds["pose_sample_interval_seconds"] = min(float(thresholds["pose_sample_interval_seconds"]), 0.12)
        thresholds["pose_window_seconds"] = min(float(thresholds["pose_window_seconds"]), 4.5)
        thresholds["min_landmark_pose_span"] = max(float(thresholds["min_landmark_pose_span"]), 0.05)
        thresholds["min_landmark_pose_samples"] = max(int(thresholds["min_landmark_pose_samples"]), 2)
        thresholds["min_patch_diversity"] = max(float(thresholds["min_patch_diversity"]), 0.09)

    thresholds["max_highlight_ratio"] = min(max(float(thresholds["max_highlight_ratio"]), 0.0), 1.0)
    return thresholds


def get_scan_liveness_profile():
    return "strict"


def set_scan_liveness_profile(profile):
    _ = profile
    normalized = "strict"
    with scan_lock:
        previous = normalize_scan_liveness_profile(scan_state.get("liveness_profile", LIVE_LIVENESS_PROFILE))
        scan_state["liveness_profile"] = normalized
        if previous != normalized:
            tracks = scan_state.get("face_tracks") or {}
            if isinstance(tracks, dict):
                for row in tracks.values():
                    if not isinstance(row, dict):
                        continue
                    row["liveness_frames"] = 0
                    row["liveness_motion_score"] = 0.0
                    row["liveness_pose_score"] = 0.0
                    row["liveness_motion_component"] = 0.0
                    row["liveness_area_component"] = 0.0
                    row["liveness_pose_component"] = 0.0
                    row["liveness_parallax_score"] = 0.0
                    row["liveness_planar_streak"] = 0
                    row["liveness_patch_delta"] = 0.0
                    row["liveness_patch_grad_delta"] = 0.0
                    row["liveness_prev_patch"] = None
                    row["liveness_prev_grad_patch"] = None
                    row["liveness_patch_ts"] = 0.0
                    row["liveness_blink_count"] = 0
                    row["liveness_last_blink_ts"] = 0.0
                    row["liveness_last_ear"] = 0.0
                    row["liveness_last_ear_ts"] = 0.0
                    row["liveness_eye_closed_frames"] = 0
                    row["liveness_last_landmark_ts"] = 0.0
                    row["liveness_landmark_yaw"] = 0.0
                    row["liveness_landmark_pitch"] = 0.0
                    row["liveness_landmark_pose_span"] = 0.0
                    row["liveness_landmark_pose_samples"] = 0
                    row["liveness_pose_history"] = []
                    row["liveness_patch_diversity"] = 0.0
                    row["liveness_display_score"] = 0.0
                    row["liveness_display_streak"] = 0
                    row["liveness_ready"] = bool(not LIVE_SCAN_LIVENESS_ENABLED)
                    row["liveness_message"] = "" if not LIVE_SCAN_LIVENESS_ENABLED else "Liveness Check Required"
    return normalized


def session_info_for_mode(dt, mode):
    normalized_mode = normalize_scan_session_mode(mode, default="auto")
    if normalized_mode == "manual_in":
        return {
            "session": "Manual In",
            "gate_action": "IN",
            "verification_label": "Verified In",
            "status": "Present",
            "display_message": "Verified In (Manual)",
            "voice_message": "Welcome",
        }
    if normalized_mode == "manual_out":
        return {
            "session": "Manual Out",
            "gate_action": "OUT",
            "verification_label": "Verified Out",
            "status": "Present",
            "display_message": "Verified Out (Manual)",
            "voice_message": "Thank you",
        }
    return session_info_for_time(dt)


def resolve_gate_session(dt=None):
    current_dt = dt or now_local()
    mode = get_scan_session_mode()
    if mode == "auto":
        status_hint = session_info_for_time(current_dt)
        session_info = {
            "session": "Real-Time Smart Tracking",
            "gate_action": "AUTO",
            "verification_label": "Smart IN/OUT",
            "status": status_hint.get("status", "Present"),
            "display_message": "Real-time face recognition is active.",
            "voice_message": "Scanning ready",
        }
    else:
        session_info = session_info_for_mode(current_dt, mode)
    return {
        **session_info,
        "mode": mode,
        "mode_label": scan_session_mode_label(mode),
    }


def normalize_gate_action_value(value, session_name=""):
    normalized = str(value or "").strip().upper()
    if normalized in {"IN", "OUT"}:
        return normalized
    session_text = str(session_name or "").strip().lower()
    if "out" in session_text:
        return "OUT"
    if "in" in session_text:
        return "IN"
    return ""


def parse_gate_record_datetime(record):
    if not isinstance(record, dict):
        return None

    timestamp_value = str(record.get("timestamp") or "").strip()
    if timestamp_value:
        normalized = timestamp_value.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(normalized)
            if parsed.tzinfo is not None:
                return parsed.astimezone().replace(tzinfo=None)
            return parsed
        except ValueError:
            pass

    date_value = str(record.get("date") or "").strip()
    time_value = str(record.get("time") or "").strip()
    if date_value and time_value:
        try:
            return datetime.fromisoformat(f"{date_value}T{time_value}")
        except ValueError:
            return None
    return None


def parse_datetime_like(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, dtime):
        return datetime.combine(now_local().date(), value)

    raw = str(value or "").strip()
    if not raw:
        return None

    normalized = raw.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
        if parsed.tzinfo is not None:
            return parsed.astimezone().replace(tzinfo=None)
        return parsed
    except ValueError:
        pass

    for pattern in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d",
        "%H:%M:%S",
        "%H:%M",
        "%I:%M %p",
        "%I:%M:%S %p",
    ):
        try:
            return datetime.strptime(raw, pattern)
        except ValueError:
            continue
    return None


def format_time_for_display(value, fallback=""):
    for candidate in (value, fallback):
        parsed = parse_datetime_like(candidate)
        if parsed is not None:
            return parsed.strftime("%I:%M %p").lstrip("0")
    return str(fallback or value or "").strip()


def format_timestamp_for_display(value, fallback=""):
    for candidate in (value, fallback):
        parsed = parse_datetime_like(candidate)
        if parsed is None:
            continue
        if parsed.year == 1900 and parsed.month == 1 and parsed.day == 1:
            return parsed.strftime("%I:%M:%S %p").lstrip("0")
        return parsed.strftime("%b %d, %Y %I:%M:%S %p").replace(" 0", " ")
    return str(fallback or value or "").strip()


def serialize_gate_log_display_row(row, profile_photo=""):
    if not row:
        return {}
    student_id = str(row.get("student_id") or "").strip()
    time_raw = str(row.get("time") or "")
    timestamp_raw = str(row.get("timestamp") or "")
    return {
        "_id": str(row.get("_id") or ""),
        "student_id": student_id,
        "name": row.get("student_name", ""),
        "action": row.get("gate_action", "IN"),
        "status": row.get("status", "Present"),
        "session": row.get("session", ""),
        "verification_label": row.get("verification_label", ""),
        "date": row.get("date", ""),
        "time": format_time_for_display(time_raw, timestamp_raw),
        "timestamp": format_timestamp_for_display(timestamp_raw, time_raw),
        "raw_time": time_raw,
        "raw_timestamp": timestamp_raw,
        "profile_photo": profile_photo,
    }


def serialize_sms_log_display_row(row):
    if not row:
        return {}
    time_raw = str(row.get("time") or "")
    timestamp_raw = str(row.get("timestamp") or "")
    status_value = str(row.get("status", "") or "")
    return {
        "_id": str(row.get("_id") or ""),
        "student_id": row.get("student_id", ""),
        "name": row.get("name", ""),
        "parent_contact": row.get("parent_contact", ""),
        "message": row.get("message", ""),
        "status": status_value.upper() if status_value else "",
        "sid": row.get("sid", ""),
        "error": row.get("error", ""),
        "date": row.get("date", ""),
        "time": format_time_for_display(time_raw, timestamp_raw),
        "timestamp": format_timestamp_for_display(timestamp_raw, time_raw),
        "raw_time": time_raw,
        "raw_timestamp": timestamp_raw,
    }


def format_wait_time_short(total_seconds):
    remaining_seconds = max(int(total_seconds or 0), 0)
    minutes, seconds = divmod(remaining_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    parts = []
    if hours:
        parts.append(f"{hours}h")
    if minutes:
        parts.append(f"{minutes}m")
    if not parts or (not hours and seconds and minutes < 2):
        parts.append(f"{seconds}s")
    return " ".join(parts[:2])


def build_gate_feedback(action, status):
    normalized_action = "OUT" if str(action or "").strip().upper() == "OUT" else "IN"
    normalized_status = str(status or "").strip().title()
    if normalized_action == "OUT":
        verification_label = "Thank You"
        display_message = "Thank You"
        voice_message = "Thank you"
    else:
        verification_label = "Welcome"
        display_message = "Welcome"
        voice_message = "Welcome"

    if normalized_status == "Late" and normalized_action == "IN":
        display_message += " - You are Late"
    elif normalized_status == "Holiday":
        display_message += " (Holiday)"

    return {
        "verification_label": verification_label,
        "display_message": display_message,
        "voice_message": voice_message,
    }


def build_gate_session_name(action, mode, dt):
    normalized_action = "OUT" if str(action or "").strip().upper() == "OUT" else "IN"
    normalized_mode = normalize_scan_session_mode(mode, default="auto")
    prefix = "Manual" if normalized_mode.startswith("manual_") else "Live"
    return f"{prefix} {normalized_action} {format_time_for_display(dt)}"


def build_scan_activity_entry(student_id, student_name, gate_action, status, verification_label, timestamp, time_str):
    return {
        "student_id": student_id,
        "name": student_name,
        "gate_action": gate_action,
        "status": status,
        "verification_label": verification_label,
        "timestamp": timestamp,
        "time": format_time_for_display(time_str, timestamp),
        "label": f"{student_name} ({gate_action})",
    }


def should_suppress_persistent_face_scan(student_id, now_ts=None, mode=None):
    normalized_student_id = str(student_id or "").strip()
    if not normalized_student_id:
        return False

    current_ts = float(now_ts if now_ts is not None else time.time())
    requested_action = scan_mode_forced_gate_action(mode)
    with scan_lock:
        lock_entry = scan_presence_locks.get(normalized_student_id)
        if not lock_entry:
            return False
        last_seen_ts = float(lock_entry.get("last_seen_ts") or 0.0)
        if current_ts - last_seen_ts > float(SCAN_FACE_PRESENCE_RESET_SECONDS):
            scan_presence_locks.pop(normalized_student_id, None)
            return False
        locked_action = normalize_gate_action_value((lock_entry or {}).get("gate_action"))
        if requested_action and locked_action and requested_action != locked_action:
            scan_presence_locks.pop(normalized_student_id, None)
            return False
        lock_entry["last_seen_ts"] = current_ts
        return True


def should_suppress_recent_live_scan(student_id, now_ts=None, mode=None):
    normalized_student_id = str(student_id or "").strip()
    if not normalized_student_id:
        return False

    current_ts = float(now_ts if now_ts is not None else time.time())
    requested_action = scan_mode_forced_gate_action(mode)
    with scan_lock:
        cooldown_entry = last_scanned.get(normalized_student_id)
        if isinstance(cooldown_entry, dict):
            cooldown_until = float(cooldown_entry.get("until_ts") or 0.0)
            cooldown_action = normalize_gate_action_value(cooldown_entry.get("gate_action"))
        else:
            cooldown_until = float(cooldown_entry or 0.0)
            cooldown_action = ""
        if current_ts < cooldown_until:
            if not (requested_action and cooldown_action and requested_action != cooldown_action):
                return True

    return should_suppress_persistent_face_scan(normalized_student_id, current_ts, mode=mode)


def mark_persistent_face_scan(student_id, now_ts=None, gate_action=None):
    normalized_student_id = str(student_id or "").strip()
    if not normalized_student_id:
        return

    current_ts = float(now_ts if now_ts is not None else time.time())
    with scan_lock:
        scan_presence_locks[normalized_student_id] = {
            "last_seen_ts": current_ts,
            "gate_action": normalize_gate_action_value(gate_action),
        }


def get_latest_gate_record(attendance_collection, student_id):
    if not student_id:
        return None
    return attendance_collection.find_one(
        {"student_id": student_id, "gate_action": {"$in": ["IN", "OUT"]}},
        sort=[("timestamp", DESCENDING), ("_id", DESCENDING)],
    )


def get_latest_gate_action_record(attendance_collection, student_id, gate_action):
    normalized_student_id = str(student_id or "").strip()
    normalized_gate_action = normalize_gate_action_value(gate_action)
    if not normalized_student_id or not normalized_gate_action:
        return None
    return attendance_collection.find_one(
        {"student_id": normalized_student_id, "gate_action": normalized_gate_action},
        sort=[("timestamp", DESCENDING), ("_id", DESCENDING)],
    )


def build_gate_duplicate_result(result, message, voice_message, reason="duplicate", cooldown_until_ts=None):
    if not isinstance(result, dict):
        return None

    resolved_until_ts = float(cooldown_until_ts if cooldown_until_ts is not None else time.time())
    return {
        **result,
        "verification_label": "Already Recorded",
        "display_message": message,
        "voice_message": voice_message,
        "duplicate": True,
        "duplicate_reason": str(reason or "duplicate"),
        "feed_update": False,
        "activity_entry": None,
        "next_allowed_scan_ts": resolved_until_ts,
    }


def should_debounce_gate_scan_event(attendance_collection, student_id, gate_action, now):
    normalized_student_id = str(student_id or "").strip()
    normalized_gate_action = normalize_gate_action_value(gate_action)
    if not normalized_student_id or not normalized_gate_action:
        return False, 0.0

    debounce_seconds = float(SCAN_DUPLICATE_EVENT_WINDOW_SECONDS or 0.0)
    if debounce_seconds <= 0.0:
        return False, 0.0

    latest_same_action = get_latest_gate_action_record(
        attendance_collection,
        normalized_student_id,
        normalized_gate_action,
    )
    latest_same_action_dt = parse_gate_record_datetime(latest_same_action)
    if latest_same_action_dt is None:
        return False, 0.0

    current_ts = float(now.timestamp())
    latest_same_action_ts = float(latest_same_action_dt.timestamp())
    if (current_ts - latest_same_action_ts) >= debounce_seconds:
        return False, 0.0

    return True, latest_same_action_ts + debounce_seconds


def build_gate_scan_result(attendance_collection, school_year_label, student, now, source="gate_scan", mode=None):
    normalized_mode = normalize_scan_session_mode(mode or get_scan_session_mode(), default="auto")
    runtime_schedule = normalize_attendance_schedule((get_active_schedule(now) or {}).get("schedule", {}))
    scan_cooldown_minutes = clamp_int_value(
        runtime_schedule.get("scan_cooldown_minutes"),
        SCAN_OUT_MINUTES,
        minimum=1,
        maximum=240,
    )
    scan_cooldown_seconds = scan_cooldown_minutes * 60
    scan_in_cooldown_seconds = clamp_int_value(
        runtime_schedule.get("scan_in_cooldown_seconds"),
        SCAN_LIVE_IN_COOLDOWN_SECONDS,
        minimum=5,
        maximum=300,
    )
    scan_out_cooldown_minutes = clamp_int_value(
        SCAN_LIVE_OUT_COOLDOWN_MINUTES,
        5,
        minimum=5,
        maximum=60,
    )
    scan_out_cooldown_seconds = scan_out_cooldown_minutes * 60
    timestamp = now.isoformat(timespec="seconds")
    now_ts = float(now.timestamp())
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    student_id = (student.get("student_id") or "").strip()
    student_name = (student.get("name") or "").strip()
    parent_contact = (student.get("parent_contact") or "").strip()
    if not student_id or not student_name:
        return None

    last_record = get_latest_gate_record(attendance_collection, student_id)
    last_action = normalize_gate_action_value(
        (last_record or {}).get("gate_action"),
        (last_record or {}).get("session"),
    )
    last_record_dt = parse_gate_record_datetime(last_record)
    last_status = str((last_record or {}).get("status") or "Present").strip().title() or "Present"
    elapsed_since_last = None
    last_record_ts = None
    if last_record_dt is not None:
        elapsed_since_last = max((now - last_record_dt).total_seconds(), 0.0)
        last_record_ts = float(last_record_dt.timestamp())
    in_transition_allowed_scan_ts = now_ts + scan_cooldown_seconds
    out_reentry_allowed_scan_ts = now_ts + scan_out_cooldown_seconds
    if last_record_ts is not None:
        in_transition_allowed_scan_ts = max(last_record_ts + scan_cooldown_seconds, now_ts)
        out_reentry_allowed_scan_ts = max(last_record_ts + scan_out_cooldown_seconds, now_ts)

    status_hint = session_info_for_time(now)
    in_status = str(status_hint.get("status") or "Present").strip().title() or "Present"
    out_status = "Holiday" if in_status == "Holiday" else "Present"

    def duplicate_result(message, voice_message, action=None, status=None, reason="duplicate", cooldown_until_ts=None):
        resolved_action = normalize_gate_action_value(action or last_action or "IN")
        resolved_status = str(status or last_status or "Present").strip().title() or "Present"
        return {
            "student_id": student_id,
            "student_name": student_name,
            "parent_contact": parent_contact,
            "school_year": school_year_label,
            "status": resolved_status,
            "session": str((last_record or {}).get("session") or build_gate_session_name(resolved_action, normalized_mode, now)),
            "source": source,
            "timestamp": timestamp,
            "date": date_str,
            "time": time_str,
            "gate_action": resolved_action,
            "verification_label": "Already Recorded",
            "display_message": message,
            "voice_message": voice_message,
            "duplicate": True,
            "duplicate_reason": reason,
            "feed_update": False,
            "activity_entry": None,
            "tracking_mode": normalized_mode,
            "scan_cooldown_minutes": scan_cooldown_minutes,
            "scan_in_cooldown_seconds": scan_in_cooldown_seconds,
            "scan_out_cooldown_minutes": scan_out_cooldown_minutes,
            "next_allowed_scan_ts": float(cooldown_until_ts if cooldown_until_ts is not None else now_ts),
        }

    manual_override_enabled = normalized_mode in {"manual_in", "manual_out"}
    allow_immediate_reentry = last_action == "OUT" and normalized_mode == "auto"

    if (
        not manual_override_enabled
        and not allow_immediate_reentry
        and last_action == "OUT"
        and last_record_dt is not None
    ):
        elapsed_since_out = max((now - last_record_dt).total_seconds(), 0.0)
        if elapsed_since_out < scan_out_cooldown_seconds:
            remaining_seconds = max(scan_out_cooldown_seconds - elapsed_since_out, 0.0)
            return duplicate_result(
                f"Already OUT - wait {format_wait_time_short(remaining_seconds)} before scanning again.",
                "Please wait before scanning again",
                action="OUT",
                status=last_status,
                reason="out_cooldown",
                cooldown_until_ts=out_reentry_allowed_scan_ts,
            )

    if (
        not manual_override_enabled
        and elapsed_since_last is not None
        and elapsed_since_last < SCAN_REPEAT_SUPPRESSION_SECONDS
        and not allow_immediate_reentry
    ):
        return duplicate_result(
            "Already recorded moments ago.",
            "Already recorded",
            action=last_action or "IN",
            status=last_status,
            reason="repeat_suppressed",
            cooldown_until_ts=max((last_record_ts or now_ts) + float(SCAN_REPEAT_SUPPRESSION_SECONDS), now_ts),
        )

    if normalized_mode == "manual_in":
        next_action = "IN"
    elif normalized_mode == "manual_out":
        next_action = "OUT"
    else:
        if last_action == "IN" and last_record_dt is not None:
            elapsed_since_in = max((now - last_record_dt).total_seconds(), 0.0)
            if elapsed_since_in < scan_cooldown_seconds:
                return duplicate_result(
                    f"Already IN - wait {format_wait_time_short(scan_cooldown_seconds - elapsed_since_in)} before OUT.",
                    "Please wait before exit",
                    action="IN",
                    status=last_status,
                    reason="out_wait_period",
                    cooldown_until_ts=in_transition_allowed_scan_ts,
                )
            next_action = "OUT"
        else:
            next_action = "IN"

    if not manual_override_enabled and next_action == "IN" and last_action == "IN":
        return duplicate_result(
            "Already marked IN.",
            "Already checked in",
            action="IN",
            status=last_status,
            reason="already_in",
        )

    if not manual_override_enabled and next_action == "OUT" and last_action == "OUT":
        return duplicate_result(
            "Already marked OUT.",
            "Already checked out",
            action="OUT",
            status=last_status,
            reason="already_out",
        )

    status = in_status if next_action == "IN" else out_status
    feedback = build_gate_feedback(next_action, status)
    session_name = build_gate_session_name(next_action, normalized_mode, now)
    activity_entry = build_scan_activity_entry(
        student_id,
        student_name,
        next_action,
        status,
        feedback["verification_label"],
        timestamp,
        time_str,
    )
    return {
        "student_id": student_id,
        "student_name": student_name,
        "parent_contact": parent_contact,
        "school_year": school_year_label,
        "status": status,
        "session": session_name,
        "source": source,
        "timestamp": timestamp,
        "date": date_str,
        "time": time_str,
        "gate_action": next_action,
        "verification_label": feedback["verification_label"],
        "display_message": feedback["display_message"],
        "voice_message": feedback["voice_message"],
        "duplicate": False,
        "duplicate_reason": "",
        "feed_update": True,
        "activity_entry": activity_entry,
        "tracking_mode": normalized_mode,
        "scan_cooldown_minutes": scan_cooldown_minutes,
        "scan_in_cooldown_seconds": scan_in_cooldown_seconds,
        "scan_out_cooldown_minutes": scan_out_cooldown_minutes,
        "next_allowed_scan_ts": now_ts + (scan_out_cooldown_seconds if next_action == "OUT" else scan_in_cooldown_seconds),
    }


def push_scan_event(event_type, payload):
    with scan_event_condition:
        scan_state["event_counter"] += 1
        event = {
            "id": scan_state["event_counter"],
            "type": event_type,
            "timestamp": now_iso(),
            **payload,
        }
        scan_state["events"].append(event)
        if len(scan_state["events"]) > 300:
            scan_state["events"] = scan_state["events"][-300:]
        scan_event_condition.notify_all()
        return dict(event)


def create_alert(level, message, category="system", meta=None):
    global alert_revision
    normalized_level = str(level or "info").strip().lower() or "info"
    normalized_category = str(category or "system").strip().lower() or "system"
    timestamp = now_iso()
    normalized_meta = dict(meta) if isinstance(meta, dict) else {}
    # Always use the current active school year from database for alerts
    school_year_label = get_current_school_year_label()
    alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)
    if school_year_label:
        normalized_meta.setdefault("school_year", school_year_label)
    payload = {
        "title": f"{normalized_level.title()} Alert",
        "level": normalized_level,
        "message": message,
        "category": normalized_category,
        "type": normalized_category,
        "source": "System",
        "meta": normalized_meta,
        "details": normalized_meta,
        "status": "unread",
        "is_read": False,
        "school_year": school_year_label,
        "timestamp": timestamp,
        "created_at": timestamp,
    }
    try:
        alerts_collection.insert_one(payload)
        with alert_lock:
            alert_revision += 1
    except Exception as exc:
        print(f"[ERROR] Failed to insert alert: {exc}")


def cleanup_notification_alerts(force=False):
    global alert_revision
    if not ALERT_NOTIFICATION_CLEANUP_ENABLED:
        return 0

    current_dt = now_local()
    current_day_key = current_dt.strftime("%Y-%m-%d")
    with alert_cleanup_lock:
        if not force and alert_cleanup_state.get("last_checked_date") == current_day_key:
            return 0
        alert_cleanup_state["last_checked_date"] = current_day_key

    cutoff_dt = month_start_months_back(ALERT_NOTIFICATION_RETENTION_MONTHS - 1, current_dt)
    cutoff_iso = cutoff_dt.isoformat(timespec="seconds")
    cleanup_query = {
        "$or": [
            {"created_at": {"$lt": cutoff_iso}},
            {"timestamp": {"$lt": cutoff_iso}},
        ]
    }

    removed_total = 0
    try:
        removed_total += int(alerts.delete_many(cleanup_query).deleted_count or 0)
    except Exception as exc:
        print(f"[WARNING] Failed cleaning active notification alerts: {exc}")

    try:
        removed_total += int(alerts_archive.delete_many(cleanup_query).deleted_count or 0)
    except Exception as exc:
        print(f"[WARNING] Failed cleaning archived notification alerts: {exc}")

    if removed_total > 0:
        with alert_lock:
            alert_revision += 1
        print(
            f"[INFO] Removed {removed_total} notification alert(s) older than {cutoff_iso} "
            f"(retention={ALERT_NOTIFICATION_RETENTION_MONTHS} month(s))."
        )
    return removed_total


def unread_notifications_query(school_year=""):
    query = {
        "$or": [
            {"status": "unread"},
            {"is_read": False},
        ]
    }
    school_year_label = normalize_school_year_value(school_year)
    if school_year_label:
        query["school_year"] = school_year_label
    return query


def normalize_notification_details(raw_details):
    if isinstance(raw_details, dict):
        items = []
        for key, value in raw_details.items():
            if value in (None, ""):
                continue
            label = str(key).replace("_", " ").strip().title() or "Detail"
            if isinstance(value, (dict, list)):
                rendered = json.dumps(value, ensure_ascii=True)
            else:
                rendered = str(value)
            items.append({"label": label, "value": rendered})
        return items

    if isinstance(raw_details, list):
        items = []
        for item in raw_details:
            if isinstance(item, dict):
                label = str(item.get("label") or item.get("key") or "Detail").strip()
                value = item.get("value")
                if value in (None, ""):
                    continue
                items.append({"label": label, "value": str(value)})
            elif item not in (None, ""):
                items.append({"label": "Detail", "value": str(item)})
        return items

    if raw_details not in (None, ""):
        return [{"label": "Detail", "value": str(raw_details)}]
    return []


def normalize_notification_doc(doc):
    if not doc:
        return None

    raw_status = str(doc.get("status") or "").strip().lower()
    is_read = raw_status == "read" or bool(doc.get("is_read"))
    category = str(doc.get("category") or doc.get("type") or "system").strip().lower() or "system"
    level = str(doc.get("level") or "").strip().lower() or "info"
    timestamp = (
        doc.get("timestamp")
        or doc.get("created_at")
        or doc.get("updatedAt")
        or doc.get("updated_at")
        or ""
    )
    title = str(doc.get("title") or "").strip()
    if not title:
        if level in {"critical", "high", "warning", "info"}:
            title = f"{level.title()} Alert"
        else:
            title = f"{category.replace('_', ' ').title()} Notification"

    return {
        "_id": str(doc.get("_id")),
        "title": title,
        "message": str(doc.get("message") or "").strip(),
        "timestamp": timestamp,
        "school_year": normalize_school_year_value(doc.get("school_year") or (doc.get("meta") or {}).get("school_year")),
        "status": "read" if is_read else "unread",
        "type": str(doc.get("type") or category).strip().lower() or "system",
        "level": level,
        "category": category,
        "source": str(doc.get("source") or category.replace("_", " ").title() or "System"),
        "details": normalize_notification_details(doc.get("details") or doc.get("meta") or {}),
        "meta": doc.get("meta") or {},
    }


def notification_summary(limit=12, school_year=""):
    try:
        safe_limit = int(limit or 12)
    except (TypeError, ValueError):
        safe_limit = 12
    safe_limit = max(1, min(safe_limit, 50))
    school_year_label = normalize_school_year_value(school_year)
    alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)
    base_query = {"school_year": school_year_label} if school_year_label else {}
    docs = list(alerts_collection.find(base_query).sort([("timestamp", -1), ("created_at", -1)]).limit(safe_limit))
    notifications = [normalize_notification_doc(doc) for doc in docs]
    unread_count = alerts_collection.count_documents(unread_notifications_query(school_year_label))
    return {"notifications": notifications, "unread": unread_count}


def data_change_snapshot():
    with data_change_lock:
        return {
            "revision": int(data_change_revision),
            "students": int(data_change_domains.get("students", 0)),
            "sections": int(data_change_domains.get("sections", 0)),
            "gate_logs": int(data_change_domains.get("gate_logs", 0)),
            "sms_logs": int(data_change_domains.get("sms_logs", 0)),
            "users": int(data_change_domains.get("users", 0)),
        }


def signal_data_change(*domains):
    global data_change_revision
    valid_domains = [domain for domain in domains if domain in data_change_domains]
    if not valid_domains:
        return data_change_snapshot()

    with data_change_lock:
        data_change_revision += 1
        for domain in set(valid_domains):
            data_change_domains[domain] = int(data_change_domains.get(domain, 0)) + 1
        return {
            "revision": int(data_change_revision),
            "students": int(data_change_domains.get("students", 0)),
            "sections": int(data_change_domains.get("sections", 0)),
            "gate_logs": int(data_change_domains.get("gate_logs", 0)),
            "sms_logs": int(data_change_domains.get("sms_logs", 0)),
            "users": int(data_change_domains.get("users", 0)),
        }


def sidebar_context(current_page, school_year=""):
    unread = 0
    theme = "light"
    selected_school_year = resolve_selected_school_year(school_year)
    role_name = current_role()
    can_manage_alerts = has_permission("alerts_manage")
    is_full_admin = role_name == ROLE_FULL_ADMIN
    try:
        if can_manage_alerts:
            alerts_collection, selected_school_year, _ = get_alerts_storage(selected_school_year)
            unread = alerts_collection.count_documents(unread_notifications_query(selected_school_year))
    except Exception:
        unread = 0

    pending_eto_count = 0
    if is_full_admin:
        try:
            eto_collection, _, _ = get_early_timeout_requests_storage(selected_school_year)
            pending_eto_count = int(eto_collection.count_documents({"status": "pending"}))
        except Exception:
            pass

    display_user = session.get("admin", "Admin")
    try:
        user_doc, profile = current_user_profile()
        if user_doc and profile:
            display_user = profile.get("fullName") or display_user
            theme = normalize_theme_value(profile.get("theme"))
    except Exception:
        pass

    return {
        "current_page": current_page,
        "current_user": display_user,
        "current_role": role_name,
        "is_full_admin": is_full_admin,
        "is_staff_dashboard": role_name == ROLE_STAFF,
        "can_manage_students": has_permission("students_write"),
        "can_register_faces": has_permission("face_register"),
        "can_view_logs": has_permission("logs"),
        "can_view_analytics": has_permission("analytics"),
        "can_manage_users": has_permission("users_manage"),
        "can_manage_alerts": can_manage_alerts,
        "alerts_unread": unread,
        "pending_eto_count": pending_eto_count,
        "current_theme": theme,
        "selected_school_year": selected_school_year,
        "current_school_year": get_current_school_year_label(),
        "school_year_query": urlencode({"school_year": selected_school_year}) if selected_school_year else "",
    }


def safe_count_documents(collection, query=None):
    try:
        return int(collection.count_documents(query or {}))
    except Exception:
        return 0


def distinct_school_year_labels_for_collection(collection):
    labels = set()
    try:
        for raw_value in collection.distinct("school_year"):
            normalized = normalize_school_year_value(raw_value)
            if normalized:
                labels.add(normalized)
    except Exception:
        return set()
    return labels


def active_storage_mismatch_query(current_school_year):
    return {
        "$or": [
            {"school_year": {"$exists": False}},
            {"school_year": ""},
            {"school_year": None},
            {"school_year": {"$ne": current_school_year}},
        ]
    }


def build_archive_summary_payload(selected_school_year=""):
    selected_school_year = resolve_selected_school_year(selected_school_year)
    current_school_year = get_current_school_year_label()
    school_year_labels = set(list_student_enrollment_school_year_labels())
    school_year_labels.add(current_school_year)

    for collection in (
        sections,
        attendance_logs,
        attendance_logs_archive,
        sms_logs,
        sms_logs_archive,
        alerts,
        alerts_archive,
        attendance_corrections,
        attendance_corrections_archive,
        early_timeout_requests,
        early_timeout_requests_archive,
        calendar_events,
        calendar_events_archive,
    ):
        school_year_labels.update(distinct_school_year_labels_for_collection(collection))

    ordered_labels = sorted(school_year_labels, key=school_year_sort_key)
    school_year_rows = []

    current_student_collection = db[student_enrollment_collection_name(current_school_year)]
    current_student_count = safe_count_documents(current_student_collection)
    archived_student_count = 0

    for label in ordered_labels:
        is_current = label == current_school_year
        student_collection_name_value = student_enrollment_collection_name(label)
        student_collection = db[student_collection_name_value]
        student_count = safe_count_documents(student_collection)
        section_count = safe_count_documents(sections, {"school_year": label})

        gate_active = safe_count_documents(attendance_logs, {"school_year": label})
        gate_archive = safe_count_documents(attendance_logs_archive, {"school_year": label})
        sms_active = safe_count_documents(sms_logs, {"school_year": label})
        sms_archive = safe_count_documents(sms_logs_archive, {"school_year": label})
        alerts_active = safe_count_documents(alerts, {"school_year": label})
        alerts_archive_count = safe_count_documents(alerts_archive, {"school_year": label})
        corrections_active = safe_count_documents(attendance_corrections, {"school_year": label})
        corrections_archive = safe_count_documents(attendance_corrections_archive, {"school_year": label})
        eto_active = safe_count_documents(early_timeout_requests, {"school_year": label})
        eto_archive = safe_count_documents(early_timeout_requests_archive, {"school_year": label})
        calendar_active = safe_count_documents(calendar_events, {"school_year": label})
        calendar_archive = safe_count_documents(calendar_events_archive, {"school_year": label})

        misplaced_total = (
            gate_archive
            + sms_archive
            + alerts_archive_count
            + corrections_archive
            + eto_archive
            + calendar_archive
            if is_current
            else gate_active + sms_active + alerts_active + corrections_active + eto_active + calendar_active
        )

        if not is_current:
            archived_student_count += student_count

        school_year_rows.append({
            "label": label,
            "is_current": is_current,
            "is_selected": label == selected_school_year,
            "workspace_label": "Current Active Year" if is_current else "Archived School Year",
            "student_collection": student_collection_name_value,
            "students": student_count,
            "sections": section_count,
            "gate_active": gate_active,
            "gate_archive": gate_archive,
            "gate_total": gate_active + gate_archive,
            "sms_active": sms_active,
            "sms_archive": sms_archive,
            "sms_total": sms_active + sms_archive,
            "alerts_active": alerts_active,
            "alerts_archive": alerts_archive_count,
            "alerts_total": alerts_active + alerts_archive_count,
            "corrections_active": corrections_active,
            "corrections_archive": corrections_archive,
            "corrections_total": corrections_active + corrections_archive,
            "eto_active": eto_active,
            "eto_archive": eto_archive,
            "eto_total": eto_active + eto_archive,
            "calendar_active": calendar_active,
            "calendar_archive": calendar_archive,
            "calendar_total": calendar_active + calendar_archive,
            "misplaced_total": misplaced_total,
            "student_page_url": url_for("students_page", school_year=label),
            "gate_logs_url": url_for("gate_logs_page", school_year=label),
            "sms_logs_url": url_for("sms_logs_page", school_year=label),
            "analytics_url": url_for("analytics", school_year=label),
            "calendar_url": url_for("admin_calendar", school_year=label),
            "eto_url": url_for("early_timeout_page", school_year=label),
        })

    active_storage_total = (
        safe_count_documents(attendance_logs)
        + safe_count_documents(sms_logs)
        + safe_count_documents(alerts)
        + safe_count_documents(attendance_corrections)
        + safe_count_documents(early_timeout_requests)
        + safe_count_documents(calendar_events)
        + current_student_count
        + safe_count_documents(sections, {"school_year": current_school_year})
    )
    archived_storage_total = (
        safe_count_documents(attendance_logs_archive)
        + safe_count_documents(sms_logs_archive)
        + safe_count_documents(alerts_archive)
        + safe_count_documents(attendance_corrections_archive)
        + safe_count_documents(early_timeout_requests_archive)
        + safe_count_documents(calendar_events_archive)
        + archived_student_count
        + safe_count_documents(sections, {"school_year": {"$ne": current_school_year}})
    )

    storage_health_rows = []
    for label, active_collection, archive_collection in (
        ("Attendance Logs", attendance_logs, attendance_logs_archive),
        ("SMS Logs", sms_logs, sms_logs_archive),
        ("Alerts", alerts, alerts_archive),
        ("Attendance Corrections", attendance_corrections, attendance_corrections_archive),
        ("Early Time-Out Requests", early_timeout_requests, early_timeout_requests_archive),
        ("Calendar Events", calendar_events, calendar_events_archive),
    ):
        active_mismatches = safe_count_documents(active_collection, active_storage_mismatch_query(current_school_year))
        archive_current = safe_count_documents(archive_collection, {"school_year": current_school_year})
        storage_health_rows.append({
            "label": label,
            "active_mismatches": active_mismatches,
            "archive_current": archive_current,
            "issue_count": active_mismatches + archive_current,
        })

    collection_inventory = [
        {
            "name": "students",
            "scope": "Shared Master",
            "category": "Profiles",
            "count": safe_count_documents(students),
            "details": "Master student identity records shared across school years.",
        },
        {
            "name": "sections",
            "scope": "Mixed by school year",
            "category": "Structure",
            "count": safe_count_documents(sections),
            "details": "Section definitions partitioned logically by the school_year field.",
        },
        {
            "name": "school_years",
            "scope": "System Metadata",
            "category": "Registry",
            "count": safe_count_documents(school_years),
            "details": "Declared school years and the current active year flag.",
        },
        {
            "name": "attendance_logs",
            "scope": "Active",
            "category": "Gate Logs",
            "count": safe_count_documents(attendance_logs),
            "details": f"Operational gate logs for the current school year {current_school_year}.",
        },
        {
            "name": "attendance_logs_archive",
            "scope": "Archive",
            "category": "Gate Logs",
            "count": safe_count_documents(attendance_logs_archive),
            "details": "Historical gate logs migrated out of the active workspace.",
        },
        {
            "name": "sms_logs",
            "scope": "Active",
            "category": "SMS Logs",
            "count": safe_count_documents(sms_logs),
            "details": f"Operational SMS logs for the current school year {current_school_year}.",
        },
        {
            "name": "sms_logs_archive",
            "scope": "Archive",
            "category": "SMS Logs",
            "count": safe_count_documents(sms_logs_archive),
            "details": "Historical SMS logs retained for reference.",
        },
        {
            "name": "alerts",
            "scope": "Active",
            "category": "Alerts",
            "count": safe_count_documents(alerts),
            "details": f"Active alert stream for the current school year {current_school_year}.",
        },
        {
            "name": "alerts_archive",
            "scope": "Archive",
            "category": "Alerts",
            "count": safe_count_documents(alerts_archive),
            "details": "Archived alert history for prior school years.",
        },
        {
            "name": "attendance_corrections",
            "scope": "Active",
            "category": "Corrections",
            "count": safe_count_documents(attendance_corrections),
            "details": f"Correction requests for the current school year {current_school_year}.",
        },
        {
            "name": "attendance_corrections_archive",
            "scope": "Archive",
            "category": "Corrections",
            "count": safe_count_documents(attendance_corrections_archive),
            "details": "Historical attendance correction requests.",
        },
        {
            "name": "early_timeout_requests",
            "scope": "Active",
            "category": "Early Time-Out",
            "count": safe_count_documents(early_timeout_requests),
            "details": f"Operational early time-out requests for the current school year {current_school_year}.",
        },
        {
            "name": "early_timeout_requests_archive",
            "scope": "Archive",
            "category": "Early Time-Out",
            "count": safe_count_documents(early_timeout_requests_archive),
            "details": "Historical early time-out requests retained outside the active workspace.",
        },
        {
            "name": "calendar_events",
            "scope": "Active",
            "category": "Calendar",
            "count": safe_count_documents(calendar_events),
            "details": f"Calendar events for the current school year {current_school_year}.",
        },
        {
            "name": "calendar_events_archive",
            "scope": "Archive",
            "category": "Calendar",
            "count": safe_count_documents(calendar_events_archive),
            "details": "Archived calendar events for prior school years.",
        },
    ]

    for row in school_year_rows:
        collection_inventory.append({
            "name": row["student_collection"],
            "scope": "Current Year" if row["is_current"] else "Archived Year",
            "category": "Enrollments",
            "count": row["students"],
            "details": f"Student enrollment records for {row['label']}.",
        })

    collection_inventory.sort(key=lambda row: (row["category"], row["name"]))

    return {
        "generated_at_label": now_local().strftime("%B %d, %Y %I:%M:%S %p"),
        "archive_totals": {
            "school_year_count": len(school_year_rows),
            "archived_year_count": sum(1 for row in school_year_rows if not row["is_current"]),
            "active_storage_total": active_storage_total,
            "archived_storage_total": archived_storage_total,
            "storage_issue_total": sum(row["issue_count"] for row in storage_health_rows),
            "current_student_count": current_student_count,
            "archived_student_count": archived_student_count,
        },
        "school_year_rows": school_year_rows,
        "storage_health_rows": storage_health_rows,
        "collection_inventory": collection_inventory,
    }


def calculate_match_confidence(distance):
    try:
        return round(max(0.0, min(100.0, (1.0 - float(distance)) * 100.0)), 2)
    except Exception:
        return 0.0

def _extract_encodings_from_student(student_doc, allow_legacy_fallback=False):
    stored = student_doc.get("face_encodings", student_doc.get("face_embeddings", []))
    encs = sanitize_face_encodings(
        stored,
        max_count=FACE_INDEX_MAX_ENCODINGS_PER_STUDENT,
    )
    if encs:
        return encs

    if not allow_legacy_fallback:
        return encs

    # Backward compatibility for legacy docs with only image data.
    faces = student_doc.get("face_data", student_doc.get("faces", []))
    if isinstance(faces, list):
        for raw in faces[:5]:
            if not raw or "," not in raw:
                continue
            try:
                img_b64 = raw.split(",", 1)[1]
                img = Image.open(BytesIO(base64.b64decode(img_b64))).convert("RGB")
                np_img = np.array(img)
                face_enc = face_recognition.face_encodings(np_img)
                if face_enc:
                    encs.extend(
                        sanitize_face_encodings(
                            [face_enc[0]],
                            max_count=FACE_INDEX_MAX_ENCODINGS_PER_STUDENT,
                        )
                    )
            except Exception:
                continue
    return encs


def _active_students_match_clause():
    return {
        "$or": [
            {"status": "Active"},
            {"status": {"$regex": "^active$", "$options": "i"}},
            {"status": {"$exists": False}},
            {"status": ""},
        ]
    }


def count_legacy_face_only_students():
    query = {
        "$and": [
            _active_students_match_clause(),
            {"$or": [{"face_data.0": {"$exists": True}}, {"faces.0": {"$exists": True}}]},
            {"face_encodings.0": {"$exists": False}},
            {"face_embeddings.0": {"$exists": False}},
        ]
    }
    try:
        return int(students.count_documents(query))
    except Exception:
        return 0


def load_face_index_from_db(allow_legacy_fallback=False, prefer_current_roster=True):
    face_source_match = {
        "$or": [
            {"face_encodings.0": {"$exists": True}},
            {"face_embeddings.0": {"$exists": True}},
        ]
    }
    if allow_legacy_fallback:
        face_source_match["$or"].extend([
            {"face_data.0": {"$exists": True}},
            {"faces.0": {"$exists": True}},
        ])

    projection = {
        "student_id": 1,
        "name": 1,
        "parent_contact": 1,
        "status": 1,
        "grade_level": 1,
        "section": 1,
        "face_encodings": 1,
        "face_embeddings": 1,
    }
    if allow_legacy_fallback:
        projection["face_data"] = 1
        projection["faces"] = 1

    def collect_face_index_rows(extra_clauses=None):
        raw_students = []
        query = {"$and": [_active_students_match_clause(), face_source_match]}
        if extra_clauses:
            query["$and"].extend(extra_clauses)

        for row in students.find(query, projection):
            sid = normalize_lrn_value(row.get("student_id"))
            name = (row.get("name") or "").strip()
            if not sid or not name:
                continue

            encs = _extract_encodings_from_student(row, allow_legacy_fallback=allow_legacy_fallback)
            if not encs:
                continue

            raw_students.append({
                "student_id": sid,
                "name": name,
                "parent_contact": row.get("parent_contact", ""),
                "grade_level": row.get("grade_level", ""),
                "section": row.get("section", ""),
                "student_ref_id": str(row.get("_id") or ""),
                "encodings": encs,
            })

        face_index = build_student_face_index(
            raw_students,
            max_encodings_per_student=FACE_INDEX_MAX_ENCODINGS_PER_STUDENT,
        )
        return face_index.get("centroids"), face_index.get("students")

    # Prefer current school year enrollments, but do not let a missing/misaligned
    # enrollment mirror collapse the live face index to zero usable encodings.
    current_sy = get_current_school_year_label()
    roster_clause = None
    roster_filter_applied = False
    roster_has_rows = False
    if current_sy and prefer_current_roster:
        try:
            enrollment_coll = get_student_enrollment_collection(current_sy)
            enrolled_ids = set()
            enrolled_ref_ids = set()
            enrolled_cursor = enrollment_coll.find(
                {"status": {"$nin": ["Dropped", "Withdrawn", "Transferred Out"]}},
                {"student_id": 1, "lrn": 1, "student_ref_id": 1},
            )
            for doc in enrolled_cursor:
                roster_has_rows = True
                normalized_student_id = normalize_lrn_value(doc.get("student_id") or doc.get("lrn"))
                if normalized_student_id:
                    enrolled_ids.add(normalized_student_id)
                student_ref_oid = parse_student_oid(doc.get("student_ref_id"))
                if student_ref_oid:
                    enrolled_ref_ids.add(student_ref_oid)

            roster_conditions = []
            if enrolled_ids:
                roster_conditions.extend([
                    {"student_id": {"$in": list(enrolled_ids)}},
                    {"lrn": {"$in": list(enrolled_ids)}},
                ])
            if enrolled_ref_ids:
                roster_conditions.append({"_id": {"$in": list(enrolled_ref_ids)}})
            if roster_conditions:
                roster_clause = {"$or": roster_conditions}
                roster_filter_applied = True
        except Exception as exc:
            print(f"[WARNING] Failed to resolve current-year enrollment roster for face index: {exc}")

    if current_sy and prefer_current_roster and not roster_filter_applied:
        print(
            f"[WARNING] Face index roster filter unavailable for {current_sy}; "
            "falling back to active registered students."
        )

    if roster_clause:
        known_db_encodings, known_db_students = collect_face_index_rows([roster_clause])
        if known_db_encodings is not None and len(known_db_encodings) > 0:
            return known_db_encodings, known_db_students
        if roster_has_rows:
            print(
                f"[WARNING] Face index roster for {current_sy} produced zero encodings; "
                "falling back to active registered students."
            )

    return collect_face_index_rows()


def record_login(username, role):
    try:
        login_history.insert_one({
            "username": username,
            "role": role,
            "timestamp": now_iso(),
            "ip": request.headers.get("X-Forwarded-For", request.remote_addr),
        })
    except Exception as exc:
        print(f"[ERROR] Failed to write login history: {exc}")


def serialize_attendance_correction(doc):
    if not doc:
        return {}
    log_timestamp = str(doc.get("log_timestamp") or "")
    requested_at = str(doc.get("requested_at") or "")
    reviewed_at = str(doc.get("reviewed_at") or "")
    payload = {
        "_id": str(doc.get("_id")),
        "attendance_log_id": str(doc.get("attendance_log_id") or ""),
        "student_id": str(doc.get("student_id") or ""),
        "student_name": str(doc.get("student_name") or ""),
        "log_timestamp": format_timestamp_for_display(log_timestamp),
        "current_status": str(doc.get("current_status") or ""),
        "requested_status": str(doc.get("requested_status") or ""),
        "reason": str(doc.get("reason") or ""),
        "status": str(doc.get("status") or "pending"),
        "requested_by": str(doc.get("requested_by") or ""),
        "requested_at": format_timestamp_for_display(requested_at),
        "reviewed_by": str(doc.get("reviewed_by") or ""),
        "reviewed_at": format_timestamp_for_display(reviewed_at),
        "review_note": str(doc.get("review_note") or ""),
        "applied": bool(doc.get("applied")),
    }
    return payload


def build_system_health_snapshot():
    now_utc = _now_utc()
    db_status = "error"
    db_message = "Unavailable"
    try:
        client.admin.command("ping")
        db_status = "ok"
        db_message = "Connected"
    except Exception as exc:
        db_status = "error"
        db_message = f"Ping failed: {exc}"

    sms_health = sms_provider.health_check()
    sms_status = "ok" if sms_health.get("status") == "ok" else "warn"
    smtp_error = smtp_configuration_error()
    email_status = "ok" if not smtp_error else "warn"

    pending_corrections = attendance_corrections.count_documents({"status": "pending"})
    queued_sms = sms_logs.count_documents({"status": sms_status_mongo_filter("queued")})
    active_lockouts = login_attempts.count_documents({"lockout_until": {"$gt": now_utc}})
    enabled_reports = scheduled_reports.count_documents({"enabled": True})
    enabled_anomaly_rules = anomaly_rules.count_documents({"enabled": True})
    uptime_seconds = max(0, int(time.time() - APP_START_TS))
    with scan_lock:
        scan_active = bool(scan_state.get("active"))
        model_status = str(scan_state.get("model_status") or "idle")
        known_faces = len(scan_state.get("known_students") or [])
        active_tracks = len(scan_state.get("face_tracks") or {})
        pending_recognition = len(scan_state.get("pending_recognition") or {})

    return {
        "generated_at": now_iso(),
        "uptime_seconds": uptime_seconds,
        "database": {"status": db_status, "message": db_message},
        "sms": {
            "status": sms_status,
            "message": sms_health.get("message", ""),
            "provider": sms_health.get("provider", ""),
        },
        "email": {
            "status": email_status,
            "message": "Configured" if email_status == "ok" else smtp_error,
        },
        "scanner": {
            "status": "active" if scan_active else "idle",
            "model_status": model_status,
            "known_faces": known_faces,
            "active_tracks": int(active_tracks),
            "pending_recognition": int(pending_recognition),
            "liveness_enabled": bool(LIVE_SCAN_LIVENESS_ENABLED),
            "liveness_profile": str(get_scan_liveness_profile()),
        },
        "queues": {
            "queued_sms": int(queued_sms),
            "pending_corrections": int(pending_corrections),
            "active_lockouts": int(active_lockouts),
            "enabled_reports": int(enabled_reports),
            "enabled_anomaly_rules": int(enabled_anomaly_rules),
        },
    }


def parse_bool_value(value, default=False):
    if isinstance(value, bool):
        return value
    if value is None:
        return bool(default)
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return bool(default)


def parse_hhmm(value, default=SCHEDULED_REPORT_DEFAULT_SEND_TIME):
    raw = str(value or "").strip()
    if not raw:
        raw = default
    try:
        parsed = datetime.strptime(raw, "%H:%M")
        return parsed.strftime("%H:%M")
    except Exception:
        return default


def parse_email_list(raw_value, max_items=SCHEDULED_REPORT_MAX_RECIPIENTS):
    candidates = []
    if isinstance(raw_value, str):
        candidates = [item.strip() for item in raw_value.split(",")]
    elif isinstance(raw_value, (list, tuple, set)):
        candidates = [str(item).strip() for item in raw_value]

    seen = set()
    valid = []
    for email in candidates:
        normalized = normalize_email_value(email)
        if not normalized or normalized in seen:
            continue
        if not validate_email_format(normalized):
            continue
        seen.add(normalized)
        valid.append(normalized)
        if len(valid) >= max_items:
            break
    return valid


def compute_next_report_run_at(frequency, send_time, now_dt=None):
    now_dt = now_dt or now_local()
    hhmm = parse_hhmm(send_time)
    run_hour, run_minute = [int(part) for part in hhmm.split(":")]
    candidate = now_dt.replace(hour=run_hour, minute=run_minute, second=0, microsecond=0)

    if frequency == "daily":
        if candidate <= now_dt:
            candidate += timedelta(days=1)
        return candidate

    if frequency == "weekly":
        if candidate <= now_dt:
            candidate += timedelta(days=7)
        return candidate

    if frequency == "monthly":
        if candidate <= now_dt:
            candidate += timedelta(days=30)
        return candidate

    return now_dt + timedelta(days=1)


def report_days_from_frequency(frequency):
    if frequency == "daily":
        return 1
    if frequency == "weekly":
        return 7
    return 30


def build_scope_student_ids(grade_value="", section_value="", cap=5000):
    grade = normalize_grade_level(grade_value)
    section = normalize_section_value(section_value)
    query = {}
    if grade:
        query["grade_level"] = grade
    if section:
        query["section"] = section
    if not query:
        return []
    rows = students.find(query, {"student_id": 1}).limit(cap)
    values = []
    for row in rows:
        sid = str(row.get("student_id") or "").strip()
        if sid:
            values.append(sid)
    return values


def build_report_snapshot(days=7, grade_value="", section_value=""):
    days = max(1, min(int(days or 7), 90))
    end_date = now_local().date()
    start_date = end_date - timedelta(days=days - 1)
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")

    attendance_query = {"date": {"$gte": start_str, "$lte": end_str}}
    grade = normalize_grade_level(grade_value)
    section = normalize_section_value(section_value)
    if grade:
        attendance_query["grade_level"] = grade
    if section:
        attendance_query["section"] = section

    scoped_student_ids = build_scope_student_ids(grade, section) if (grade or section) else []

    total_students_query = {}
    if grade:
        total_students_query["grade_level"] = grade
    if section:
        total_students_query["section"] = section

    sms_query = {
        "date": {"$gte": start_str, "$lte": end_str},
        "status": sms_status_mongo_filter("failed"),
    }
    if scoped_student_ids:
        sms_query["student_id"] = {"$in": scoped_student_ids}

    unknown_scan_query = {
        "date": {"$gte": start_str, "$lte": end_str},
        "reason": {"$in": ["unknown_face", "not_registered"]},
    }
    if scoped_student_ids:
        unknown_scan_query["student_id"] = {"$in": scoped_student_ids}

    correction_query = {"status": "pending"}
    if scoped_student_ids:
        correction_query["student_id"] = {"$in": scoped_student_ids}

    total_students = students.count_documents(total_students_query or {})
    present_ids = attendance_logs.distinct("student_id", attendance_query)
    present_count = len([sid for sid in present_ids if sid])
    late_count = attendance_logs.count_documents({**attendance_query, "status": "Late"})
    gate_entries = attendance_logs.count_documents(attendance_query)
    failed_sms = sms_logs.count_documents(sms_query)
    unknown_scans = failed_scans.count_documents(unknown_scan_query)
    pending_corrections = attendance_corrections.count_documents(correction_query)

    top_sections = list(
        attendance_logs.aggregate([
            {"$match": {**attendance_query, "status": "Late"}},
            {"$group": {"_id": {"$ifNull": ["$section", "N/A"]}, "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 3},
        ])
    )

    return {
        "range_start": start_str,
        "range_end": end_str,
        "days": days,
        "scope": {
            "grade": grade or "",
            "section": section or "",
        },
        "stats": {
            "total_students": int(total_students),
            "present_students": int(present_count),
            "late_count": int(late_count),
            "gate_entries": int(gate_entries),
            "failed_sms": int(failed_sms),
            "unknown_scans": int(unknown_scans),
            "pending_corrections": int(pending_corrections),
        },
        "top_late_sections": [
            {"section": str(item.get("_id") or "N/A"), "count": int(item.get("count") or 0)}
            for item in top_sections
        ],
    }


def render_report_email_body(report_doc, snapshot):
    report_name = str(report_doc.get("name") or "Scheduled Report").strip()
    frequency = str(report_doc.get("frequency") or "weekly").strip().lower()
    scope = snapshot.get("scope") or {}
    stats = snapshot.get("stats") or {}
    late_sections = snapshot.get("top_late_sections") or []
    late_lines = [f"- {row.get('section', 'N/A')}: {row.get('count', 0)} late logs" for row in late_sections]
    if not late_lines:
        late_lines = ["- No late section spikes detected."]

    return (
        f"{report_name}\n"
        f"Frequency: {frequency.title()}\n"
        f"Range: {snapshot.get('range_start', '')} to {snapshot.get('range_end', '')}\n"
        f"Scope: Grade={scope.get('grade') or 'All'} | Section={scope.get('section') or 'All'}\n\n"
        "Summary:\n"
        f"- Total Students: {stats.get('total_students', 0)}\n"
        f"- Present Students: {stats.get('present_students', 0)}\n"
        f"- Gate Entries: {stats.get('gate_entries', 0)}\n"
        f"- Late Count: {stats.get('late_count', 0)}\n"
        f"- Failed SMS: {stats.get('failed_sms', 0)}\n"
        f"- Unknown Scans: {stats.get('unknown_scans', 0)}\n"
        f"- Pending Corrections: {stats.get('pending_corrections', 0)}\n\n"
        "Top Late Sections:\n"
        f"{chr(10).join(late_lines)}\n\n"
        "This message was generated automatically by CHS Gate Access."
    )


def serialize_scheduled_report(doc):
    if not doc:
        return {}
    next_run = doc.get("next_run_at")
    last_run = doc.get("last_run_at")
    return {
        "_id": str(doc.get("_id")),
        "name": str(doc.get("name") or ""),
        "frequency": str(doc.get("frequency") or "weekly"),
        "send_time": parse_hhmm(doc.get("send_time") or SCHEDULED_REPORT_DEFAULT_SEND_TIME),
        "enabled": bool(doc.get("enabled", True)),
        "recipients": list(doc.get("recipients") or []),
        "filters": doc.get("filters") if isinstance(doc.get("filters"), dict) else {"grade": "", "section": ""},
        "last_status": str(doc.get("last_status") or ""),
        "last_error": str(doc.get("last_error") or ""),
        "last_run_at": normalize_timestamp_value(last_run),
        "next_run_at": normalize_timestamp_value(next_run),
        "created_by": str(doc.get("created_by") or ""),
        "updated_by": str(doc.get("updated_by") or ""),
        "updated_at": normalize_timestamp_value(doc.get("updated_at")),
    }


def serialize_scheduled_report_run(doc):
    if not doc:
        return {}
    return {
        "_id": str(doc.get("_id")),
        "report_id": str(doc.get("report_id") or ""),
        "report_name": str(doc.get("report_name") or ""),
        "status": str(doc.get("status") or ""),
        "trigger": str(doc.get("trigger") or ""),
        "started_at": normalize_timestamp_value(doc.get("started_at")),
        "completed_at": normalize_timestamp_value(doc.get("completed_at")),
        "error": str(doc.get("error") or ""),
        "recipients_count": len(doc.get("recipients") or []),
    }


def run_single_scheduled_report(report_doc, trigger="manual"):
    if not report_doc:
        return {"status": "error", "message": "Report not found."}

    frequency = str(report_doc.get("frequency") or "weekly").strip().lower()
    if frequency not in SCHEDULED_REPORT_ALLOWED_FREQUENCIES:
        frequency = "weekly"
    send_time = parse_hhmm(report_doc.get("send_time") or SCHEDULED_REPORT_DEFAULT_SEND_TIME)
    recipients = parse_email_list(report_doc.get("recipients") or [])
    now_dt = now_local()

    run_doc = {
        "report_id": str(report_doc.get("_id")),
        "report_name": str(report_doc.get("name") or "Scheduled Report"),
        "status": "failed",
        "trigger": str(trigger or "manual"),
        "started_at": now_dt,
        "completed_at": None,
        "error": "",
        "recipients": recipients,
    }

    if not recipients:
        run_doc["error"] = "No valid recipients configured."
        run_doc["completed_at"] = now_local()
        scheduled_report_runs.insert_one(run_doc)
        scheduled_reports.update_one(
            {"_id": report_doc["_id"]},
            {"$set": {
                "last_status": "failed",
                "last_error": run_doc["error"],
                "last_run_at": now_dt,
                "next_run_at": compute_next_report_run_at(frequency, send_time, now_dt=now_dt),
                "updated_at": now_dt,
            }},
        )
        return {"status": "error", "message": run_doc["error"]}

    filters = report_doc.get("filters") if isinstance(report_doc.get("filters"), dict) else {}
    snapshot = build_report_snapshot(
        days=report_days_from_frequency(frequency),
        grade_value=filters.get("grade") or "",
        section_value=filters.get("section") or "",
    )
    email_subject = f"[CHS] {str(report_doc.get('name') or 'Scheduled Report').strip()} ({frequency.title()})"
    email_body = render_report_email_body(report_doc, snapshot)
    sent_ok, send_error = send_email_message(
        subject=email_subject,
        body_text=email_body,
        recipients=recipients,
        from_name="CHS Reports",
    )

    run_doc["completed_at"] = now_local()
    run_doc["snapshot"] = snapshot

    if sent_ok:
        run_doc["status"] = "success"
        create_alert(
            "info",
            f"Scheduled report '{report_doc.get('name', 'Report')}' sent to {len(recipients)} recipient(s).",
            "analytics",
            {"report_id": str(report_doc.get("_id")), "trigger": trigger},
        )
        log_audit_event(
            action="analytics.scheduled_report_run",
            outcome="success",
            severity="info",
            target_type="scheduled_report",
            target_id=str(report_doc.get("_id")),
            details={"trigger": trigger, "recipients_count": len(recipients)},
        )
    else:
        run_doc["status"] = "failed"
        run_doc["error"] = send_error
        log_audit_event(
            action="analytics.scheduled_report_run",
            outcome="failed",
            severity="warn",
            target_type="scheduled_report",
            target_id=str(report_doc.get("_id")),
            details={"trigger": trigger, "error": send_error},
        )

    scheduled_report_runs.insert_one(run_doc)
    scheduled_reports.update_one(
        {"_id": report_doc["_id"]},
        {"$set": {
            "last_status": run_doc["status"],
            "last_error": run_doc.get("error", ""),
            "last_run_at": run_doc["completed_at"],
            "next_run_at": compute_next_report_run_at(frequency, send_time, now_dt=run_doc["completed_at"]),
            "updated_at": run_doc["completed_at"],
        }},
    )
    return {
        "status": "ok" if sent_ok else "error",
        "message": "Report sent successfully." if sent_ok else (send_error or "Failed to send report."),
        "run": serialize_scheduled_report_run(run_doc),
    }


def run_due_scheduled_reports(max_reports=5):
    now_dt = now_local()
    try:
        limit_value = max(1, min(int(max_reports or 5), 20))
    except Exception:
        limit_value = 5
    due_reports = list(
        scheduled_reports.find({
            "enabled": True,
            "next_run_at": {"$lte": now_dt},
        }).sort("next_run_at", 1).limit(limit_value)
    )
    results = []
    for report_doc in due_reports:
        results.append(run_single_scheduled_report(report_doc, trigger="scheduler"))
    return results


def compare_anomaly_value(metric_value, operator_value, threshold_value):
    if operator_value == "gt":
        return metric_value > threshold_value
    if operator_value == "gte":
        return metric_value >= threshold_value
    if operator_value == "lt":
        return metric_value < threshold_value
    return metric_value <= threshold_value


def compute_anomaly_metric_value(rule_doc):
    metric = str(rule_doc.get("metric") or "").strip().lower()
    window_days = max(1, min(int(rule_doc.get("window_days") or 1), 90))
    filters = rule_doc.get("filters") if isinstance(rule_doc.get("filters"), dict) else {}
    grade = normalize_grade_level(filters.get("grade") or "")
    section = normalize_section_value(filters.get("section") or "")

    end_date = now_local().date()
    start_date = end_date - timedelta(days=window_days - 1)
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")
    attendance_query = {"date": {"$gte": start_str, "$lte": end_str}}
    if grade:
        attendance_query["grade_level"] = grade
    if section:
        attendance_query["section"] = section

    scoped_student_ids = build_scope_student_ids(grade, section) if (grade or section) else []

    if metric == "late_count":
        value = attendance_logs.count_documents({**attendance_query, "status": "Late"})
    elif metric == "failed_sms_count":
        sms_query = {
            "date": {"$gte": start_str, "$lte": end_str},
            "status": sms_status_mongo_filter("failed"),
        }
        if scoped_student_ids:
            sms_query["student_id"] = {"$in": scoped_student_ids}
        value = sms_logs.count_documents(sms_query)
    elif metric == "unknown_scan_count":
        failed_query = {
            "date": {"$gte": start_str, "$lte": end_str},
            "reason": {"$in": ["unknown_face", "not_registered"]},
        }
        if scoped_student_ids:
            failed_query["student_id"] = {"$in": scoped_student_ids}
        value = failed_scans.count_documents(failed_query)
    else:
        correction_query = {"status": "pending"}
        if scoped_student_ids:
            correction_query["student_id"] = {"$in": scoped_student_ids}
        value = attendance_corrections.count_documents(correction_query)

    return {
        "metric": metric,
        "value": float(value),
        "window_days": window_days,
        "range_start": start_str,
        "range_end": end_str,
        "filters": {"grade": grade or "", "section": section or ""},
    }


def serialize_anomaly_rule(doc):
    if not doc:
        return {}
    return {
        "_id": str(doc.get("_id")),
        "name": str(doc.get("name") or ""),
        "metric": str(doc.get("metric") or ""),
        "operator": str(doc.get("operator") or ""),
        "threshold": float(doc.get("threshold") or 0),
        "window_days": int(doc.get("window_days") or 1),
        "severity": str(doc.get("severity") or "warn"),
        "enabled": bool(doc.get("enabled", True)),
        "cooldown_minutes": int(doc.get("cooldown_minutes") or ANOMALY_DEFAULT_COOLDOWN_MINUTES),
        "filters": doc.get("filters") if isinstance(doc.get("filters"), dict) else {"grade": "", "section": ""},
        "notify_emails": list(doc.get("notify_emails") or []),
        "last_evaluated_at": normalize_timestamp_value(doc.get("last_evaluated_at")),
        "last_triggered_at": normalize_timestamp_value(doc.get("last_triggered_at")),
        "last_value": float(doc.get("last_value") or 0),
        "last_result": str(doc.get("last_result") or ""),
    }


def serialize_anomaly_event(doc):
    if not doc:
        return {}
    return {
        "_id": str(doc.get("_id")),
        "rule_id": str(doc.get("rule_id") or ""),
        "rule_name": str(doc.get("rule_name") or ""),
        "metric": str(doc.get("metric") or ""),
        "value": float(doc.get("value") or 0),
        "operator": str(doc.get("operator") or ""),
        "threshold": float(doc.get("threshold") or 0),
        "severity": str(doc.get("severity") or "warn"),
        "triggered_at": normalize_timestamp_value(doc.get("triggered_at")),
        "trigger": str(doc.get("trigger") or ""),
    }


def evaluate_anomaly_rule(rule_doc, trigger="manual"):
    if not rule_doc:
        return {"status": "error", "message": "Rule not found."}

    metric = str(rule_doc.get("metric") or "").strip().lower()
    operator_value = str(rule_doc.get("operator") or "").strip().lower()
    threshold = float(rule_doc.get("threshold") or 0)
    severity = str(rule_doc.get("severity") or "warn").strip().lower()
    if metric not in ANOMALY_ALLOWED_METRICS or operator_value not in ANOMALY_ALLOWED_OPERATORS:
        return {"status": "error", "message": "Invalid anomaly rule definition."}
    if severity not in ANOMALY_ALLOWED_SEVERITIES:
        severity = "warn"

    eval_payload = compute_anomaly_metric_value(rule_doc)
    metric_value = float(eval_payload.get("value") or 0)
    matched = compare_anomaly_value(metric_value, operator_value, threshold)
    now_dt = now_local()

    update_payload = {
        "last_evaluated_at": now_dt,
        "last_value": metric_value,
        "last_result": "matched" if matched else "normal",
        "updated_at": now_dt,
    }

    event_doc = None
    if matched:
        cooldown_minutes = max(5, int(rule_doc.get("cooldown_minutes") or ANOMALY_DEFAULT_COOLDOWN_MINUTES))
        last_triggered = rule_doc.get("last_triggered_at")
        cooldown_ok = not isinstance(last_triggered, datetime) or (now_dt - last_triggered) >= timedelta(minutes=cooldown_minutes)

        if cooldown_ok:
            event_doc = {
                "rule_id": str(rule_doc.get("_id")),
                "rule_name": str(rule_doc.get("name") or ""),
                "metric": metric,
                "value": metric_value,
                "operator": operator_value,
                "threshold": threshold,
                "severity": severity,
                "triggered_at": now_dt,
                "trigger": str(trigger or "manual"),
                "context": eval_payload,
            }
            anomaly_events.insert_one(event_doc)
            update_payload["last_triggered_at"] = now_dt

            create_alert(
                severity,
                f"Anomaly rule '{rule_doc.get('name', 'Rule')}' triggered ({metric_value:.2f} {operator_value} {threshold:.2f}).",
                "analytics",
                {
                    "rule_id": str(rule_doc.get("_id")),
                    "metric": metric,
                    "value": metric_value,
                    "threshold": threshold,
                    "operator": operator_value,
                },
            )

            notify_emails = parse_email_list(rule_doc.get("notify_emails") or [])
            if notify_emails:
                subject = f"[CHS] Anomaly Triggered: {rule_doc.get('name', 'Rule')}"
                body = (
                    f"Rule: {rule_doc.get('name', '')}\n"
                    f"Metric: {metric}\n"
                    f"Observed Value: {metric_value:.2f}\n"
                    f"Condition: {operator_value} {threshold:.2f}\n"
                    f"Window: last {eval_payload.get('window_days', 1)} day(s)\n"
                    f"Range: {eval_payload.get('range_start', '')} to {eval_payload.get('range_end', '')}\n"
                    f"Severity: {severity}\n"
                )
                send_email_message(subject=subject, body_text=body, recipients=notify_emails, from_name="CHS Alerts")

            log_audit_event(
                action="analytics.anomaly_triggered",
                outcome="success",
                severity="warn" if severity == "warn" else severity,
                target_type="anomaly_rule",
                target_id=str(rule_doc.get("_id")),
                details={
                    "metric": metric,
                    "value": metric_value,
                    "operator": operator_value,
                    "threshold": threshold,
                    "trigger": trigger,
                },
            )

    anomaly_rules.update_one({"_id": rule_doc["_id"]}, {"$set": update_payload})
    return {
        "status": "ok",
        "rule": serialize_anomaly_rule({**rule_doc, **update_payload}),
        "matched": bool(matched),
        "event": serialize_anomaly_event(event_doc) if event_doc else None,
    }


def evaluate_all_anomaly_rules(trigger="manual", max_rules=50):
    try:
        safe_limit = max(1, min(int(max_rules or 50), 200))
    except Exception:
        safe_limit = 50
    docs = list(anomaly_rules.find({"enabled": True}).sort("updated_at", -1).limit(safe_limit))
    results = [evaluate_anomaly_rule(doc, trigger=trigger) for doc in docs]
    triggered = len([row for row in results if row.get("matched") and row.get("event")])
    return {
        "status": "ok",
        "evaluated": len(results),
        "triggered": triggered,
        "results": results,
    }


def background_jobs_worker_loop():
    while True:
        try:
            cleanup_notification_alerts()
            process_pending_sms_retries(max_logs=5)
            run_due_scheduled_reports(max_reports=3)
            evaluate_all_anomaly_rules(trigger="scheduler", max_rules=50)
        except Exception as exc:
            print(f"[WARNING] Background jobs loop error: {exc}")
        time.sleep(BACKGROUND_JOB_INTERVAL_SECONDS)


def sms_status_filter_values(*statuses):
    values = []
    for status in statuses:
        norm = (status or "").strip()
        if not norm:
            continue
        lower = norm.lower()
        upper = lower.upper()
        if lower not in values:
            values.append(lower)
        if upper not in values:
            values.append(upper)
    return values


def sms_status_mongo_filter(*statuses):
    values = sms_status_filter_values(*statuses)
    if not values:
        return {}
    return {"$in": values}


def _coerce_balance_units(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return float(value)
        except (TypeError, ValueError, InvalidOperation):
            return None
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group(0))
        except (TypeError, ValueError):
            return None
    return None


def invalidate_sms_balance_cache(reason=""):
    with sms_balance_lock:
        sms_balance_cache["checked_ts"] = 0.0
        sms_balance_cache["checked_at"] = ""
        sms_balance_cache["status"] = "stale"
        if reason:
            sms_balance_cache["message"] = reason


def get_sms_balance_snapshot(force=False):
    now_ts = time.time()
    with sms_balance_lock:
        cached_ts = float(sms_balance_cache.get("checked_ts") or 0.0)
        if not force and cached_ts and (now_ts - cached_ts) < SMS_BALANCE_CACHE_TTL_SECONDS:
            cached_units = sms_balance_cache.get("units")
            is_low = cached_units is not None and cached_units < SMS_BALANCE_LOW_THRESHOLD
            return {
                **sms_balance_cache,
                "cached": True,
                "low_threshold": SMS_BALANCE_LOW_THRESHOLD,
                "is_low": is_low,
            }

    provider_result = {}
    try:
        balance_method = getattr(sms_provider, "get_balance", None)
        if callable(balance_method):
            provider_result = balance_method() or {}
        else:
            provider_result = sms_provider.health_check() or {}
    except Exception as exc:
        provider_result = {
            "status": "failed",
            "message": f"Failed to fetch balance: {exc}",
        }

    raw_units = provider_result.get("units")
    units = _coerce_balance_units(raw_units)
    status = str(provider_result.get("status") or "failed").lower()
    message = str(provider_result.get("message") or "").strip()
    if status == "ok" and units is None:
        status = "warn"
        if not message:
            message = "Balance is not available from provider."

    checked_at = now_iso()
    with sms_balance_lock:
        sms_balance_cache.update({
            "status": status,
            "units": units,
            "message": message,
            "provider": str(provider_result.get("provider") or "PHILSMS"),
            "http_status": provider_result.get("http_status"),
            "probe_path": str(provider_result.get("probe_path") or ""),
            "checked_at": checked_at,
            "checked_ts": now_ts,
        })
        cached = dict(sms_balance_cache)

    is_low = units is not None and units < SMS_BALANCE_LOW_THRESHOLD
    return {
        **cached,
        "cached": False,
        "low_threshold": SMS_BALANCE_LOW_THRESHOLD,
        "is_low": is_low,
    }


def normalize_sms_template_text(value):
    return sanitize_profile_text(value, SMS_TEMPLATE_MAX_LENGTH, allow_newlines=False)


def normalize_boolean_setting_value(value, default=False):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    normalized = str(value or "").strip().lower()
    if normalized in {"1", "true", "yes", "on", "enabled"}:
        return True
    if normalized in {"0", "false", "no", "off", "disabled"}:
        return False
    return bool(default)


def parse_boolean_setting_input(value, field_name="enabled"):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    normalized = str(value or "").strip().lower()
    if normalized in {"1", "true", "yes", "on", "enabled"}:
        return True
    if normalized in {"0", "false", "no", "off", "disabled"}:
        return False
    raise ValueError(f"{field_name.replace('_', ' ').title()} must be true or false.")


def get_default_attendance_sms_template():
    fallback = "CHS Gate Access: {student_name} {movement_text} the gate ({status}) at {time} on {date}."
    from_env = normalize_sms_template_text(ATTENDANCE_SMS_TEMPLATE_DEFAULT)
    return from_env or fallback


def ensure_attendance_sms_notification_defaults():
    now_ts = now_iso()
    try:
        system_settings.update_one(
            {"key": ATTENDANCE_SMS_NOTIFICATION_SETTINGS_KEY},
            {
                "$setOnInsert": {
                    "name": "Attendance SMS Notifications",
                    "context": "attendance_gate_scan",
                    "enabled": True,
                    "createdAt": now_ts,
                    "updatedAt": now_ts,
                    "updatedBy": {"username": "system", "role": "System"},
                }
            },
            upsert=True,
        )
    except Exception as exc:
        print(f"[WARNING] Failed ensuring attendance SMS notification default: {exc}")


def ensure_sms_template_defaults():
    now_ts = now_iso()
    default_template = get_default_attendance_sms_template()
    try:
        sms_templates.update_one(
            {"_id": ATTENDANCE_SMS_TEMPLATE_DOC_ID},
            {
                "$setOnInsert": {
                    "name": "Attendance Gate Scan Notification",
                    "context": "attendance_gate_scan",
                    "template": default_template,
                    "variables": list(ATTENDANCE_SMS_TEMPLATE_VARIABLES),
                    "maxLength": SMS_TEMPLATE_MAX_LENGTH,
                    "createdAt": now_ts,
                    "updatedAt": now_ts,
                    "updatedBy": {"username": "system", "role": "System"},
                }
            },
            upsert=True,
        )
    except Exception as exc:
        print(f"[WARNING] Failed ensuring SMS template default: {exc}")


def get_attendance_sms_notification_settings_payload():
    payload = {
        "enabled": True,
        "updated_at": "",
        "updated_by": "",
        "status_label": "Enabled",
    }

    try:
        doc = system_settings.find_one({"key": ATTENDANCE_SMS_NOTIFICATION_SETTINGS_KEY}) or {}
    except Exception as exc:
        print(f"[WARNING] Failed reading attendance SMS notification settings: {exc}")
        return payload

    payload["enabled"] = normalize_boolean_setting_value(doc.get("enabled"), default=True)
    payload["status_label"] = "Enabled" if payload["enabled"] else "Disabled"
    payload["updated_at"] = str(doc.get("updatedAt") or "").strip()

    updated_by_value = doc.get("updatedBy")
    if isinstance(updated_by_value, dict):
        payload["updated_by"] = str(updated_by_value.get("username") or "").strip()
    else:
        payload["updated_by"] = str(updated_by_value or "").strip()

    return payload


def attendance_sms_notifications_enabled():
    return bool(get_attendance_sms_notification_settings_payload().get("enabled"))


def get_attendance_sms_template_payload():
    default_template = get_default_attendance_sms_template()
    payload = {
        "template": default_template,
        "default_template": default_template,
        "updated_at": "",
        "updated_by": "",
        "max_length": SMS_TEMPLATE_MAX_LENGTH,
        "variables": list(ATTENDANCE_SMS_TEMPLATE_VARIABLES),
    }

    try:
        doc = sms_templates.find_one({"_id": ATTENDANCE_SMS_TEMPLATE_DOC_ID}) or {}
    except Exception as exc:
        print(f"[WARNING] Failed reading SMS template from MongoDB: {exc}")
        return payload

    stored_template = normalize_sms_template_text(doc.get("template", ""))
    payload["template"] = stored_template or default_template
    payload["updated_at"] = str(doc.get("updatedAt") or "").strip()

    updated_by_value = doc.get("updatedBy")
    if isinstance(updated_by_value, dict):
        payload["updated_by"] = str(updated_by_value.get("username") or "").strip()
    else:
        payload["updated_by"] = str(updated_by_value or "").strip()

    return payload


def save_attendance_sms_template(template_text, actor_username="", actor_role=""):
    cleaned_template = normalize_sms_template_text(template_text)
    if not cleaned_template:
        raise ValueError("SMS template cannot be empty.")

    now_ts = now_iso()
    actor_name = sanitize_profile_text(actor_username, 64)
    actor_role_text = sanitize_profile_text(actor_role, 64)

    sms_templates.update_one(
        {"_id": ATTENDANCE_SMS_TEMPLATE_DOC_ID},
        {
            "$set": {
                "name": "Attendance Gate Scan Notification",
                "context": "attendance_gate_scan",
                "template": cleaned_template,
                "variables": list(ATTENDANCE_SMS_TEMPLATE_VARIABLES),
                "maxLength": SMS_TEMPLATE_MAX_LENGTH,
                "updatedAt": now_ts,
                "updatedBy": {
                    "username": actor_name or "system",
                    "role": actor_role_text or "System",
                },
            },
            "$setOnInsert": {
                "createdAt": now_ts,
            },
        },
        upsert=True,
    )

    payload = get_attendance_sms_template_payload()
    payload["template"] = cleaned_template
    payload["updated_at"] = now_ts
    payload["updated_by"] = actor_name or "system"
    return payload


def save_attendance_sms_notification_settings(enabled, actor_username="", actor_role=""):
    normalized_enabled = parse_boolean_setting_input(enabled, field_name="enabled")
    now_ts = now_iso()
    actor_name = sanitize_profile_text(actor_username, 64)
    actor_role_text = sanitize_profile_text(actor_role, 64)

    system_settings.update_one(
        {"key": ATTENDANCE_SMS_NOTIFICATION_SETTINGS_KEY},
        {
            "$set": {
                "name": "Attendance SMS Notifications",
                "context": "attendance_gate_scan",
                "enabled": normalized_enabled,
                "updatedAt": now_ts,
                "updatedBy": {
                    "username": actor_name or "system",
                    "role": actor_role_text or "System",
                },
            },
            "$setOnInsert": {
                "createdAt": now_ts,
            },
        },
        upsert=True,
    )

    payload = get_attendance_sms_notification_settings_payload()
    payload["enabled"] = normalized_enabled
    payload["status_label"] = "Enabled" if normalized_enabled else "Disabled"
    payload["updated_at"] = now_ts
    payload["updated_by"] = actor_name or "system"
    return payload


def log_skipped_sms(student_id="", student_name="", parent_contact="", message="", reason="skipped", sms_type="transactional", metadata=None):
    now = now_local()
    timestamp = now_iso()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    metadata_payload = metadata if isinstance(metadata, dict) else {}
    # Always use the current active school year from database for SMS logging
    school_year_label = get_current_school_year_label()
    sms_collection, school_year_label, _ = get_sms_logs_storage(school_year_label)
    raw_to = str(parent_contact or "").strip()
    normalized_to = ""
    if raw_to:
        try:
            normalized_to = SmsProvider.normalize_phone_number(raw_to)
        except Exception:
            normalized_to = raw_to

    doc = {
        "to": normalized_to or raw_to,
        "message": str(message or "").strip(),
        "type": (sms_type or "transactional").strip().lower() or "transactional",
        "status": "skipped",
        "provider": "PHILSMS",
        "providerMessageId": "",
        "providerResponse": {
            "phase": "skipped",
            "reason": str(reason or "skipped"),
            "meta": metadata_payload,
        },
        "error": "",
        "httpStatus": None,
        "errorCode": "SKIPPED",
        "errorMessage": str(reason or "skipped"),
        "createdAt": timestamp,
        "updatedAt": timestamp,
        "school_year": school_year_label,
        # Legacy compatibility fields
        "student_id": (student_id or "").strip(),
        "name": (student_name or "").strip(),
        "parent_contact": raw_to,
        "parent_contact_raw": raw_to,
        "retryEligible": False,
        "retryCount": 0,
        "retryMaxAttempts": 0,
        "nextRetryAt": None,
        "lastRetryError": None,
        "sid": "",
        "timestamp": timestamp,
        "date": date_str,
        "time": time_str,
    }
    try:
        sms_collection.insert_one(doc)
        signal_data_change("sms_logs")
    except Exception as exc:
        print(f"[ERROR] Failed to persist skipped SMS log: {exc}")


def send_sms(to_number, message, sms_type="transactional", metadata=None, student_id="", student_name="", parent_contact="", persist=True):
    now = now_local()
    timestamp = now_iso()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    normalized_to = ""
    queued_doc_id = None
    provider = "PHILSMS"
    msg = str(message or "").strip()
    raw_to = str(to_number or "").strip()
    metadata_payload = metadata if isinstance(metadata, dict) else {}
    # Always use the current active school year from database for SMS sending
    school_year_label = get_current_school_year_label()
    sms_collection, school_year_label, _ = get_sms_logs_storage(school_year_label)
    try:
        retry_count = max(int(metadata_payload.get("retry_count", 0) or 0), 0)
    except Exception:
        retry_count = 0
    try:
        retry_delay_seconds = max(int(os.getenv("SMS_RETRY_DELAY_SECONDS", "300")), 60)
    except Exception:
        retry_delay_seconds = 300
    try:
        retry_max_attempts = max(int(os.getenv("SMS_RETRY_MAX_ATTEMPTS", "3")), 1)
    except Exception:
        retry_max_attempts = 3

    def persist_failed_input(error_message):
        if not persist:
            return
        try:
            sms_collection.insert_one({
                "to": raw_to,
                "message": msg,
                "type": (sms_type or "transactional").strip().lower() or "transactional",
                "status": "failed",
                "provider": provider,
                "providerMessageId": "",
                "providerResponse": {"error": error_message, "phase": "input_validation"},
                "error": error_message,
                "httpStatus": None,
                "errorCode": "INPUT_VALIDATION",
                "errorMessage": error_message,
                "createdAt": timestamp,
                "updatedAt": timestamp,
                "school_year": school_year_label,
                # Legacy compatibility fields
                "student_id": (student_id or "").strip(),
                "name": (student_name or "").strip(),
                "parent_contact": (parent_contact or raw_to).strip(),
                "parent_contact_raw": (parent_contact or raw_to).strip(),
                "retryEligible": False,
                "retryCount": retry_count,
                "retryMaxAttempts": retry_max_attempts,
                "nextRetryAt": None,
                "lastRetryError": error_message,
                "sid": "",
                "timestamp": timestamp,
                "date": date_str,
                "time": time_str,
            })
            signal_data_change("sms_logs")
        except Exception as exc:
            print(f"[ERROR] Failed to persist invalid SMS attempt: {exc}")

    try:
        normalized_to = SmsProvider.normalize_phone_number(to_number)
    except Exception as exc:
        err = str(exc)
        persist_failed_input(err)
        return {"status": "failed", "error": err, "sid": "", "provider_message_id": "", "provider_response": {"error": err}}

    if not msg:
        err = "Message is required."
        persist_failed_input(err)
        return {"status": "failed", "error": err, "sid": "", "provider_message_id": "", "provider_response": {"error": err}}

    if persist:
        queued_doc = {
            "to": normalized_to,
            "message": msg,
            "type": (sms_type or "transactional").strip().lower() or "transactional",
            "status": "sending",
            "provider": provider,
            "providerMessageId": "",
            "providerResponse": {"phase": "sending"},
            "error": "",
            "httpStatus": None,
            "errorCode": None,
            "errorMessage": None,
            "createdAt": timestamp,
            "updatedAt": timestamp,
            "school_year": school_year_label,
            # Legacy compatibility fields
            "student_id": (student_id or "").strip(),
            "name": (student_name or "").strip(),
            "parent_contact": (parent_contact or raw_to or normalized_to).strip(),
            "parent_contact_raw": (parent_contact or raw_to).strip(),
            "retryEligible": False,
            "retryCount": retry_count,
            "retryMaxAttempts": retry_max_attempts,
            "nextRetryAt": None,
            "lastRetryError": None,
            "sid": "",
            "timestamp": timestamp,
            "date": date_str,
            "time": time_str,
        }
        try:
            queued_doc_id = sms_collection.insert_one(queued_doc).inserted_id
        except Exception as exc:
            print(f"[ERROR] Failed to persist sending SMS log: {exc}")

    try:
        result = sms_provider.send_sms(
            to_number=normalized_to,
            message=msg,
            sms_type=(sms_type or "transactional").strip().lower() or "transactional",
            metadata=metadata_payload,
        )
    except Exception as exc:
        result = {
            "status": "failed",
            "provider": provider,
            "provider_message_id": "",
            "provider_response": {"exception": str(exc), "phase": "provider_send"},
            "http_status": None,
            "to": normalized_to,
            "error": str(exc),
            "error_code": "PROVIDER_CONFIGURATION",
            "error_message": str(exc),
        }

    log_fields = SmsProvider.map_result_to_log_fields(result)
    delivery_status = log_fields["status"]
    provider_message_id = (log_fields.get("providerMessageId") or "").strip()
    provider_response = log_fields.get("providerResponse") or {}
    error = (result.get("error") or "").strip() or (log_fields.get("errorMessage") or "")

    if persist and queued_doc_id:
        update_doc = {
            **log_fields,
            "updatedAt": now_iso(),
        }
        if delivery_status == "failed":
            can_retry = retry_count < retry_max_attempts
            update_doc.update({
                "retryEligible": can_retry,
                "retryCount": retry_count,
                "retryMaxAttempts": retry_max_attempts,
                "nextRetryAt": (now_local() + timedelta(seconds=retry_delay_seconds)).isoformat() if can_retry else None,
                "lastRetryError": error or (log_fields.get("errorMessage") or None),
            })
        else:
            update_doc.update({
                "retryEligible": False,
                "retryCount": retry_count,
                "retryMaxAttempts": retry_max_attempts,
                "nextRetryAt": None,
                "lastRetryError": None,
            })
        try:
            sms_collection.update_one({"_id": queued_doc_id}, {"$set": update_doc})
            signal_data_change("sms_logs")
        except Exception as exc:
            print(f"[ERROR] Failed to update SMS log status: {exc}")

    if delivery_status == "sent":
        invalidate_sms_balance_cache("Balance refresh pending after SMS send.")

    return {
        "status": delivery_status,
        "sid": provider_message_id,
        "provider_message_id": provider_message_id,
        "provider_response": provider_response,
        "error": error,
        "http_status": result.get("http_status"),
        "error_code": result.get("error_code", ""),
        "error_message": result.get("error_message", ""),
        "to": normalized_to,
        "log_id": str(queued_doc_id) if queued_doc_id else "",
    }


def process_pending_sms_retries(max_logs=5):
    try:
        safe_limit = max(1, min(int(max_logs or 5), 25))
    except Exception:
        safe_limit = 5

    school_year_label = get_current_school_year_label()
    sms_collection, school_year_label, _ = get_sms_logs_storage(school_year_label)
    now_dt = now_local()
    now_value = now_iso()

    try:
        retry_delay_seconds = max(int(os.getenv("SMS_RETRY_DELAY_SECONDS", "300")), 60)
    except Exception:
        retry_delay_seconds = 300
    try:
        default_retry_max_attempts = max(int(os.getenv("SMS_RETRY_MAX_ATTEMPTS", "3")), 1)
    except Exception:
        default_retry_max_attempts = 3

    due_query = {
        "retryEligible": True,
        "status": sms_status_mongo_filter("failed"),
        "nextRetryAt": {"$ne": None, "$lte": now_value},
    }
    if school_year_label:
        due_query["school_year"] = school_year_label

    try:
        due_docs = list(sms_collection.find(due_query).sort("nextRetryAt", 1).limit(safe_limit))
    except Exception as exc:
        print(f"[WARNING] Failed reading retryable SMS logs: {exc}")
        return {"processed": 0, "sent": 0, "failed": 0, "remaining": 0}

    summary = {"processed": 0, "sent": 0, "failed": 0, "remaining": 0}

    if not attendance_sms_notifications_enabled():
        try:
            summary["remaining"] = int(sms_collection.count_documents(due_query))
        except Exception:
            summary["remaining"] = 0
        return summary

    for doc in due_docs:
        doc_id = doc.get("_id")
        recipient = str(doc.get("to") or doc.get("parent_contact") or "").strip()
        message = str(doc.get("message") or "").strip()
        if not doc_id or not recipient or not message:
            continue

        try:
            previous_retry_count = max(int(doc.get("retryCount") or 0), 0)
        except Exception:
            previous_retry_count = 0
        try:
            retry_max_attempts = max(int(doc.get("retryMaxAttempts") or default_retry_max_attempts), 1)
        except Exception:
            retry_max_attempts = default_retry_max_attempts
        attempt_number = previous_retry_count + 1

        metadata_payload = {}
        provider_response = doc.get("providerResponse")
        if isinstance(provider_response, dict):
            response_meta = provider_response.get("meta")
            if isinstance(response_meta, dict):
                metadata_payload.update(response_meta)
        metadata_payload.update({
            "context": "auto_retry_sms",
            "retry_of": str(doc_id),
            "school_year": school_year_label,
        })
        metadata_payload["retry_count"] = attempt_number

        sms_result = send_sms(
            recipient,
            message,
            sms_type=doc.get("type", "transactional"),
            metadata=metadata_payload,
            student_id=doc.get("student_id", ""),
            student_name=doc.get("name", ""),
            parent_contact=doc.get("parent_contact", ""),
            persist=False,
        )

        log_fields = SmsProvider.map_result_to_log_fields(sms_result)
        delivery_status = log_fields["status"]
        error_message = (
            str(sms_result.get("error") or "").strip()
            or str(sms_result.get("error_message") or "").strip()
            or str(log_fields.get("errorMessage") or "").strip()
        )
        can_retry_again = delivery_status != "sent" and attempt_number < retry_max_attempts
        next_retry_at = (now_dt + timedelta(seconds=retry_delay_seconds)).isoformat() if can_retry_again else None

        update_doc = {
            **log_fields,
            "updatedAt": now_iso(),
            "retryEligible": can_retry_again,
            "retryCount": attempt_number,
            "retryMaxAttempts": retry_max_attempts,
            "nextRetryAt": next_retry_at,
            "lastRetryError": error_message or None,
        }
        if delivery_status == "sent":
            update_doc["lastRetryError"] = None
            invalidate_sms_balance_cache("Balance refresh pending after SMS retry.")

        try:
            sms_collection.update_one({"_id": doc_id}, {"$set": update_doc})
            signal_data_change("sms_logs")
        except Exception as exc:
            print(f"[WARNING] Failed updating retried SMS log {doc_id}: {exc}")
            continue

        summary["processed"] += 1
        if delivery_status == "sent":
            summary["sent"] += 1
        else:
            summary["failed"] += 1
            if not can_retry_again:
                create_alert(
                    level="high",
                    message=f"SMS delivery failed after retries for {doc.get('name') or doc.get('student_id') or 'Unknown student'}.",
                    category="sms",
                    meta={
                        "student_id": str(doc.get("student_id") or ""),
                        "error": error_message,
                        "school_year": school_year_label,
                    },
                )

    try:
        remaining_query = dict(due_query)
        remaining_query["nextRetryAt"] = {"$ne": None, "$lte": now_iso()}
        summary["remaining"] = int(sms_collection.count_documents(remaining_query))
    except Exception:
        summary["remaining"] = 0

    return summary


def ensure_default_admin_user():
    try:
        admin = users.find_one({"username": "admin"})
        if not admin:
            created = now_iso()
            users.insert_one({
                "username": "admin",
                "password_hash": hash_password("admin123"),
                "role": ROLE_FULL_ADMIN,
                "fullName": "System Administrator",
                "email": "admin@chs.local",
                "phone": "",
                "address": "",
                "bio": "",
                "avatarUrl": "",
                "twoFactorEnabled": False,
                "theme": "light",
                "created_at": created,
                "updated_at": created,
                "updatedAt": created,
            })
        else:
            updates = {}
            if "password_hash" not in admin:
                legacy_pwd = admin.get("password", "admin123")
                updates["password_hash"] = hash_password(legacy_pwd)
            if normalize_account_role(admin.get("role"), "admin") != ROLE_FULL_ADMIN:
                updates["role"] = ROLE_FULL_ADMIN
            if "fullName" not in admin:
                updates["fullName"] = "System Administrator"
            if "email" not in admin:
                updates["email"] = "admin@chs.local"
            if "phone" not in admin:
                updates["phone"] = ""
            if "address" not in admin:
                updates["address"] = ""
            if "bio" not in admin:
                updates["bio"] = ""
            if "avatarUrl" not in admin:
                updates["avatarUrl"] = ""
            if "twoFactorEnabled" not in admin:
                updates["twoFactorEnabled"] = False
            if normalize_theme_value(admin.get("theme"), default="") == "":
                updates["theme"] = "light"
            if "updatedAt" not in admin:
                updates["updatedAt"] = (admin.get("updated_at") or admin.get("created_at") or now_iso())
            if updates:
                users.update_one({"_id": admin["_id"]}, {"$set": updates})
    except Exception as exc:
        print(f"[ERROR] Failed ensuring default admin user: {exc}")


def migrate_user_roles():
    try:
        updated = now_iso()
        users.update_many(
            {"role": LEGACY_LIMITED_ACCESS_ROLE},
            {"$set": {"role": ROLE_STAFF, "updated_at": updated, "updatedAt": updated}},
        )
        users.update_many(
            {
                "username": {"$ne": "admin"},
                "$or": [
                    {"role": {"$exists": False}},
                    {"role": ""},
                    {"role": {"$nin": [ROLE_FULL_ADMIN, ROLE_STAFF, LEGACY_LIMITED_ACCESS_ROLE]}},
                ],
            },
            {"$set": {"role": ROLE_STAFF, "updated_at": updated, "updatedAt": updated}},
        )
        users.update_one(
            {"username": "admin"},
            {"$set": {"role": ROLE_FULL_ADMIN, "updated_at": updated, "updatedAt": updated}},
        )
    except Exception as exc:
        print(f"[ERROR] Failed migrating user roles: {exc}")


def migrate_plaintext_user_passwords():
    try:
        cursor = users.find({"password": {"$exists": True}})
        for user in cursor:
            if user.get("password_hash"):
                continue
            plain = user.get("password")
            if not plain:
                continue
            users.update_one(
                {"_id": user["_id"]},
                {
                    "$set": {"password_hash": hash_password(plain)},
                    "$unset": {"password": ""},
                },
            )
    except Exception as exc:
        print(f"[ERROR] Failed migrating plaintext passwords: {exc}")


def ensure_user_theme_defaults():
    try:
        users.update_many(
            {
                "$or": [
                    {"theme": {"$exists": False}},
                    {"theme": ""},
                    {"theme": {"$nin": ["light", "dark"]}},
                ]
            },
            {"$set": {"theme": "light"}},
        )
    except Exception as exc:
        print(f"[ERROR] Failed ensuring user theme defaults: {exc}")


def ensure_user_profile_defaults():
    try:
        for doc in users.find({}):
            username = (doc.get("username") or "").strip()
            fallback_name = username or "User"
            fallback_email = build_internal_account_email(username) if username else ""
            updated = (doc.get("updatedAt") or doc.get("updated_at") or doc.get("created_at") or now_iso())

            patch = {}
            if not (doc.get("fullName") or "").strip():
                patch["fullName"] = fallback_name
            if not (doc.get("email") or "").strip():
                patch["email"] = fallback_email
            if "phone" not in doc:
                patch["phone"] = ""
            if "address" not in doc:
                patch["address"] = ""
            if "bio" not in doc:
                patch["bio"] = ""
            if "avatarUrl" not in doc:
                patch["avatarUrl"] = ""
            if "twoFactorEnabled" not in doc:
                patch["twoFactorEnabled"] = False
            elif not isinstance(doc.get("twoFactorEnabled"), bool):
                patch["twoFactorEnabled"] = bool(doc.get("twoFactorEnabled"))
            if "updatedAt" not in doc:
                patch["updatedAt"] = updated
            if patch:
                users.update_one({"_id": doc["_id"]}, {"$set": patch})
    except Exception as exc:
        print(f"[ERROR] Failed ensuring user profile defaults: {exc}")


def maybe_create_absence_alerts():
    today = now_local().date()
    # Always use the current active school year from database for absence alerts
    school_year_label = get_current_school_year_label()
    school_days = []
    cursor = today
    while len(school_days) < 7:
        if cursor.weekday() < 5:
            school_days.append(cursor.strftime("%Y-%m-%d"))
        cursor = cursor - timedelta(days=1)

    attendance_map = {}
    for row in attendance_logs.aggregate([
        {
            "$match": {
                "school_year": school_year_label,
                "date": {"$in": school_days},
                "student_id": {"$nin": ["", None]},
            }
        },
        {"$group": {"_id": "$student_id", "present_days": {"$addToSet": "$date"}}},
    ]):
        sid = (row.get("_id") or "").strip()
        if sid:
            attendance_map[sid] = set(row.get("present_days") or [])

    enrollment_collection = get_school_year_enrollment_collection(school_year_label)
    for student in enrollment_collection.find({"school_year": school_year_label}, {"student_id": 1, "name": 1}):
        sid = (student.get("student_id") or "").strip()
        if not sid:
            continue

        present_days = attendance_map.get(sid, set())

        absences = len(school_days) - len(present_days)
        if absences >= 3:
            alert_key = f"absence-{school_year_label}-{sid}-{today.isoformat()}"
            exists = alerts.count_documents({"school_year": school_year_label, "meta.alert_key": alert_key})
            if exists == 0:
                create_alert(
                    level="warning",
                    message=f"{student.get('name', sid)} reached {absences} absences in the last 7 school days.",
                    category="attendance",
                    meta={
                        "student_id": sid,
                        "alert_key": alert_key,
                        "absences": absences,
                        "school_year": school_year_label,
                    },
                )


def _active_student_query(student_id):
    return {
        "student_id": student_id,
        "$or": [
            {"status": "Active"},
            {"status": {"$exists": False}},
            {"status": ""},
        ],
    }


def log_attendance_and_sms(student, source="gate_scan", send_notifications=True, mode=None):
    now = now_local()
    # Always use the current active school year from database for gate scanning
    school_year_label = get_current_school_year_label()
    attendance_collection, school_year_label, _ = get_attendance_logs_storage(school_year_label)
    result = build_gate_scan_result(
        attendance_collection,
        school_year_label,
        student,
        now,
        source=source,
        mode=mode,
    )
    if not result:
        return None

    if result["duplicate"]:
        return {
            **result,
            "sms_status": "skipped",
        }

    should_debounce, debounce_until_ts = should_debounce_gate_scan_event(
        attendance_collection,
        result.get("student_id"),
        result.get("gate_action"),
        now,
    )
    if should_debounce:
        remaining_seconds = max(float(debounce_until_ts) - float(now.timestamp()), 0.0)
        action_phrase = "checked out" if normalize_gate_action_value(result.get("gate_action")) == "OUT" else "checked in"
        duplicate_result = build_gate_duplicate_result(
            result,
            f"Already {action_phrase} - wait {format_wait_time_short(remaining_seconds)} before scanning again.",
            "Please wait before scanning again",
            reason="event_debounce",
            cooldown_until_ts=debounce_until_ts,
        )
        return {
            **duplicate_result,
            "sms_status": "skipped",
        }

    attendance_doc = {
        "student_id": result["student_id"],
        "student_name": result["student_name"],
        "school_year": school_year_label,
        "status": result["status"],
        "session": result["session"],
        "source": source,
        "timestamp": result["timestamp"],
        "date": result["date"],
        "time": result["time"],
        "gate_action": result["gate_action"],
        "verification_label": result["verification_label"],
        "tracking_mode": result["tracking_mode"],
    }
    attendance_unique_filter = {
        "student_id": result["student_id"],
        "date": result["date"],
        "session": result["session"],
    }

    try:
        upsert_result = attendance_collection.update_one(
            attendance_unique_filter,
            {"$setOnInsert": attendance_doc},
            upsert=True,
        )
        inserted_new_record = bool(upsert_result.upserted_id)
    except DuplicateKeyError:
        inserted_new_record = False
    except Exception as exc:
        print(f"[ERROR] Failed to sync attendance log: {exc}")
        return None

    if not inserted_new_record:
        existing_record = attendance_collection.find_one(attendance_unique_filter) or attendance_doc
        return {
            **result,
            "status": existing_record.get("status", result["status"]),
            "sms_status": "skipped",
            "gate_action": existing_record.get("gate_action", result["gate_action"]),
            "verification_label": existing_record.get("verification_label", result["verification_label"]),
            "session": existing_record.get("session", result["session"]),
            "display_message": "Already recorded moments ago.",
            "voice_message": "Already recorded",
            "duplicate": True,
            "duplicate_reason": "duplicate_key",
            "feed_update": False,
            "activity_entry": None,
        }
    signal_data_change("gate_logs")

    sms_status = "skipped"
    parent_contact = result["parent_contact"]
    notifications_enabled = attendance_sms_notifications_enabled()

    if send_notifications and not notifications_enabled:
        log_skipped_sms(
            student_id=result["student_id"],
            student_name=result["student_name"],
            parent_contact=parent_contact,
            message=f"Attendance SMS notifications are disabled. SMS not sent for {result['student_name']}.",
            reason="notifications_disabled",
            sms_type="transactional",
            metadata={
                "context": "attendance_gate_scan",
                "session": result["session"],
                "school_year": school_year_label,
                "notifications_enabled": False,
            },
        )
        sms_status = "disabled"
    elif send_notifications and parent_contact:
        movement_text = "entered" if result["gate_action"] == "IN" else "exited"
        display_time = format_time_for_display(result.get("time"), result.get("timestamp"))
        template_payload = get_attendance_sms_template_payload()
        template_text = template_payload.get("template") or get_default_attendance_sms_template()
        template_variables = {
            "student_name": result["student_name"],
            "student_id": result["student_id"],
            "movement_text": movement_text,
            "gate_action": result["gate_action"],
            "status": result["status"],
            "session": result["session"],
            "time": display_time,
            "date": result["date"],
        }
        try:
            msg_text = SmsProvider.render_template(template_text, template_variables)
        except Exception:
            msg_text = SmsProvider.render_template(get_default_attendance_sms_template(), template_variables)
        msg_text = normalize_sms_template_text(msg_text) or SmsProvider.render_template(
            get_default_attendance_sms_template(),
            template_variables,
        )

        def _send_sms_async():
            sms_result = send_sms(
                parent_contact,
                msg_text,
                sms_type="transactional",
                metadata={
                    "context": "attendance_gate_scan",
                    "session": result["session"],
                    "school_year": school_year_label,
                    "template_id": ATTENDANCE_SMS_TEMPLATE_DOC_ID,
                    "template_updated_at": template_payload.get("updated_at", ""),
                },
                student_id=result["student_id"],
                student_name=result["student_name"],
                parent_contact=parent_contact,
            )
            status_async = "sent" if sms_result.get("status") == "sent" else "failed"
            error_async = sms_result.get("error", "")

            if status_async == "failed":
                create_alert(
                    level="high",
                    message=f"Failed SMS notification for {result['student_name']}.",
                    category="sms",
                    meta={"student_id": result["student_id"], "error": error_async, "school_year": school_year_label},
                )

        import threading
        threading.Thread(target=_send_sms_async, daemon=True).start()
        sms_status = "queued"
    elif send_notifications:
        log_skipped_sms(
            student_id=result["student_id"],
            student_name=result["student_name"],
            parent_contact=parent_contact,
            message=f"No parent contact configured for {result['student_name']}. SMS not sent.",
            reason="missing_parent_contact",
            sms_type="transactional",
            metadata={"context": "attendance_gate_scan", "session": result["session"], "school_year": school_year_label},
        )

    return {
        **result,
        "sms_status": sms_status,
    }


def resolve_live_scan_repeat_hold_seconds(result, now_ts=None):
    if not isinstance(result, dict):
        return 0.0

    current_ts = float(now_ts if now_ts is not None else time.time())
    duplicate_reason = str(result.get("duplicate_reason") or "").strip().lower()
    next_allowed_scan_ts = float(result.get("next_allowed_scan_ts") or 0.0)
    remaining_seconds = max(next_allowed_scan_ts - current_ts, 0.0)

    if not result.get("duplicate"):
        return remaining_seconds
    if duplicate_reason in {"out_cooldown", "event_debounce"}:
        return remaining_seconds
    if duplicate_reason in {"already_out", "out_requires_in"}:
        return 2.0
    return float(SCAN_COOLDOWN_SECONDS)


def handle_verified_student(student, confidence=0.0):
    now_ts = time.time()
    student_id = (student.get("student_id") or "").strip()
    student_name = (student.get("name") or "").strip()
    if not student_id or not student_name:
        return None

    current_mode = get_scan_session_mode()
    if should_suppress_recent_live_scan(student_id, now_ts, mode=current_mode):
        return None

    result = log_attendance_and_sms(student, mode=current_mode)
    if not result:
        return None

    repeat_hold_seconds = resolve_live_scan_repeat_hold_seconds(result, now_ts=now_ts)
    result_gate_action = normalize_gate_action_value(result.get("gate_action"))

    with scan_lock:
        last_scanned[student_id] = {
            "until_ts": now_ts + max(repeat_hold_seconds, 0.0),
            "gate_action": result_gate_action,
            "mode": normalize_scan_session_mode(current_mode, default="auto"),
        }

    if result["duplicate"]:
        return None

    mark_persistent_face_scan(student_id, now_ts, gate_action=result_gate_action)

    result_session = str(result.get("session") or "")
    push_scan_event("verified", {
        "student_id": student_id,
        "name": student_name,
        "verified": True,
        "attendance_status": result["status"],
        "sms_status": result["sms_status"],
        "gate_action": result["gate_action"],
        "verification_label": result["verification_label"],
        "session": result_session,
        "display_message": result["display_message"],
        "voice_message": result["voice_message"],
        "voice_key": f"{student_id}:{result['gate_action']}:{result['timestamp']}",
        "confidence": confidence,
        "confidence_pct": confidence,
        "duplicate": False,
        "duplicate_reason": "",
        "time": format_time_for_display(result.get("time"), result.get("timestamp")),
        "timestamp_display": format_timestamp_for_display(result.get("timestamp"), result.get("time")),
        "feed_update": bool(result.get("feed_update")),
        "activity_entry": result.get("activity_entry"),
        "tracking_mode": result.get("tracking_mode", get_scan_session_mode()),
    })
    return result


def push_not_registered_event(reason="no_match", confidence=0.0):
    now_ts = time.time()
    with scan_lock:
        if now_ts - scan_state["last_not_registered_ts"] < UNREGISTERED_EVENT_COOLDOWN_SECONDS:
            return
        scan_state["last_not_registered_ts"] = now_ts
    push_scan_event("not_registered", {
        "verified": False,
        "message": "Not Registered!",
        "reason": reason,
        "confidence": confidence,
        "confidence_pct": confidence,
    })


def push_multi_face_event(face_count):
    now_ts = time.time()
    with scan_lock:
        if now_ts - scan_state["last_multi_face_ts"] < UNREGISTERED_EVENT_COOLDOWN_SECONDS:
            return
        scan_state["last_multi_face_ts"] = now_ts
    push_scan_event("scan_warning", {
        "verified": False,
        "message": "Multiple faces detected. Please scan one person at a time.",
        "reason": "multiple_faces",
        "face_count": int(face_count or 0),
    })


def safe_capture_set(capture, prop_name, value):
    prop = getattr(cv2, prop_name, None)
    if prop is None:
        return
    try:
        capture.set(prop, value)
    except Exception:
        pass


def configure_capture_device(capture):
    safe_capture_set(capture, "CAP_PROP_BUFFERSIZE", 1)
    safe_capture_set(capture, "CAP_PROP_FRAME_WIDTH", SCAN_FRAME_WIDTH)
    safe_capture_set(capture, "CAP_PROP_FRAME_HEIGHT", SCAN_FRAME_HEIGHT)
    safe_capture_set(capture, "CAP_PROP_FPS", SCAN_TARGET_FPS)
    if hasattr(cv2, "VideoWriter_fourcc"):
        try:
            mjpg = cv2.VideoWriter_fourcc(*"MJPG")
            safe_capture_set(capture, "CAP_PROP_FOURCC", mjpg)
        except Exception:
            pass


def open_capture_device():
    if os.name == "nt" and hasattr(cv2, "CAP_DSHOW"):
        return cv2.VideoCapture(0, cv2.CAP_DSHOW)
    return cv2.VideoCapture(0)


def _refresh_face_index_worker():
    started_at = time.time()
    encoding_matrix = np.empty((0, 128), dtype=np.float64)
    db_students = []
    model_status = "model_not_ready"
    try:
        db_encodings, db_students = load_face_index_from_db(
            allow_legacy_fallback=SCAN_FACE_INDEX_ALLOW_LEGACY_IMAGE_FALLBACK
        )
        if db_encodings is not None and len(db_encodings) > 0:
            encoding_matrix = np.asarray(db_encodings, dtype=np.float64)
            if encoding_matrix.ndim != 2:
                encoding_matrix = np.empty((0, 128), dtype=np.float64)
        else:
            encoding_matrix = np.empty((0, 128), dtype=np.float64)
        if len(encoding_matrix) > 0:
            model_status = "ready"
        else:
            legacy_only = count_legacy_face_only_students()
            if legacy_only > 0 and not SCAN_FACE_INDEX_ALLOW_LEGACY_IMAGE_FALLBACK:
                model_status = "legacy_faces_only"
            else:
                model_status = "no_registered_students"
        elapsed = time.time() - started_at
        total_samples = sum(int(row.get("encoding_count") or 0) for row in db_students)
        print(
            f"[INFO] Face index loaded: status={model_status}, "
            f"profiles={len(db_students)}, centroids={len(encoding_matrix)}, samples={total_samples}, "
            f"legacy_fallback={'on' if SCAN_FACE_INDEX_ALLOW_LEGACY_IMAGE_FALLBACK else 'off'}, "
            f"elapsed={elapsed:.2f}s"
        )
    except Exception as exc:
        print(f"[ERROR] Failed loading face index from MongoDB: {exc}")
        model_status = "model_not_ready"
    finally:
        with scan_lock:
            if not scan_state.get("active"):
                scan_state["face_index_loading"] = False
                return
            scan_state["known_encodings"] = encoding_matrix
            scan_state["known_students"] = db_students
            scan_state["model_status"] = model_status
            scan_state["face_index_loading"] = False


def refresh_face_index_async():
    with scan_lock:
        if scan_state.get("face_index_loading"):
            return
        scan_state["face_index_loading"] = True

    worker = threading.Thread(target=_refresh_face_index_worker, name="scan-face-index-loader", daemon=True)
    worker.start()


def start_scan_capture():
    """
    Start scan mode. In the new client-side camera approach, this initializes
    the face recognition system without requiring a server-side camera.
    Frames will be sent by client devices.
    """
    with scan_lock:
        if scan_state["active"]:
            return True, "Scan already running"

    with scan_lock:
        if scan_state["active"]:
            try:
                capture = scan_state.get("capture")
                if capture:
                    capture.release()
            except Exception:
                pass
            return True, "Scan already running"
        
        # In client-side frame mode, we don't need a server-side capture device
        # Just mark the system as active
        last_scanned.clear()
        scan_presence_locks.clear()
        scan_state["capture"] = None  # No server-side camera needed
        scan_state["active"] = True
        scan_state["events"] = []
        scan_state["event_counter"] = 0
        scan_state["known_encodings"] = np.empty((0, 128), dtype=np.float64)
        scan_state["known_students"] = []
        scan_state["model_status"] = "loading"
        scan_state["face_index_loading"] = False
        scan_state["last_not_registered_ts"] = 0.0
        scan_state["last_multi_face_ts"] = 0.0
        scan_state["pending_recognition"] = {}
        scan_state["face_tracks"] = {}
        scan_state["next_face_track_id"] = 1
        scan_state["face_track_cursor"] = 0
        scan_state["last_faces_payload"] = []

    refresh_face_index_async()
    return True, "Scan started (waiting for client frames)"


def stop_scan_capture():
    with scan_lock:
        scan_state["active"] = False
        scan_presence_locks.clear()
        capture = scan_state.get("capture")
        scan_state["capture"] = None
        scan_state["known_encodings"] = np.empty((0, 128), dtype=np.float64)
        scan_state["known_students"] = []
        scan_state["model_status"] = "idle"
        scan_state["face_index_loading"] = False
        scan_state["pending_recognition"] = {}
        scan_state["face_tracks"] = {}
        scan_state["next_face_track_id"] = 1
        scan_state["face_track_cursor"] = 0
        scan_state["last_faces_payload"] = []

    if capture is not None:
        try:
            capture.release()
        except Exception as exc:
            print(f"[WARNING] Failed to release capture cleanly: {exc}")


def generate_frames():
    """
    Legacy frame generator. With client-side camera, this function is not actively used
    for getting frames (which now come from client via /process_scan_frame). 
    This remains for backward compatibility and can provide default images if needed.
    """
    banner = ""
    banner_until = 0.0
    cached_overlays = []
    frame_counter = 0
    last_analysis_ts = 0.0
    recognition_interval = max(SCAN_RECOGNITION_INTERVAL_MS, 50) / 1000.0
    process_every_n = max(1, SCAN_PROCESS_EVERY_N_FRAMES)
    target_frame_interval = 1.0 / max(1, SCAN_TARGET_FPS)
    scale = min(max(SCAN_RECOGNITION_SCALE, 0.25), 1.0)
    flush_grabs = max(0, int(SCAN_CAPTURE_FLUSH_GRABS))
    target_w = max(320, int(SCAN_FRAME_WIDTH))
    target_h = max(240, int(SCAN_FRAME_HEIGHT))
    jpeg_quality_params = [int(cv2.IMWRITE_JPEG_QUALITY), int(SCAN_JPEG_QUALITY)]

    # Create a placeholder frame since we're using client-side cameras
    placeholder_frame = np.zeros((target_h, target_w, 3), dtype=np.uint8)
    cv2.rectangle(placeholder_frame, (0, 0), (target_w, target_h), (40, 40, 40), -1)
    cv2.putText(
        placeholder_frame, 
        "Waiting for client camera frames...", 
        (int(target_w * 0.1), int(target_h // 2)),
        cv2.FONT_HERSHEY_SIMPLEX, 
        0.6, 
        (150, 150, 150), 
        2
    )

    # Stream placeholder frames
    while True:
        frame_start = time.time()
        with scan_lock:
            active = scan_state["active"]

        if not active:
            break

        ret, buffer = cv2.imencode(".jpg", placeholder_frame, jpeg_quality_params)
        if not ret:
            continue

        yield (
            b"--frame\r\n"
            b"Content-Type: image/jpeg\r\n\r\n" + buffer.tobytes() + b"\r\n"
        )

        elapsed = time.time() - frame_start
        sleep_for = target_frame_interval - elapsed
        if sleep_for > 0:
            time.sleep(sleep_for)

    stop_scan_capture()


def compute_dashboard_data(args, school_year=""):
    school_year_label = resolve_selected_school_year(school_year or args.get("school_year", ""))
    enrollment_collection = get_school_year_enrollment_collection(school_year_label)
    attendance_collection, school_year_label, _ = get_attendance_logs_storage(school_year_label)
    sms_collection, school_year_label, _ = get_sms_logs_storage(school_year_label)
    alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)
    current_role_name = current_role()
    can_manage_alerts = has_permission("alerts_manage")
    is_full_admin = current_role_name == ROLE_FULL_ADMIN
    today_date = now_local().date()
    today = today_date.strftime("%Y-%m-%d")
    total_students = enrollment_collection.count_documents({"school_year": school_year_label})
    total_male_students = enrollment_collection.count_documents({
        "school_year": school_year_label,
        "$or": [{"gender": "Male"}, {"sex": "Male"}],
    })
    total_female_students = enrollment_collection.count_documents({
        "school_year": school_year_label,
        "$or": [{"gender": "Female"}, {"sex": "Female"}],
    })
    attendance_today_query = {"school_year": school_year_label, "date": today}
    present_today_ids = attendance_collection.distinct("student_id", attendance_today_query)
    present_today = len([sid for sid in present_today_ids if sid])
    sms_sent_today = sms_collection.count_documents({
        "school_year": school_year_label,
        "date": today,
        "status": sms_status_mongo_filter("sent"),
    })
    late_today = attendance_collection.count_documents({**attendance_today_query, "status": "Late"})
    session_counts = {"Morning In": 0, "Noon Out": 0, "Afternoon In": 0, "Afternoon Out": 0}
    for row in attendance_collection.aggregate([
        {"$match": attendance_today_query},
        {"$group": {"_id": "$session", "count": {"$sum": 1}}},
    ]):
        session_name = row.get("_id")
        if session_name in session_counts:
            session_counts[session_name] = row.get("count", 0)

    unread_alerts = 0
    alert_docs = []
    if can_manage_alerts:
        unread_alerts = alerts_collection.count_documents(unread_notifications_query(school_year_label))
        alert_docs = [
            normalize_notification_doc(doc)
            for doc in alerts_collection.find({"school_year": school_year_label}).sort([("timestamp", -1), ("created_at", -1)]).limit(25)
        ]

    users_list = []
    users_stats = {"total": 0, "full_admin": 0, "staff": 0}
    login_rows = []
    if is_full_admin:
        users_list = build_dashboard_users_list()
        users_stats = build_dashboard_user_stats(users_list)

        login_rows = list(login_history.find().sort("timestamp", -1).limit(20))
        for r in login_rows:
            r["_id"] = str(r["_id"])
            r["role"] = normalize_role_value(r.get("role"), ROLE_STAFF)
            r["timestamp"] = format_timestamp_for_display(r.get("timestamp"))

    q = args.get("q", "").strip()
    log_type = args.get("log_type", "all")
    status_filter = args.get("status", "")
    date_filter = args.get("date", "")
    class_filter = args.get("student_class", "").strip()
    q_regex = contains_regex_filter(q)
    class_regex = contains_regex_filter(class_filter)

    student_filters = []
    if q_regex:
        student_filters.append({
            "$or": [
                {"name": q_regex},
                {"student_id": q_regex},
            ]
        })
    if class_regex:
        student_filters.append({
            "$or": [
                {"grade_level": class_regex},
                {"grade": class_regex},
            ]
        })
    student_query = {"$and": student_filters} if student_filters else {}

    if student_query:
        student_query = {"$and": [{"school_year": school_year_label}, student_query]}
    else:
        student_query = {"school_year": school_year_label}
    students_result = [normalize_student_doc(s) for s in enrollment_collection.find(student_query).limit(15)]

    gate_query = {"school_year": school_year_label}
    if q_regex:
        gate_query["$or"] = [
            {"student_name": q_regex},
            {"student_id": q_regex},
        ]
    if date_filter:
        gate_query["date"] = date_filter
    if status_filter:
        gate_query["status"] = status_filter

    sms_query = {"school_year": school_year_label}
    if q_regex:
        sms_query["$or"] = [
            {"name": q_regex},
            {"student_id": q_regex},
        ]
    if date_filter:
        sms_query["date"] = date_filter
    if status_filter:
        sms_query["status"] = sms_status_mongo_filter(status_filter)

    gate_results = []
    sms_results = []
    if log_type in ("all", "gate"):
        gate_results = list(attendance_collection.find(gate_query).sort("timestamp", -1).limit(20))
        for g in gate_results:
            g["_id"] = str(g["_id"])
            g["time"] = format_time_for_display(g.get("time"), g.get("timestamp"))
            g["timestamp"] = format_timestamp_for_display(g.get("timestamp"), g.get("time"))
    if log_type in ("all", "sms"):
        sms_results = list(sms_collection.find(sms_query).sort("timestamp", -1).limit(20))
        for s in sms_results:
            s["_id"] = str(s["_id"])
            s["time"] = format_time_for_display(s.get("time"), s.get("timestamp"))
            s["timestamp"] = format_timestamp_for_display(s.get("timestamp"), s.get("time"))

    return {
        "selected_school_year": school_year_label,
        "total_students": total_students,
        "total_male_students": total_male_students,
        "total_female_students": total_female_students,
        "present_today": present_today,
        "sms_sent_today": sms_sent_today,
        "late_today": late_today,
        "session_counts": session_counts,
        "alerts_unread": unread_alerts,
        "alert_rows": alert_docs,
        "users_list": users_list,
        "users_stats": users_stats,
        "login_rows": login_rows,
        "search_students": students_result,
        "search_gate_logs": gate_results,
        "search_sms_logs": sms_results,
        "filters": {
            "q": q,
            "log_type": log_type,
            "status": status_filter,
            "date": date_filter,
            "student_class": class_filter,
        },
    }


ensure_default_admin_user()
migrate_user_roles()
migrate_plaintext_user_passwords()
ensure_user_theme_defaults()
ensure_user_profile_defaults()
ensure_sms_template_defaults()
ensure_attendance_sms_notification_defaults()


# =====================================
# ROUTES
# =====================================
@app.route("/")
def home():
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    reset_notice = ""
    if request.method == "GET" and request.args.get("reset") == "success":
        reset_notice = "Password updated successfully. You can now sign in with your new password."

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        remember_me = False
        client_ip = _safe_client_ip()

        if not username or not password:
            return render_template(
                "login.html",
                current_year=datetime.now().year,
                error="Username and password are required.",
                success=reset_notice,
                entered_username=username,
                remember_me=remember_me,
            )

        lockout_seconds = get_login_lockout_seconds(username, client_ip)
        if lockout_seconds > 0:
            wait_minutes = max(1, (lockout_seconds + 59) // 60)
            log_audit_event(
                action="auth.login_blocked",
                outcome="blocked",
                severity="warn",
                target_type="user",
                target_id=username,
                details={"reason": "lockout", "wait_minutes": wait_minutes},
            )
            return render_template(
                "login.html",
                current_year=datetime.now().year,
                error=f"Too many failed login attempts. Try again in about {wait_minutes} minute(s).",
                success=reset_notice,
                entered_username=username,
                remember_me=remember_me,
            )

        user = users.find_one({"username": username})
        if user:
            password_ok = False
            stored_hash = user.get("password_hash")
            if stored_hash:
                password_ok = check_password_hash(stored_hash, password)

            if password_ok:
                role = normalize_account_role(user.get("role"), username)
                session.clear()
                session.permanent = remember_me
                session["admin"] = username
                session["role"] = role
                session["theme"] = normalize_theme_value(user.get("theme"))
                record_login(username, role)
                clear_login_attempts(username, client_ip)
                log_audit_event(
                    action="auth.login_success",
                    outcome="success",
                    severity="info",
                    target_type="user",
                    target_id=username,
                    details={"role": role},
                )
                return redirect(post_login_redirect(role))

        register_failed_login_attempt(username, client_ip)
        log_audit_event(
            action="auth.login_failed",
            outcome="failed",
            severity="warn",
            target_type="user",
            target_id=username,
        )
        return render_template(
            "login.html",
            current_year=datetime.now().year,
            error="Invalid credentials.",
            success=reset_notice,
            entered_username=username,
            remember_me=remember_me,
        )

    return render_template(
        "login.html",
        current_year=datetime.now().year,
        success=reset_notice,
        entered_username="",
        remember_me=False,
    )


@app.route("/api/auth/forgot-password/request", methods=["POST"])
def forgot_password_request_api():
    payload = request.get_json(silent=True) or request.form.to_dict(flat=True)
    email = normalize_email_value(payload.get("email", ""))

    if not validate_email_format(email):
        return jsonify({"status": "error", "message": "Enter a valid registered email address.", "field": "email"}), 400

    generic_message = "Password reset link has been sent to your email."
    client_ip = get_request_client_ip()
    user = find_user_by_email(email)

    preview_link = ""
    allow_preview = PASSWORD_RESET_DEV_LINK_FALLBACK and os.getenv("FLASK_ENV", "production").strip().lower() != "production"
    config_error = smtp_configuration_error()
    if config_error and not allow_preview:
        return jsonify({"status": "error", "message": config_error}), 503

    if user and PASSWORD_RESET_RATE_LIMIT_ENABLED and password_reset_request_rate_limited(email, client_ip):
        return jsonify({
            "status": "error",
            "message": "Too many password reset requests. Please try again later.",
        }), 429

    if user:
        password_reset_tokens.update_many(
            {"user_id": user["_id"], "used": False},
            {"$set": {"used": True, "invalidatedAt": datetime.utcnow()}},
        )

        raw_token = secrets.token_urlsafe(32)
        reset_link = build_password_reset_link(raw_token)
        expires_at = datetime.utcnow() + timedelta(minutes=PASSWORD_RESET_TOKEN_TTL_MINUTES)
        password_reset_tokens.insert_one({
            "user_id": user["_id"],
            "email": email,
            "token_hash": hash_password_reset_token(raw_token),
            "used": False,
            "createdAt": datetime.utcnow(),
            "expiresAt": expires_at,
            "requestIp": client_ip,
            "usedAt": None,
        })

        sent, send_error = send_password_reset_email(email, reset_link)
        if allow_preview and not sent:
            preview_link = reset_link
        if not sent:
            password_reset_tokens.update_one(
                {"email": email, "token_hash": hash_password_reset_token(raw_token)},
                {"$set": {"used": True, "usedAt": datetime.utcnow(), "invalidReason": "email_send_failed"}},
            )
            if not allow_preview:
                return jsonify({
                    "status": "error",
                    "message": send_error or "Unable to send reset email right now. Please try again.",
                }), 502

    body = {"status": "ok", "message": generic_message}
    if preview_link:
        body["previewLink"] = preview_link
        body["previewNote"] = "Development preview link (email service not configured)."
    return jsonify(body)


@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    if request.method == "GET":
        token = (request.args.get("token") or "").strip()
        if not token:
            return render_template(
                "reset_password.html",
                current_year=datetime.now().year,
                token="",
                error="This password reset link is invalid.",
                success="",
            )

        record = get_password_reset_record(token)
        if not record:
            return render_template(
                "reset_password.html",
                current_year=datetime.now().year,
                token="",
                error="This password reset link is invalid or has expired.",
                success="",
            )

        return render_template(
            "reset_password.html",
            current_year=datetime.now().year,
            token=token,
            error="",
            success="",
        )

    token = (request.form.get("token") or "").strip()
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")

    record = get_password_reset_record(token)
    if not record:
        return render_template(
            "reset_password.html",
            current_year=datetime.now().year,
            token="",
            error="This password reset link is invalid or has expired.",
            success="",
        )

    validation_message, validation_field = validate_password_reset_input(new_password, confirm_password)
    if validation_message:
        field_label = "New password" if validation_field == "newPassword" else "Password confirmation"
        return render_template(
            "reset_password.html",
            current_year=datetime.now().year,
            token=token,
            error=f"{field_label}: {validation_message}",
            success="",
        )

    user_doc = users.find_one({"_id": record.get("user_id")})
    if not user_doc:
        password_reset_tokens.update_one(
            {"_id": record["_id"]},
            {"$set": {"used": True, "usedAt": datetime.utcnow(), "invalidReason": "missing_user"}},
        )
        return render_template(
            "reset_password.html",
            current_year=datetime.now().year,
            token="",
            error="Unable to reset password for this account.",
            success="",
        )

    stored_hash = user_doc.get("password_hash")
    if stored_hash and check_password_hash(stored_hash, new_password):
        return render_template(
            "reset_password.html",
            current_year=datetime.now().year,
            token=token,
            error="New password must be different from the current password.",
            success="",
        )

    updated = now_iso()
    users.update_one(
        {"_id": user_doc["_id"]},
        {
            "$set": {
                "password_hash": hash_password(new_password),
                "updatedAt": updated,
                "updated_at": updated,
            },
            "$unset": {"password": ""},
        },
    )
    password_reset_tokens.update_one(
        {"_id": record["_id"]},
        {"$set": {"used": True, "usedAt": datetime.utcnow()}},
    )
    create_alert("info", f"Password reset completed for user '{user_doc.get('username', 'unknown')}'.", "security")
    return redirect(url_for("login", reset="success"))


@app.route("/logout", methods=["POST"])
def logout():
    log_audit_event(
        action="auth.logout",
        outcome="success",
        severity="info",
        target_type="user",
        target_id=session.get("admin", ""),
    )
    session.clear()
    stop_scan_capture()
    return redirect(url_for("login"))


LIVE_MONITORING_DEFAULT_CHANNEL = os.getenv("LIVE_MONITORING_CHANNEL", "main-gate").strip() or "main-gate"
LIVE_MONITORING_SIGNALING_PORT = env_int("LIVE_MONITORING_SIGNALING_PORT", FLASK_PORT + 1, minimum=1, maximum=65535)
LIVE_MONITORING_TOKEN_TTL_SECONDS = env_int("LIVE_MONITORING_TOKEN_TTL_SECONDS", 120, minimum=30, maximum=900)
LIVE_MONITORING_MANAGED = env_bool(
    "LIVE_MONITORING_MANAGED",
    not bool(os.getenv("LIVE_MONITORING_SIGNALING_URL", "").strip()),
)
LIVE_MONITORING_AUTO_START = env_bool("LIVE_MONITORING_AUTO_START", LIVE_MONITORING_MANAGED)
LIVE_MONITORING_STARTUP_TIMEOUT_SECONDS = env_float(
    "LIVE_MONITORING_STARTUP_TIMEOUT_SECONDS",
    6.0,
    minimum=1.0,
    maximum=30.0,
)
LIVE_MONITORING_NODE_BIN = os.getenv("LIVE_MONITORING_NODE_BIN", "node").strip() or "node"
LIVE_MONITORING_HOST = os.getenv("LIVE_MONITORING_HOST", "0.0.0.0").strip() or "0.0.0.0"
LIVE_MONITORING_SERVER_SCRIPT = os.path.join(BASE_DIR, "live_monitor_server.js")
live_monitoring_process_lock = threading.Lock()
live_monitoring_process = None
live_monitoring_process_started_by_app = False


def live_monitoring_secret_bytes():
    secret = (
        os.getenv("LIVE_MONITORING_TOKEN_SECRET", "").strip()
        or str(app.secret_key or "").strip()
        or "live-monitoring-secret"
    )
    return secret.encode("utf-8")


def resolve_live_monitoring_health_hosts():
    candidates = []
    host = LIVE_MONITORING_HOST.strip()
    if host and host not in {"0.0.0.0", "::", "[::]"}:
        candidates.append(host.strip("[]"))

    configured_url = os.getenv("LIVE_MONITORING_SIGNALING_URL", "").strip()
    if configured_url:
        try:
            configured_host = (urlparse(configured_url).hostname or "").strip()
            if configured_host:
                candidates.append(configured_host)
        except Exception:
            pass

    candidates.extend(["127.0.0.1", "localhost"])

    unique_hosts = []
    seen = set()
    for candidate in candidates:
        normalized = str(candidate or "").strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        unique_hosts.append(normalized)
    return unique_hosts


def is_live_monitoring_service_available(timeout=0.35):
    timeout_seconds = max(float(timeout or 0.0), 0.1)
    for host in resolve_live_monitoring_health_hosts():
        try:
            with socket.create_connection((host, LIVE_MONITORING_SIGNALING_PORT), timeout=timeout_seconds):
                return True
        except OSError:
            continue
    return False


def tracked_live_monitoring_process_running():
    return bool(live_monitoring_process is not None and live_monitoring_process.poll() is None)


def resolve_live_monitoring_start_command():
    node_binary = shutil.which(LIVE_MONITORING_NODE_BIN)
    if not node_binary:
        raise RuntimeError(
            f"Live monitoring requires Node.js. Could not find '{LIVE_MONITORING_NODE_BIN}' in PATH."
        )
    if not os.path.exists(LIVE_MONITORING_SERVER_SCRIPT):
        raise RuntimeError(
            f"Live monitoring server script was not found at '{LIVE_MONITORING_SERVER_SCRIPT}'."
        )
    return [node_binary, LIVE_MONITORING_SERVER_SCRIPT]


def wait_for_live_monitoring_service(timeout_seconds=None):
    deadline = time.time() + max(
        float(timeout_seconds if timeout_seconds is not None else LIVE_MONITORING_STARTUP_TIMEOUT_SECONDS),
        0.5,
    )
    last_error = "Live monitoring service did not become ready in time."

    while time.time() < deadline:
        if is_live_monitoring_service_available(timeout=0.25):
            return True, "Live monitoring service is ready."

        process = live_monitoring_process
        if process is not None:
            exit_code = process.poll()
            if exit_code is not None:
                return False, f"Live monitoring service exited with code {exit_code}."

        time.sleep(0.15)

    return False, last_error


def start_live_monitoring_service(reason="runtime"):
    global live_monitoring_process, live_monitoring_process_started_by_app

    if is_live_monitoring_service_available():
        return True, "Live monitoring service already available."

    if tracked_live_monitoring_process_running():
        return wait_for_live_monitoring_service()

    try:
        command = resolve_live_monitoring_start_command()
    except Exception as exc:
        return False, str(exc)

    creation_flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)

    try:
        live_monitoring_process = subprocess.Popen(
            command,
            cwd=BASE_DIR,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            env=os.environ.copy(),
            creationflags=creation_flags,
        )
        live_monitoring_process_started_by_app = True
        print(
            f"[INFO] Starting live monitoring service via Flask bootstrap "
            f"(reason={reason}, port={LIVE_MONITORING_SIGNALING_PORT})"
        )
    except Exception as exc:
        live_monitoring_process = None
        live_monitoring_process_started_by_app = False
        return False, f"Unable to start live monitoring service: {exc}"

    return wait_for_live_monitoring_service()


def ensure_live_monitoring_service_ready(reason="runtime", fail_hard=False):
    if not LIVE_MONITORING_MANAGED:
        return True, "Live monitoring service is managed externally."

    with live_monitoring_process_lock:
        if is_live_monitoring_service_available():
            return True, "Live monitoring service already available."

        if tracked_live_monitoring_process_running():
            ok, message = wait_for_live_monitoring_service()
        else:
            ok, message = start_live_monitoring_service(reason=reason)

    if not ok:
        print(f"[WARNING] {message}")
        if fail_hard:
            return False, message
    return ok, message


def stop_managed_live_monitoring_service():
    global live_monitoring_process, live_monitoring_process_started_by_app

    with live_monitoring_process_lock:
        process = live_monitoring_process
        started_by_app = live_monitoring_process_started_by_app
        live_monitoring_process = None
        live_monitoring_process_started_by_app = False

    if not started_by_app or process is None:
        return

    if process.poll() is not None:
        return

    try:
        process.terminate()
        process.wait(timeout=3)
    except Exception:
        try:
            process.kill()
        except Exception:
            pass


def should_bootstrap_live_monitoring_on_startup():
    if not LIVE_MONITORING_AUTO_START:
        return False

    if not LIVE_MONITORING_MANAGED:
        return False

    if FLASK_DEBUG_MODE or DEV_AUTO_RELOAD:
        return os.environ.get("WERKZEUG_RUN_MAIN") == "true"

    return True


atexit.register(stop_managed_live_monitoring_service)


def resolve_live_monitoring_signaling_url():
    configured = os.getenv("LIVE_MONITORING_SIGNALING_URL", "").strip()
    if configured:
        return configured.rstrip("/")

    host = request.host.split(":", 1)[0].strip() or request.host.strip()
    scheme = request.scheme or ("https" if HTTPS_ENABLED else "http")
    return f"{scheme}://{host}:{LIVE_MONITORING_SIGNALING_PORT}"


def resolve_live_monitoring_ice_servers():
    raw_urls = os.getenv(
        "LIVE_MONITORING_STUN_URLS",
        "stun:stun.l.google.com:19302,stun:stun1.l.google.com:19302",
    )
    urls = [value.strip() for value in raw_urls.split(",") if value.strip()]
    return [{"urls": urls}] if urls else []


def create_live_monitoring_token(payload):
    import hmac

    payload_json = json.dumps(payload, separators=(",", ":"), sort_keys=True).encode("utf-8")
    payload_b64 = base64.urlsafe_b64encode(payload_json).decode("ascii").rstrip("=")
    signature = hmac.new(
        live_monitoring_secret_bytes(),
        payload_b64.encode("ascii"),
        hashlib.sha256,
    ).digest()
    signature_b64 = base64.urlsafe_b64encode(signature).decode("ascii").rstrip("=")
    return f"{payload_b64}.{signature_b64}"


@app.route("/api/live-monitoring/token")
@require_permission("dashboard")
def live_monitoring_token():
    role_name = current_role()
    if role_name not in {ROLE_FULL_ADMIN, ROLE_STAFF}:
        return jsonify({"status": "error", "message": "Live monitoring is unavailable for this role."}), 403

    live_monitoring_ready, live_monitoring_message = ensure_live_monitoring_service_ready(
        reason="token_request",
        fail_hard=True,
    )
    if not live_monitoring_ready:
        return jsonify({
            "status": "error",
            "message": live_monitoring_message,
        }), 503

    issued_at = int(time.time())
    expires_at = issued_at + LIVE_MONITORING_TOKEN_TTL_SECONDS
    username = (session.get("admin") or "").strip()
    channel = LIVE_MONITORING_DEFAULT_CHANNEL
    token = create_live_monitoring_token({
        "sub": username,
        "role": role_name,
        "channel": channel,
        "iat": issued_at,
        "exp": expires_at,
    })

    return jsonify({
        "status": "ok",
        "token": token,
        "role": role_name,
        "channel": channel,
        "signaling_url": resolve_live_monitoring_signaling_url(),
        "ice_servers": resolve_live_monitoring_ice_servers(),
        "expires_at": expires_at,
    })


@app.route("/dashboard")
@require_permission("dashboard")
def dashboard():
    maybe_create_absence_alerts()
    selected_school_year = resolve_selected_school_year(request.args.get("school_year", ""))
    payload = compute_dashboard_data(request.args, selected_school_year)
    payload.update(sidebar_context("dashboard", selected_school_year))
    payload["message"] = request.args.get("message", "").strip()
    payload["message_type"] = request.args.get("message_type", "success").strip() or "success"
    # Inject today's active calendar event so the template can show a banner
    today_active = get_active_schedule(now_local())
    payload["today_event_type"] = today_active.get("type", "regular")
    payload["today_special_condition"] = today_active.get("special_condition") or ""
    today_date_str = now_local().strftime("%Y-%m-%d")
    today_ev_doc = calendar_events.find_one({"date": today_date_str})
    payload["today_event_title"] = today_ev_doc.get("title", "") if today_ev_doc else ""
    return render_template("dashboard.html", **payload)


@app.route("/live-gate-monitoring")
@require_permission("dashboard")
def live_monitoring_page():
    if current_role() != ROLE_FULL_ADMIN:
        return redirect(url_for("dashboard"))

    selected_school_year = resolve_selected_school_year(request.args.get("school_year", ""))
    payload = sidebar_context("live_monitoring", selected_school_year)
    return render_template("live_monitoring.html", **payload)


@app.route("/test_enhanced_scanning")
@require_permission("scan")
def test_enhanced_scanning():
    """Enhanced scanning test page"""
    return render_template("test_enhanced_scanning.html")


@app.route("/simple_demo")
def simple_demo():
    """Simple demo page that shows green boxes without server dependencies"""
    return render_template("simple_demo.html")


@app.route("/developers")
@require_permission("dashboard")
def developers_page():
    developers = [
        {
            "name": "CORDOVA, APRIL BRYAN C.",
            "role": "Full Stack Developer",
            "contribution": "Worked across frontend and backend modules, integrating core system workflows and feature delivery.",
            "email": "aprilbryancordova@gmail.com",
            "profile_photo": url_for("static", filename="developer_photos/real/cordova-april-bryan.jpg"),
            "fallback_photo": url_for("static", filename="developer_photos/dev-cordova.svg"),
            "links": [{"label": "Email", "url": "mailto:aprilbryancordova@gmail.com"}],
        },
        {
            "name": "PILONGO, RON ALLEN R.",
            "role": "Backend Developer",
            "contribution": "Implemented and maintained API services, database transactions, and backend integration for key modules.",
            "email": "pilongoronallen@gmail.com",
            "profile_photo": url_for("static", filename="developer_photos/real/pilongo-ron-allen.png"),
            "fallback_photo": url_for("static", filename="developer_photos/dev-pilongo.svg"),
            "links": [{"label": "Email", "url": "mailto:pilongoronallen@gmail.com"}],
        },
        {
            "name": "ZAMORA, ANGEL V.",
            "role": "System Developer",
            "contribution": "Supported system design, architecture alignment, and module-level technical implementation.",
            "email": "angelvillanueva0456@gmail.com",
            "profile_photo": url_for("static", filename="developer_photos/real/zamora-angel.jpg"),
            "fallback_photo": url_for("static", filename="developer_photos/dev-zamora.svg"),
            "links": [{"label": "Email", "url": "mailto:angelvillanueva0456@gmail.com"}],
        },
        {
            "name": "ANLAP, GIAN EUGENE R.",
            "role": "Project Contributor",
            "contribution": "Contributed to system implementation, testing support, and project development tasks.",
            "email": "gianeugene.anlap@chs-gate.local",
            "profile_photo": url_for("static", filename="developer_photos/real/anlap-gian-eugene.jpg"),
            "fallback_photo": url_for("static", filename="developer_photos/dev-anlap.svg"),
            "links": [{"label": "Email", "url": "mailto:gianeugene.anlap@chs-gate.local"}],
        },
        {
            "name": "RAMIREZ, ELMER D.",
            "role": "Project Contributor",
            "contribution": "Contributed to system implementation, testing support, and project development tasks.",
            "email": "ramirezelmer113003@gmail.com",
            "profile_photo": url_for("static", filename="developer_photos/real/ramirez-elmer.jpg"),
            "fallback_photo": url_for("static", filename="developer_photos/dev-ramirez.svg"),
            "links": [{"label": "Email", "url": "mailto:ramirezelmer113003@gmail.com"}],
        },
        {
            "name": "CLANZA, ROSME A.",
            "role": "Project Contributor",
            "contribution": "Contributed to system implementation, testing support, and project development tasks.",
            "email": "clanzarosme97@gmail.com",
            "profile_photo": url_for("static", filename="developer_photos/real/clanza-rosme.jpg"),
            "fallback_photo": url_for("static", filename="developer_photos/dev-clanza.svg"),
            "links": [{"label": "Email", "url": "mailto:clanzarosme97@gmail.com"}],
        },
        {
            "name": "GELLA, BRENUS C.",
            "role": "Project Contributor",
            "contribution": "Contributed to project implementation and supported interface and usability refinements for the platform.",
            "email": "brenusgella09@gmail.com",
            "profile_photo": url_for("static", filename="developer_photos/real/gella-brenus.jpg"),
            "fallback_photo": url_for("static", filename="developer_photos/dev-gella.svg"),
            "links": [{"label": "Email", "url": "mailto:brenusgella09@gmail.com"}],
        },
    ]

    return render_template(
        "developers.html",
        developers=developers,
        system_info={
            "version": os.getenv("SYSTEM_VERSION", "v1.0.0"),
            "last_update": os.getenv("SYSTEM_LAST_UPDATE", now_local().strftime("%B %d, %Y")),
            "environment": os.getenv("FLASK_ENV", "production").strip() or "production",
        },
        **sidebar_context("developers"),
    )


@app.route("/admin/archive-summary")
@require_permission("users_manage")
def admin_archive_summary_page():
    selected_school_year = resolve_selected_school_year(request.args.get("school_year", ""))
    payload = build_archive_summary_payload(selected_school_year)
    return render_template(
        "archive_summary.html",
        **payload,
        **sidebar_context("archive_summary", selected_school_year),
    )


@app.route("/api/dashboard/stats")
@require_permission("dashboard", api=True)
def dashboard_stats_api():
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    today = now_local().strftime("%Y-%m-%d")
    enrollment_collection = get_school_year_enrollment_collection(school_year_label)
    attendance_collection, school_year_label, _ = get_attendance_logs_storage(school_year_label)
    sms_collection, school_year_label, _ = get_sms_logs_storage(school_year_label)
    total_students = enrollment_collection.count_documents({"school_year": school_year_label})
    total_male_students = enrollment_collection.count_documents({
        "school_year": school_year_label,
        "$or": [{"gender": "Male"}, {"sex": "Male"}],
    })
    total_female_students = enrollment_collection.count_documents({
        "school_year": school_year_label,
        "$or": [{"gender": "Female"}, {"sex": "Female"}],
    })
    present_today_ids = attendance_collection.distinct("student_id", {"school_year": school_year_label, "date": today})
    present_today = len([sid for sid in present_today_ids if sid])
    sms_sent_today = sms_collection.count_documents({
        "school_year": school_year_label,
        "date": today,
        "status": sms_status_mongo_filter("sent"),
    })

    return jsonify({
        "status": "ok",
        "school_year": school_year_label,
        "total_students": total_students,
        "present_today": present_today,
        "sms_sent_today": sms_sent_today,
        "total_male_students": total_male_students,
        "total_female_students": total_female_students,
    })


@app.route("/api/system/health", methods=["GET"])
@require_permission("dashboard", api=True)
def system_health_api():
    try:
        return jsonify({"status": "ok", "health": build_system_health_snapshot()})
    except Exception as exc:
        return jsonify({"status": "error", "message": f"Failed to build system health snapshot: {exc}"}), 500


@app.route("/api/audit/recent", methods=["GET"])
@require_permission("users_manage", api=True)
def audit_recent_api():
    try:
        limit_value = int(request.args.get("limit", "30"))
    except (TypeError, ValueError):
        limit_value = 30
    limit_value = max(1, min(limit_value, 100))

    docs = list(audit_logs.find().sort("createdAt", -1).limit(limit_value))
    rows = []
    for doc in docs:
        rows.append({
            "_id": str(doc.get("_id")),
            "action": str(doc.get("action") or ""),
            "outcome": str(doc.get("outcome") or ""),
            "severity": str(doc.get("severity") or ""),
            "target_type": str(doc.get("target_type") or ""),
            "target_id": str(doc.get("target_id") or ""),
            "actor": str((doc.get("actor") or {}).get("username") or ""),
            "role": str((doc.get("actor") or {}).get("role") or ""),
            "ip": str(doc.get("ip") or ""),
            "created_at": str(doc.get("created_at") or ""),
            "details": doc.get("details") or {},
        })
    return jsonify({"status": "ok", "rows": rows})


@app.route("/api/admin/users/roster", methods=["GET"])
@require_permission("users_manage", api=True)
def admin_users_roster_api():
    if current_role() != ROLE_FULL_ADMIN:
        return jsonify({"status": "error", "message": "Only Full Admin can view the identities roster."}), 403

    users_list = build_dashboard_users_list()
    return jsonify({
        "status": "ok",
        "users": users_list,
        "stats": build_dashboard_user_stats(users_list),
    })


@app.route("/api/profile", methods=["GET"])
@require_permission("dashboard", api=True)
def profile_get_api():
    user_doc, profile = current_user_profile()
    if not user_doc or not profile:
        return jsonify({"status": "error", "message": "User session is invalid."}), 401
    return jsonify({"status": "ok", "profile": profile})


@app.route("/api/profile", methods=["PUT"])
@require_permission("dashboard", api=True)
def profile_update_api():
    user_doc, profile = current_user_profile()
    if not user_doc or not profile:
        return jsonify({"status": "error", "message": "User session is invalid."}), 401
    if current_role() == ROLE_STAFF:
        return jsonify({"status": "error", "message": "Staff profile details cannot be edited from live recognition mode."}), 403

    payload = request.get_json(silent=True) or {}
    full_name = sanitize_profile_text(payload.get("fullName"), 120)
    email = sanitize_profile_text(payload.get("email"), 160).lower()
    phone = sanitize_profile_text(payload.get("phone"), 32)
    address = sanitize_profile_text(payload.get("address"), 240)
    bio = sanitize_profile_text(payload.get("bio"), 800, allow_newlines=True)
    two_factor_value = payload.get("twoFactorEnabled", profile.get("twoFactorEnabled"))
    remove_avatar = payload.get("removeAvatar", False)

    if not full_name:
        return jsonify({"status": "error", "message": "Full Name is required.", "field": "fullName"}), 400
    if not email:
        return jsonify({"status": "error", "message": "Email is required.", "field": "email"}), 400
    if not validate_email_format(email):
        return jsonify({"status": "error", "message": "Invalid email format.", "field": "email"}), 400
    if not validate_phone_format(phone):
        return jsonify({"status": "error", "message": "Invalid phone number format.", "field": "phone"}), 400
    if not isinstance(two_factor_value, bool):
        return jsonify({"status": "error", "message": "Two-factor setting must be true or false.", "field": "twoFactorEnabled"}), 400
    if not isinstance(remove_avatar, bool):
        return jsonify({"status": "error", "message": "Avatar remove flag must be true or false.", "field": "removeAvatar"}), 400

    duplicate_email_user = users.find_one({
        "email": email,
        "_id": {"$ne": user_doc["_id"]},
    })
    if duplicate_email_user:
        return jsonify({"status": "error", "message": "Email is already in use.", "field": "email"}), 400

    updated = now_iso()
    update_payload = {
        "fullName": full_name,
        "email": email[:160],
        "phone": phone,
        "address": address,
        "bio": bio,
        "twoFactorEnabled": two_factor_value,
        "updatedAt": updated,
        "updated_at": updated,
    }
    if remove_avatar:
        old_avatar = (user_doc.get("avatarUrl") or "").strip()
        if old_avatar.startswith("/static/avatars/"):
            old_name = old_avatar.split("/static/avatars/", 1)[1]
            old_path = os.path.join(AVATAR_UPLOAD_DIR, old_name)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except Exception:
                    pass
        update_payload["avatarUrl"] = ""

    users.update_one({"_id": user_doc["_id"]}, {"$set": update_payload})
    signal_data_change("users")
    refreshed = users.find_one({"_id": user_doc["_id"]})
    return jsonify({"status": "ok", "profile": normalize_profile_user_doc(refreshed)})


@app.route("/api/profile/theme", methods=["GET"])
@require_permission("dashboard", api=True)
def profile_theme_get_api():
    _user_doc, profile = current_user_profile()
    if not profile:
        return jsonify({"status": "error", "message": "User session is invalid."}), 401
    return jsonify({"status": "ok", "theme": normalize_theme_value(profile.get("theme"))})


@app.route("/api/profile/theme", methods=["PUT"])
@require_permission("dashboard", api=True)
def profile_theme_update_api():
    user_doc, profile = current_user_profile()
    if not user_doc or not profile:
        return jsonify({"status": "error", "message": "User session is invalid."}), 401

    payload = request.get_json(silent=True) or {}
    theme = normalize_theme_value(payload.get("theme"), default="")
    if theme not in ("light", "dark"):
        return jsonify({"status": "error", "message": "Theme must be 'light' or 'dark'."}), 400

    updated = now_iso()
    users.update_one({"_id": user_doc["_id"]}, {"$set": {"theme": theme, "updatedAt": updated, "updated_at": updated}})
    session["theme"] = theme
    return jsonify({"status": "ok", "theme": theme})


@app.route("/api/profile/photo", methods=["POST"])
@require_permission("dashboard", api=True)
def profile_photo_upload_api():
    user_doc, profile = current_user_profile()
    if not user_doc or not profile:
        return jsonify({"status": "error", "message": "User session is invalid."}), 401
    if current_role() == ROLE_STAFF:
        return jsonify({"status": "error", "message": "Staff profile photos cannot be edited from live recognition mode.", "field": "avatar"}), 403

    file = request.files.get("avatar")
    if not file or not file.filename:
        return jsonify({"status": "error", "message": "No image selected.", "field": "avatar"}), 400

    filename = secure_filename(file.filename)
    if "." not in filename:
        return jsonify({"status": "error", "message": "Invalid file format.", "field": "avatar"}), 400
    ext = filename.rsplit(".", 1)[1].lower()
    if ext not in ALLOWED_AVATAR_EXTENSIONS:
        return jsonify({"status": "error", "message": "Only JPG, PNG, and WEBP are allowed.", "field": "avatar"}), 400

    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > MAX_AVATAR_SIZE_BYTES:
        return jsonify({"status": "error", "message": "Image size exceeds 5MB limit.", "field": "avatar"}), 400

    # Verify the uploaded payload is a real image, not only extension-based.
    try:
        image_bytes = file.read()
        if not image_bytes:
            return jsonify({"status": "error", "message": "Uploaded image is empty.", "field": "avatar"}), 400
        with Image.open(BytesIO(image_bytes)) as img:
            img.verify()
        file.seek(0)
    except Exception:
        return jsonify({"status": "error", "message": "Uploaded file is not a valid image.", "field": "avatar"}), 400

    unique_name = f"{user_doc.get('username', 'user')}_{uuid.uuid4().hex[:12]}.{ext}"
    save_path = os.path.join(AVATAR_UPLOAD_DIR, unique_name)

    try:
        file.save(save_path)
    except Exception:
        return jsonify({"status": "error", "message": "Failed to save uploaded image.", "field": "avatar"}), 500

    old_avatar = (user_doc.get("avatarUrl") or "").strip()
    if old_avatar.startswith("/static/avatars/"):
        old_name = old_avatar.split("/static/avatars/", 1)[1]
        old_path = os.path.join(AVATAR_UPLOAD_DIR, old_name)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except Exception:
                pass

    avatar_url = f"/static/avatars/{unique_name}"
    updated = now_iso()
    users.update_one(
        {"_id": user_doc["_id"]},
        {"$set": {"avatarUrl": avatar_url, "updatedAt": updated, "updated_at": updated}},
    )
    signal_data_change("users")
    refreshed = users.find_one({"_id": user_doc["_id"]})

    return jsonify({
        "status": "ok",
        "avatarUrl": avatar_url,
        "profile": normalize_profile_user_doc(refreshed),
    })


@app.route("/api/profile/password", methods=["PUT"])
@require_permission("dashboard", api=True)
def profile_password_update_api():
    user_doc, profile = current_user_profile()
    if not user_doc or not profile:
        return jsonify({"status": "error", "message": "User session is invalid."}), 401

    payload = request.get_json(silent=True) or {}
    current_password = str(payload.get("currentPassword") or "")
    new_password = str(payload.get("newPassword") or "")
    confirm_password = str(payload.get("confirmPassword") or "")

    if not current_password:
        return jsonify({"status": "error", "message": "Current password is required.", "field": "currentPassword"}), 400
    if not new_password:
        return jsonify({"status": "error", "message": "New password is required.", "field": "newPassword"}), 400
    if not confirm_password:
        return jsonify({"status": "error", "message": "Please confirm your new password.", "field": "confirmPassword"}), 400
    if len(new_password) < MIN_PASSWORD_LENGTH or len(new_password) > MAX_PASSWORD_LENGTH:
        return jsonify({
            "status": "error",
            "message": f"Password must be between {MIN_PASSWORD_LENGTH} and {MAX_PASSWORD_LENGTH} characters.",
            "field": "newPassword",
        }), 400
    if new_password != confirm_password:
        return jsonify({"status": "error", "message": "Passwords do not match.", "field": "confirmPassword"}), 400

    checks = [
        bool(re.search(r"[A-Z]", new_password)),
        bool(re.search(r"[a-z]", new_password)),
        bool(re.search(r"[0-9]", new_password)),
        bool(re.search(r"[^A-Za-z0-9]", new_password)),
    ]
    if sum(checks) < 3:
        return jsonify({
            "status": "error",
            "message": "Use at least 3 of: uppercase, lowercase, number, special character.",
            "field": "newPassword",
        }), 400

    stored_hash = user_doc.get("password_hash")
    if not stored_hash:
        return jsonify({
            "status": "error",
            "message": "Password record is invalid for this account. Contact an administrator.",
        }), 400
    current_ok = check_password_hash(stored_hash, current_password)

    if not current_ok:
        return jsonify({"status": "error", "message": "Current password is incorrect.", "field": "currentPassword"}), 400

    if stored_hash and check_password_hash(stored_hash, new_password):
        return jsonify({"status": "error", "message": "New password must be different from the current password.", "field": "newPassword"}), 400
    updated = now_iso()
    users.update_one(
        {"_id": user_doc["_id"]},
        {
            "$set": {
                "password_hash": hash_password(new_password),
                "updatedAt": updated,
                "updated_at": updated,
            },
            "$unset": {"password": ""},
        },
    )
    refreshed = users.find_one({"_id": user_doc["_id"]})
    return jsonify({"status": "ok", "profile": normalize_profile_user_doc(refreshed)})


# =====================================
# SMS / OTP API
# =====================================
def get_client_ip():
    xff = request.headers.get("X-Forwarded-For", "").strip()
    if xff:
        return xff.split(",")[0].strip()
    return (request.remote_addr or "").strip()


def parse_json_payload():
    if request.is_json:
        return request.get_json(silent=True) or {}
    return request.form.to_dict() if request.form else {}


def validate_message_type(value):
    normalized = (value or "transactional").strip().lower()
    if normalized not in {"transactional", "otp"}:
        raise ValueError("Invalid message type. Allowed values: transactional, otp.")
    return normalized


def otp_rate_limit_check(phone, client_ip):
    now = now_local()
    throttle_after = (now - timedelta(seconds=OTP_THROTTLE_SECONDS)).isoformat(timespec="seconds")
    hour_after = (now - timedelta(hours=1)).isoformat(timespec="seconds")

    recent_phone = otp_requests.count_documents({"phone": phone, "createdAt": {"$gte": throttle_after}})
    if recent_phone > 0:
        return False, f"Please wait {OTP_THROTTLE_SECONDS} seconds before requesting another OTP."

    recent_hourly = otp_requests.count_documents({"phone": phone, "createdAt": {"$gte": hour_after}})
    if recent_hourly >= OTP_MAX_PER_HOUR:
        return False, "Hourly OTP limit reached. Please try again later."

    if client_ip:
        recent_ip = otp_requests.count_documents({"requestIp": client_ip, "createdAt": {"$gte": hour_after}})
        if recent_ip >= (OTP_MAX_PER_HOUR * 3):
            return False, "Too many OTP requests from this IP. Please try again later."

    return True, ""


@app.route("/api/sms/send", methods=["POST"])
@require_permission("users_manage", api=True)
def api_sms_send():
    payload = parse_json_payload()
    to_raw = (payload.get("to") or "").strip()
    message = (payload.get("message") or "").strip()
    template = payload.get("template")
    variables = payload.get("variables") if isinstance(payload.get("variables"), dict) else {}

    try:
        sms_type = validate_message_type(payload.get("type", "transactional"))
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    if template:
        try:
            message = SmsProvider.render_template(template, variables)
        except ValueError as exc:
            return jsonify({"status": "error", "message": str(exc)}), 400

    if not to_raw:
        return jsonify({"status": "error", "message": "Recipient phone number is required."}), 400
    if not message:
        return jsonify({"status": "error", "message": "Message is required."}), 400

    sms_result = send_sms(
        to_raw,
        message,
        sms_type=sms_type,
        metadata={"context": "api_sms_send"},
        parent_contact=to_raw,
    )
    if sms_result.get("status") != "sent":
        return jsonify({
            "status": "error",
            "message": sms_result.get("error", "SMS failed to send."),
            "data": {
                "provider": "PHILSMS",
                "to": sms_result.get("to", ""),
                "providerMessageId": sms_result.get("provider_message_id", ""),
                "logId": sms_result.get("log_id", ""),
            },
        }), 502

    return jsonify({
        "status": "ok",
        "message": "SMS sent successfully.",
        "data": {
            "provider": "PHILSMS",
            "to": sms_result.get("to", ""),
            "providerMessageId": sms_result.get("provider_message_id", ""),
            "logId": sms_result.get("log_id", ""),
        },
    })


@app.route("/api/sms/health", methods=["GET"])
@require_permission("users_manage", api=True)
def api_sms_health():
    health = sms_provider.health_check()
    if health.get("status") == "ok":
        return jsonify({"status": "ok", "data": health})
    return jsonify({"status": "error", "message": health.get("message", "SMS provider unhealthy."), "data": health}), 503


@app.route("/api/sms/auth-check", methods=["GET"])
@require_permission("scan", api=True)
def api_sms_auth_check():
    checker = getattr(sms_provider, "auth_check", None)
    if callable(checker):
        result = checker()
    else:
        result = sms_provider.health_check()

    if result.get("status") == "ok":
        return jsonify({"status": "ok", "data": result})
    return jsonify({
        "status": "error",
        "message": result.get("message", "SMS auth check failed."),
        "data": result,
    }), 503


@app.route("/api/sms/balance", methods=["GET"])
@require_permission("dashboard", api=True)
def api_sms_balance():
    try:
        force = str(request.args.get("force", "") or "").strip().lower() in {"1", "true", "yes", "on"}
        snapshot = get_sms_balance_snapshot(force=force)
    except Exception as exc:
        return jsonify({"status": "error", "message": f"Failed to fetch SMS balance: {exc}"}), 500

    if snapshot.get("status") == "ok":
        return jsonify({"status": "ok", "data": snapshot})
    return jsonify({
        "status": "error",
        "message": snapshot.get("message", "SMS balance is unavailable."),
        "data": snapshot,
    }), 503


@app.route("/api/auth/otp/request", methods=["POST"])
def api_otp_request():
    payload = parse_json_payload()
    phone_raw = (payload.get("phone") or "").strip()
    if not phone_raw:
        return jsonify({"status": "error", "message": "Phone number is required."}), 400

    try:
        normalized_phone = SmsProvider.normalize_phone_number(phone_raw)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    client_ip = get_client_ip()
    allowed, reason = otp_rate_limit_check(normalized_phone, client_ip)
    if not allowed:
        return jsonify({"status": "error", "message": reason}), 429

    now = now_local()
    created_at = now_iso()
    expires_at_dt = now + timedelta(minutes=OTP_EXPIRES_MINUTES)
    expires_at = expires_at_dt.isoformat(timespec="seconds")
    otp_code = generate_otp_code(OTP_CODE_LENGTH)
    otp_hash = hash_otp_code(otp_code)

    # Keep only one active OTP per phone.
    otp_requests.update_many(
        {"phone": normalized_phone, "status": "pending"},
        {"$set": {"status": "replaced", "updatedAt": created_at}},
    )

    otp_doc = {
        "phone": normalized_phone,
        "otpHash": otp_hash,
        "expiresAt": expires_at,
        "attempts": 0,
        "verifiedAt": None,
        "status": "pending",
        "requestIp": client_ip,
        "createdAt": created_at,
        "updatedAt": created_at,
        "type": "otp",
    }
    otp_insert = otp_requests.insert_one(otp_doc)
    otp_id = str(otp_insert.inserted_id)

    otp_template = os.getenv(
        "OTP_MESSAGE_TEMPLATE",
        "Your CHS Gate Access OTP is {code}. It expires in {minutes} minutes.",
    )
    message = SmsProvider.render_template(
        otp_template,
        {"code": otp_code, "minutes": OTP_EXPIRES_MINUTES},
    )

    sms_result = send_sms(
        normalized_phone,
        message,
        sms_type="otp",
        metadata={"context": "otp_request", "otpRequestId": otp_id},
        parent_contact=normalized_phone,
    )
    if sms_result.get("status") != "sent":
        otp_requests.update_one(
            {"_id": otp_insert.inserted_id},
            {"$set": {
                "status": "failed",
                "updatedAt": now_iso(),
                "error": sms_result.get("error", "Failed to dispatch OTP SMS."),
            }},
        )
        return jsonify({
            "status": "error",
            "message": "OTP dispatch failed.",
            "error": sms_result.get("error", "Failed to dispatch OTP SMS."),
            "data": {"phone": normalized_phone, "otpRequestId": otp_id},
        }), 502

    return jsonify({
        "status": "ok",
        "message": "OTP sent successfully.",
        "data": {
            "phone": normalized_phone,
            "otpRequestId": otp_id,
            "expiresAt": expires_at,
        },
    })


@app.route("/api/auth/otp/verify", methods=["POST"])
def api_otp_verify():
    payload = parse_json_payload()
    phone_raw = (payload.get("phone") or "").strip()
    otp_code = (payload.get("otp") or "").strip()
    if not phone_raw or not otp_code:
        return jsonify({"status": "error", "message": "Phone number and OTP are required."}), 400

    try:
        normalized_phone = SmsProvider.normalize_phone_number(phone_raw)
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

    otp_record = otp_requests.find_one(
        {"phone": normalized_phone, "status": "pending"},
        sort=[("createdAt", -1)],
    )
    if not otp_record:
        return jsonify({"status": "error", "message": "No active OTP request found."}), 404

    now_ts = now_iso()
    expires_at = (otp_record.get("expiresAt") or "").strip()
    if expires_at and now_ts > expires_at:
        otp_requests.update_one(
            {"_id": otp_record["_id"]},
            {"$set": {"status": "expired", "updatedAt": now_ts}},
        )
        return jsonify({"status": "error", "message": "OTP has expired."}), 400

    attempts = int(otp_record.get("attempts") or 0)
    if attempts >= OTP_MAX_ATTEMPTS:
        otp_requests.update_one(
            {"_id": otp_record["_id"]},
            {"$set": {"status": "locked", "updatedAt": now_ts}},
        )
        return jsonify({"status": "error", "message": "OTP attempts exceeded."}), 429

    if verify_otp_code(otp_record.get("otpHash", ""), otp_code):
        otp_requests.update_one(
            {"_id": otp_record["_id"]},
            {"$set": {"status": "verified", "verifiedAt": now_ts, "updatedAt": now_ts}},
        )
        return jsonify({
            "status": "ok",
            "message": "OTP verified successfully.",
            "data": {"phone": normalized_phone, "verifiedAt": now_ts},
        })

    attempts += 1
    updated_status = "locked" if attempts >= OTP_MAX_ATTEMPTS else "pending"
    otp_requests.update_one(
        {"_id": otp_record["_id"]},
        {"$set": {"attempts": attempts, "status": updated_status, "updatedAt": now_ts}},
    )
    return jsonify({
        "status": "error",
        "message": "Invalid OTP.",
        "data": {"attempts": attempts, "maxAttempts": OTP_MAX_ATTEMPTS},
    }), 400


# =====================================
# SCANNING ROUTES
# =====================================
@app.route("/start_scan", methods=["POST"])
@require_permission("scan", api=True)
def start_scan():
    requested_mode = None
    requested_liveness_profile = None
    if request.method == "POST":
        payload = request_payload()
        requested_mode = payload.get("session_mode") or payload.get("mode")
        requested_liveness_profile = payload.get("liveness_profile") or payload.get("profile")
    if requested_mode is not None:
        try:
            set_scan_session_mode(requested_mode)
        except ValueError as exc:
            return jsonify({"status": "error", "message": str(exc)}), 400
    if requested_liveness_profile is not None:
        try:
            set_scan_liveness_profile(requested_liveness_profile)
        except ValueError as exc:
            return jsonify({"status": "error", "message": str(exc)}), 400

    validator = getattr(sms_provider, "validate_configuration", None)
    if callable(validator):
        sms_status = validator(raise_on_error=False)
    else:
        sms_status = sms_provider.health_check()
    sms_warning = ""
    if sms_status.get("status") != "ok":
        sms_warning = sms_status.get("message", "SMS provider is unavailable. Scanning will continue without SMS delivery.")
        print(f"[WARNING] SMS provider not ready during scan start: {sms_warning}")

    live_monitoring_ready, live_monitoring_message = ensure_live_monitoring_service_ready(
        reason="scan_start",
        fail_hard=False,
    )

    ok, message = start_scan_capture()
    if not ok:
        payload = {"status": "failed", "message": message, "sms_auth": sms_status}
        if sms_warning:
            payload["sms_warning"] = sms_warning
        status_code = 503 if message == "Webcam could not be opened" else 500
        return jsonify(payload), status_code

    with scan_lock:
        model_status = scan_state.get("model_status", "idle")
        registered_faces = len(scan_state.get("known_students", []))
        session_mode = normalize_scan_session_mode(scan_state.get("session_mode", "auto"), default="auto")
        liveness_profile = normalize_scan_liveness_profile(scan_state.get("liveness_profile", LIVE_LIVENESS_PROFILE))
        face_index_loading = bool(scan_state.get("face_index_loading"))
    effective_session = resolve_gate_session(now_local())
    payload = {
        "status": "ok",
        "message": message,
        "model_status": model_status,
        "face_index_loading": face_index_loading,
        "registered_faces": registered_faces,
        "sms_auth": sms_status,
        "scan_session_mode": session_mode,
        "session_mode_label": scan_session_mode_label(session_mode),
        "scan_liveness_profile": liveness_profile,
        "liveness_profile_label": scan_liveness_profile_label(liveness_profile),
        "effective_session": {
            "session": effective_session.get("session", ""),
            "gate_action": effective_session.get("gate_action", ""),
            "verification_label": effective_session.get("verification_label", ""),
            "status": effective_session.get("status", ""),
            "display_message": effective_session.get("display_message", ""),
            "voice_message": effective_session.get("voice_message", ""),
        },
    }
    if sms_warning:
        payload["sms_warning"] = sms_warning
    if not live_monitoring_ready:
        payload["live_monitoring_warning"] = live_monitoring_message
    return jsonify(payload)


@app.route("/stop_scan", methods=["POST"])
@require_permission("scan", api=True)
def stop_scan():
    stop_scan_capture()
    return jsonify({"status": "ok", "message": "Scan stopped"})


def filter_live_face_locations(face_locations, frame_shape):
    if not isinstance(face_locations, list):
        return []

    frame_height = max(int((frame_shape or [0, 0])[0] or 0), 1)
    frame_width = max(int((frame_shape or [0, 0])[1] or 0), 1)
    min_face_height = max(frame_height * 0.08, 20.0)
    min_face_width = max(frame_width * 0.07, 20.0)
    filtered = []

    for location in face_locations:
        if not isinstance(location, (list, tuple)) or len(location) != 4:
            continue
        top, right, bottom, left = location
        face_height = max(float(bottom) - float(top), 0.0)
        face_width = max(float(right) - float(left), 0.0)
        if face_height < min_face_height or face_width < min_face_width:
            continue
        filtered.append(location)
    return filtered


def prioritize_live_face_locations(face_locations):
    normalized = [
        location
        for location in list(face_locations or [])
        if isinstance(location, (list, tuple)) and len(location) == 4
    ]
    normalized.sort(
        key=lambda location: max(float(location[2]) - float(location[0]), 0.0) * max(float(location[1]) - float(location[3]), 0.0),
        reverse=True,
    )

    if len(normalized) <= 1:
        return normalized, False

    def face_area(location):
        return max(float(location[2]) - float(location[0]), 0.0) * max(float(location[1]) - float(location[3]), 0.0)

    largest_area = face_area(normalized[0])
    second_area = face_area(normalized[1])
    if largest_area <= 0:
        return normalized[:1], False

    if second_area <= largest_area * float(LIVE_RECOGNITION_DOMINANT_FACE_RATIO):
        return [normalized[0]], False

    return normalized, True


def clear_pending_live_recognition(student_id=None):
    normalized_student_id = str(student_id or "").strip()
    with scan_lock:
        pending = scan_state.setdefault("pending_recognition", {})
        if not isinstance(pending, dict):
            scan_state["pending_recognition"] = {}
            return
        if normalized_student_id:
            pending.pop(normalized_student_id, None)
            return
        pending.clear()


def measure_live_face_quality(frame, face_location, scale_back=1.0):
    metrics = {
        "brightness": 0.0,
        "contrast": 0.0,
        "sharpness": 0.0,
        "texture": 0.0,
        "highlights": 0.0,
        "area_ratio": 0.0,
    }

    if frame is None or frame.size == 0:
        return metrics

    try:
        frame_height, frame_width = frame.shape[:2]
        top, right, bottom, left = face_location
        top = max(0, min(frame_height, int(round(float(top) * scale_back))))
        right = max(0, min(frame_width, int(round(float(right) * scale_back))))
        bottom = max(0, min(frame_height, int(round(float(bottom) * scale_back))))
        left = max(0, min(frame_width, int(round(float(left) * scale_back))))

        if bottom <= top or right <= left:
            return metrics

        roi = frame[top:bottom, left:right]
        if roi.size == 0:
            return metrics

        gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        metrics["brightness"] = round(float(np.mean(gray)), 2)
        metrics["contrast"] = round(float(np.std(gray)), 2)
        metrics["sharpness"] = round(float(cv2.Laplacian(gray, cv2.CV_64F).var()), 2)
        sobel_x = cv2.Sobel(gray, cv2.CV_32F, 1, 0, ksize=3)
        sobel_y = cv2.Sobel(gray, cv2.CV_32F, 0, 1, ksize=3)
        texture_strength = float(np.mean(cv2.magnitude(sobel_x, sobel_y)))
        metrics["texture"] = round(texture_strength, 2)
        metrics["highlights"] = round(float(np.mean(gray >= 245)), 4)
        metrics["area_ratio"] = round(float(((bottom - top) * (right - left)) / max(frame_height * frame_width, 1)), 6)
        return metrics
    except Exception:
        return metrics


def _live_track_cache_is_stable(track):
    row = track if isinstance(track, dict) else {}
    motion_component = abs(float(row.get("liveness_motion_component") or 0.0))
    area_component = abs(float(row.get("liveness_area_component") or 0.0))
    pose_component = abs(float(row.get("liveness_pose_component") or 0.0))
    return (
        motion_component <= float(LIVE_RECOGNITION_CACHE_MAX_MOTION_COMPONENT)
        and area_component <= float(LIVE_RECOGNITION_CACHE_MAX_AREA_COMPONENT)
        and pose_component <= float(LIVE_RECOGNITION_CACHE_MAX_POSE_COMPONENT)
    )


def resolve_live_track_face_quality(track, frame, face_location, scale_back=1.0, now_ts=None):
    row = track if isinstance(track, dict) else {}
    current_ts = float(now_ts if now_ts is not None else time.time())
    cached_quality = row.get("face_quality")
    cached_quality_ts = float(row.get("face_quality_ts") or 0.0)
    cache_window = max(float(LIVE_RECOGNITION_FACE_QUALITY_CACHE_SECONDS or 0.0), 0.05)

    if (
        isinstance(cached_quality, dict)
        and cached_quality
        and (current_ts - cached_quality_ts) <= cache_window
        and _live_track_cache_is_stable(row)
    ):
        return dict(cached_quality)

    quality = measure_live_face_quality(frame, face_location, scale_back=scale_back)
    normalized_quality = dict(quality) if isinstance(quality, dict) else {}
    updates = {
        "face_quality": normalized_quality,
        "face_quality_ts": float(current_ts),
    }
    if isinstance(track, dict):
        track.update(updates)

    track_id = row.get("track_id")
    if track_id not in (None, ""):
        set_live_face_track_liveness_state(track_id, **updates)
    return normalized_quality


def get_cached_live_track_encoding(track, now_ts=None):
    row = track if isinstance(track, dict) else {}
    cached_encoding = row.get("last_encoding")
    if cached_encoding is None:
        return None

    current_ts = float(now_ts if now_ts is not None else time.time())
    cached_encoding_ts = float(row.get("last_encoding_ts") or 0.0)
    cache_window = max(float(LIVE_RECOGNITION_ENCODING_CACHE_SECONDS or 0.0), 0.05)
    if (current_ts - cached_encoding_ts) > cache_window:
        return None
    if not _live_track_cache_is_stable(row):
        return None

    try:
        normalized_encoding = np.asarray(cached_encoding, dtype=np.float64).reshape(-1)
    except Exception:
        return None
    if normalized_encoding.size != 128:
        return None
    return normalized_encoding


def cache_live_track_encoding(track, encoding, now_ts=None):
    if encoding is None:
        return None
    try:
        normalized_encoding = np.asarray(encoding, dtype=np.float64).reshape(-1)
    except Exception:
        return None
    if normalized_encoding.size != 128:
        return None

    row = track if isinstance(track, dict) else {}
    current_ts = float(now_ts if now_ts is not None else time.time())
    updates = {
        "last_encoding": normalized_encoding,
        "last_encoding_ts": float(current_ts),
    }
    if isinstance(track, dict):
        track.update(updates)

    track_id = row.get("track_id")
    if track_id not in (None, ""):
        set_live_face_track_liveness_state(track_id, **updates)
    return normalized_encoding


def get_cached_live_track_match_result(track, now_ts=None):
    row = track if isinstance(track, dict) else {}
    cached_match = row.get("last_match_result")
    if not isinstance(cached_match, dict) or not cached_match:
        return None

    current_ts = float(now_ts if now_ts is not None else time.time())
    cached_match_ts = float(row.get("last_match_ts") or 0.0)
    cache_window = max(float(LIVE_RECOGNITION_ENCODING_CACHE_SECONDS or 0.0), 0.05)
    if (current_ts - cached_match_ts) > cache_window:
        return None
    if not _live_track_cache_is_stable(row):
        return None

    cached_match_encoding_ts = float(row.get("last_match_encoding_ts") or 0.0)
    cached_encoding_ts = float(row.get("last_encoding_ts") or 0.0)
    if cached_encoding_ts <= 0.0 or abs(cached_match_encoding_ts - cached_encoding_ts) > 1e-6:
        return None

    return cached_match


def cache_live_track_match_result(track, match_result, now_ts=None):
    if not isinstance(match_result, dict) or not match_result:
        return None

    row = track if isinstance(track, dict) else {}
    cached_encoding_ts = float(row.get("last_encoding_ts") or 0.0)
    if cached_encoding_ts <= 0.0:
        return None

    candidate = match_result.get("candidate")
    runner_up = match_result.get("runner_up")
    normalized_match = {
        "recognized": bool(match_result.get("recognized")),
        "student": dict(match_result.get("student") or {}) if isinstance(match_result.get("student"), dict) else match_result.get("student"),
        "confidence": float(match_result.get("confidence") or 0.0),
        "distance": float(match_result.get("distance") or 0.0),
        "score_margin": float(match_result.get("score_margin") or 0.0),
        "distance_margin": float(match_result.get("distance_margin") or 0.0),
        "reason": str(match_result.get("reason") or ""),
        "candidate": dict(candidate) if isinstance(candidate, dict) else candidate,
        "runner_up": dict(runner_up) if isinstance(runner_up, dict) else runner_up,
    }

    current_ts = float(now_ts if now_ts is not None else time.time())
    updates = {
        "last_match_result": normalized_match,
        "last_match_ts": float(current_ts),
        "last_match_encoding_ts": float(cached_encoding_ts),
    }
    if isinstance(track, dict):
        track.update(updates)

    track_id = row.get("track_id")
    if track_id not in (None, ""):
        set_live_face_track_liveness_state(track_id, **updates)
    return normalized_match


def upscale_face_locations(face_locations, scale_back, frame_shape):
    frame_height = max(int((frame_shape or [0, 0])[0] or 0), 1)
    frame_width = max(int((frame_shape or [0, 0])[1] or 0), 1)
    scaled_locations = []

    for location in list(face_locations or []):
        if not isinstance(location, (list, tuple)) or len(location) != 4:
            continue
        top, right, bottom, left = location
        scaled_top = max(0, min(frame_height, int(round(float(top) * scale_back))))
        scaled_right = max(0, min(frame_width, int(round(float(right) * scale_back))))
        scaled_bottom = max(0, min(frame_height, int(round(float(bottom) * scale_back))))
        scaled_left = max(0, min(frame_width, int(round(float(left) * scale_back))))
        if scaled_bottom <= scaled_top or scaled_right <= scaled_left:
            continue
        scaled_locations.append((scaled_top, scaled_right, scaled_bottom, scaled_left))

    return scaled_locations


def _normalized_face_location(location):
    if not isinstance(location, (list, tuple)) or len(location) != 4:
        return None
    try:
        top = int(round(float(location[0])))
        right = int(round(float(location[1])))
        bottom = int(round(float(location[2])))
        left = int(round(float(location[3])))
    except (TypeError, ValueError):
        return None
    if bottom <= top or right <= left:
        return None
    return (top, right, bottom, left)


def _face_location_center(location):
    top, right, bottom, left = location
    return ((float(left) + float(right)) / 2.0, (float(top) + float(bottom)) / 2.0)


def _face_location_area(location):
    top, right, bottom, left = location
    return max(float(bottom) - float(top), 0.0) * max(float(right) - float(left), 0.0)


def _face_location_dimensions(location):
    top, right, bottom, left = location
    width = max(float(right) - float(left), 0.0)
    height = max(float(bottom) - float(top), 0.0)
    return width, height


def _face_track_match_limit(location):
    top, right, bottom, left = location
    face_width = max(float(right) - float(left), 1.0)
    face_height = max(float(bottom) - float(top), 1.0)
    longest_edge = max(face_width, face_height)
    return max(20.0, longest_edge * float(LIVE_RECOGNITION_TRACK_MATCH_DISTANCE_RATIO))


def build_live_face_payload_from_tracks(face_tracks, frame_shape):
    frame_height = max(int((frame_shape or [0, 0])[0] or 0), 1)
    frame_width = max(int((frame_shape or [0, 0])[1] or 0), 1)
    payload_faces = []

    for track in list(face_tracks or []):
        location = _normalized_face_location(track.get("full_location"))
        if not location:
            continue
        top, right, bottom, left = location
        payload_faces.append({
            "id": f"scan-face-{int(track.get('track_id') or 0)}",
            "track_id": int(track.get("track_id") or 0),
            "x": int(left),
            "y": int(top),
            "width": int(max(right - left, 0)),
            "height": int(max(bottom - top, 0)),
            "frame_width": int(frame_width),
            "frame_height": int(frame_height),
            "stability": int(track.get("stability") or 0),
            "student_id": str(track.get("student_id") or "").strip(),
            "liveness_ready": bool(track.get("liveness_ready", not LIVE_SCAN_LIVENESS_ENABLED)),
            "liveness_motion_score": round(float(track.get("liveness_motion_score") or 0.0), 3),
            "liveness_pose_score": round(float(track.get("liveness_pose_score") or 0.0), 3),
            "liveness_parallax_score": round(float(track.get("liveness_parallax_score") or 0.0), 3),
            "liveness_landmark_pose_span": round(float(track.get("liveness_landmark_pose_span") or 0.0), 3),
            "liveness_patch_diversity": round(float(track.get("liveness_patch_diversity") or 0.0), 3),
            "liveness_display_score": round(float(track.get("liveness_display_score") or 0.0), 3),
            "liveness_message": str(track.get("liveness_message") or ""),
        })

    payload_faces.sort(
        key=lambda row: float(row.get("width", 0) * row.get("height", 0)),
        reverse=True,
    )
    return payload_faces[: max(int(LIVE_RECOGNITION_MAX_TRACKED_FACES or 1), 1)]


def get_latest_live_faces_payload():
    with scan_lock:
        faces = scan_state.get("last_faces_payload", [])
        if not isinstance(faces, list):
            return []
        return [dict(row) for row in faces if isinstance(row, dict)]


def update_live_face_tracks(face_locations_small, scale_back, frame_shape, now_ts=None):
    current_ts = float(now_ts if now_ts is not None else time.time())
    normalized_locations = []
    for location in list(face_locations_small or []):
        normalized = _normalized_face_location(location)
        if normalized:
            normalized_locations.append(normalized)

    normalized_locations.sort(key=_face_location_area, reverse=True)
    if len(normalized_locations) > max(int(LIVE_RECOGNITION_MAX_TRACKED_FACES or 1), 1):
        normalized_locations = normalized_locations[: max(int(LIVE_RECOGNITION_MAX_TRACKED_FACES or 1), 1)]

    liveness_config = resolve_live_liveness_profile_thresholds() if LIVE_SCAN_LIVENESS_ENABLED else {}
    liveness_decay = min(max(float(LIVE_LIVENESS_SCORE_DECAY or 0.88), 0.5), 0.99)
    min_liveness_frames = max(int(liveness_config.get("min_track_frames") or 1), 1)
    min_motion_score = max(float(liveness_config.get("min_motion_score") or 0.0), 0.0)
    min_pose_score = max(float(liveness_config.get("min_pose_score") or 0.0), 0.0)
    pose_grace_seconds = max(float(liveness_config.get("pose_grace_seconds") or 0.0), 0.0)

    with scan_lock:
        tracks = scan_state.setdefault("face_tracks", {})
        if not isinstance(tracks, dict):
            tracks = {}
            scan_state["face_tracks"] = tracks

        timeout_seconds = max(float(LIVE_RECOGNITION_TRACK_TIMEOUT_SECONDS), 0.1)
        stale_track_ids = [
            track_id
            for track_id, row in tracks.items()
            if current_ts - float((row or {}).get("last_seen_ts") or 0.0) > timeout_seconds
        ]
        for track_id in stale_track_ids:
            tracks.pop(track_id, None)

        try:
            next_track_id = max(int(scan_state.get("next_face_track_id") or 1), 1)
        except (TypeError, ValueError):
            next_track_id = 1

        matched_track_ids = set()
        active_tracks = []

        for location in normalized_locations:
            center_x, center_y = _face_location_center(location)
            match_limit = _face_track_match_limit(location)
            best_track_id = None
            best_distance = float("inf")

            for track_id, existing in tracks.items():
                if track_id in matched_track_ids:
                    continue
                track_center = existing.get("center")
                if not isinstance(track_center, (list, tuple)) or len(track_center) != 2:
                    continue
                distance = float(np.hypot(center_x - float(track_center[0]), center_y - float(track_center[1])))
                existing_limit = max(float(existing.get("match_limit") or 0.0), 20.0)
                allowed_distance = max(match_limit, existing_limit)
                if distance <= allowed_distance and distance < best_distance:
                    best_distance = distance
                    best_track_id = track_id

            scaled_location = upscale_face_locations([location], scale_back, frame_shape)
            if not scaled_location:
                continue
            full_location = scaled_location[0]

            if best_track_id is None:
                best_track_id = next_track_id
                next_track_id += 1
                previous = {}
                stability = 1
                center_distance = 0.0
            else:
                previous = tracks.get(best_track_id) or {}
                stability = int(previous.get("stability") or 0) + 1
                center_distance = float(best_distance if np.isfinite(best_distance) else 0.0)

            area = float(_face_location_area(location))
            face_width, face_height = _face_location_dimensions(location)
            aspect_ratio = float(face_width / max(face_height, 1.0))
            face_diagonal = max(float(np.hypot(face_width, face_height)), 1.0)

            first_seen_ts = float(previous.get("first_seen_ts") or current_ts)
            previous_area = max(float(previous.get("area") or area or 1.0), 1.0)
            previous_aspect_ratio = max(float(previous.get("aspect_ratio") or aspect_ratio or 1.0), 0.01)
            motion_component = min(center_distance / face_diagonal, 1.5)
            area_component = min(abs(float(np.log(max(area, 1.0) / previous_area))), 1.2)
            pose_component = min(abs(aspect_ratio - previous_aspect_ratio) / previous_aspect_ratio, 0.6)

            motion_signal = min(motion_component + (area_component * 0.45), 1.5)
            liveness_motion_score = max(float(previous.get("liveness_motion_score") or 0.0) * liveness_decay + motion_signal, 0.0)
            liveness_pose_score = max(float(previous.get("liveness_pose_score") or 0.0) * liveness_decay + pose_component, 0.0)
            liveness_frames = int(previous.get("liveness_frames") or 0) + 1
            track_age_seconds = max(current_ts - first_seen_ts, 0.0)
            liveness_ready = (
                (not LIVE_SCAN_LIVENESS_ENABLED)
                or (
                    liveness_frames >= min_liveness_frames
                    and liveness_motion_score >= min_motion_score
                    and (
                        liveness_pose_score >= min_pose_score
                        or track_age_seconds >= pose_grace_seconds
                    )
                )
            )

            if not LIVE_SCAN_LIVENESS_ENABLED:
                liveness_message = ""
            elif liveness_frames < min_liveness_frames:
                liveness_message = "Liveness check required"
            elif liveness_motion_score < min_motion_score:
                liveness_message = "Liveness check required: move slightly"
            elif liveness_pose_score < min_pose_score and track_age_seconds < pose_grace_seconds:
                liveness_message = "Liveness check required: turn head slightly"
            else:
                liveness_message = ""

            updated_track = {
                "track_id": int(best_track_id),
                "first_seen_ts": first_seen_ts,
                "last_seen_ts": current_ts,
                "stability": int(stability),
                "small_location": location,
                "full_location": full_location,
                "center": (center_x, center_y),
                "area": area,
                "aspect_ratio": aspect_ratio,
                "match_limit": float(match_limit),
                "next_attempt_ts": float(previous.get("next_attempt_ts") or 0.0),
                "student_id": str(previous.get("student_id") or "").strip(),
                "last_result": str(previous.get("last_result") or ""),
                "last_confidence": float(previous.get("last_confidence") or 0.0),
                "face_quality": dict(previous.get("face_quality") or {}) if isinstance(previous.get("face_quality"), dict) else {},
                "face_quality_ts": float(previous.get("face_quality_ts") or 0.0),
                "last_encoding": previous.get("last_encoding"),
                "last_encoding_ts": float(previous.get("last_encoding_ts") or 0.0),
                "liveness_frames": int(liveness_frames),
                "liveness_motion_score": float(round(liveness_motion_score, 6)),
                "liveness_pose_score": float(round(liveness_pose_score, 6)),
                "liveness_motion_component": float(round(motion_component, 6)),
                "liveness_area_component": float(round(area_component, 6)),
                "liveness_pose_component": float(round(pose_component, 6)),
                "liveness_parallax_score": float(previous.get("liveness_parallax_score") or 0.0),
                "liveness_planar_streak": int(previous.get("liveness_planar_streak") or 0),
                "liveness_patch_delta": float(previous.get("liveness_patch_delta") or 0.0),
                "liveness_patch_grad_delta": float(previous.get("liveness_patch_grad_delta") or 0.0),
                "liveness_prev_patch": previous.get("liveness_prev_patch"),
                "liveness_prev_grad_patch": previous.get("liveness_prev_grad_patch"),
                "liveness_patch_ts": float(previous.get("liveness_patch_ts") or 0.0),
                "liveness_patch_diversity": float(previous.get("liveness_patch_diversity") or 0.0),
                "liveness_last_landmark_ts": float(previous.get("liveness_last_landmark_ts") or 0.0),
                "liveness_landmark_yaw": float(previous.get("liveness_landmark_yaw") or 0.0),
                "liveness_landmark_pitch": float(previous.get("liveness_landmark_pitch") or 0.0),
                "liveness_landmark_pose_span": float(previous.get("liveness_landmark_pose_span") or 0.0),
                "liveness_landmark_pose_samples": int(previous.get("liveness_landmark_pose_samples") or 0),
                "liveness_pose_history": list(previous.get("liveness_pose_history") or []),
                "liveness_display_score": float(previous.get("liveness_display_score") or 0.0),
                "liveness_display_streak": int(previous.get("liveness_display_streak") or 0),
                "liveness_ready": bool(liveness_ready),
                "liveness_message": str(liveness_message or ""),
            }
            tracks[best_track_id] = updated_track
            matched_track_ids.add(best_track_id)
            active_tracks.append(updated_track)

        scan_state["face_tracks"] = tracks
        scan_state["next_face_track_id"] = next_track_id

    active_tracks.sort(key=lambda row: float(row.get("area") or 0.0), reverse=True)
    return active_tracks


def evaluate_live_texture_liveness(face_quality):
    if not LIVE_LIVENESS_TEXTURE_CHECK_ENABLED:
        return {"accepted": True, "reason": "texture_check_disabled", "message": ""}

    liveness_config = resolve_live_liveness_profile_thresholds()
    min_texture = max(float(liveness_config.get("min_texture") or 0.0), 0.0)
    min_contrast = max(float(liveness_config.get("min_contrast") or 0.0), 0.0)
    max_highlight_ratio = min(max(float(liveness_config.get("max_highlight_ratio") or 0.0), 0.0), 1.0)

    quality = face_quality if isinstance(face_quality, dict) else {}
    brightness = float(quality.get("brightness") or 0.0)
    contrast = float(quality.get("contrast") or 0.0)
    sharpness = float(quality.get("sharpness") or 0.0)
    texture = float(quality.get("texture") or 0.0)
    highlights = float(quality.get("highlights") or 0.0)
    area_ratio = float(quality.get("area_ratio") or 0.0)

    if area_ratio <= 0.0:
        return {"accepted": True, "reason": "texture_unavailable", "message": ""}

    # Avoid penalizing low-light scenes where texture extraction is less reliable.
    if brightness > 0.0 and brightness < float(LIVE_RECOGNITION_LOW_LIGHT_BRIGHTNESS) * 0.85:
        return {"accepted": True, "reason": "texture_low_light", "message": ""}

    # Strong detail is typically enough evidence against flat replay media.
    if sharpness >= float(LIVE_RECOGNITION_MIN_SHARPNESS) * 1.2 and texture >= min_texture:
        return {"accepted": True, "reason": "texture_strong_detail", "message": ""}

    low_texture_contrast = (
        texture < min_texture
        and contrast < min_contrast
    )
    heavy_glare = (
        highlights > max_highlight_ratio
        and contrast < min_contrast * 1.2
    )

    if low_texture_contrast or heavy_glare:
        return {
            "accepted": False,
            "reason": "liveness_texture_check",
            "message": "Verification Failed",
        }

    return {"accepted": True, "reason": "texture_ok", "message": ""}


def _compute_eye_aspect_ratio(eye_points):
    if not isinstance(eye_points, (list, tuple)) or len(eye_points) < 6:
        return 0.0
    try:
        pts = np.array(eye_points[:6], dtype=np.float32)
        vertical_a = float(np.linalg.norm(pts[1] - pts[5]))
        vertical_b = float(np.linalg.norm(pts[2] - pts[4]))
        horizontal = float(np.linalg.norm(pts[0] - pts[3]))
    except Exception:
        return 0.0
    if horizontal <= 1e-6:
        return 0.0
    return max((vertical_a + vertical_b) / (2.0 * horizontal), 0.0)


def _mean_landmark_point(points):
    if not isinstance(points, (list, tuple)) or not points:
        return None
    try:
        pts = np.array(points, dtype=np.float32)
    except Exception:
        return None
    if pts.ndim != 2 or pts.shape[1] != 2:
        return None
    center = np.mean(pts, axis=0)
    return float(center[0]), float(center[1])


def _extract_live_landmark_snapshot(rgb_frame, face_location):
    if rgb_frame is None or getattr(rgb_frame, "size", 0) == 0:
        return None
    location = _normalized_face_location(face_location)
    if not location:
        return None
    try:
        landmarks = face_recognition.face_landmarks(
            rgb_frame,
            [location],
            model="small",
        )
    except Exception:
        landmarks = []
    if not landmarks:
        return None

    face_landmarks = landmarks[0] if isinstance(landmarks[0], dict) else {}
    left_eye = face_landmarks.get("left_eye") or []
    right_eye = face_landmarks.get("right_eye") or []
    nose_tip = face_landmarks.get("nose_tip") or []
    nose_bridge = face_landmarks.get("nose_bridge") or []
    top_lip = face_landmarks.get("top_lip") or []
    bottom_lip = face_landmarks.get("bottom_lip") or []

    left_eye_center = _mean_landmark_point(left_eye)
    right_eye_center = _mean_landmark_point(right_eye)
    mouth_upper = _mean_landmark_point(top_lip)
    mouth_lower = _mean_landmark_point(bottom_lip)
    mouth_center = None
    if mouth_upper and mouth_lower:
        mouth_center = (
            (mouth_upper[0] + mouth_lower[0]) / 2.0,
            (mouth_upper[1] + mouth_lower[1]) / 2.0,
        )
    elif mouth_upper or mouth_lower:
        mouth_center = mouth_upper or mouth_lower
    nose_center = _mean_landmark_point(nose_tip) or _mean_landmark_point(nose_bridge[-2:] if len(nose_bridge) >= 2 else nose_bridge)

    left_ear = _compute_eye_aspect_ratio(left_eye)
    right_ear = _compute_eye_aspect_ratio(right_eye)
    if left_ear <= 0.0 and right_ear <= 0.0:
        ear_value = None
    elif left_ear <= 0.0:
        ear_value = right_ear
    elif right_ear <= 0.0:
        ear_value = left_ear
    else:
        ear_value = (left_ear + right_ear) / 2.0

    if not left_eye_center or not right_eye_center:
        return {"ear": ear_value}

    eye_distance = float(np.hypot(
        float(right_eye_center[0]) - float(left_eye_center[0]),
        float(right_eye_center[1]) - float(left_eye_center[1]),
    ))
    if eye_distance <= 1e-6:
        return {"ear": ear_value}

    eye_mid_x = (float(left_eye_center[0]) + float(right_eye_center[0])) / 2.0
    eye_mid_y = (float(left_eye_center[1]) + float(right_eye_center[1])) / 2.0
    yaw_value = 0.0
    if nose_center:
        yaw_value = (float(nose_center[0]) - eye_mid_x) / eye_distance

    pitch_value = 0.0
    if nose_center and mouth_center:
        pitch_value = (float(mouth_center[1]) - float(nose_center[1])) / eye_distance
    elif mouth_center:
        pitch_value = (float(mouth_center[1]) - eye_mid_y) / eye_distance

    return {
        "ear": ear_value,
        "yaw": float(round(yaw_value, 6)),
        "pitch": float(round(pitch_value, 6)),
        "eye_distance": float(round(eye_distance, 6)),
    }


def extract_live_landmark_signature(rgb_small, face_location, rgb_full=None, face_location_full=None):
    signature = _extract_live_landmark_snapshot(rgb_small, face_location)
    if signature and (
        signature.get("ear") is not None
        or signature.get("yaw") is not None
        or signature.get("pitch") is not None
    ):
        return signature

    normalized_full_location = _normalized_face_location(face_location_full)
    if rgb_full is None or getattr(rgb_full, "size", 0) == 0 or not normalized_full_location:
        return signature

    full_top, full_right, full_bottom, full_left = normalized_full_location
    if (full_bottom - full_top) < 56 or (full_right - full_left) < 56:
        return signature

    fallback_signature = _extract_live_landmark_snapshot(rgb_full, normalized_full_location)
    return fallback_signature or signature


def sample_live_landmark_liveness(track, rgb_small=None, face_location=None, now_ts=None, rgb_full=None, face_location_full=None):
    if not LIVE_LIVENESS_ENABLED:
        return {"signature": None, "updates": {}}
    if not LIVE_LIVENESS_BLINK_ENABLED and not LIVE_LIVENESS_LANDMARK_POSE_ENABLED:
        return {"signature": None, "updates": {}}

    row = track if isinstance(track, dict) else {}
    current_ts = float(now_ts if now_ts is not None else time.time())
    liveness_config = resolve_live_liveness_profile_thresholds()
    sample_interval = max(float(liveness_config.get("pose_sample_interval_seconds") or 0.0), 0.05)
    pose_samples = int(row.get("liveness_landmark_pose_samples") or 0)
    first_seen_ts = float(row.get("first_seen_ts") or current_ts)
    track_age_seconds = max(current_ts - first_seen_ts, 0.0)
    last_blink_ts = float(row.get("liveness_last_blink_ts") or 0.0)
    blink_recent_window = max(float(liveness_config.get("blink_recent_seconds") or 0.0), 0.5)
    has_recent_blink = (current_ts - last_blink_ts) <= blink_recent_window
    if track_age_seconds <= 1.6:
        if pose_samples < 2:
            sample_interval = min(sample_interval, 0.09)
        elif not has_recent_blink:
            sample_interval = min(sample_interval, 0.1)
    last_sample_ts = float(row.get("liveness_last_landmark_ts") or 0.0)
    if (current_ts - last_sample_ts) < sample_interval:
        return {"signature": None, "updates": {}}

    signature = extract_live_landmark_signature(
        rgb_small,
        face_location,
        rgb_full=rgb_full,
        face_location_full=face_location_full,
    )
    if not isinstance(signature, dict) or not signature:
        return {"signature": None, "updates": {}}

    pose_history = list(row.get("liveness_pose_history") or [])
    if LIVE_LIVENESS_LANDMARK_POSE_ENABLED:
        yaw_value = signature.get("yaw")
        pitch_value = signature.get("pitch")
        if yaw_value is not None or pitch_value is not None:
            pose_history.append((
                float(current_ts),
                float(yaw_value if yaw_value is not None else 0.0),
                float(pitch_value if pitch_value is not None else 0.0),
            ))
            pose_window_seconds = max(float(liveness_config.get("pose_window_seconds") or 0.0), 1.0)
            pose_history = [
                row_value
                for row_value in pose_history
                if isinstance(row_value, (list, tuple))
                and len(row_value) >= 3
                and (current_ts - float(row_value[0])) <= pose_window_seconds
            ][-10:]
    else:
        pose_history = []

    pose_span = 0.0
    if pose_history:
        yaw_values = [float(row_value[1]) for row_value in pose_history]
        pitch_values = [float(row_value[2]) for row_value in pose_history]
        yaw_span = max(yaw_values) - min(yaw_values)
        pitch_span = max(pitch_values) - min(pitch_values)
        pose_span = max(float(yaw_span), float(pitch_span) * 0.82)

    updates = {
        "liveness_last_landmark_ts": float(current_ts),
        "liveness_landmark_yaw": float(signature.get("yaw") or 0.0),
        "liveness_landmark_pitch": float(signature.get("pitch") or 0.0),
        "liveness_landmark_pose_span": float(round(pose_span, 6)),
        "liveness_landmark_pose_samples": int(len(pose_history)),
        "liveness_pose_history": pose_history,
    }
    return {"signature": signature, "updates": updates}


def extract_live_blink_ear(rgb_small, face_location):
    signature = extract_live_landmark_signature(rgb_small, face_location)
    if not isinstance(signature, dict):
        return None
    ear_value = signature.get("ear")
    return float(ear_value) if ear_value is not None else None


def evaluate_live_blink_liveness(
    track,
    rgb_small=None,
    face_location=None,
    now_ts=None,
    rgb_full=None,
    face_location_full=None,
    landmark_signature=None,
):
    if not LIVE_LIVENESS_ENABLED or not LIVE_LIVENESS_BLINK_ENABLED:
        return {"accepted": True, "reason": "blink_disabled", "message": "", "updates": {}, "blink_detected": False}

    liveness_config = resolve_live_liveness_profile_thresholds()
    blink_required = bool(liveness_config.get("blink_required"))
    current_ts = float(now_ts if now_ts is not None else time.time())
    row = track if isinstance(track, dict) else {}
    updates = {}

    blink_count = int(row.get("liveness_blink_count") or 0)
    eye_closed_frames = int(row.get("liveness_eye_closed_frames") or 0)
    last_blink_ts = float(row.get("liveness_last_blink_ts") or 0.0)
    last_ear_ts = float(row.get("liveness_last_ear_ts") or 0.0)
    blink_detected = False
    ear_value = None

    min_interval = max(float(liveness_config.get("blink_interval_seconds") or 0.0), 0.05)
    required_recent_seconds = max(float(liveness_config.get("blink_recent_seconds") or 0.0), 0.5)
    has_recent_blink = (current_ts - last_blink_ts) <= required_recent_seconds
    effective_interval = min_interval
    if has_recent_blink:
        effective_interval = max(min_interval, min(required_recent_seconds * 0.25, 0.55))
    else:
        effective_interval = min(min_interval, 0.1)

    if (current_ts - last_ear_ts) >= effective_interval:
        if isinstance(landmark_signature, dict):
            ear_value = landmark_signature.get("ear")
        if ear_value is None and rgb_small is not None and face_location is not None:
            ear_value = extract_live_blink_ear(rgb_small, face_location)
        if ear_value is None and rgb_full is not None and face_location_full is not None:
            normalized_full_location = _normalized_face_location(face_location_full)
            if normalized_full_location:
                full_top, full_right, full_bottom, full_left = normalized_full_location
                full_height = max(full_bottom - full_top, 0)
                full_width = max(full_right - full_left, 0)
                if full_height >= 56 and full_width >= 56:
                    ear_value = extract_live_blink_ear(rgb_full, normalized_full_location)
        if ear_value is not None:
            closed_threshold = float(LIVE_LIVENESS_BLINK_EAR_CLOSED)
            open_threshold = max(float(LIVE_LIVENESS_BLINK_EAR_OPEN), closed_threshold + 0.01)
            min_closed_frames = max(int(LIVE_LIVENESS_BLINK_MIN_CLOSED_FRAMES or 1), 1)

            if ear_value <= closed_threshold:
                eye_closed_frames += 1
            elif ear_value >= open_threshold:
                if eye_closed_frames >= min_closed_frames and (current_ts - last_blink_ts) >= 0.18:
                    blink_count += 1
                    blink_detected = True
                    last_blink_ts = current_ts
                eye_closed_frames = 0
            else:
                eye_closed_frames = max(eye_closed_frames - 1, 0)

            updates["liveness_last_ear"] = float(round(float(ear_value), 6))
            updates["liveness_last_ear_ts"] = float(current_ts)
            updates["liveness_eye_closed_frames"] = int(eye_closed_frames)
            updates["liveness_blink_count"] = int(blink_count)
            updates["liveness_last_blink_ts"] = float(last_blink_ts)

    has_recent_blink = (current_ts - last_blink_ts) <= required_recent_seconds

    if blink_required and not has_recent_blink:
        return {
            "accepted": True,
            "reason": "liveness_blink_pending",
            "message": "",
            "updates": updates,
            "blink_detected": blink_detected,
        }

    if blink_detected:
        return {
            "accepted": True,
            "reason": "blink_detected",
            "message": "Blink Detected",
            "updates": updates,
            "blink_detected": True,
        }

    return {
        "accepted": True,
        "reason": "blink_ok",
        "message": "",
        "updates": updates,
        "blink_detected": False,
    }


def summarize_live_track_liveness_support(track, liveness_config=None, now_ts=None):
    row = track if isinstance(track, dict) else {}
    config = liveness_config if isinstance(liveness_config, dict) else resolve_live_liveness_profile_thresholds()
    current_ts = float(now_ts if now_ts is not None else time.time())

    liveness_frames = int(row.get("liveness_frames") or 0)
    motion_score = float(row.get("liveness_motion_score") or 0.0)
    pose_score = float(row.get("liveness_pose_score") or 0.0)
    first_seen_ts = float(row.get("first_seen_ts") or current_ts)
    track_age_seconds = max(current_ts - first_seen_ts, 0.0)

    min_frames = max(int(config.get("min_track_frames") or 1), 1)
    min_motion = max(float(config.get("min_motion_score") or 0.0), 0.0)
    min_pose = max(float(config.get("min_pose_score") or 0.0), 0.0)
    pose_grace_seconds = max(float(config.get("pose_grace_seconds") or 0.0), 0.0)
    min_parallax = max(float(config.get("min_parallax_score") or 0.0), 0.0)
    required_streak = max(int(config.get("planar_streak_frames") or 4), 2)
    min_landmark_pose_span = max(float(config.get("min_landmark_pose_span") or 0.0), 0.0)
    min_landmark_pose_samples = max(int(config.get("min_landmark_pose_samples") or 1), 1)
    min_patch_diversity = max(float(config.get("min_patch_diversity") or 0.0), 0.0)

    parallax_score = float(row.get("liveness_parallax_score") or 0.0)
    planar_streak = int(row.get("liveness_planar_streak") or 0)
    patch_diversity = float(row.get("liveness_patch_diversity") or 0.0)
    landmark_pose_span = float(row.get("liveness_landmark_pose_span") or 0.0)
    landmark_pose_samples = int(row.get("liveness_landmark_pose_samples") or 0)
    last_blink_ts = float(row.get("liveness_last_blink_ts") or 0.0)
    blink_recent_seconds = max(float(config.get("blink_recent_seconds") or 0.0), 0.5)
    has_recent_blink = (current_ts - last_blink_ts) <= blink_recent_seconds

    age_relax = min(max(track_age_seconds / max(pose_grace_seconds + 0.6, 0.9), 0.0), 1.0)
    effective_min_motion = max(min_motion * (1.0 - (0.28 * age_relax)), min_motion * 0.68)
    effective_min_pose = max(min_pose * (1.0 - (0.22 * age_relax)), min_pose * 0.65)

    landmark_pose_ok = (
        landmark_pose_samples >= min_landmark_pose_samples
        and landmark_pose_span >= min_landmark_pose_span
    )
    partial_landmark_pose = (
        landmark_pose_samples >= max(min_landmark_pose_samples - 1, 1)
        and landmark_pose_span >= max(min_landmark_pose_span * 0.72, min_landmark_pose_span - 0.02, 0.03)
    )
    strong_depth_signal = (
        (not LIVE_LIVENESS_PATCH_PARALLAX_ENABLED)
        or parallax_score >= max((min_parallax * 0.86), min_parallax - 0.006, 0.026)
        or planar_streak < max(required_streak - 1, 2)
    )
    strong_motion_signal = motion_score >= max((effective_min_motion * 1.08), effective_min_motion + 0.015)
    strong_pose_signal = (
        pose_score >= max((effective_min_pose * 0.94), effective_min_pose - 0.004, 0.012)
        or landmark_pose_ok
        or track_age_seconds >= pose_grace_seconds
    )
    patch_diversity_ok = (
        patch_diversity >= max((min_patch_diversity * 0.84), 0.08)
        or parallax_score >= max(min_parallax * 0.92, min_parallax - 0.004, 0.03)
    )
    strong_non_blink_liveness = (
        liveness_frames >= (min_frames + 1)
        and strong_motion_signal
        and strong_pose_signal
        and strong_depth_signal
        and patch_diversity_ok
        and (
            landmark_pose_ok
            or (
                partial_landmark_pose
                and parallax_score >= max(min_parallax * 0.82, min_parallax - 0.008, 0.024)
            )
            or (
                pose_score >= max((effective_min_pose * 1.08), effective_min_pose + 0.004, 0.02)
                and parallax_score >= max(min_parallax * 1.02, min_parallax + 0.002, 0.032)
            )
        )
    )
    fast_ready_age_seconds = min(
        max(float(config.get("pose_sample_interval_seconds") or 0.0) * 5.0, 0.42),
        0.95,
    )
    fast_motion_signal = motion_score >= max((effective_min_motion * 1.14), effective_min_motion + 0.02, 0.16)
    fast_pose_signal = (
        landmark_pose_ok
        or (
            partial_landmark_pose
            and pose_score >= max((effective_min_pose * 0.98), effective_min_pose - 0.002, 0.013)
        )
    )
    fast_depth_signal = (
        (not LIVE_LIVENESS_PATCH_PARALLAX_ENABLED)
        or (
            parallax_score >= max(min_parallax * 0.9, min_parallax - 0.004, 0.026)
            and patch_diversity_ok
        )
    )
    fast_non_blink_liveness = (
        liveness_frames >= min_frames
        and track_age_seconds >= fast_ready_age_seconds
        and fast_motion_signal
        and fast_pose_signal
        and fast_depth_signal
    )

    return {
        "current_ts": current_ts,
        "liveness_frames": liveness_frames,
        "motion_score": motion_score,
        "pose_score": pose_score,
        "track_age_seconds": track_age_seconds,
        "min_frames": min_frames,
        "effective_min_motion": effective_min_motion,
        "effective_min_pose": effective_min_pose,
        "pose_grace_seconds": pose_grace_seconds,
        "min_parallax": min_parallax,
        "required_streak": required_streak,
        "parallax_score": parallax_score,
        "planar_streak": planar_streak,
        "patch_diversity": patch_diversity,
        "min_patch_diversity": min_patch_diversity,
        "landmark_pose_span": landmark_pose_span,
        "landmark_pose_samples": landmark_pose_samples,
        "min_landmark_pose_span": min_landmark_pose_span,
        "min_landmark_pose_samples": min_landmark_pose_samples,
        "has_recent_blink": has_recent_blink,
        "landmark_pose_ok": landmark_pose_ok,
        "partial_landmark_pose": partial_landmark_pose,
        "strong_depth_signal": strong_depth_signal,
        "strong_motion_signal": strong_motion_signal,
        "strong_pose_signal": strong_pose_signal,
        "patch_diversity_ok": patch_diversity_ok,
        "strong_non_blink_liveness": strong_non_blink_liveness,
        "fast_ready_age_seconds": fast_ready_age_seconds,
        "fast_motion_signal": fast_motion_signal,
        "fast_pose_signal": fast_pose_signal,
        "fast_depth_signal": fast_depth_signal,
        "fast_non_blink_liveness": fast_non_blink_liveness,
    }


def evaluate_live_landmark_pose_liveness(track, now_ts=None):
    if not LIVE_LIVENESS_ENABLED or not LIVE_LIVENESS_LANDMARK_POSE_ENABLED:
        return {"accepted": True, "reason": "landmark_pose_disabled", "message": ""}

    liveness_config = resolve_live_liveness_profile_thresholds()
    support = summarize_live_track_liveness_support(track, liveness_config=liveness_config, now_ts=now_ts)
    row = track if isinstance(track, dict) else {}
    current_ts = float(support.get("current_ts") or (now_ts if now_ts is not None else time.time()))
    min_pose_span = max(float(liveness_config.get("min_landmark_pose_span") or 0.0), 0.0)
    min_pose_samples = max(int(liveness_config.get("min_landmark_pose_samples") or 1), 1)
    pose_span = float(row.get("liveness_landmark_pose_span") or 0.0)
    pose_samples = int(row.get("liveness_landmark_pose_samples") or 0)
    last_landmark_ts = float(row.get("liveness_last_landmark_ts") or 0.0)
    pose_window_seconds = max(float(liveness_config.get("pose_window_seconds") or 0.0), 1.0)

    if (current_ts - last_landmark_ts) > pose_window_seconds:
        if bool(support.get("has_recent_blink")) and bool(support.get("strong_depth_signal")):
            return {"accepted": True, "reason": "landmark_pose_temporarily_unavailable", "message": ""}
        return {
            "accepted": True,
            "reason": "landmark_pose_building",
            "message": "",
        }
    if pose_samples < min_pose_samples:
        if bool(support.get("fast_non_blink_liveness")) or (
            bool(support.get("has_recent_blink")) and bool(support.get("strong_depth_signal"))
        ):
            return {"accepted": True, "reason": "landmark_pose_building", "message": ""}
        return {
            "accepted": True,
            "reason": "landmark_pose_building",
            "message": "",
        }
    if pose_span < min_pose_span:
        if bool(support.get("strong_non_blink_liveness")) or bool(support.get("fast_non_blink_liveness")) or (
            bool(support.get("has_recent_blink"))
            and bool(support.get("strong_depth_signal"))
            and float(support.get("pose_score") or 0.0) >= max(float(support.get("effective_min_pose") or 0.0) * 0.9, 0.012)
        ):
            return {"accepted": True, "reason": "landmark_pose_compensated", "message": ""}
        return {
            "accepted": False,
            "reason": "liveness_pose_landmarks_required",
            "message": "Liveness check required: blink or turn head slightly",
        }
    return {"accepted": True, "reason": "landmark_pose_ok", "message": ""}


def extract_live_liveness_patch(frame, face_location, scale_back=1.0):
    if frame is None or frame.size == 0:
        return None
    if not isinstance(face_location, (list, tuple)) or len(face_location) != 4:
        return None

    try:
        frame_height, frame_width = frame.shape[:2]
        top, right, bottom, left = face_location
        top = max(0, min(frame_height, int(round(float(top) * scale_back))))
        right = max(0, min(frame_width, int(round(float(right) * scale_back))))
        bottom = max(0, min(frame_height, int(round(float(bottom) * scale_back))))
        left = max(0, min(frame_width, int(round(float(left) * scale_back))))
    except Exception:
        return None

    if bottom <= top or right <= left:
        return None

    pad_y = int(round((bottom - top) * 0.12))
    pad_x = int(round((right - left) * 0.12))
    top = max(0, top - pad_y)
    left = max(0, left - pad_x)
    bottom = min(frame_height, bottom + pad_y)
    right = min(frame_width, right + pad_x)
    if bottom <= top or right <= left:
        return None

    roi = frame[top:bottom, left:right]
    if roi is None or roi.size == 0:
        return None

    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    if gray.size == 0:
        return None
    gray = cv2.equalizeHist(gray)
    patch_size = max(int(LIVE_LIVENESS_PATCH_SIZE or 24), 12)
    patch = cv2.resize(gray, (patch_size, patch_size), interpolation=cv2.INTER_AREA)
    grad_patch = cv2.Laplacian(patch, cv2.CV_32F)
    return patch, grad_patch


def evaluate_live_patch_parallax_liveness(track, frame, face_location, scale_back=1.0, now_ts=None):
    if not LIVE_LIVENESS_ENABLED or not LIVE_LIVENESS_PATCH_PARALLAX_ENABLED:
        return {"accepted": True, "reason": "patch_parallax_disabled", "message": "", "updates": {}}

    liveness_config = resolve_live_liveness_profile_thresholds()
    patch_bundle = extract_live_liveness_patch(frame, face_location, scale_back=scale_back)
    if not patch_bundle:
        return {"accepted": True, "reason": "patch_unavailable", "message": "", "updates": {}}

    patch, grad_patch = patch_bundle
    previous_patch = track.get("liveness_prev_patch")
    previous_grad_patch = track.get("liveness_prev_grad_patch")
    patch_delta = 0.0
    grad_delta = 0.0
    patch_diversity = float(track.get("liveness_patch_diversity") or 0.0)

    if isinstance(previous_patch, np.ndarray) and previous_patch.shape == patch.shape:
        patch_delta = float(np.mean(cv2.absdiff(patch, previous_patch)) / 255.0)
        grid_rows = max(2, min(3, patch.shape[0] // 8 or 2))
        grid_cols = max(2, min(3, patch.shape[1] // 8 or 2))
        cell_deltas = []
        row_edges = np.linspace(0, patch.shape[0], grid_rows + 1, dtype=int)
        col_edges = np.linspace(0, patch.shape[1], grid_cols + 1, dtype=int)
        for row_idx in range(grid_rows):
            for col_idx in range(grid_cols):
                row_start = int(row_edges[row_idx])
                row_end = int(row_edges[row_idx + 1])
                col_start = int(col_edges[col_idx])
                col_end = int(col_edges[col_idx + 1])
                if row_end <= row_start or col_end <= col_start:
                    continue
                current_cell = patch[row_start:row_end, col_start:col_end]
                previous_cell = previous_patch[row_start:row_end, col_start:col_end]
                if current_cell.size == 0 or previous_cell.size == 0:
                    continue
                cell_delta = float(np.mean(cv2.absdiff(current_cell, previous_cell)) / 255.0)
                cell_deltas.append(cell_delta)
        if cell_deltas:
            mean_cell_delta = float(np.mean(cell_deltas))
            patch_diversity = float(
                np.std(cell_deltas) / max(mean_cell_delta, 1e-4)
            ) if mean_cell_delta > 1e-4 else 0.0
    if isinstance(previous_grad_patch, np.ndarray) and previous_grad_patch.shape == grad_patch.shape:
        grad_delta = float(np.mean(np.abs(grad_patch - previous_grad_patch)) / 255.0)

    motion_component = float(track.get("liveness_motion_component") or 0.0)
    pose_component = float(track.get("liveness_pose_component") or 0.0)
    planar_streak = int(track.get("liveness_planar_streak") or 0)
    previous_parallax_score = float(track.get("liveness_parallax_score") or 0.0)
    decay = min(max(float(LIVE_LIVENESS_SCORE_DECAY or 0.88), 0.5), 0.99)

    parallax_step = max(
        (patch_delta * 1.18)
        + (grad_delta * 0.82)
        + (pose_component * 0.35)
        - 0.018,
        0.0,
    )
    parallax_score = max((previous_parallax_score * decay) + parallax_step, 0.0)

    min_patch_diversity = max(float(liveness_config.get("min_patch_diversity") or 0.0), 0.0)
    suspicious_uniform_motion = (
        motion_component >= max(float(LIVE_LIVENESS_MOTION_COMPONENT_THRESHOLD) * 1.08, 0.03)
        and pose_component <= max(float(liveness_config.get("max_planar_pose_component") or 0.0), 0.028)
        and patch_diversity <= max(min_patch_diversity * 0.82, 0.055)
        and patch_delta >= max(float(liveness_config.get("max_planar_patch_delta") or 0.0) * 0.82, 0.01)
    )
    planar_motion_detected = (
        motion_component >= float(LIVE_LIVENESS_MOTION_COMPONENT_THRESHOLD)
        and pose_component <= float(liveness_config.get("max_planar_pose_component") or 0.0)
        and patch_delta <= float(liveness_config.get("max_planar_patch_delta") or 0.0)
        and grad_delta <= float(liveness_config.get("max_planar_grad_delta") or 0.0)
    )
    if planar_motion_detected or suspicious_uniform_motion:
        planar_streak += 1
    else:
        planar_streak = max(planar_streak - 1, 0)

    updates = {
        "liveness_prev_patch": patch,
        "liveness_prev_grad_patch": grad_patch,
        "liveness_patch_delta": float(round(patch_delta, 6)),
        "liveness_patch_grad_delta": float(round(grad_delta, 6)),
        "liveness_patch_diversity": float(round(patch_diversity, 6)),
        "liveness_parallax_score": float(round(parallax_score, 6)),
        "liveness_planar_streak": int(planar_streak),
        "liveness_patch_ts": float(now_ts if now_ts is not None else time.time()),
    }

    required_streak = max(int(liveness_config.get("planar_streak_frames") or 4), 2)
    min_parallax = max(float(liveness_config.get("min_parallax_score") or 0.0), 0.0)
    last_blink_ts = float(track.get("liveness_last_blink_ts") or 0.0)
    blink_recent_window = max(float(liveness_config.get("blink_recent_seconds") or 0.0), 0.5)
    has_recent_blink = (float(now_ts if now_ts is not None else time.time()) - last_blink_ts) <= blink_recent_window
    near_parallax_threshold = parallax_score >= (min_parallax * 0.74)
    blink_backed_motion = (
        has_recent_blink
        and motion_component >= max(float(LIVE_LIVENESS_MOTION_COMPONENT_THRESHOLD) * 0.82, 0.02)
        and pose_component >= max(float(liveness_config.get("max_planar_pose_component") or 0.0) * 0.4, 0.004)
        and near_parallax_threshold
    )
    if planar_streak >= required_streak and parallax_score < min_parallax and not blink_backed_motion:
        return {
            "accepted": False,
            "reason": "liveness_flat_motion_pattern",
            "message": "Liveness check failed: flat motion detected",
            "updates": updates,
        }

    return {
        "accepted": True,
        "reason": "patch_parallax_ok",
        "message": "",
        "updates": updates,
    }


def evaluate_live_display_liveness(track, frame, face_location, face_quality=None, scale_back=1.0):
    if not LIVE_LIVENESS_ENABLED or not LIVE_LIVENESS_DISPLAY_CHECK_ENABLED:
        return {"accepted": True, "reason": "display_check_disabled", "message": "", "updates": {}}
    if frame is None or getattr(frame, "size", 0) == 0:
        return {"accepted": True, "reason": "display_frame_missing", "message": "", "updates": {}}
    if not isinstance(face_location, (list, tuple)) or len(face_location) != 4:
        return {"accepted": True, "reason": "display_location_missing", "message": "", "updates": {}}

    try:
        frame_height, frame_width = frame.shape[:2]
        top, right, bottom, left = face_location
        top = max(0, min(frame_height, int(round(float(top) * scale_back))))
        right = max(0, min(frame_width, int(round(float(right) * scale_back))))
        bottom = max(0, min(frame_height, int(round(float(bottom) * scale_back))))
        left = max(0, min(frame_width, int(round(float(left) * scale_back))))
    except Exception:
        return {"accepted": True, "reason": "display_crop_invalid", "message": "", "updates": {}}

    if bottom <= top or right <= left:
        return {"accepted": True, "reason": "display_crop_invalid", "message": "", "updates": {}}

    face_width = max(right - left, 1)
    face_height = max(bottom - top, 1)
    pad_x = int(round(face_width * 0.7))
    pad_y = int(round(face_height * 0.7))
    roi_left = max(0, left - pad_x)
    roi_top = max(0, top - pad_y)
    roi_right = min(frame_width, right + pad_x)
    roi_bottom = min(frame_height, bottom + pad_y)
    if roi_bottom <= roi_top or roi_right <= roi_left:
        return {"accepted": True, "reason": "display_roi_invalid", "message": "", "updates": {}}

    roi = frame[roi_top:roi_bottom, roi_left:roi_right]
    if roi is None or roi.size == 0:
        return {"accepted": True, "reason": "display_roi_missing", "message": "", "updates": {}}

    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(gray, 60, 150)
    edge_density = float(np.mean(edges > 0))

    face_center_local = (
        int((left + right) / 2) - roi_left,
        int((top + bottom) / 2) - roi_top,
    )
    face_area = float(face_width * face_height)
    min_area_ratio = max(float(LIVE_LIVENESS_DISPLAY_MIN_RECT_AREA_RATIO or 0.0), 1.0)
    max_area_ratio = max(float(LIVE_LIVENESS_DISPLAY_MAX_RECT_AREA_RATIO or 0.0), min_area_ratio + 0.5)
    best_rect_score = 0.0

    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    for contour in contours:
        contour_area = float(cv2.contourArea(contour))
        if contour_area <= 0.0:
            continue
        perimeter = float(cv2.arcLength(contour, True))
        if perimeter <= 0.0:
            continue
        approx = cv2.approxPolyDP(contour, 0.04 * perimeter, True)
        if len(approx) != 4 or not cv2.isContourConvex(approx):
            continue
        x, y, width, height = cv2.boundingRect(approx)
        if width <= face_width or height <= face_height:
            continue
        rect_area = float(width * height)
        rect_area_ratio = rect_area / max(face_area, 1.0)
        if rect_area_ratio < min_area_ratio or rect_area_ratio > max_area_ratio:
            continue
        if not (x <= face_center_local[0] <= (x + width) and y <= face_center_local[1] <= (y + height)):
            continue
        aspect_ratio = float(width / max(height, 1))
        if aspect_ratio < 0.42 or aspect_ratio > 2.25:
            continue
        rectangularity = contour_area / max(rect_area, 1.0)
        if rectangularity < 0.72:
            continue
        border_mask = np.zeros_like(gray)
        cv2.drawContours(border_mask, [approx], -1, 255, 2)
        border_edge_density = float(np.mean(edges[border_mask > 0] > 0)) if np.any(border_mask > 0) else 0.0
        rect_score = (rectangularity * 0.48) + (min(border_edge_density, 1.0) * 0.52)
        if rect_score > best_rect_score:
            best_rect_score = rect_score

    quality = face_quality if isinstance(face_quality, dict) else {}
    texture_value = float(quality.get("texture") or 0.0)
    contrast_value = float(quality.get("contrast") or 0.0)
    highlight_ratio = float(quality.get("highlights") or 0.0)
    low_detail_surface = (
        texture_value < max(float(LIVE_LIVENESS_MIN_TEXTURE) * 1.12, 6.6)
        and contrast_value < max(float(LIVE_LIVENESS_MIN_CONTRAST) * 1.08, 13.5)
    )
    reflective_surface = (
        highlight_ratio > min(float(LIVE_LIVENESS_MAX_HIGHLIGHT_RATIO) * 0.88, 0.36)
        and contrast_value < max(float(LIVE_LIVENESS_MIN_CONTRAST) * 1.12, 14.5)
    )
    suspicious_surface = low_detail_surface or reflective_surface
    display_suspected = best_rect_score >= 0.65 and edge_density >= 0.02 and suspicious_surface

    display_streak = int((track or {}).get("liveness_display_streak") or 0)
    if display_suspected:
        display_streak += 1
    else:
        display_streak = max(display_streak - 1, 0)

    updates = {
        "liveness_display_score": float(round(best_rect_score, 6)),
        "liveness_display_streak": int(display_streak),
    }
    if display_streak >= max(int(LIVE_LIVENESS_DISPLAY_MIN_STREAK or 1), 3):
        return {
            "accepted": False,
            "reason": "liveness_display_medium_detected",
            "message": "Liveness check failed: remove photo or screen",
            "updates": updates,
        }
    return {"accepted": True, "reason": "display_ok", "message": "", "updates": updates}


def get_liveness_ai_model():
    global liveness_ai_model, liveness_ai_model_loaded, liveness_ai_model_error
    if not LIVE_LIVENESS_AI_MODEL_ENABLED:
        return None
    if liveness_ai_model_loaded:
        return liveness_ai_model

    with liveness_ai_model_lock:
        if liveness_ai_model_loaded:
            return liveness_ai_model
        liveness_ai_model_loaded = True
        liveness_ai_model = None
        liveness_ai_model_error = ""

        model_path = str(LIVE_LIVENESS_AI_MODEL_PATH or "").strip()
        if not model_path:
            liveness_ai_model_error = "AI model path missing"
            return None
        if not os.path.exists(model_path):
            liveness_ai_model_error = f"AI model not found: {model_path}"
            return None
        try:
            if model_path.lower().endswith(".onnx"):
                liveness_ai_model = cv2.dnn.readNetFromONNX(model_path)
            else:
                liveness_ai_model_error = "Unsupported AI model format (use .onnx)"
        except Exception as exc:
            liveness_ai_model_error = f"AI model load failed: {exc}"
            liveness_ai_model = None
    return liveness_ai_model


def evaluate_live_ai_liveness(frame, face_location, scale_back=1.0):
    if not LIVE_LIVENESS_ENABLED or not LIVE_LIVENESS_AI_MODEL_ENABLED:
        return {"accepted": True, "reason": "ai_disabled", "message": ""}

    model = get_liveness_ai_model()
    if model is None:
        return {"accepted": True, "reason": "ai_unavailable", "message": ""}

    if frame is None or getattr(frame, "size", 0) == 0:
        return {"accepted": True, "reason": "ai_frame_missing", "message": ""}
    location = _normalized_face_location(face_location)
    if not location:
        return {"accepted": True, "reason": "ai_location_missing", "message": ""}

    frame_height, frame_width = frame.shape[:2]
    top, right, bottom, left = location
    top = max(0, min(frame_height, int(round(float(top) * scale_back))))
    right = max(0, min(frame_width, int(round(float(right) * scale_back))))
    bottom = max(0, min(frame_height, int(round(float(bottom) * scale_back))))
    left = max(0, min(frame_width, int(round(float(left) * scale_back))))
    if bottom <= top or right <= left:
        return {"accepted": True, "reason": "ai_crop_invalid", "message": ""}

    crop = frame[top:bottom, left:right]
    if crop is None or crop.size == 0:
        return {"accepted": True, "reason": "ai_crop_missing", "message": ""}

    try:
        resized = cv2.resize(crop, (128, 128), interpolation=cv2.INTER_LINEAR)
        blob = cv2.dnn.blobFromImage(
            resized,
            scalefactor=1.0 / 255.0,
            size=(128, 128),
            mean=(0.0, 0.0, 0.0),
            swapRB=True,
            crop=False,
        )
        model.setInput(blob)
        output = model.forward()
        vector = np.array(output).reshape(-1)
        if vector.size == 0:
            return {"accepted": True, "reason": "ai_output_empty", "message": ""}
        if vector.size == 1:
            live_score = float(vector[0])
        else:
            # Common binary heads: [spoof, live] or logits; use second index when available.
            live_score = float(vector[1] if vector.size > 1 else vector[-1])
        if live_score < float(LIVE_LIVENESS_AI_MIN_LIVE_SCORE):
            return {"accepted": False, "reason": "liveness_ai_spoof", "message": "Verification Failed"}
        return {"accepted": True, "reason": "ai_live_ok", "message": ""}
    except Exception:
        return {"accepted": True, "reason": "ai_inference_error", "message": ""}


def evaluate_live_track_liveness(track, face_quality=None, now_ts=None):
    if not LIVE_LIVENESS_ENABLED:
        return {"accepted": True, "reason": "liveness_disabled", "message": ""}

    liveness_config = resolve_live_liveness_profile_thresholds()
    support = summarize_live_track_liveness_support(track, liveness_config=liveness_config, now_ts=now_ts)
    liveness_frames = int(support.get("liveness_frames") or 0)
    motion_score = float(support.get("motion_score") or 0.0)
    pose_score = float(support.get("pose_score") or 0.0)
    track_age_seconds = float(support.get("track_age_seconds") or 0.0)
    min_frames = int(support.get("min_frames") or 1)
    effective_min_motion = float(support.get("effective_min_motion") or 0.0)
    effective_min_pose = float(support.get("effective_min_pose") or 0.0)
    pose_grace_seconds = float(support.get("pose_grace_seconds") or 0.0)
    min_parallax = float(support.get("min_parallax") or 0.0)
    required_streak = int(support.get("required_streak") or 2)
    planar_streak = int(support.get("planar_streak") or 0)
    parallax_score = float(support.get("parallax_score") or 0.0)
    has_recent_blink = bool(support.get("has_recent_blink"))
    landmark_pose_ok = bool(support.get("landmark_pose_ok"))
    strong_non_blink_liveness = bool(support.get("strong_non_blink_liveness"))
    fast_non_blink_liveness = bool(support.get("fast_non_blink_liveness"))

    if liveness_frames < min_frames:
        return {"accepted": False, "reason": "liveness_building", "message": "Liveness Check Required"}
    if motion_score < effective_min_motion and not fast_non_blink_liveness:
        return {"accepted": False, "reason": "liveness_motion_required", "message": "Liveness Check Required"}
    if (
        pose_score < effective_min_pose
        and track_age_seconds < pose_grace_seconds
        and not landmark_pose_ok
        and not fast_non_blink_liveness
    ):
        return {"accepted": False, "reason": "liveness_pose_required", "message": "Liveness Check Required"}
    if (
        LIVE_LIVENESS_PATCH_PARALLAX_ENABLED
        and planar_streak >= required_streak
        and parallax_score < min_parallax
        and liveness_frames >= (min_frames + 1)
        and not has_recent_blink
        and not strong_non_blink_liveness
        and not fast_non_blink_liveness
    ):
        return {"accepted": False, "reason": "liveness_flat_motion_pattern", "message": "Verification Failed"}

    if LIVE_LIVENESS_BLINK_ENABLED and bool(liveness_config.get("blink_required")):
        if not has_recent_blink and not strong_non_blink_liveness and not fast_non_blink_liveness:
            return {
                "accepted": False,
                "reason": "liveness_blink_or_motion_required",
                "message": "Liveness check required: blink or turn head slightly",
            }

    if face_quality is None:
        if strong_non_blink_liveness and not has_recent_blink:
            return {"accepted": True, "reason": "liveness_strong_non_blink_ok", "message": ""}
        if fast_non_blink_liveness and not has_recent_blink:
            return {"accepted": True, "reason": "liveness_fast_non_blink_ok", "message": ""}
        return {"accepted": True, "reason": "liveness_motion_pose_ok", "message": ""}

    texture_gate = evaluate_live_texture_liveness(face_quality)
    if not texture_gate.get("accepted"):
        return texture_gate
    if strong_non_blink_liveness and not has_recent_blink:
        return {"accepted": True, "reason": "liveness_strong_non_blink_ok", "message": ""}
    if fast_non_blink_liveness and not has_recent_blink:
        return {"accepted": True, "reason": "liveness_fast_non_blink_ok", "message": ""}
    return {"accepted": True, "reason": "liveness_ok", "message": ""}


def set_live_face_track_state(track_id, *, next_attempt_ts=None, student_id=None, last_result=None, last_confidence=None):
    try:
        normalized_track_id = int(track_id)
    except (TypeError, ValueError):
        return

    with scan_lock:
        tracks = scan_state.get("face_tracks")
        if not isinstance(tracks, dict):
            return
        track = tracks.get(normalized_track_id)
        if not isinstance(track, dict):
            return
        if next_attempt_ts is not None:
            track["next_attempt_ts"] = float(next_attempt_ts)
        if student_id is not None:
            track["student_id"] = str(student_id or "").strip()
        if last_result is not None:
            track["last_result"] = str(last_result)
        if last_confidence is not None:
            try:
                track["last_confidence"] = float(last_confidence)
            except (TypeError, ValueError):
                track["last_confidence"] = 0.0


def set_live_face_track_liveness_state(track_id, **updates):
    if not updates:
        return
    try:
        normalized_track_id = int(track_id)
    except (TypeError, ValueError):
        return

    with scan_lock:
        tracks = scan_state.get("face_tracks")
        if not isinstance(tracks, dict):
            return
        track = tracks.get(normalized_track_id)
        if not isinstance(track, dict):
            return
        for key, value in updates.items():
            if not key:
                continue
            track[str(key)] = value


def prepare_live_encoding_frame(frame, face_locations):
    if frame is None or frame.size == 0:
        return frame, list(face_locations or [])

    frame_height, frame_width = frame.shape[:2]
    max_width = max(int(LIVE_RECOGNITION_ENCODING_MAX_WIDTH or 0), 1)
    if frame_width <= max_width:
        return frame, list(face_locations or [])

    resize_scale = max_width / float(frame_width)
    resized_height = max(1, int(round(frame_height * resize_scale)))
    resized_frame = cv2.resize(frame, (max_width, resized_height), interpolation=cv2.INTER_LINEAR)

    resized_locations = []
    for location in list(face_locations or []):
        if not isinstance(location, (list, tuple)) or len(location) != 4:
            continue
        top, right, bottom, left = location
        resized_locations.append((
            max(0, min(resized_height, int(round(float(top) * resize_scale)))),
            max(0, min(max_width, int(round(float(right) * resize_scale)))),
            max(0, min(resized_height, int(round(float(bottom) * resize_scale)))),
            max(0, min(max_width, int(round(float(left) * resize_scale)))),
        ))

    return resized_frame, resized_locations


def build_live_encoding_variants(frame, low_light_hint=False):
    variants = []
    if frame is None or frame.size == 0:
        return variants
    variants.append(frame)
    if not low_light_hint and not LIVE_RECOGNITION_LOW_LIGHT_ENCODING_RETRY_ENABLED:
        return variants

    try:
        lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
        l_channel, a_channel, b_channel = cv2.split(lab)
        clip_limit = 2.6 if low_light_hint else 2.0
        clahe = cv2.createCLAHE(clipLimit=clip_limit, tileGridSize=(8, 8))
        balanced = cv2.cvtColor(cv2.merge((clahe.apply(l_channel), a_channel, b_channel)), cv2.COLOR_LAB2BGR)
        variants.append(balanced)

        softened = cv2.GaussianBlur(balanced, (0, 0), 1.0 if low_light_hint else 0.85)
        sharpened = cv2.addWeighted(balanced, 1.24 if low_light_hint else 1.18, softened, -0.24 if low_light_hint else -0.18, 0)
        variants.append(sharpened)
    except Exception:
        pass
    return variants


def encode_live_face_locations(frame, face_locations, *, num_jitters=1, low_light_hint=False):
    normalized_locations = [
        _normalized_face_location(location)
        for location in list(face_locations or [])
    ]
    normalized_locations = [location for location in normalized_locations if location]
    if frame is None or frame.size == 0 or not normalized_locations:
        return []

    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    try:
        batch_rows = face_recognition.face_encodings(
            rgb_frame,
            normalized_locations,
            num_jitters=num_jitters,
            model="small",
        )
    except Exception:
        batch_rows = []

    if len(batch_rows) == len(normalized_locations):
        return list(batch_rows)

    per_location_rows = [None] * len(normalized_locations)
    variant_frames = build_live_encoding_variants(frame, low_light_hint=low_light_hint)
    variant_rgbs = []
    for variant_frame in variant_frames:
        try:
            variant_rgbs.append(cv2.cvtColor(variant_frame, cv2.COLOR_BGR2RGB))
        except Exception:
            continue
    if not variant_rgbs:
        variant_rgbs = [rgb_frame]

    for index, location in enumerate(normalized_locations):
        for variant_rgb in variant_rgbs:
            try:
                encoded_rows = face_recognition.face_encodings(
                    variant_rgb,
                    [location],
                    num_jitters=num_jitters,
                    model="small",
                )
            except Exception:
                encoded_rows = []
            if encoded_rows:
                per_location_rows[index] = encoded_rows[0]
                break

    return per_location_rows


def evaluate_live_recognition_match(match_result, confidence_pct, face_quality):
    candidate = match_result.get("candidate") or {}
    runner_up = match_result.get("runner_up")
    best_distance = float(match_result.get("distance") or candidate.get("best_distance") or 1.0)
    score_margin = float(match_result.get("score_margin") or 0.0)
    distance_margin = float(match_result.get("distance_margin") or 0.0)
    brightness = float(face_quality.get("brightness") or 0.0)
    sharpness = float(face_quality.get("sharpness") or 0.0)
    strong_match = best_distance <= float(RECOGNITION_STRONG_MATCH_DISTANCE)
    strict_distance = min(float(RECOGNITION_TOLERANCE), float(LIVE_RECOGNITION_DISTANCE_TOLERANCE))
    strict_confidence = max(float(MIN_RECOGNITION_CONFIDENCE), float(LIVE_RECOGNITION_MIN_CONFIDENCE))
    strict_score_margin = max(float(RECOGNITION_MARGIN_THRESHOLD), float(LIVE_RECOGNITION_MARGIN_THRESHOLD))
    has_margin_data = runner_up is not None or "score_margin" in match_result or "distance_margin" in match_result
    strict_sharpness = float(LIVE_RECOGNITION_MIN_SHARPNESS)

    if brightness > 0.0 and brightness < float(LIVE_RECOGNITION_LOW_LIGHT_BRIGHTNESS):
        strict_confidence = max(float(MIN_RECOGNITION_CONFIDENCE), strict_confidence - 2.0)
        strict_sharpness = min(strict_sharpness, float(LIVE_RECOGNITION_LOW_LIGHT_MIN_SHARPNESS))

    if strong_match:
        strict_confidence = max(float(MIN_RECOGNITION_CONFIDENCE), strict_confidence - 1.0)
        strict_sharpness = max(float(LIVE_RECOGNITION_LOW_LIGHT_MIN_SHARPNESS), strict_sharpness - 0.5)

    if confidence_pct < strict_confidence:
        return {
            "accepted": False,
            "reason": "awaiting_clearer_frame",
            "message": "Holding for a clearer face sample.",
        }

    if best_distance > strict_distance:
        return {
            "accepted": False,
            "reason": "awaiting_better_match",
            "message": "Holding for a stronger face match.",
        }

    if sharpness < strict_sharpness:
        return {
            "accepted": False,
            "reason": "motion_blur",
            "message": "Hold still for a moment to confirm identity.",
        }

    if has_margin_data and not strong_match and score_margin < strict_score_margin:
        return {
            "accepted": False,
            "reason": "awaiting_clearer_separation",
            "message": "Holding for a more distinct face angle.",
        }

    if has_margin_data and not strong_match and distance_margin < float(LIVE_RECOGNITION_DISTANCE_MARGIN_THRESHOLD):
        return {
            "accepted": False,
            "reason": "awaiting_clearer_separation",
            "message": "Holding for a more distinct face angle.",
        }

    return {
        "accepted": True,
        "reason": "stable_match",
        "message": "Stable face match ready.",
    }


def track_pending_live_recognition(student, confidence_pct, match_result, face_quality, now_ts=None):
    student_id = str((student or {}).get("student_id") or "").strip()
    if not student_id:
        return {
            "confirmed": False,
            "observed_frames": 0,
            "required_frames": int(LIVE_RECOGNITION_CONFIRMATION_FRAMES),
        }

    current_ts = float(now_ts if now_ts is not None else time.time())
    confirmation_window_seconds = max(float(LIVE_RECOGNITION_CONFIRMATION_WINDOW_MS) / 1000.0, 0.1)
    candidate = match_result.get("candidate") or {}
    best_distance = float(match_result.get("distance") or candidate.get("best_distance") or 1.0)
    score_margin = float(match_result.get("score_margin") or 0.0)
    distance_margin = float(match_result.get("distance_margin") or 0.0)
    required_frames = max(int(LIVE_RECOGNITION_CONFIRMATION_FRAMES or 0), 1)
    if (
        best_distance <= float(LIVE_RECOGNITION_IMMEDIATE_MATCH_DISTANCE)
        and float(confidence_pct or 0.0) >= max(float(MIN_RECOGNITION_CONFIDENCE), float(LIVE_RECOGNITION_IMMEDIATE_CONFIDENCE))
    ):
        required_frames = 1
    elif (
        best_distance <= float(RECOGNITION_STRONG_MATCH_DISTANCE)
        and float(confidence_pct or 0.0) >= max(float(MIN_RECOGNITION_CONFIDENCE), float(LIVE_RECOGNITION_MIN_CONFIDENCE))
    ):
        required_frames = min(required_frames, max(int(LIVE_RECOGNITION_STRONG_MATCH_CONFIRMATION_FRAMES or 0), 1))

    with scan_lock:
        pending = scan_state.setdefault("pending_recognition", {})
        if not isinstance(pending, dict):
            pending = {}
            scan_state["pending_recognition"] = pending

        stale_student_ids = [
            pending_student_id
            for pending_student_id, entry in pending.items()
            if current_ts - float((entry or {}).get("last_seen_ts") or 0.0) > confirmation_window_seconds
        ]
        for pending_student_id in stale_student_ids:
            pending.pop(pending_student_id, None)

        existing = pending.get(student_id)
        if existing and current_ts - float(existing.get("last_seen_ts") or 0.0) <= confirmation_window_seconds:
            observed_frames = int(existing.get("observed_frames") or 0) + 1
        else:
            observed_frames = 1

        snapshot = {
            "last_seen_ts": current_ts,
            "observed_frames": observed_frames,
            "confidence_pct": max(float(confidence_pct or 0.0), float((existing or {}).get("confidence_pct") or 0.0)),
            "best_distance": min(best_distance, float((existing or {}).get("best_distance") or best_distance)),
            "score_margin": max(score_margin, float((existing or {}).get("score_margin") or 0.0)),
            "distance_margin": max(distance_margin, float((existing or {}).get("distance_margin") or 0.0)),
            "sharpness": max(float(face_quality.get("sharpness") or 0.0), float((existing or {}).get("sharpness") or 0.0)),
        }
        pending[student_id] = snapshot

        confirmed = observed_frames >= required_frames
        if confirmed:
            pending.pop(student_id, None)

    snapshot["confirmed"] = confirmed
    snapshot["required_frames"] = required_frames
    return snapshot


def process_client_frame(frame_bytes):
    """
    Process a frame sent from the client device's camera.
    Performs face recognition and pushes events.
    
    Returns: (success: bool, message: str, payload: dict)
    """
    payload = {"faces": []}
    try:
        # Decode image from bytes
        nparr = np.frombuffer(frame_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if frame is None or frame.size == 0:
            return False, "Failed to decode image", payload
        
        # Get current scan state
        with scan_lock:
            active = scan_state.get("active", False)
            db_encodings = scan_state.get("known_encodings", np.empty((0, 128), dtype=np.float64))
            db_students = scan_state.get("known_students", [])
            model_status = scan_state.get("model_status", "idle")
            session_mode = normalize_scan_session_mode(scan_state.get("session_mode", "auto"), default="auto")
        
        if not active:
            return False, "Scan not active", payload
        
        # Process frame for face recognition
        scale = min(max(SCAN_RECOGNITION_SCALE, 0.25), 1.0)
        
        if scale < 1.0:
            small_frame = cv2.resize(frame, (0, 0), fx=scale, fy=scale, interpolation=cv2.INTER_LINEAR)
        else:
            small_frame = frame
        
        rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        rgb_full_for_blink = (
            cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            if LIVE_SCAN_LIVENESS_ENABLED and (LIVE_LIVENESS_BLINK_ENABLED or LIVE_LIVENESS_LANDMARK_POSE_ENABLED)
            else None
        )
        
        # Detect faces
        face_locations_small = face_recognition.face_locations(
            rgb_small,
            number_of_times_to_upsample=0,
            model="hog",
        )
        retry_scale = min(max(float(SCAN_MULTI_FACE_RETRY_SCALE_PERCENT or 0) / 100.0, scale + 0.1), 0.82)
        should_retry_multi_face_detection = len(face_locations_small) == 0
        if len(face_locations_small) == 1:
            frame_area = max(float(rgb_small.shape[0] * rgb_small.shape[1]), 1.0)
            primary_face_area_ratio = float(_face_location_area(face_locations_small[0])) / frame_area
            should_retry_multi_face_detection = (
                primary_face_area_ratio <= float(SCAN_MULTI_FACE_RETRY_MAX_FACE_AREA_RATIO or 0.16)
            )
        if should_retry_multi_face_detection and retry_scale > (scale + 0.02):
            retry_frame = cv2.resize(frame, (0, 0), fx=retry_scale, fy=retry_scale, interpolation=cv2.INTER_LINEAR)
            retry_rgb = cv2.cvtColor(retry_frame, cv2.COLOR_BGR2RGB)
            retry_locations = face_recognition.face_locations(
                retry_rgb,
                number_of_times_to_upsample=0,
                model="hog",
            )
            if len(retry_locations) > len(face_locations_small):
                small_frame = retry_frame
                rgb_small = retry_rgb
                face_locations_small = retry_locations
                scale = retry_scale
        scale_back = 1.0 / scale if scale > 0 else 1.0

        if len(face_locations_small) == 0:
            clear_pending_live_recognition()
            with scan_lock:
                scan_state["last_faces_payload"] = []
            return True, "No faces detected", payload

        face_locations_small = filter_live_face_locations(face_locations_small, rgb_small.shape)
        if len(face_locations_small) == 0:
            push_not_registered_event("face_too_small", 0.0)
            clear_pending_live_recognition()
            with scan_lock:
                scan_state["last_faces_payload"] = []
            return True, "Face detected but too small for reliable recognition", payload

        frame_now_ts = time.time()
        face_tracks = update_live_face_tracks(
            face_locations_small,
            scale_back,
            frame.shape,
            now_ts=frame_now_ts,
        )
        payload["faces"] = build_live_face_payload_from_tracks(face_tracks, frame.shape)
        with scan_lock:
            scan_state["last_faces_payload"] = [dict(row) for row in payload["faces"]]

        db_encoding_count = int(len(db_encodings)) if db_encodings is not None else 0
        legacy_flat_index = any("encodings" not in (student or {}) for student in db_students)
        face_index = {
            "students": db_students,
            "centroids": db_encodings,
        }

        if model_status == "loading":
            return True, "Model still loading", payload
        elif model_status != "ready" or db_encoding_count == 0:
            reason = "model_not_ready" if model_status == "model_not_ready" else "no_registered_students"
            push_not_registered_event(reason, 0.0)
            return True, f"Not ready: {reason}", payload

        eligible_tracks = []
        liveness_holds = []
        required_stability = max(int(LIVE_RECOGNITION_TRACK_STABILITY_FRAMES or 1), 1)
        base_candidate_limit = max(int(LIVE_RECOGNITION_MAX_RECOGNITIONS_PER_FRAME or 1), 1)
        multi_face_candidate_limit = max(int(LIVE_RECOGNITION_MULTI_FACE_MAX_CANDIDATES or base_candidate_limit), base_candidate_limit)
        max_candidates = (
            min(multi_face_candidate_limit, max(int(LIVE_RECOGNITION_MAX_TRACKED_FACES or 1), 1))
            if len(payload["faces"]) > 1
            else base_candidate_limit
        )
        max_new_encodings = min(max(int(LIVE_RECOGNITION_MAX_NEW_ENCODINGS_PER_FRAME or 1), 1), max_candidates)
        ordered_tracks = sorted(
            face_tracks,
            key=lambda row: -float((row or {}).get("area") or 0.0),
        )
        if ordered_tracks:
            with scan_lock:
                try:
                    track_cursor = int(scan_state.get("face_track_cursor") or 0)
                except (TypeError, ValueError):
                    track_cursor = 0
            track_cursor %= len(ordered_tracks)
            rotated_tracks = ordered_tracks[track_cursor:] + ordered_tracks[:track_cursor]
        else:
            track_cursor = 0
            rotated_tracks = []

        for track in rotated_tracks:
            if len(eligible_tracks) >= max_candidates:
                break
            if int(track.get("stability") or 0) < required_stability:
                continue
            if frame_now_ts < float(track.get("next_attempt_ts") or 0.0):
                continue
            full_location = _normalized_face_location(track.get("full_location"))
            if full_location is None:
                continue
            face_quality = None
            if LIVE_SCAN_LIVENESS_ENABLED:
                landmark_sample = sample_live_landmark_liveness(
                    track,
                    rgb_small=rgb_small,
                    face_location=track.get("small_location"),
                    now_ts=frame_now_ts,
                    rgb_full=rgb_full_for_blink,
                    face_location_full=full_location,
                )
                landmark_updates = landmark_sample.get("updates") or {}
                if landmark_updates:
                    set_live_face_track_liveness_state(track.get("track_id"), **landmark_updates)
                    track.update(landmark_updates)
                blink_liveness_gate = evaluate_live_blink_liveness(
                    track,
                    rgb_small=rgb_small,
                    face_location=track.get("small_location"),
                    now_ts=frame_now_ts,
                    rgb_full=rgb_full_for_blink,
                    face_location_full=full_location,
                    landmark_signature=landmark_sample.get("signature"),
                )
                blink_updates = blink_liveness_gate.get("updates") or {}
                if blink_updates:
                    set_live_face_track_liveness_state(track.get("track_id"), **blink_updates)
                    track.update(blink_updates)
                pose_landmark_gate = evaluate_live_landmark_pose_liveness(track, now_ts=frame_now_ts)

                patch_liveness_gate = None
                if LIVE_LIVENESS_PATCH_PARALLAX_ENABLED:
                    patch_liveness_gate = evaluate_live_patch_parallax_liveness(
                        track,
                        frame,
                        track.get("small_location"),
                        scale_back=scale_back,
                        now_ts=frame_now_ts,
                    )
                    patch_updates = patch_liveness_gate.get("updates") or {}
                    if patch_updates:
                        set_live_face_track_liveness_state(track.get("track_id"), **patch_updates)
                        track.update(patch_updates)

                liveness_gate = evaluate_live_track_liveness(track, face_quality=None, now_ts=frame_now_ts)
                if liveness_gate.get("accepted") and not blink_liveness_gate.get("accepted"):
                    liveness_gate = blink_liveness_gate
                if liveness_gate.get("accepted") and not pose_landmark_gate.get("accepted"):
                    liveness_gate = pose_landmark_gate
                if liveness_gate.get("accepted") and patch_liveness_gate is not None and not patch_liveness_gate.get("accepted"):
                    liveness_gate = patch_liveness_gate
                if liveness_gate.get("accepted") and LIVE_LIVENESS_TEXTURE_CHECK_ENABLED:
                    face_quality = resolve_live_track_face_quality(
                        track,
                        frame,
                        track.get("small_location"),
                        scale_back=scale_back,
                        now_ts=frame_now_ts,
                    )
                    texture_gate = evaluate_live_texture_liveness(face_quality)
                    if not texture_gate.get("accepted"):
                        liveness_gate = texture_gate
                if liveness_gate.get("accepted") and LIVE_LIVENESS_DISPLAY_CHECK_ENABLED:
                    if face_quality is None:
                        face_quality = resolve_live_track_face_quality(
                            track,
                            frame,
                            track.get("small_location"),
                            scale_back=scale_back,
                            now_ts=frame_now_ts,
                        )
                    display_gate = evaluate_live_display_liveness(
                        track,
                        frame,
                        track.get("small_location"),
                        face_quality=face_quality,
                        scale_back=scale_back,
                    )
                    display_updates = display_gate.get("updates") or {}
                    if display_updates:
                        set_live_face_track_liveness_state(track.get("track_id"), **display_updates)
                        track.update(display_updates)
                    if not display_gate.get("accepted"):
                        liveness_gate = display_gate
                if liveness_gate.get("accepted") and LIVE_LIVENESS_AI_MODEL_ENABLED:
                    ai_gate = evaluate_live_ai_liveness(
                        frame,
                        track.get("small_location"),
                        scale_back=scale_back,
                    )
                    if not ai_gate.get("accepted"):
                        liveness_gate = ai_gate
            else:
                blink_liveness_gate = {"accepted": True, "blink_detected": False}
                liveness_gate = {"accepted": True, "reason": "liveness_bypassed", "message": ""}
            if not liveness_gate.get("accepted"):
                hold_message = str(liveness_gate.get("message") or "Liveness Check Required")
                if blink_liveness_gate.get("blink_detected") and hold_message.lower().startswith("liveness check"):
                    hold_message = "Blink Detected"
                liveness_holds.append(hold_message)
                set_live_face_track_state(
                    track.get("track_id"),
                    next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_LIVENESS_RETRY_SECONDS),
                    student_id="",
                    last_result=str(liveness_gate.get("reason") or "liveness_check_required"),
                    last_confidence=0.0,
                )
                continue
            queued_track = dict(track)
            queued_track["track_id"] = int(track.get("track_id") or 0)
            queued_track["small_location"] = track.get("small_location")
            queued_track["full_location"] = full_location
            queued_track["area"] = float(track.get("area") or 0.0)
            queued_track["face_quality"] = (
                dict(face_quality)
                if isinstance(face_quality, dict)
                else dict(track.get("face_quality") or {}) if isinstance(track.get("face_quality"), dict) else None
            )
            eligible_tracks.append(queued_track)

        if ordered_tracks:
            with scan_lock:
                scan_state["face_track_cursor"] = (track_cursor + max(len(eligible_tracks), 1)) % len(ordered_tracks)

        if not eligible_tracks:
            if liveness_holds:
                return True, liveness_holds[0], payload
            if len(payload["faces"]) > 1:
                return True, f"Tracking {len(payload['faces'])} faces", payload
            return True, "Face detected", payload

        cached_face_encodings = {}
        recognition_candidates = []
        uncached_tracks = []
        for track in eligible_tracks:
            cached_encoding = get_cached_live_track_encoding(track, now_ts=frame_now_ts)
            if cached_encoding is None:
                uncached_tracks.append(track)
                continue
            cached_face_encodings[int(track.get("track_id") or 0)] = cached_encoding
            recognition_candidates.append(track)

        remaining_candidate_slots = max(max_candidates - len(recognition_candidates), 0)
        tracks_needing_encoding = uncached_tracks[: min(max_new_encodings, remaining_candidate_slots)]
        recognition_candidates.extend(tracks_needing_encoding)

        new_face_encodings = {}
        if tracks_needing_encoding:
            face_locations_full = [row["full_location"] for row in tracks_needing_encoding]
            encoding_frame, encoding_face_locations = prepare_live_encoding_frame(frame, face_locations_full)
            low_light_retry_hint = False
            if LIVE_RECOGNITION_LOW_LIGHT_ENCODING_RETRY_ENABLED:
                for track in tracks_needing_encoding:
                    face_quality = track.get("face_quality")
                    if not isinstance(face_quality, dict) or not face_quality:
                        face_quality = resolve_live_track_face_quality(
                            track,
                            frame,
                            track.get("small_location"),
                            scale_back=scale_back,
                            now_ts=frame_now_ts,
                        )
                        track["face_quality"] = dict(face_quality) if isinstance(face_quality, dict) else {}
                    brightness = float((face_quality or {}).get("brightness") or 0.0)
                    if brightness > 0.0 and brightness < float(LIVE_RECOGNITION_LOW_LIGHT_ENCODING_BRIGHTNESS):
                        low_light_retry_hint = True
                        break

            face_encs = encode_live_face_locations(
                encoding_frame,
                encoding_face_locations,
                num_jitters=LIVE_RECOGNITION_ENCODING_JITTERS,
                low_light_hint=low_light_retry_hint,
            )

            if len(face_encs) < len(tracks_needing_encoding):
                face_encs.extend([None] * (len(tracks_needing_encoding) - len(face_encs)))

            missing_indices = [
                index
                for index, encoding_row in enumerate(face_encs[:len(tracks_needing_encoding)])
                if encoding_row is None
            ]
            if missing_indices:
                fallback_locations = []
                fallback_index_map = []
                for index in missing_indices:
                    fallback_location = _normalized_face_location(tracks_needing_encoding[index].get("small_location"))
                    if not fallback_location:
                        continue
                    fallback_index_map.append(index)
                    fallback_locations.append(fallback_location)
                fallback_face_encs = encode_live_face_locations(
                    small_frame,
                    fallback_locations,
                    num_jitters=LIVE_RECOGNITION_ENCODING_JITTERS,
                    low_light_hint=low_light_retry_hint,
                ) if fallback_locations else []
                for position, encoding_row in enumerate(fallback_face_encs):
                    if position >= len(fallback_index_map):
                        break
                    face_encs[fallback_index_map[position]] = encoding_row

            for track, enc in zip(tracks_needing_encoding, face_encs):
                if enc is None:
                    continue
                normalized_encoding = cache_live_track_encoding(track, enc, now_ts=frame_now_ts)
                if normalized_encoding is None:
                    continue
                new_face_encodings[int(track.get("track_id") or 0)] = normalized_encoding

        available_encoding_count = len(cached_face_encodings) + len(new_face_encodings)
        if available_encoding_count == 0:
            for track in recognition_candidates:
                set_live_face_track_state(
                    track.get("track_id"),
                    next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_TRACK_UNKNOWN_RETRY_SECONDS),
                    last_result="face_not_encoded",
                    last_confidence=0.0,
                )
            push_not_registered_event("face_not_encoded", 0.0)
            return True, "Face(s) not encoded", payload

        results = []
        verified_students = []
        duplicate_students = []
        unknown_confidences = []
        unknown_reasons = []
        seen_student_ids = set()
        encoded_track_ids = set()
        for track in recognition_candidates:
            track_id = track.get("track_id")
            enc = cached_face_encodings.get(int(track_id or 0))
            if enc is None:
                enc = new_face_encodings.get(int(track_id or 0))
            if enc is None:
                continue
            encoded_track_ids.add(int(track_id or 0))
            face_location_small = track.get("small_location")
            cached_match_result = get_cached_live_track_match_result(track, now_ts=frame_now_ts)
            if cached_match_result is not None:
                match_result = cached_match_result
            elif legacy_flat_index:
                distances = face_recognition.face_distance(db_encodings, enc)
                if len(distances) > 0:
                    best_idx = int(np.argmin(distances))
                    best_distance = float(distances[best_idx])
                    confidence_pct = calculate_match_confidence(best_distance)
                    is_match = (
                        best_distance <= RECOGNITION_TOLERANCE
                        and confidence_pct >= MIN_RECOGNITION_CONFIDENCE
                        and best_idx < len(db_students)
                    )
                    match_result = {
                        "recognized": is_match,
                        "student": db_students[best_idx] if is_match else None,
                        "confidence": confidence_pct,
                        "distance": best_distance,
                        "candidate": {
                            "student": db_students[best_idx],
                            "best_distance": best_distance,
                        } if best_idx < len(db_students) else None,
                        "reason": "match" if is_match else "low_confidence",
                    }
                else:
                    match_result = {"recognized": False, "candidate": None, "reason": "no_face_index", "confidence": 0.0}
                cache_live_track_match_result(track, match_result, now_ts=frame_now_ts)
            else:
                match_result = match_face_probe(
                    enc,
                    face_index,
                    shortlist_size=RECOGNITION_SHORTLIST_SIZE,
                    match_distance_threshold=RECOGNITION_TOLERANCE,
                    score_threshold=RECOGNITION_SCORE_TOLERANCE,
                    support_distance_threshold=RECOGNITION_SUPPORT_TOLERANCE,
                    strong_match_distance=RECOGNITION_STRONG_MATCH_DISTANCE,
                    margin_threshold=RECOGNITION_MARGIN_THRESHOLD,
                    min_support_count=RECOGNITION_MIN_SUPPORT_COUNT,
                )
                cache_live_track_match_result(track, match_result, now_ts=frame_now_ts)
            if not isinstance(match_result, dict):
                match_result = {"recognized": False, "candidate": None, "reason": "no_face_index", "confidence": 0.0}

            if match_result.get("recognized"):
                candidate = match_result.get("student") or {}
                confidence_pct = float(match_result.get("confidence") or 0.0)
                if confidence_pct < MIN_RECOGNITION_CONFIDENCE:
                    unknown_confidences.append(confidence_pct)
                    unknown_reasons.append("low_confidence")
                    results.append(f"Low confidence: {confidence_pct:.1f}%")
                    set_live_face_track_state(
                        track_id,
                        next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_TRACK_UNKNOWN_RETRY_SECONDS),
                        student_id="",
                        last_result="low_confidence",
                        last_confidence=confidence_pct,
                    )
                    continue

                candidate_student_id = str(candidate.get("student_id") or "").strip()
                if candidate_student_id and should_suppress_recent_live_scan(candidate_student_id, frame_now_ts, mode=session_mode):
                    clear_pending_live_recognition(candidate_student_id)
                    set_live_face_track_state(
                        track_id,
                        next_attempt_ts=frame_now_ts + max(float(LIVE_RECOGNITION_TRACK_RETRY_SECONDS), float(SCAN_FACE_PRESENCE_RESET_SECONDS)),
                        student_id=candidate_student_id,
                        last_result="suppressed_recent_scan",
                        last_confidence=confidence_pct,
                    )
                    continue
                face_quality = resolve_live_track_face_quality(
                    track,
                    frame,
                    face_location_small,
                    scale_back=scale_back,
                    now_ts=frame_now_ts,
                )
                live_match_gate = evaluate_live_recognition_match(match_result, confidence_pct, face_quality)
                if not live_match_gate.get("accepted"):
                    results.append(str(live_match_gate.get("message") or "Hold still for identity confirmation."))
                    set_live_face_track_state(
                        track_id,
                        next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_CONFIRMATION_RETRY_SECONDS),
                        student_id=candidate_student_id,
                        last_result=str(live_match_gate.get("reason") or "awaiting_confirmation"),
                        last_confidence=confidence_pct,
                    )
                    continue

                confirmation_state = track_pending_live_recognition(
                    candidate,
                    confidence_pct,
                    match_result,
                    face_quality,
                    now_ts=frame_now_ts,
                )
                if not confirmation_state.get("confirmed"):
                    observed_frames = int(confirmation_state.get("observed_frames") or 0)
                    required_frames = int(confirmation_state.get("required_frames") or LIVE_RECOGNITION_CONFIRMATION_FRAMES)
                    results.append(
                        f"Confirming {candidate.get('name', 'identity')} ({observed_frames}/{required_frames})"
                    )
                    set_live_face_track_state(
                        track_id,
                        next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_CONFIRMATION_RETRY_SECONDS),
                        student_id=candidate_student_id,
                        last_result="confirming_identity",
                        last_confidence=confidence_pct,
                    )
                    continue

                if candidate_student_id and candidate_student_id in seen_student_ids:
                    duplicate_students.append(candidate.get("name", "Unknown"))
                    set_live_face_track_state(
                        track_id,
                        next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_CONFIRMATION_RETRY_SECONDS),
                        student_id=candidate_student_id,
                        last_result="duplicate_in_frame",
                        last_confidence=confidence_pct,
                    )
                    continue
                if candidate_student_id:
                    seen_student_ids.add(candidate_student_id)
                verification = handle_verified_student(candidate, confidence_pct)
                if verification:
                    verified_students.append(candidate.get("name", "Unknown"))
                    results.append(f"Verified: {candidate.get('name', 'Unknown')}")
                    set_live_face_track_state(
                        track_id,
                        next_attempt_ts=frame_now_ts + max(float(LIVE_RECOGNITION_TRACK_RETRY_SECONDS), float(SCAN_FACE_PRESENCE_RESET_SECONDS)),
                        student_id=candidate_student_id,
                        last_result="verified",
                        last_confidence=confidence_pct,
                    )
                else:
                    duplicate_students.append(candidate.get("name", "Unknown"))
                    results.append(f"Duplicate scan (cooldown) - {candidate.get('name', 'Unknown')}")
                    set_live_face_track_state(
                        track_id,
                        next_attempt_ts=frame_now_ts + max(float(LIVE_RECOGNITION_TRACK_RETRY_SECONDS), float(SCAN_FACE_PRESENCE_RESET_SECONDS)),
                        student_id=candidate_student_id,
                        last_result="duplicate_cooldown",
                        last_confidence=confidence_pct,
                    )
            elif match_result.get("candidate"):
                confidence_pct = float(match_result.get("confidence") or 0.0)
                unknown_confidences.append(confidence_pct)
                unknown_reasons.append(str(match_result.get("reason") or "low_confidence"))
                results.append(
                    f"No reliable match ({str(match_result.get('reason') or 'unmatched').replace('_', ' ')})"
                )
                set_live_face_track_state(
                    track_id,
                    next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_TRACK_UNKNOWN_RETRY_SECONDS),
                    student_id="",
                    last_result=str(match_result.get("reason") or "unmatched"),
                    last_confidence=confidence_pct,
                )
            else:
                unknown_confidences.append(0.0)
                unknown_reasons.append("no_face_index")
                results.append("No face index")
                set_live_face_track_state(
                    track_id,
                    next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_TRACK_UNKNOWN_RETRY_SECONDS),
                    student_id="",
                    last_result="no_face_index",
                    last_confidence=0.0,
                )

        if len(encoded_track_ids) < len(recognition_candidates):
            for track in recognition_candidates:
                if int(track.get("track_id") or 0) in encoded_track_ids:
                    continue
                set_live_face_track_state(
                    track.get("track_id"),
                    next_attempt_ts=frame_now_ts + float(LIVE_RECOGNITION_TRACK_UNKNOWN_RETRY_SECONDS),
                    student_id="",
                    last_result="encoding_missing",
                    last_confidence=0.0,
                )

        if not verified_students and unknown_confidences:
            clear_pending_live_recognition()
            rejection_reason = "low_confidence"
            if "ambiguous_match" in unknown_reasons:
                rejection_reason = "ambiguous_match"
            push_not_registered_event(rejection_reason, max(unknown_confidences))

        if verified_students:
            verified_count = len(verified_students)
            duplicate_count = len(duplicate_students)
            unknown_count = len(unknown_confidences)
            summary = f"Verified {verified_count} student{'s' if verified_count != 1 else ''}: " + ", ".join(verified_students)
            if duplicate_count:
                summary += f" | {duplicate_count} duplicate scan{'s' if duplicate_count != 1 else ''} ignored"
            if unknown_count:
                summary += f" | {unknown_count} unmatched face{'s' if unknown_count != 1 else ''}"
            return True, summary, payload

        if len(payload["faces"]) > 1 and not results:
            return True, f"Tracking {len(payload['faces'])} faces", payload
        return True, results[0] if results else "Frame processed", payload
        
    except Exception as exc:
        error_msg = f"Frame processing error: {str(exc)}"
        print(f"[ERROR] {error_msg}")
        return False, error_msg, payload


@app.route("/process_scan_frame", methods=["POST"])
@require_permission("scan", api=True)
def process_scan_frame():
    """
    Endpoint to receive and process frames from client device camera.
    Accepts either multipart/form-data with a 'frame' file field or a raw image body.
    """
    try:
        frame_bytes = b""
        frame_file = request.files.get('frame')
        if frame_file is not None:
            if frame_file.filename == '':
                return jsonify({"status": "error", "message": "Empty frame"}), 400
            frame_bytes = frame_file.read()
        else:
            frame_bytes = request.get_data(cache=False)
        if not frame_bytes:
            return jsonify({"status": "error", "message": "No frame content"}), 400
        
        if not frame_processing_lock.acquire(blocking=False):
            return jsonify({
                "status": "ok",
                "message": "Frame skipped: processor busy",
                "processed": False,
                "busy": True,
                "server_processing_ms": 0.0,
                "faces": get_latest_live_faces_payload(),
            }), 200

        try:
            # Process the frame
            process_started_at = time.perf_counter()
            success, message, payload = process_client_frame(frame_bytes)
        finally:
            frame_processing_lock.release()

        payload = dict(payload or {})
        payload["server_processing_ms"] = round(max((time.perf_counter() - process_started_at) * 1000.0, 0.0), 2)
        
        return jsonify({
            "status": "ok" if success else "error",
            "message": message,
            "processed": success,
            "busy": False,
            "faces": payload.get("faces", []),
            "server_processing_ms": payload.get("server_processing_ms", 0.0),
        }), (200 if success else 400)
        
    except Exception as exc:
        error_msg = f"Failed to process frame: {str(exc)}"
        print(f"[ERROR] {error_msg}")
        return jsonify({"status": "error", "message": error_msg}), 500


@app.route("/detect_faces", methods=["POST"])
@require_permission("scan", api=True)
def detect_faces():
    """
    Enhanced face detection endpoint for multi-face scanning.
    Optimized for high-performance real-time face detection.
    """
    try:
        # Import enhanced face processor
        from enhanced_face_processor import face_processor, detect_faces_endpoint
        
        # Use enhanced face processor
        result = detect_faces_endpoint()
        return result
        
    except ImportError:
        # Fallback to basic face detection if enhanced processor not available
        return jsonify({
            "status": "error", 
            "message": "Enhanced face processor not available"
        }), 503
    except Exception as exc:
        error_msg = f"Face detection error: {str(exc)}"
        print(f"[ERROR] {error_msg}")
        return jsonify({"status": "error", "message": error_msg}), 500


@app.route("/recognize_face", methods=["POST"])
@require_permission("scan", api=True)
def recognize_face():
    """
    Enhanced face recognition endpoint with caching and duplicate prevention.
    Optimized for high-volume face recognition with performance tracking.
    """
    try:
        # Import enhanced face processor
        from enhanced_face_processor import face_processor, recognize_face_endpoint
        
        # Use enhanced face processor
        result = recognize_face_endpoint()
        return result
        
    except ImportError:
        # Fallback to basic recognition if enhanced processor not available
        return jsonify({
            "status": "error", 
            "message": "Enhanced face processor not available"
        }), 503
    except Exception as exc:
        error_msg = f"Face recognition error: {str(exc)}"
        print(f"[ERROR] {error_msg}")
        return jsonify({"status": "error", "message": error_msg}), 500


@app.route("/face_metrics", methods=["GET"])
@require_permission("scan", api=True)
def face_metrics():
    """
    Endpoint for real-time performance metrics of face processing system.
    """
    try:
        # Import enhanced face processor
        from enhanced_face_processor import face_processor
        
        # Get metrics from enhanced processor
        result = face_processor.get_metrics_endpoint()
        return result
        
    except ImportError:
        # Return basic metrics if enhanced processor not available
        with scan_lock:
            return jsonify({
                "status": "ok",
                "metrics": {
                    "detections": scan_state.get("events_processed", 0),
                    "recognitions": len(scan_state.get("events", [])),
                    "fps": 0,
                    "active_tracks": 0,
                    "average_processing_time": 0,
                    "cache_size": 0,
                    "known_faces": len(scan_state.get("known_students", []))
                }
            })
    except Exception as exc:
        error_msg = f"Metrics error: {str(exc)}"
        print(f"[ERROR] {error_msg}")
        return jsonify({"status": "error", "message": error_msg}), 500


@app.route("/video_feed")
@require_permission("scan", api=True)
def video_feed():
    return Response(generate_frames(), mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/scan_events")
@require_permission("scan", api=True)
def scan_events():
    try:
        since = max(int(request.args.get("since", "0")), 0)
    except (TypeError, ValueError):
        since = 0
    return jsonify(build_scan_events_payload(since))


def build_scan_events_payload(since=0):
    try:
        normalized_since = max(int(since), 0)
    except (TypeError, ValueError):
        normalized_since = 0

    with scan_lock:
        payload = _build_scan_events_payload_locked(normalized_since)

    effective_session = resolve_gate_session(now_local())
    payload["effective_session"] = {
        "session": effective_session.get("session", ""),
        "gate_action": effective_session.get("gate_action", ""),
        "verification_label": effective_session.get("verification_label", ""),
        "status": effective_session.get("status", ""),
    }
    return payload


def _build_scan_events_payload_locked(since=0):
    events = [dict(e) for e in scan_state.get("events", []) if int(e.get("id") or 0) > since]
    last_event_id = int(since)
    if events:
        last_event_id = max(int(row.get("id") or 0) for row in events)
    active = bool(scan_state.get("active"))
    session_mode = normalize_scan_session_mode(scan_state.get("session_mode", "auto"), default="auto")
    liveness_profile = normalize_scan_liveness_profile(scan_state.get("liveness_profile", LIVE_LIVENESS_PROFILE))
    model_status = str(scan_state.get("model_status") or "idle")
    face_index_loading = bool(scan_state.get("face_index_loading"))
    registered_faces = len(scan_state.get("known_students") or [])
    active_tracks = len(scan_state.get("face_tracks") or {})
    return {
        "events": events,
        "active": active,
        "scan_session_mode": session_mode,
        "scan_liveness_profile": liveness_profile,
        "model_status": model_status,
        "face_index_loading": face_index_loading,
        "registered_faces": int(registered_faces),
        "active_tracks": int(active_tracks),
        "session_mode_label": scan_session_mode_label(session_mode),
        "liveness_profile_label": scan_liveness_profile_label(liveness_profile),
        "last_event_id": int(last_event_id),
        "server_time": now_iso(),
    }


@app.route("/api/scan/stream")
@require_permission("scan", api=True)
def scan_events_stream():
    try:
        since = max(int(request.args.get("since", "0")), 0)
    except (TypeError, ValueError):
        since = 0

    def generate(initial_since):
        last_seen = int(initial_since)
        yield "retry: 1200\n\n"
        while True:
            try:
                payload = None
                with scan_event_condition:
                    payload = _build_scan_events_payload_locked(last_seen)
                    if not payload.get("events"):
                        scan_event_condition.wait(timeout=1.2)
                        payload = _build_scan_events_payload_locked(last_seen)
                    if payload.get("events"):
                        last_seen = int(payload.get("last_event_id") or last_seen)

                if payload and payload.get("events"):
                    effective_session = resolve_gate_session(now_local())
                    payload["effective_session"] = {
                        "session": effective_session.get("session", ""),
                        "gate_action": effective_session.get("gate_action", ""),
                        "verification_label": effective_session.get("verification_label", ""),
                        "status": effective_session.get("status", ""),
                    }
                    yield f"event: scan_event\ndata: {json.dumps(payload)}\n\n"
                else:
                    yield ": keep-alive\n\n"
            except GeneratorExit:
                break
            except Exception:
                time.sleep(1.2)

    return Response(
        stream_with_context(generate(since)),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/scan/session-mode", methods=["GET", "POST"])
@require_permission("scan", api=True)
def api_scan_session_mode():
    if request.method == "POST":
        payload = request_payload()
        requested_mode = payload.get("mode") or payload.get("session_mode")
        try:
            mode = set_scan_session_mode(requested_mode)
        except ValueError as exc:
            return jsonify({"status": "error", "message": str(exc)}), 400
    else:
        mode = get_scan_session_mode()

    with scan_lock:
        active = bool(scan_state.get("active"))

    effective_session = resolve_gate_session(now_local())
    return jsonify({
        "status": "ok",
        "mode": mode,
        "mode_label": scan_session_mode_label(mode),
        "active": active,
        "effective_session": {
            "session": effective_session.get("session", ""),
            "gate_action": effective_session.get("gate_action", ""),
            "verification_label": effective_session.get("verification_label", ""),
            "status": effective_session.get("status", ""),
            "display_message": effective_session.get("display_message", ""),
            "voice_message": effective_session.get("voice_message", ""),
        },
    })


@app.route("/api/scan/liveness-profile", methods=["GET", "POST"])
@require_permission("scan", api=True)
def api_scan_liveness_profile():
    if request.method == "POST":
        payload = request_payload()
        requested_profile = payload.get("profile") or payload.get("liveness_profile")
        try:
            profile = set_scan_liveness_profile(requested_profile)
        except ValueError as exc:
            return jsonify({"status": "error", "message": str(exc)}), 400
    else:
        profile = get_scan_liveness_profile()

    with scan_lock:
        active = bool(scan_state.get("active"))

    return jsonify({
        "status": "ok",
        "profile": profile,
        "profile_label": scan_liveness_profile_label(profile),
        "active": active,
    })


# =====================================
# NOTIFICATION ROUTES
# =====================================
def mark_notifications_read_in_db(object_ids=None, mark_all=False, school_year=""):
    global alert_revision
    school_year_label = normalize_school_year_value(school_year)
    alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)

    update_payload = {
        "$set": {
            "status": "read",
            "is_read": True,
            "read_at": now_iso(),
        }
    }
    modified = 0
    if mark_all:
        result = alerts_collection.update_many(unread_notifications_query(school_year_label), update_payload)
        modified = int(result.modified_count or 0)
    elif object_ids:
        update_query = {
            "_id": {"$in": object_ids},
            "$or": [
                {"status": {"$ne": "read"}},
                {"is_read": {"$ne": True}},
            ],
        }
        if school_year_label:
            update_query["school_year"] = school_year_label
        result = alerts_collection.update_many(
            update_query,
            update_payload,
        )
        modified = int(result.modified_count or 0)

    if modified > 0:
        with alert_lock:
            alert_revision += 1
    return modified


@app.route("/api/notifications", methods=["GET"])
@require_permission("alerts_manage", api=True)
def notifications_list_api():
    limit = request.args.get("limit", 12)
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    try:
        payload = notification_summary(limit=limit, school_year=school_year_label)
    except Exception as exc:
        return jsonify({"status": "error", "message": f"Failed to load notifications: {exc}"}), 500
    return jsonify({"status": "ok", "school_year": school_year_label, **payload})


@app.route("/api/notifications/<notification_id>", methods=["GET"])
@require_permission("alerts_manage", api=True)
def notification_detail_api(notification_id):
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)
    try:
        object_id = ObjectId(notification_id)
    except Exception:
        return jsonify({"status": "error", "message": "Notification not found."}), 404

    doc = alerts_collection.find_one({"_id": object_id, "school_year": school_year_label})
    if not doc:
        return jsonify({"status": "error", "message": "Notification not found."}), 404
    return jsonify({"status": "ok", "notification": normalize_notification_doc(doc)})


@app.route("/api/notifications/mark-read", methods=["POST"])
@require_permission("alerts_manage", api=True)
def notifications_mark_read_api():
    data = request.get_json(silent=True) or {}
    school_year_label = resolve_selected_school_year(
        request.args.get("school_year", "") or data.get("school_year", "")
    )
    if data.get("all"):
        mark_notifications_read_in_db(mark_all=True, school_year=school_year_label)
        alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)
        return jsonify({
            "status": "ok",
            "school_year": school_year_label,
            "unread": alerts_collection.count_documents(unread_notifications_query(school_year_label)),
        })

    ids = data.get("ids", [])
    object_ids = []
    for item in ids:
        try:
            object_ids.append(ObjectId(str(item)))
        except Exception:
            continue

    if not object_ids:
        return jsonify({"status": "error", "message": "No valid notification IDs provided."}), 400

    mark_notifications_read_in_db(object_ids=object_ids, school_year=school_year_label)
    alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)
    return jsonify({
        "status": "ok",
        "school_year": school_year_label,
        "unread": alerts_collection.count_documents(unread_notifications_query(school_year_label)),
    })


@app.route("/alerts/mark-read", methods=["POST"])
@require_permission("alerts_manage", api=True)
def mark_alerts_read():
    return notifications_mark_read_api()


@app.route("/alerts/unread-count")
@app.route("/api/notifications/unread-count")
@require_permission("alerts_manage", api=True)
def unread_alert_count():
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)
    unread = alerts_collection.count_documents(unread_notifications_query(school_year_label))
    return jsonify({"unread": unread, "school_year": school_year_label})


@app.route("/alerts/stream")
@app.route("/api/notifications/stream")
@require_permission("alerts_manage", api=True)
def alerts_stream():
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    alerts_collection, school_year_label, _ = get_alerts_storage(school_year_label)
    def generate():
        last_seen = -1
        while True:
            try:
                with alert_lock:
                    current_rev = alert_revision

                if current_rev != last_seen:
                    unread = alerts_collection.count_documents(unread_notifications_query(school_year_label))
                    payload = json.dumps({
                        "revision": current_rev,
                        "unread": unread,
                        "school_year": school_year_label,
                    })
                    yield f"event: alerts\ndata: {payload}\n\n"
                    last_seen = current_rev
                else:
                    # keep-alive for intermediaries/proxies
                    yield ": keep-alive\n\n"

                time.sleep(1.5)
            except GeneratorExit:
                break
            except Exception:
                time.sleep(2)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/changes/stream")
@require_permission("dashboard", api=True)
def data_changes_stream():
    def generate():
        last_seen = -1
        while True:
            try:
                snapshot = data_change_snapshot()
                current_rev = int(snapshot.get("revision", 0))
                if current_rev != last_seen:
                    payload = json.dumps({**snapshot, "server_time": now_iso()})
                    yield f"event: data_change\ndata: {payload}\n\n"
                    last_seen = current_rev
                else:
                    yield ": keep-alive\n\n"
                time.sleep(1.5)
            except GeneratorExit:
                break
            except Exception:
                time.sleep(2)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# =====================================
# USER MANAGEMENT
# =====================================
def dashboard_management_redirect(message="", message_type="success", school_year=""):
    params = {}
    school_year_label = resolve_selected_school_year(school_year or request.values.get("school_year", ""))
    if school_year_label:
        params["school_year"] = school_year_label
    if message:
        params["message"] = message
    if message_type:
        params["message_type"] = message_type
    return redirect(f"{url_for('dashboard', **params)}#user-management")


@app.route("/admin/users/add", methods=["POST"])
@require_permission("users_manage")
def add_user():
    if current_role() != ROLE_FULL_ADMIN:
        return dashboard_management_redirect("Only Full Admin can manage system accounts.", "error")

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    role = normalize_account_role(request.form.get("role", ROLE_STAFF), username)
    email = build_internal_account_email(username) if username else ""

    if not username or not password:
        create_alert("warning", "User creation requires username and password.", "system")
        log_audit_event(
            action="admin.user_create",
            outcome="failed",
            severity="warn",
            target_type="user",
            target_id=username,
            details={"reason": "missing_fields"},
        )
        return dashboard_management_redirect("Username and password are required.", "error")

    if users.count_documents({"username": username}) > 0:
        create_alert("warning", f"User creation skipped: {username} already exists.", "system")
        log_audit_event(
            action="admin.user_create",
            outcome="failed",
            severity="warn",
            target_type="user",
            target_id=username,
            details={"reason": "username_exists"},
        )
        return dashboard_management_redirect(f"Username '{username}' already exists.", "error")

    duplicate_email_user = users.find_one({
        "email": {"$regex": f"^{re.escape(email)}$", "$options": "i"},
    })
    if duplicate_email_user:
        create_alert("warning", f"User creation skipped: generated system address for '{username}' is already in use.", "system")
        log_audit_event(
            action="admin.user_create",
            outcome="failed",
            severity="warn",
            target_type="user",
            target_id=username,
            details={"reason": "system_identity_conflict", "system_email": email},
        )
        return dashboard_management_redirect("Choose a different username for this account.", "error")

    password_error, _ = validate_password_reset_input(password, password)
    if password_error:
        return dashboard_management_redirect(password_error, "error")

    created = now_iso()
    users.insert_one({
        "username": username,
        "password_hash": hash_password(password),
        "role": role,
        "fullName": username,
        "email": email,
        "phone": "",
        "address": "",
        "bio": "",
        "avatarUrl": "",
        "twoFactorEnabled": False,
        "theme": "light",
        "created_at": created,
        "updated_at": created,
        "updatedAt": created,
    })
    signal_data_change("users")
    create_alert("info", f"New user '{username}' added with role {role}.", "system")
    log_audit_event(
        action="admin.user_create",
        outcome="success",
        severity="info",
        target_type="user",
        target_id=username,
        details={"role": role, "system_email": email},
    )
    return dashboard_management_redirect(f"Account '{username}' created successfully.", "success")


@app.route("/admin/users/<user_id>/delete", methods=["POST"])
@require_permission("users_manage")
def delete_staff_user(user_id):
    if current_role() != ROLE_FULL_ADMIN:
        return dashboard_management_redirect("Only Full Admin can delete Staff accounts.", "error")

    school_year_label = request.form.get("school_year", "")
    try:
        user_oid = ObjectId(str(user_id))
    except Exception:
        return dashboard_management_redirect("Invalid Staff account.", "error", school_year_label)

    user_doc = users.find_one({"_id": user_oid})
    if not user_doc:
        return dashboard_management_redirect("Staff account not found.", "error", school_year_label)

    username = str(user_doc.get("username") or "").strip()
    role_name = normalize_account_role(user_doc.get("role"), username)
    if role_name != ROLE_STAFF:
        return dashboard_management_redirect("Only Staff accounts can be deleted here.", "error", school_year_label)
    if username.lower() == str(session.get("admin") or "").strip().lower():
        return dashboard_management_redirect("You cannot delete your own account.", "error", school_year_label)

    users.delete_one({"_id": user_oid})
    signal_data_change("users")
    user_email = normalize_email_value(user_doc.get("email", ""))
    if user_email:
        password_reset_tokens.delete_many({"email": user_email})

    create_alert("warning", f"Staff account '{username}' was deleted by {session.get('admin', 'system')}.", "security")
    log_audit_event(
        action="admin.user_delete",
        outcome="success",
        severity="warn",
        target_type="user",
        target_id=username,
        details={"role": role_name},
    )
    return dashboard_management_redirect(f"Staff account '{username}' deleted.", "success", school_year_label)


@app.route("/admin/users/<user_id>/password", methods=["POST"])
@require_permission("users_manage")
def reset_staff_user_password(user_id):
    if current_role() != ROLE_FULL_ADMIN:
        return dashboard_management_redirect("Only Full Admin can change Staff passwords.", "error")

    school_year_label = request.form.get("school_year", "")
    try:
        user_oid = ObjectId(str(user_id))
    except Exception:
        return dashboard_management_redirect("Invalid Staff account.", "error", school_year_label)

    user_doc = users.find_one({"_id": user_oid})
    if not user_doc:
        return dashboard_management_redirect("Staff account not found.", "error", school_year_label)

    username = str(user_doc.get("username") or "").strip()
    role_name = normalize_account_role(user_doc.get("role"), username)
    if role_name != ROLE_STAFF:
        return dashboard_management_redirect("Only Staff accounts can be updated here.", "error", school_year_label)

    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")
    validation_error, _ = validate_password_reset_input(new_password, confirm_password)
    if validation_error:
        log_audit_event(
            action="admin.user_password_reset",
            outcome="failed",
            severity="warn",
            target_type="user",
            target_id=username,
            details={"reason": "validation_failed"},
        )
        return dashboard_management_redirect(validation_error, "error", school_year_label)

    updated = now_iso()
    users.update_one(
        {"_id": user_oid},
        {
            "$set": {
                "password_hash": hash_password(new_password),
                "updated_at": updated,
                "updatedAt": updated,
            },
            "$unset": {"password": ""},
        },
    )
    create_alert("info", f"Password updated for Staff account '{username}'.", "security")
    log_audit_event(
        action="admin.user_password_reset",
        outcome="success",
        severity="info",
        target_type="user",
        target_id=username,
        details={"role": role_name},
    )
    return dashboard_management_redirect(f"Password updated for Staff account '{username}'.", "success", school_year_label)


# =====================================
# STUDENTS CRUD
# =====================================
def api_success(payload=None, status_code=200):
    body = {"status": "ok"}
    if isinstance(payload, dict):
        body.update(payload)
    return jsonify(body), status_code


def api_error(message, status_code=400, field=None, extra=None):
    body = {"status": "error", "message": message}
    if field:
        body["field"] = field
    if isinstance(extra, dict):
        body.update(extra)
    return jsonify(body), status_code


def request_payload():
    if request.is_json:
        data = request.get_json(silent=True) or {}
    else:
        data = request.form.to_dict(flat=True)
    return data if isinstance(data, dict) else {}


def parse_student_oid(raw_id):
    try:
        return ObjectId(raw_id)
    except Exception:
        return None


def extract_grade_number(value):
    grade_label = normalize_grade_level(value)
    if not grade_label:
        return ""
    lower = grade_label.lower()
    if lower.startswith("grade "):
        return grade_label.split(" ", 1)[1].strip()
    if grade_label.isdigit():
        return grade_label
    match = re.search(r"\d+", grade_label)
    return match.group(0) if match else grade_label


def promote_grade_level(value):
    grade_label = normalize_grade_level(value)
    grade_number = extract_grade_number(grade_label)
    if not grade_number.isdigit():
        return grade_label
    next_grade = min(int(grade_number) + 1, 12)
    return f"Grade {next_grade}"


def resolve_student_grade_and_section(grade_value, section_value, require_existing_section=False, school_year=""):
    section_clean = normalize_section_value(section_value)
    if not section_clean:
        return "", "", "Section is required.", "section"

    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    requested_grade_level = normalize_grade_level(grade_value)
    predefined = PREDEFINED_SECTION_LOOKUP.get(section_clean.lower())
    if predefined:
        predefined_grade_level = normalize_grade_level(predefined.get("grade_level"))
        if require_existing_section and requested_grade_level and predefined_grade_level and requested_grade_level != predefined_grade_level:
            return "", "", f"Section '{predefined['section']}' is registered under {predefined_grade_level}, not {requested_grade_level}.", "section"
        return predefined_grade_level, predefined["section"], "", ""

    grade_level = requested_grade_level
    section_normalized = section_clean.lower()

    if not grade_level:
        matches = list(sections.find(
            {"school_year": school_year_label, "section_normalized": section_normalized},
            {"grade_level": 1, "grade_key": 1, "section": 1},
        ).limit(2))
        if len(matches) == 1:
            inferred_grade = normalize_grade_level(matches[0].get("grade_level") or matches[0].get("grade_key"))
            if inferred_grade:
                grade_level = inferred_grade
                inferred_section = normalize_section_value(matches[0].get("section"))
                if inferred_section:
                    section_clean = inferred_section
        elif require_existing_section and len(matches) > 1:
            return "", "", f"Grade level is required for section '{section_clean}'.", "grade_level"

    if not grade_level:
        return "", "", "Grade level is required.", "grade_level"

    grade_key = extract_grade_number(grade_level)
    if grade_key:
        existing = sections.find_one(
            {"school_year": school_year_label, "grade_key": str(grade_key), "section_normalized": section_normalized},
            {"section": 1},
        )
        existing_section = normalize_section_value(existing.get("section")) if existing else ""
        if existing_section:
            return grade_level, existing_section, "", ""

    if require_existing_section:
        return "", "", f"Section '{section_clean}' is not registered under {grade_level}. Please use an existing section name.", "section"

    return grade_level, section_clean, "", ""


def build_grade_filter(grade_value):
    grade_level = normalize_grade_level(grade_value)
    if not grade_level:
        return None, ""

    grade_number = extract_grade_number(grade_level)
    grade_candidates = [grade_level]
    if grade_number:
        if grade_number not in grade_candidates:
            grade_candidates.append(grade_number)
        grade_prefixed = f"Grade {grade_number}"
        if grade_prefixed not in grade_candidates:
            grade_candidates.append(grade_prefixed)

    return {
        "$or": [
            {"grade_level": {"$in": grade_candidates}},
            {"grade": {"$in": grade_candidates}},
        ]
    }, grade_level


def normalize_face_status_filter(value=""):
    raw = str(value or "").strip().lower()
    if raw in {"registered", "yes", "true", "1"}:
        return "registered"
    if raw in {"unregistered", "not_registered", "no", "false", "0"}:
        return "unregistered"
    return ""


def build_student_enrollment_document(student_doc, school_year, grade_level="", section="", status="", source_school_year=""):
    normalized_student = normalize_student_doc(student_doc)
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    grade_label = normalize_grade_level(grade_level or normalized_student.get("grade_level") or normalized_student.get("grade"))
    section_value = normalize_section_value(section or normalized_student.get("section"))
    status_value = normalize_student_status(status or normalized_student.get("status", "Active"), default=normalized_student.get("status", "Active"))
    now_value = now_iso()
    return {
        "student_ref_id": str(normalized_student.get("_id") or ""),
        "student_id": normalized_student.get("student_id") or normalized_student.get("lrn") or "",
        "lrn": normalized_student.get("lrn") or normalized_student.get("student_id") or "",
        "name": normalized_student.get("name") or "",
        "gender": normalized_student.get("gender") or "",
        "sex": normalized_student.get("gender") or "",
        "parent_contact": normalized_student.get("parent_contact") or "",
        "profile_photo": normalized_student.get("profile_photo") or "",
        "face_registered": bool(normalized_student.get("face_registered")),
        "school_year": school_year_label,
        "grade_level": grade_label,
        "grade": grade_label,
        "section": section_value,
        "status": status_value,
        "source_school_year": normalize_school_year_value(source_school_year),
        "created_at": now_value,
        "updated_at": now_value,
    }


def sync_student_base_fields_to_enrollments(student_doc):
    normalized_student = normalize_student_doc(student_doc)
    student_id = normalized_student.get("student_id") or normalized_student.get("lrn") or ""
    if not student_id:
        return
    update_student_base_fields_across_enrollments(
        student_id,
        {
            "student_ref_id": str(normalized_student.get("_id") or ""),
            "lrn": normalized_student.get("lrn") or student_id,
            "name": normalized_student.get("name") or "",
            "gender": normalized_student.get("gender") or "",
            "sex": normalized_student.get("gender") or "",
            "parent_contact": normalized_student.get("parent_contact") or "",
            "profile_photo": normalized_student.get("profile_photo") or "",
            "face_registered": bool(normalized_student.get("face_registered")),
            "updated_at": now_iso(),
        },
    )


def upsert_student_enrollment(student_doc, school_year, grade_level="", section="", status="Active", source_school_year="", update_existing=False):
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    collection = get_school_year_enrollment_collection(school_year_label)
    enrollment_doc = build_student_enrollment_document(
        student_doc,
        school_year_label,
        grade_level=grade_level,
        section=section,
        status=status,
        source_school_year=source_school_year,
    )
    student_id = enrollment_doc.get("student_id")
    if not student_id:
        raise ValueError("Student id is required for enrollment.")
    query = {"student_id": student_id}
    existing = collection.find_one(query)
    if existing and not update_existing:
        raise ValueError(f"Student is already enrolled for School Year {school_year_label}.")

    if existing:
        update_fields = dict(enrollment_doc)
        update_fields.pop("created_at", None)
        collection.update_one({"_id": existing["_id"]}, {"$set": update_fields})
        return collection.find_one({"_id": existing["_id"]})

    collection.insert_one(enrollment_doc)
    return collection.find_one(query)


def build_section_canonical_fields(section_doc, default_school_year=""):
    default_label = normalize_school_year_value(default_school_year) or get_current_school_year_label()
    school_year_label = normalize_school_year_value(section_doc.get("school_year")) or default_label
    grade_level = normalize_grade_level(section_doc.get("grade_level") or section_doc.get("grade_key"))
    grade_key = str(section_doc.get("grade_key") or extract_grade_number(grade_level) or "").strip()
    section_value = normalize_section_value(section_doc.get("section"))
    if not school_year_label or not grade_key or not section_value:
        return {}
    return {
        "school_year": school_year_label,
        "grade_key": grade_key,
        "grade_level": grade_level or f"Grade {grade_key}",
        "section": section_value,
        "section_normalized": section_value.lower(),
    }


def deduplicate_sections_for_school_year(default_school_year=""):
    default_label = normalize_school_year_value(default_school_year) or get_current_school_year_label()
    projection = {
        "school_year": 1,
        "grade_key": 1,
        "grade_level": 1,
        "section": 1,
        "section_normalized": 1,
        "created_at": 1,
        "updated_at": 1,
    }
    grouped_rows = {}
    for row in sections.find({}, projection):
        canonical = build_section_canonical_fields(row, default_label)
        if not canonical:
            continue
        key = (
            canonical.get("school_year", ""),
            canonical.get("grade_key", ""),
            canonical.get("section_normalized", ""),
        )
        grouped_rows.setdefault(key, []).append((row, canonical))

    removed_count = 0
    for items in grouped_rows.values():
        if len(items) < 2:
            continue
        items.sort(
            key=lambda item: (
                1 if normalize_school_year_value(item[0].get("school_year")) else 0,
                str(item[0].get("updated_at") or ""),
                str(item[0].get("created_at") or ""),
                str(item[0].get("_id") or ""),
            ),
            reverse=True,
        )
        keeper_row, keeper_fields = items[0]
        duplicate_ids = [entry[0].get("_id") for entry in items[1:] if entry[0].get("_id") is not None]
        if duplicate_ids:
            sections.delete_many({"_id": {"$in": duplicate_ids}})
            removed_count += len(duplicate_ids)
        sections.update_one(
            {"_id": keeper_row["_id"]},
            {"$set": {**keeper_fields, "updated_at": now_iso()}},
        )

    for row in sections.find({}, projection):
        canonical = build_section_canonical_fields(row, default_label)
        if not canonical:
            continue
        updates = {}
        for key, value in canonical.items():
            if row.get(key) != value:
                updates[key] = value
        if updates:
            updates["updated_at"] = now_iso()
            sections.update_one({"_id": row["_id"]}, {"$set": updates})

    return removed_count


def ensure_sections_school_year_defaults():
    school_year_label = get_current_school_year_label()
    removed_count = deduplicate_sections_for_school_year(school_year_label)
    if removed_count > 0:
        print(f"[INFO] Removed {removed_count} duplicate section record(s) during school year migration.")


def cleanup_empty_school_year_section_mismatches(school_year=""):
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    if not school_year_label:
        return 0
    if count_school_year_enrollments(school_year_label) > 0:
        return 0

    mismatched_ids = []
    projection = {"grade_level": 1, "grade_key": 1, "section": 1}
    for row in sections.find({"school_year": school_year_label}, projection):
        section_value = normalize_section_value(row.get("section"))
        if not section_value:
            continue
        canonical = PREDEFINED_SECTION_LOOKUP.get(section_value.lower())
        if not canonical:
            continue
        current_grade = normalize_grade_level(row.get("grade_level") or row.get("grade_key"))
        canonical_grade = normalize_grade_level(canonical.get("grade_level"))
        if canonical_grade and current_grade and canonical_grade != current_grade:
            mismatched_ids.append(row.get("_id"))

    mismatched_ids = [section_id for section_id in mismatched_ids if section_id is not None]
    if not mismatched_ids:
        return 0

    delete_result = sections.delete_many({"_id": {"$in": mismatched_ids}})
    removed_count = int(delete_result.deleted_count or 0)
    if removed_count > 0:
        print(
            f"[INFO] Removed {removed_count} cross-grade section record(s) from {school_year_label} "
            "because the school year has no enrollments yet."
        )
        signal_data_change("sections")
    return removed_count


def ensure_student_enrollment_defaults():
    if any(count_school_year_enrollments(label) > 0 for label in list_student_enrollment_school_year_labels()):
        return
    school_year_label = get_current_school_year_label()
    collection = get_school_year_enrollment_collection(school_year_label)
    projection = {
        "student_id": 1,
        "lrn": 1,
        "name": 1,
        "grade_level": 1,
        "grade": 1,
        "section": 1,
        "status": 1,
        "gender": 1,
        "sex": 1,
        "parent_contact": 1,
        "profile_photo": 1,
        "face_registered": 1,
    }
    for row in students.find({}, projection):
        student_id = normalize_lrn_value(row.get("student_id") or row.get("lrn"))
        if not student_id:
            continue
        if collection.count_documents({"student_id": student_id}, limit=1):
            continue
        enrollment_doc = build_student_enrollment_document(row, school_year_label)
        collection.insert_one(enrollment_doc)


def previous_school_year_label(reference_school_year=""):
    current_label = normalize_school_year_value(reference_school_year) or get_current_school_year_label()
    rows = sorted(
        [row for row in list_school_year_docs() if normalize_school_year_value(row.get("label"))],
        key=lambda row: int(str(row.get("label")).split("-", 1)[0]),
        reverse=True,
    )
    for index, row in enumerate(rows):
        if row.get("label") == current_label and index + 1 < len(rows):
            return normalize_school_year_value(rows[index + 1].get("label"))
    return ""


def cleanup_accidental_current_school_year_seed():
    current_school_year = get_current_school_year_label()
    previous_school_year = previous_school_year_label(current_school_year)
    if not current_school_year or not previous_school_year:
        return 0

    current_collection = get_school_year_enrollment_collection(current_school_year)
    previous_collection = get_school_year_enrollment_collection(previous_school_year)

    current_count = int(current_collection.count_documents({}))
    previous_count = int(previous_collection.count_documents({}))
    if current_count == 0 or previous_count == 0 or current_count != previous_count:
        return 0

    blank_source_count = int(current_collection.count_documents({
        "$or": [
            {"source_school_year": {"$exists": False}},
            {"source_school_year": ""},
        ],
    }))
    if blank_source_count != current_count:
        return 0

    projection = {"student_id": 1, "grade_level": 1, "grade": 1, "section": 1, "status": 1}
    previous_map = {}
    for row in previous_collection.find({}, projection):
        student_id = normalize_lrn_value(row.get("student_id") or row.get("lrn"))
        if not student_id:
            continue
        previous_map[student_id] = (
            normalize_grade_level(row.get("grade_level") or row.get("grade")),
            normalize_section_value(row.get("section")),
            normalize_student_status(row.get("status", "Active"), default="Active"),
        )

    identical_count = 0
    for row in current_collection.find({}, projection):
        student_id = normalize_lrn_value(row.get("student_id") or row.get("lrn"))
        if not student_id or student_id not in previous_map:
            return 0
        current_signature = (
            normalize_grade_level(row.get("grade_level") or row.get("grade")),
            normalize_section_value(row.get("section")),
            normalize_student_status(row.get("status", "Active"), default="Active"),
        )
        if current_signature != previous_map[student_id]:
            return 0
        identical_count += 1

    if identical_count != current_count:
        return 0

    delete_result = current_collection.delete_many({})
    deleted_count = int(delete_result.deleted_count or 0)
    if deleted_count > 0:
        print(
            f"[INFO] Removed {deleted_count} accidental auto-created enrollment(s) from {current_school_year}. "
            f"Students remain archived in {previous_school_year} until manual re-enrollment."
        )
    return deleted_count


def build_students_query(q_value="", grade_value="", section_value="", face_status_value="", school_year=""):
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    q_text = (q_value or "").strip()
    section_text = (section_value or "").strip()
    face_status = normalize_face_status_filter(face_status_value)

    clauses = [{"school_year": school_year_label}]
    if q_text:
        q_regex = contains_regex_filter(q_text)
        clauses.append({
            "$or": [
                {"name": q_regex},
                {"lrn": q_regex},
                {"student_id": q_regex},
                {"section": q_regex},
            ]
        })

    grade_clause, grade_level = build_grade_filter(grade_value)
    if grade_clause:
        clauses.append(grade_clause)
    else:
        grade_level = ""

    if section_text:
        clauses.append({"section": section_text})

    if face_status == "registered":
        clauses.append({"face_registered": True})
    elif face_status == "unregistered":
        clauses.append({
            "$or": [
                {"face_registered": False},
                {"face_registered": {"$exists": False}},
            ]
        })

    query = {"$and": clauses} if clauses else {}
    return query, q_text, grade_level, section_text, face_status, school_year_label


def grade_sort_key(raw_key):
    key = str(raw_key)
    return (0, int(key)) if key.isdigit() else (1, key.lower())


def build_sections_by_grade(grade_filter="", school_year=""):
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    collection = get_school_year_enrollment_collection(school_year_label)
    grade_clause, _normalized_grade = build_grade_filter(grade_filter)
    query = {"school_year": school_year_label}
    if grade_clause:
        query.update(grade_clause)

    sections_by_grade = {}
    projection = {"grade_level": 1, "grade": 1, "section": 1}
    for row in collection.find(query, projection):
        grade_label = normalize_grade_level(row.get("grade_level") or row.get("grade"))
        grade_key = extract_grade_number(grade_label)
        section_value = normalize_section_value(row.get("section"))
        if not grade_key or not section_value:
            continue
        if grade_key not in sections_by_grade:
            sections_by_grade[grade_key] = set()
        sections_by_grade[grade_key].add(section_value)

    manual_query = {"school_year": school_year_label}
    manual_grade = extract_grade_number(grade_filter)
    if manual_grade:
        manual_query["grade_key"] = str(manual_grade)

    for row in sections.find(manual_query, {"grade_key": 1, "grade_level": 1, "section": 1}):
        grade_key = str(row.get("grade_key") or extract_grade_number(row.get("grade_level")))
        section_value = normalize_section_value(row.get("section"))
        if not grade_key or not section_value:
            continue
        if grade_key not in sections_by_grade:
            sections_by_grade[grade_key] = set()
        sections_by_grade[grade_key].add(section_value)

    ordered = {}
    for grade_key in sorted(sections_by_grade.keys(), key=grade_sort_key):
        ordered[grade_key] = sorted(sections_by_grade[grade_key], key=str.lower)
    return ordered


def upsert_manual_section(grade_value, section_value, created_by="", school_year=""):
    grade_level = normalize_grade_level(grade_value)
    if not grade_level:
        raise ValueError("Grade level is required.")

    grade_key = extract_grade_number(grade_level)
    if not grade_key:
        raise ValueError("Invalid grade level.")

    section_clean = normalize_section_value(section_value)
    if not section_clean:
        raise ValueError("Section is required.")

    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    section_normalized = section_clean.lower()
    query = {"school_year": school_year_label, "grade_key": str(grade_key), "section_normalized": section_normalized}
    existing = sections.find_one(query, {"_id": 1, "grade_level": 1, "section": 1})

    if existing:
        existing_section = normalize_section_value(existing.get("section"))
        existing_grade = str(existing.get("grade_level") or "").strip()
        if existing_section != section_clean or existing_grade != grade_level:
            sections.update_one(
                {"_id": existing["_id"]},
                {"$set": {"grade_level": grade_level, "section": section_clean, "updated_at": now_iso()}},
            )
            signal_data_change("sections")
        return {
            "school_year": school_year_label,
            "grade_key": str(grade_key),
            "grade_level": grade_level,
            "section": section_clean,
        }

    now_value = now_iso()
    sections.insert_one({
        "school_year": school_year_label,
        "grade_key": str(grade_key),
        "grade_level": grade_level,
        "section": section_clean,
        "section_normalized": section_normalized,
        "created_at": now_value,
        "updated_at": now_value,
        "created_by": (created_by or "").strip(),
    })
    signal_data_change("sections")
    return {
        "school_year": school_year_label,
        "grade_key": str(grade_key),
        "grade_level": grade_level,
        "section": section_clean,
    }


def ensure_predefined_sections(school_year="", allow_create=True):
    school_year_label = normalize_school_year_value(school_year) or get_current_school_year_label()
    
    # Only create sections if school year exists and creation is allowed
    if not allow_create:
        return
        
    # Check if school year actually exists before creating sections
    if not school_years.find_one({"label": school_year_label}):
        return
        
    for grade_level, section_values in PREDEFINED_SECTIONS_BY_GRADE.items():
        for section_name in section_values:
            try:
                upsert_manual_section(grade_level, section_name, created_by="system", school_year=school_year_label)
            except Exception as exc:
                print(f"[WARNING] Could not upsert predefined section {school_year_label} {grade_level} - {section_name}: {exc}")


def build_lrn_duplicate_query(lrn_value, exclude_oid=None):
    query = {
        "$or": [
            {"lrn": lrn_value},
            {"student_id": lrn_value},
        ]
    }
    if exclude_oid is not None:
        query["_id"] = {"$ne": exclude_oid}
    return query


def normalize_student_name_value(value):
    cleaned = re.sub(r"\s+", " ", str(value or "")).strip()
    return cleaned[:120]


def normalize_student_status(value, default="Active"):
    status_value = str(value or default).strip()
    return "Inactive" if status_value == "Inactive" else "Active"


def sanitize_personal_student_payload(data, existing_doc=None, require_existing_section=False, school_year=""):
    current = existing_doc or {}
    lrn_input = data.get("lrn", data.get("student_id", current.get("lrn", current.get("student_id", ""))))
    lrn, lrn_error = validate_lrn_value(lrn_input)
    student_name = normalize_student_name_value(data.get("name", current.get("name", "")))
    resolved_grade, resolved_section, grade_section_error, grade_section_field = resolve_student_grade_and_section(
        data.get("grade_level") or data.get("grade") or current.get("grade_level") or current.get("grade"),
        data.get("section", current.get("section", "")),
        require_existing_section=require_existing_section,
        school_year=school_year,
    )
    parent_contact_raw = str(data.get("parent_contact", current.get("parent_contact", "")) or "").strip()
    gender = normalize_gender_value(data.get("gender") or data.get("sex") or current.get("gender") or current.get("sex"))
    status = normalize_student_status(data.get("status", current.get("status", "Active")), default=current.get("status", "Active"))

    if lrn_error:
        return None, lrn_error, "lrn"
    if not student_name:
        return None, "Name is required.", "name"
    if grade_section_error:
        return None, grade_section_error, grade_section_field
    if not gender:
        return None, "Sex/Gender is required.", "gender"
    try:
        parent_contact = normalize_parent_contact_value(parent_contact_raw)
    except ValueError as exc:
        return None, str(exc), "parent_contact"

    return {
        "lrn": lrn,
        "student_id": lrn,
        "name": student_name,
        "grade_level": resolved_grade,
        "grade": resolved_grade,
        "section": resolved_section,
        "parent_contact": parent_contact,
        "gender": gender,
        "sex": gender,
        "status": status,
    }, "", ""


def build_new_student_document(student_data):
    now_value = now_iso()
    payload = dict(student_data or {})
    payload.update({
        "face_registered": False,
        "face_updated_at": None,
        "face_data": [],
        "faces": [],
        "face_encodings": [],
        "face_embeddings": [],
        "profile_photo": "",
        "created_at": now_value,
        "updated_at": now_value,
    })
    return payload


def normalize_face_capture_profile(value):
    text = str(value or "standard").strip().lower().replace("-", "_").replace(" ", "_")
    if text in {"similar_faces", "similar", "twins", "twin"}:
        return "similar_faces"
    return "standard"


def parse_face_capture_meta(raw_meta):
    if isinstance(raw_meta, str):
        try:
            raw_meta = json.loads(raw_meta)
        except Exception:
            raw_meta = []
    if not isinstance(raw_meta, list):
        return []

    parsed = []
    for row in raw_meta:
        if not isinstance(row, dict):
            parsed.append({})
            continue
        parsed.append({
            "step_key": str(row.get("step_key") or row.get("key") or "").strip(),
            "label": str(row.get("label") or "").strip(),
            "instruction": str(row.get("instruction") or "").strip(),
            "yaw": row.get("yaw"),
            "pitch": row.get("pitch"),
        })
    return parsed


def measure_face_capture_quality(img_np):
    gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
    brightness = float(np.mean(gray))
    contrast = float(np.std(gray))
    sharpness = float(cv2.Laplacian(gray, cv2.CV_64F).var())
    low_light = brightness < FACE_REGISTRATION_LOW_LIGHT_BRIGHTNESS
    min_sharpness = FACE_REGISTRATION_LOW_LIGHT_SHARPNESS_MIN if low_light else FACE_REGISTRATION_SHARPNESS_MIN

    reason = ""
    hint = ""
    if brightness < FACE_REGISTRATION_BRIGHTNESS_MIN:
        reason = "too_dark"
        hint = "Increase front lighting or face the screen light before saving."
    elif brightness > FACE_REGISTRATION_BRIGHTNESS_MAX and contrast < FACE_REGISTRATION_CONTRAST_MIN * 1.35:
        reason = "too_bright"
        hint = "Reduce glare or step away from strong backlight."
    elif contrast < FACE_REGISTRATION_CONTRAST_MIN and sharpness < max(min_sharpness * 1.1, FACE_REGISTRATION_SHARPNESS_MIN):
        reason = "low_contrast"
        hint = "Use steadier lighting so the face stands out from the background."
    elif sharpness < min_sharpness:
        reason = "blurry"
        hint = "Hold still for a slightly sharper face image."

    return {
        "ok": not reason,
        "reason": reason,
        "hint": hint,
        "brightness": round(brightness, 2),
        "contrast": round(contrast, 2),
        "sharpness": round(sharpness, 2),
    }


def build_face_capture_variants(img_np):
    variants = [img_np]
    try:
        lab = cv2.cvtColor(img_np, cv2.COLOR_RGB2LAB)
        l_channel, a_channel, b_channel = cv2.split(lab)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        balanced = cv2.cvtColor(cv2.merge((clahe.apply(l_channel), a_channel, b_channel)), cv2.COLOR_LAB2RGB)
        variants.append(balanced)

        softened = cv2.GaussianBlur(balanced, (0, 0), 1.15)
        sharpened = cv2.addWeighted(balanced, 1.32, softened, -0.32, 0)
        variants.append(sharpened)
    except Exception:
        pass
    return variants


def _normalize_registration_face_location(location):
    if not isinstance(location, (list, tuple)) or len(location) != 4:
        return None
    try:
        top = int(round(float(location[0])))
        right = int(round(float(location[1])))
        bottom = int(round(float(location[2])))
        left = int(round(float(location[3])))
    except (TypeError, ValueError):
        return None
    if bottom <= top or right <= left:
        return None
    return (top, right, bottom, left)


def _scale_face_location_to_image(location, source_shape, target_shape, padding_ratio=0.0):
    normalized = _normalize_registration_face_location(location)
    if normalized is None:
        return None

    source_height = max(int((source_shape or [0, 0])[0] or 0), 1)
    source_width = max(int((source_shape or [0, 0])[1] or 0), 1)
    target_height = max(int((target_shape or [0, 0])[0] or 0), 1)
    target_width = max(int((target_shape or [0, 0])[1] or 0), 1)
    scale_y = float(target_height) / float(source_height)
    scale_x = float(target_width) / float(source_width)

    top, right, bottom, left = normalized
    scaled_top = float(top) * scale_y
    scaled_right = float(right) * scale_x
    scaled_bottom = float(bottom) * scale_y
    scaled_left = float(left) * scale_x

    if padding_ratio > 0.0:
        width = max(scaled_right - scaled_left, 1.0)
        height = max(scaled_bottom - scaled_top, 1.0)
        pad_x = width * float(padding_ratio)
        pad_y = height * float(padding_ratio)
        scaled_top -= pad_y
        scaled_bottom += pad_y
        scaled_left -= pad_x
        scaled_right += pad_x

    scaled_top = max(0, min(target_height, int(round(scaled_top))))
    scaled_right = max(0, min(target_width, int(round(scaled_right))))
    scaled_bottom = max(0, min(target_height, int(round(scaled_bottom))))
    scaled_left = max(0, min(target_width, int(round(scaled_left))))
    if scaled_bottom <= scaled_top or scaled_right <= scaled_left:
        return None
    return (scaled_top, scaled_right, scaled_bottom, scaled_left)


def _coerce_capture_pose_value(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _classify_capture_pose_bucket(yaw_value, pitch_value):
    yaw = float(yaw_value or 0.0)
    pitch = float(pitch_value or 0.0)

    if yaw <= -0.24:
        yaw_band = "right_profile"
    elif yaw <= -0.09:
        yaw_band = "right"
    elif yaw >= 0.24:
        yaw_band = "left_profile"
    elif yaw >= 0.09:
        yaw_band = "left"
    else:
        yaw_band = "center"

    if pitch <= -0.16:
        pitch_band = "up_high"
    elif pitch <= -0.07:
        pitch_band = "up"
    elif pitch >= 0.16:
        pitch_band = "down_low"
    elif pitch >= 0.07:
        pitch_band = "down"
    else:
        pitch_band = "level"
    return f"{yaw_band}:{pitch_band}"


def summarize_face_capture_pose_diversity(capture_meta):
    yaw_values = []
    pitch_values = []
    bucket_counts = {}
    for row in list(capture_meta or []):
        if not isinstance(row, dict):
            continue
        yaw = _coerce_capture_pose_value(row.get("yaw"))
        pitch = _coerce_capture_pose_value(row.get("pitch"))
        if yaw is None and pitch is None:
            continue
        yaw = 0.0 if yaw is None else yaw
        pitch = 0.0 if pitch is None else pitch
        yaw_values.append(yaw)
        pitch_values.append(pitch)
        bucket = _classify_capture_pose_bucket(yaw, pitch)
        bucket_counts[bucket] = int(bucket_counts.get(bucket, 0) or 0) + 1

    yaw_span = max(yaw_values) - min(yaw_values) if yaw_values else 0.0
    pitch_span = max(pitch_values) - min(pitch_values) if pitch_values else 0.0
    unique_buckets = sorted(bucket_counts.keys())
    return {
        "sample_count": len(yaw_values),
        "unique_bucket_count": len(unique_buckets),
        "unique_buckets": unique_buckets,
        "yaw_span": round(float(yaw_span), 6),
        "pitch_span": round(float(pitch_span), 6),
    }


def validate_face_registration_diversity(face_encodings, capture_details):
    metadata = build_face_registration_metadata(face_encodings)
    capture_details = capture_details if isinstance(capture_details, dict) else {}
    capture_profile = normalize_face_capture_profile(capture_details.get("capture_profile"))
    pose_summary = summarize_face_capture_pose_diversity(capture_details.get("capture_meta", []))

    min_pose_buckets = (
        int(FACE_REGISTRATION_MIN_POSE_BUCKETS_SIMILAR)
        if capture_profile == "similar_faces"
        else int(FACE_REGISTRATION_MIN_POSE_BUCKETS_STANDARD)
    )
    min_pose_samples = 8 if capture_profile == "similar_faces" else 5
    min_yaw_span = 0.24 if capture_profile == "similar_faces" else 0.18
    min_pitch_span = 0.18 if capture_profile == "similar_faces" else 0.12
    encoding_spread = float(metadata.get("face_encoding_spread") or 0.0)

    if metadata.get("face_encoding_count", 0) >= 2 and encoding_spread < float(FACE_REGISTRATION_MIN_ENCODING_SPREAD):
        return {
            "ok": False,
            "message": "The captures are too similar. Follow the guided angles more distinctly before saving.",
            "details": {
                "face_encoding_spread": round(encoding_spread, 6),
                "min_encoding_spread": round(float(FACE_REGISTRATION_MIN_ENCODING_SPREAD), 6),
                "pose_bucket_count": int(pose_summary.get("unique_bucket_count") or 0),
                "pose_sample_count": int(pose_summary.get("sample_count") or 0),
                "yaw_span": float(pose_summary.get("yaw_span") or 0.0),
                "pitch_span": float(pose_summary.get("pitch_span") or 0.0),
            },
        }

    if (
        min_pose_buckets > 0
        and int(pose_summary.get("sample_count") or 0) >= min_pose_samples
    ):
        unique_bucket_count = int(pose_summary.get("unique_bucket_count") or 0)
        yaw_span = float(pose_summary.get("yaw_span") or 0.0)
        pitch_span = float(pose_summary.get("pitch_span") or 0.0)
        if unique_bucket_count < min_pose_buckets:
            return {
                "ok": False,
                "message": "Need more face-angle variety across the guided sequence before saving.",
                "details": {
                    "pose_bucket_count": unique_bucket_count,
                    "min_pose_bucket_count": int(min_pose_buckets),
                    "pose_sample_count": int(pose_summary.get("sample_count") or 0),
                    "yaw_span": yaw_span,
                    "pitch_span": pitch_span,
                    "face_encoding_spread": round(encoding_spread, 6),
                },
            }
        if yaw_span < min_yaw_span and pitch_span < min_pitch_span:
            return {
                "ok": False,
                "message": "Need clearer head-angle movement across the guided captures before saving.",
                "details": {
                    "pose_bucket_count": unique_bucket_count,
                    "pose_sample_count": int(pose_summary.get("sample_count") or 0),
                    "yaw_span": yaw_span,
                    "pitch_span": pitch_span,
                    "min_yaw_span": round(float(min_yaw_span), 6),
                    "min_pitch_span": round(float(min_pitch_span), 6),
                    "face_encoding_spread": round(encoding_spread, 6),
                },
            }

    return {
        "ok": True,
        "metadata": metadata,
        "pose_summary": pose_summary,
    }


def validate_face_capture_image(raw_face):
    try:
        img_b64 = raw_face.split(",", 1)[1]
        img_bytes = base64.b64decode(img_b64)
        img = Image.open(BytesIO(img_bytes)).convert("RGB")
        img_np_full = np.array(img)
    except Exception as exc:
        print(f"[WARNING] Face capture decode skipped: {exc}")
        return None

    try:
        original_height, original_width = img_np_full.shape[:2]
        img_np = img_np_full
        height, width = img_np.shape[:2]
        max_dimension = max(height, width)
        if max_dimension > 448:
            scale = 448 / float(max_dimension)
            resized_width = max(1, int(round(width * scale)))
            resized_height = max(1, int(round(height * scale)))
            img_np = cv2.resize(img_np, (resized_width, resized_height), interpolation=cv2.INTER_AREA)

        quality = measure_face_capture_quality(img_np)
        face_locations = []
        multiple_faces_detected = False
        for candidate_img in build_face_capture_variants(img_np):
            for upsample in (0, 1, 2):
                face_locations = face_recognition.face_locations(
                    candidate_img,
                    number_of_times_to_upsample=upsample,
                    model="hog",
                )
                if len(face_locations) > 1:
                    multiple_faces_detected = True
                if len(face_locations) == 1:
                    break
            if len(face_locations) == 1:
                break
        if len(face_locations) != 1:
            reason = "multiple_faces" if multiple_faces_detected else (quality.get("reason") or "face_not_found")
            hint = (
                "Keep only one face inside the guide while registering."
                if reason == "multiple_faces"
                else quality.get("hint") or "Keep the face centered and use a slightly straighter angle."
            )
            return {
                "ok": False,
                "reason": reason,
                "hint": hint,
                "brightness": quality["brightness"],
                "contrast": quality["contrast"],
                "sharpness": quality["sharpness"],
            }

        top, right, bottom, left = face_locations[0]
        face_height = max(0, bottom - top)
        face_width = max(0, right - left)
        if face_height < img_np.shape[0] * 0.13 or face_width < img_np.shape[1] * 0.11:
            return {
                "ok": False,
                "reason": "face_too_small",
                "hint": "Move a little closer so the face fills more of the guide.",
                "brightness": quality["brightness"],
                "contrast": quality["contrast"],
                "sharpness": quality["sharpness"],
            }

        if not quality["ok"]:
            return {
                "ok": False,
                "reason": quality["reason"],
                "hint": quality["hint"],
                "brightness": quality["brightness"],
                "contrast": quality["contrast"],
                "sharpness": quality["sharpness"],
            }

        full_resolution_location = _scale_face_location_to_image(
            face_locations[0],
            img_np.shape,
            img_np_full.shape,
            padding_ratio=0.08,
        )
        enc_rows = []
        if full_resolution_location is not None:
            for candidate_img in build_face_capture_variants(img_np_full):
                enc_rows = face_recognition.face_encodings(
                    candidate_img,
                    known_face_locations=[full_resolution_location],
                    num_jitters=FACE_REGISTRATION_ENCODING_JITTERS,
                )
                if enc_rows:
                    break
        if not enc_rows:
            enc_rows = face_recognition.face_encodings(
                img_np,
                known_face_locations=face_locations,
                num_jitters=FACE_REGISTRATION_ENCODING_JITTERS,
            )
        if not enc_rows:
            return {
                "ok": False,
                "reason": "encoding_failed",
                "hint": "Use a slightly straighter angle with the full face visible.",
                "brightness": quality["brightness"],
                "contrast": quality["contrast"],
                "sharpness": quality["sharpness"],
            }
        return {
            "ok": True,
            "encoding": enc_rows[0],
            "brightness": quality["brightness"],
            "contrast": quality["contrast"],
            "sharpness": quality["sharpness"],
        }
    except Exception as exc:
        print(f"[WARNING] Face encoding skipped: {exc}")
        return {
            "ok": False,
            "reason": "validation_failed",
            "hint": "Retake this angle with the full face visible and steady.",
        }


def parse_faces_payload(data):
    raw_faces = data.get("faces", data.get("face_data", []))
    if isinstance(raw_faces, str):
        try:
            raw_faces = json.loads(raw_faces)
        except Exception:
            raw_faces = []

    if not isinstance(raw_faces, list):
        return None, None, None, "Face data must be an array.", "faces", {}

    capture_profile = normalize_face_capture_profile(data.get("capture_profile"))
    required_count = 20 if capture_profile == "similar_faces" else 10
    raw_meta = parse_face_capture_meta(data.get("capture_meta", []))

    faces_array = []
    for raw_face in raw_faces:
        if not isinstance(raw_face, str):
            continue
        raw = raw_face.strip()
        if raw and "," in raw:
            faces_array.append(raw)
        if len(faces_array) >= required_count:
            break

    if not faces_array:
        return None, None, None, "Capture at least one face image.", "faces", {}
    if len(faces_array) < required_count:
        return None, None, None, f"Capture all required guided angles ({required_count} images).", "faces", {
            "required_count": required_count,
            "valid_capture_count": 0,
            "invalid_capture_indices": [],
            "invalid_captures": [],
            "missing_capture_indices": list(range(len(faces_array), required_count)),
        }

    validated_faces = []
    face_encodings = []
    capture_meta = []
    seen_encodings = []
    accepted_capture_indices = []
    invalid_captures = []

    for index, raw_face in enumerate(faces_array):
        meta_row = raw_meta[index] if index < len(raw_meta) else {}
        row_detail = {
            "index": index,
            "step_key": meta_row.get("step_key") or f"capture_{index + 1}",
            "label": meta_row.get("label") or f"Capture {index + 1}",
            "instruction": meta_row.get("instruction") or "",
        }
        validated = validate_face_capture_image(raw_face)
        validated_payload = validated if isinstance(validated, dict) else {}
        encoding = validated_payload.get("encoding")
        if (
            not validated_payload
            or validated_payload.get("ok") is False
            or encoding is None
        ):
            invalid_captures.append({
                **row_detail,
                "reason": validated_payload.get("reason") or "validation_failed",
                "hint": validated_payload.get("hint") or "",
                "quality": {
                    "brightness": validated_payload.get("brightness"),
                    "contrast": validated_payload.get("contrast"),
                    "sharpness": validated_payload.get("sharpness"),
                },
            })
            continue

        if seen_encodings:
            try:
                distances = face_distance(seen_encodings, encoding)
                if len(distances) and float(np.min(distances)) < FACE_CAPTURE_DUPLICATE_DISTANCE:
                    invalid_captures.append({
                        **row_detail,
                        "reason": "duplicate",
                        "hint": "Shift to the next guided angle so this capture is more distinct.",
                    })
                    continue
            except Exception:
                pass

        seen_encodings.append(encoding)
        accepted_capture_indices.append(index)
        validated_faces.append(raw_face)
        face_encodings.append(encoding.tolist())
        capture_meta.append({
            "step_key": row_detail["step_key"],
            "label": row_detail["label"],
            "instruction": row_detail["instruction"],
            "yaw": meta_row.get("yaw"),
            "pitch": meta_row.get("pitch"),
            "quality": {
                "brightness": validated_payload.get("brightness"),
                "contrast": validated_payload.get("contrast"),
                "sharpness": validated_payload.get("sharpness"),
            },
        })

    if len(validated_faces) < required_count:
        return None, None, None, (
            f"Need {required_count} clear and unique captures inside the oval guide. "
            f"Only {len(validated_faces)} passed validation."
        ), "faces", {
            "required_count": required_count,
            "valid_capture_count": len(validated_faces),
            "accepted_capture_indices": accepted_capture_indices,
            "invalid_capture_indices": [row["index"] for row in invalid_captures],
            "invalid_captures": invalid_captures,
        }

    return validated_faces, face_encodings, {
        "capture_profile": capture_profile,
        "capture_count": len(validated_faces),
        "capture_meta": capture_meta,
    }, "", "", {}


def refresh_scan_face_index_if_active():
    with scan_lock:
        is_active = bool(scan_state.get("active"))

    if not is_active:
        return
    refresh_face_index_async()


def ensure_student_lrn_defaults():
    try:
        cursor = students.find({}, {"lrn": 1, "student_id": 1})
        for row in cursor:
            normalized_lrn = normalize_lrn_value(row.get("lrn") or row.get("student_id"))
            if not normalized_lrn:
                continue

            current_lrn = normalize_lrn_value(row.get("lrn"))
            current_student_id = normalize_lrn_value(row.get("student_id"))
            patch = {}
            if current_lrn != normalized_lrn:
                patch["lrn"] = normalized_lrn
            if current_student_id != normalized_lrn:
                patch["student_id"] = normalized_lrn
            if patch:
                try:
                    students.update_one({"_id": row["_id"]}, {"$set": patch})
                except DuplicateKeyError:
                    print(f"[WARNING] Skipped duplicate LRN during startup backfill: {normalized_lrn}")
    except Exception as exc:
        print(f"[WARNING] Could not backfill student LRN values: {exc}")


def ensure_student_face_defaults():
    try:
        students.update_many(
            {
                "face_registered": {"$exists": False},
                "$or": [
                    {"face_data.0": {"$exists": True}},
                    {"faces.0": {"$exists": True}},
                    {"face_encodings.0": {"$exists": True}},
                    {"face_embeddings.0": {"$exists": True}},
                ],
            },
            {"$set": {"face_registered": True}},
        )
        students.update_many(
            {"face_registered": {"$exists": False}},
            {"$set": {"face_registered": False}},
        )
    except Exception as exc:
        print(f"[WARNING] Could not ensure face_registered defaults: {exc}")


# Only ensure basic system structure, don't auto-create data that was intentionally deleted
ensure_indexes()
ensure_student_lrn_defaults()
ensure_student_face_defaults()
# Only migrate legacy data if it exists, don't force creation
if student_enrollments.count_documents({}, limit=1) > 0:
    migrate_legacy_student_enrollments()
# Only cleanup accidental duplicates, don't restore deleted data
cleanup_accidental_current_school_year_seed()
for _school_year_row in list_school_year_docs():
    cleanup_empty_school_year_section_mismatches(_school_year_row.get("label"))
# Don't run school year scope defaults at startup as it can recreate deleted school years
archive_historical_school_year_records()
cleanup_obsolete_collections()
cleanup_notification_alerts(force=True)


def build_students_stats_payload(school_year=""):
    school_year_label = normalize_school_year_value(school_year) or resolve_selected_school_year(school_year)
    collection = get_school_year_enrollment_collection(school_year_label)
    base_query = {"school_year": school_year_label}
    active_count = collection.count_documents({
        **base_query,
        "$or": [
            {"status": "Active"},
            {"status": {"$exists": False}},
            {"status": ""},
        ]
    })
    inactive_count = collection.count_documents({**base_query, "status": "Inactive"})
    today_prefix = now_local().strftime("%Y-%m-%d")
    new_today_count = collection.count_documents({
        **base_query,
        "created_at": {
            "$gte": f"{today_prefix}T00:00:00",
            "$lte": f"{today_prefix}T23:59:59",
        }
    })
    return {
        "school_year": school_year_label,
        "total": collection.count_documents(base_query),
        "active": active_count,
        "inactive": inactive_count,
        "new_today": new_today_count,
    }


@app.route("/students", methods=["GET"])
@require_permission("students_read")
def students_page():
    selected_school_year = resolve_selected_school_year(request.args.get("school_year", ""))
    stats_payload = build_students_stats_payload(selected_school_year)
    template_context = {
        **sidebar_context("students", selected_school_year),
        "message": request.args.get("message", "").strip(),
        "message_type": request.args.get("message_type", "success").strip() or "success",
        "stats": stats_payload,
        "grade_options": list(GRADE_LEVEL_OPTIONS),
        "selected_school_year": selected_school_year,
        "current_school_year": get_current_school_year_label(),
        "school_year_options": list_school_year_docs(),
        "archived_view": selected_school_year != get_current_school_year_label(),
    }

    return render_template("students.html", **template_context)


def next_school_year_label(label):
    normalized = normalize_school_year_value(label)
    if not normalized:
        return derive_default_school_year_label()
    start_year = int(normalized.split("-", 1)[0]) + 1
    return build_school_year_label(start_year)


@app.route("/api/school-years", methods=["GET", "POST"])
@require_permission("students_read", api=True)
def api_school_years():
    if request.method == "GET":
        selected_school_year = resolve_selected_school_year(request.args.get("school_year", ""))
        rows = []
        for doc in list_school_year_docs():
            label = doc.get("label") or ""
            rows.append({
                **doc,
                "is_selected": label == selected_school_year,
                "student_count": count_school_year_enrollments(label, {"school_year": label}),
            })
        return api_success({
            "school_years": rows,
            "selected_school_year": selected_school_year,
            "current_school_year": get_current_school_year_label(),
            "next_school_year": next_school_year_label(get_current_school_year_label()),
        })

    if not has_permission("students_write"):
        return api_error("Forbidden", 403)

    payload = request_payload()
    label = normalize_school_year_value(payload.get("label") or payload.get("school_year"))
    if not label:
        return api_error("Invalid school year format. Use YYYY-YYYY.", 400, "school_year")

    created = ensure_school_year_exists(label, set_current=True, created_by=session.get("admin", ""))
    ensure_predefined_sections(label)
    remember_selected_school_year(label)
    signal_data_change("students", "sections")
    return api_success({
        "message": f"School Year {label} is ready for enrollment.",
        "school_year": {
            "_id": str(created.get("_id") or ""),
            "label": label,
            "is_current": True,
        },
        "selected_school_year": label,
        "current_school_year": label,
        "next_school_year": next_school_year_label(label),
    }, 201)


@app.route("/api/students/reenrollment-candidates", methods=["GET"])
@require_permission("students_write", api=True)
def api_students_reenrollment_candidates():
    source_school_year = normalize_school_year_value(request.args.get("source_school_year", ""))
    target_school_year = resolve_selected_school_year(request.args.get("target_school_year", ""))
    if not source_school_year:
        return api_error("Source school year is required.", 400, "source_school_year")
    if source_school_year == target_school_year:
        return api_error("Source and target school years must be different.", 400, "target_school_year")

    source_collection = get_school_year_enrollment_collection(source_school_year)
    target_collection = get_school_year_enrollment_collection(target_school_year)
    rows = list(source_collection.find({"school_year": source_school_year}).sort([("grade_level", 1), ("section", 1), ("name", 1)]))
    candidates = []
    for row in rows:
        normalized = normalize_enrollment_doc(row)
        next_grade = promote_grade_level(normalized.get("grade_level"))
        already_enrolled = target_collection.count_documents(
            {"school_year": target_school_year, "student_id": normalized.get("student_id")},
            limit=1,
        ) > 0
        candidates.append({
            **normalized,
            "source_school_year": source_school_year,
            "target_school_year": target_school_year,
            "promoted_grade_level": next_grade,
            "already_enrolled": already_enrolled,
        })

    return api_success({
        "source_school_year": source_school_year,
        "target_school_year": target_school_year,
        "candidates": candidates,
        "sections_by_grade": build_sections_by_grade(school_year=target_school_year),
    })


@app.route("/api/students/reenroll", methods=["POST"])
@require_permission("students_write", api=True)
def api_students_reenroll():
    payload = request_payload()
    source_school_year = normalize_school_year_value(payload.get("source_school_year", ""))
    target_school_year = resolve_selected_school_year(payload.get("target_school_year", ""))
    if not source_school_year:
        return api_error("Source school year is required.", 400, "source_school_year")
    if not target_school_year:
        return api_error("Target school year is required.", 400, "target_school_year")
    if source_school_year == target_school_year:
        return api_error("Source and target school years must be different.", 400, "target_school_year")
    if target_school_year != get_current_school_year_label():
        return api_error("Only the current school year can accept new enrollments.", 403, "target_school_year")

    selections = payload.get("students")
    if not isinstance(selections, list) or not selections:
        return api_error("Select at least one student to enroll.", 400, "students")

    enrolled_count = 0
    skipped_count = 0
    errors = []
    source_collection = get_school_year_enrollment_collection(source_school_year)

    for item in selections:
        if not isinstance(item, dict):
            continue
        source_record_id = str(item.get("record_id") or item.get("_id") or "").strip()
        selected = parse_bool_value(item.get("selected"), default=True)
        if not selected or not source_record_id:
            continue

        source_oid = parse_student_oid(source_record_id)
        if not source_oid:
            skipped_count += 1
            errors.append(f"Invalid source record id: {source_record_id}.")
            continue

        source_doc = source_collection.find_one({"_id": source_oid, "school_year": source_school_year})
        if not source_doc:
            skipped_count += 1
            errors.append(f"Source record {source_record_id} was not found in {source_school_year}.")
            continue

        grade_level = normalize_grade_level(item.get("grade_level") or item.get("promoted_grade_level") or source_doc.get("grade_level"))
        section_value = normalize_section_value(item.get("section") or "")
        status_value = normalize_student_status(item.get("status", "Active"), default="Active")
        if not grade_level:
            skipped_count += 1
            errors.append(f"{source_doc.get('name') or source_doc.get('student_id')}: Grade level is required.")
            continue
        if not section_value:
            skipped_count += 1
            errors.append(f"{source_doc.get('name') or source_doc.get('student_id')}: Section is required.")
            continue

        student_ref_id = parse_student_oid(source_doc.get("student_ref_id"))
        student_doc = students.find_one({"_id": student_ref_id}) if student_ref_id else students.find_one(build_lrn_duplicate_query(source_doc.get("student_id")))
        if not student_doc:
            skipped_count += 1
            errors.append(f"{source_doc.get('name') or source_doc.get('student_id')}: Student profile was not found.")
            continue

        try:
            upsert_manual_section(
                grade_level,
                section_value,
                created_by=session.get("admin", ""),
                school_year=target_school_year,
            )
            upsert_student_enrollment(
                student_doc,
                target_school_year,
                grade_level=grade_level,
                section=section_value,
                status=status_value,
                source_school_year=source_school_year,
                update_existing=False,
            )
            sync_student_base_fields_to_enrollments(student_doc)
            enrolled_count += 1
        except ValueError as exc:
            skipped_count += 1
            errors.append(f"{source_doc.get('name') or source_doc.get('student_id')}: {exc}")

    if enrolled_count > 0:
        signal_data_change("students", "sections")

    return api_success({
        "message": f"Enrollment completed for School Year {target_school_year}. Enrolled: {enrolled_count}. Skipped: {skipped_count}.",
        "source_school_year": source_school_year,
        "target_school_year": target_school_year,
        "enrolled_count": enrolled_count,
        "skipped_count": skipped_count,
        "errors": errors[:100],
        "sections_by_grade": build_sections_by_grade(school_year=target_school_year),
    })


def resolve_students_export_logo_path():
    candidate_paths = [
        os.path.join(app.root_path, "static", "DepED-Logo (1).png"),
        os.path.join(app.root_path, "static", "logo.png"),
        os.path.join(app.root_path, "static", "deped-logo.png"),
        os.path.join(app.root_path, "static", "favicon-32x32.png"),
    ]
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    return ""


def format_students_export_face_status(face_status):
    if face_status == "registered":
        return "Registered"
    if face_status == "unregistered":
        return "Not Registered"
    return "All"


def format_students_export_timestamp(value, multiline=False):
    raw = normalize_timestamp_value(value).strip()
    if not raw:
        return "N/A"

    parsed = None
    normalized = raw.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except Exception:
        for pattern in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(raw[:19] if "H" in pattern else raw[:10], pattern)
                break
            except Exception:
                continue

    if parsed is not None:
        if multiline:
            return f"{parsed.strftime('%Y-%m-%d')}<br/>{parsed.strftime('%H:%M')}"
        return parsed.strftime("%B %d, %Y %I:%M:%S %p")

    fallback = xml_escape(raw)
    if multiline:
        parts = raw.replace("T", " ", 1).split()
        if len(parts) >= 2:
            return f"{xml_escape(parts[0])}<br/>{xml_escape(parts[1][:5])}"
    return fallback


def build_students_export_section_breakdown(rows):
    section_counts = {}
    for row in rows:
        grade_label = row.get("grade_level") or "Unassigned"
        section_label = normalize_section_value(row.get("section")) or "Unassigned"
        key = (grade_label, section_label)
        section_counts[key] = section_counts.get(key, 0) + 1

    ordered_rows = []
    for (grade_label, section_label), count in sorted(
        section_counts.items(),
        key=lambda item: (grade_sort_key(extract_grade_number(item[0][0]) or item[0][0]), item[0][1].lower()),
    ):
        ordered_rows.append((grade_label, section_label, count))
    return ordered_rows


def resolve_current_admin_export_name(default_name="Admin"):
    admin_name = str(default_name or session.get("admin", "Admin") or "Admin").strip() or "Admin"
    try:
        _user_doc, profile = current_user_profile()
        full_name = str((profile or {}).get("fullName") or "").strip()
        if full_name:
            admin_name = full_name
    except Exception:
        pass
    return admin_name


def resolve_report_principal_name():
    try:
        signatory_settings = system_settings.find_one({"key": "report_signatories"}) or {}
        principal_name = str(signatory_settings.get("principal_name") or "").strip()
        if principal_name:
            return principal_name
    except Exception:
        pass
    return REPORT_PRINCIPAL_NAME or REPORT_PRINCIPAL_TITLE or "School Principal"


def build_report_signatories(prepared_by_name=""):
    prepared_name = str(prepared_by_name or "").strip() or "Admin"
    return {
        "prepared_by_label": "Prepared by",
        "prepared_by_name": prepared_name,
        "prepared_by_title": REPORT_PREPARED_BY_TITLE,
        "approved_by_label": "Approved by",
        "approved_by_name": resolve_report_principal_name(),
        "approved_by_title": REPORT_PRINCIPAL_TITLE,
    }


def resolve_pdf_export_brand_font_name():
    if PDF_EXPORT_OLD_ENGLISH_FONT_NAME in pdfmetrics.getRegisteredFontNames():
        return PDF_EXPORT_OLD_ENGLISH_FONT_NAME

    for font_path in PDF_EXPORT_OLD_ENGLISH_FONT_PATHS:
        if not os.path.exists(font_path):
            continue
        try:
            pdfmetrics.registerFont(TTFont(PDF_EXPORT_OLD_ENGLISH_FONT_NAME, font_path))
            return PDF_EXPORT_OLD_ENGLISH_FONT_NAME
        except Exception:
            continue

    return "Times-Bold"


def resolve_pdf_export_header_logo_path():
    candidate_paths = [
        os.path.join(app.root_path, "static", "DepED-Logo (1).png"),
        os.path.join(app.root_path, "static", "deped-seal.png"),
        os.path.join(app.root_path, "static", "deped-seal.jpg"),
        os.path.join(app.root_path, "static", "deped-seal.jpeg"),
        os.path.join(app.root_path, "static", "deped-logo.png"),
        os.path.join(app.root_path, "static", "logo.png"),
    ]
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    return ""


def build_pdf_export_logo_flowable(max_width=1.16 * inch, max_height=0.58 * inch):
    logo_path = resolve_pdf_export_header_logo_path()
    if not logo_path:
        return None

    logo_width = max_width
    logo_height = max_height
    try:
        with Image.open(logo_path) as logo_image:
            src_width, src_height = logo_image.size
            if src_width and src_height:
                ratio = min(float(max_width) / float(src_width), float(max_height) / float(src_height))
                logo_width = max(src_width * ratio, 1)
                logo_height = max(src_height * ratio, 1)
    except Exception:
        pass

    logo_flowable = RLImage(logo_path, width=logo_width, height=logo_height)
    logo_flowable.hAlign = "CENTER"
    return logo_flowable


def build_pdf_export_header_styles():
    styles = getSampleStyleSheet()
    brand_font_name = resolve_pdf_export_brand_font_name()
    return {
        "republic": ParagraphStyle(
            "PdfExportHeaderRepublic",
            parent=styles["Normal"],
            fontName=brand_font_name,
            fontSize=15.2,
            leading=16.4,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#111111"),
        ),
        "department": ParagraphStyle(
            "PdfExportHeaderDepartment",
            parent=styles["Normal"],
            fontName=brand_font_name,
            fontSize=14.6,
            leading=15.8,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#111111"),
        ),
        "region": ParagraphStyle(
            "PdfExportHeaderRegion",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=12.8,
            leading=14.0,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#1f2937"),
        ),
        "division": ParagraphStyle(
            "PdfExportHeaderDivision",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10.7,
            leading=11.8,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#1f2937"),
        ),
        "district": ParagraphStyle(
            "PdfExportHeaderDistrict",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10.2,
            leading=11.2,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#374151"),
        ),
        "school": ParagraphStyle(
            "PdfExportHeaderSchool",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=15.4,
            leading=16.8,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#1f2937"),
        ),
        "address": ParagraphStyle(
            "PdfExportHeaderAddress",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10.4,
            leading=11.5,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#4b5563"),
        ),
        "report_title": ParagraphStyle(
            "PdfExportHeaderReportTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11.2,
            leading=13.0,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#111827"),
        ),
        "report_caption": ParagraphStyle(
            "PdfExportHeaderReportCaption",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.1,
            leading=10.8,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#4b5563"),
        ),
    }


def build_pdf_export_header_block(document_title, header_caption=""):
    styles = build_pdf_export_header_styles()
    header_story = []
    logo_flowable = build_pdf_export_logo_flowable()
    if logo_flowable is not None:
        header_story.append(logo_flowable)
        header_story.append(Spacer(1, 0.03 * inch))

    header_story.extend([
        Paragraph(xml_escape(PDF_EXPORT_HEADER_REPUBLIC), styles["republic"]),
        Paragraph(xml_escape(PDF_EXPORT_HEADER_DEPARTMENT), styles["department"]),
        Spacer(1, 0.015 * inch),
        Paragraph(xml_escape(PDF_EXPORT_HEADER_REGION), styles["region"]),
        Paragraph(xml_escape(PDF_EXPORT_HEADER_DIVISION), styles["division"]),
        Paragraph(xml_escape(PDF_EXPORT_HEADER_DISTRICT), styles["district"]),
        Paragraph(xml_escape(PDF_EXPORT_HEADER_SCHOOL), styles["school"]),
        Paragraph(xml_escape(PDF_EXPORT_HEADER_ADDRESS), styles["address"]),
        Spacer(1, 0.06 * inch),
        HRFlowable(width="100%", thickness=0.9, color=colors.HexColor("#4b5563")),
        Spacer(1, 0.08 * inch),
        Paragraph(xml_escape(str(document_title or "").strip()), styles["report_title"]),
    ])

    caption_text = str(header_caption or "").strip()
    title_text = str(document_title or "").strip()
    if caption_text and caption_text.lower() != title_text.lower():
        header_story.append(Spacer(1, 0.02 * inch))
        header_story.append(Paragraph(xml_escape(caption_text), styles["report_caption"]))

    header_story.append(Spacer(1, 0.14 * inch))
    return KeepTogether(header_story)


def build_school_export_footer(canvas_obj, doc):
    footer_title = getattr(doc, "export_footer_title", "Cawitan High School Export")
    document_title = getattr(doc, "export_title", footer_title)
    document_subject = getattr(doc, "export_subject", document_title)
    canvas_obj.saveState()
    if canvas_obj.getPageNumber() == 1:
        canvas_obj.setTitle(document_title)
        canvas_obj.setAuthor(getattr(doc, "export_author", "Cawitan High School"))
        canvas_obj.setSubject(document_subject)
    canvas_obj.setFont("Helvetica", 8.5)
    canvas_obj.setFillColor(colors.black)
    canvas_obj.drawString(doc.leftMargin, 36, footer_title)
    canvas_obj.drawRightString(doc.pagesize[0] - doc.rightMargin, 36, f"Page {canvas_obj.getPageNumber()}")
    canvas_obj.restoreState()


def build_school_export_styles():
    styles = getSampleStyleSheet()
    return {
        "header_title": ParagraphStyle(
            "SchoolExportHeaderTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "subtitle": ParagraphStyle(
            "SchoolExportSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=12,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "doc_title": ParagraphStyle(
            "SchoolExportDocTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            alignment=TA_CENTER,
            textColor=colors.black,
            spaceBefore=2,
        ),
        "metadata_label": ParagraphStyle(
            "SchoolExportMetaLabel",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=11,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
        "metadata_value": ParagraphStyle(
            "SchoolExportMetaValue",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
        "section_title": ParagraphStyle(
            "SchoolExportSectionTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=15,
            alignment=TA_LEFT,
            textColor=colors.black,
            spaceAfter=2,
        ),
        "table_header": ParagraphStyle(
            "SchoolExportTableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7.2,
            leading=8.6,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "table_cell": ParagraphStyle(
            "SchoolExportTableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7,
            leading=8.4,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
        "table_cell_center": ParagraphStyle(
            "SchoolExportTableCellCenter",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7,
            leading=8.4,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "footer_value": ParagraphStyle(
            "SchoolExportFooterValue",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
        "signature": ParagraphStyle(
            "SchoolExportSignature",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
        "signature_block": ParagraphStyle(
            "SchoolExportSignatureBlock",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "note": ParagraphStyle(
            "SchoolExportNote",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
    }


def build_school_export_document(document_title, header_caption, metadata_items, footer_title, export_subject, export_author="Cawitan High School"):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=42,
        leftMargin=42,
        topMargin=42,
        bottomMargin=42,
        pageCompression=0,
    )
    doc.export_author = export_author
    doc.export_title = document_title
    doc.export_subject = export_subject
    doc.export_footer_title = footer_title

    styles = build_school_export_styles()
    story = [build_pdf_export_header_block(document_title, header_caption)]

    metadata_rows = []
    items = list(metadata_items or [])
    if len(items) % 2 != 0:
        items.append(("", ""))
    for index in range(0, len(items), 2):
        left_label, left_value = items[index]
        right_label, right_value = items[index + 1]
        metadata_rows.append([
            Paragraph(f"<b>{xml_escape(str(left_label or ''))}</b>", styles["metadata_label"]),
            Paragraph(xml_escape(str(left_value or "")), styles["metadata_value"]),
            Paragraph(f"<b>{xml_escape(str(right_label or ''))}</b>", styles["metadata_label"]),
            Paragraph(xml_escape(str(right_value or "")), styles["metadata_value"]),
        ])

    if metadata_rows:
        metadata_table = Table(
            metadata_rows,
            colWidths=[1.15 * inch, 3.35 * inch, 1.15 * inch, 3.35 * inch],
        )
        metadata_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f7f7f7")),
            ("BOX", (0, 0), (-1, -1), 0.75, colors.black),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(metadata_table)
        story.append(Spacer(1, 0.18 * inch))

    return buffer, doc, story, styles


def append_school_export_section_title(story, title, styles):
    story.append(Paragraph(xml_escape(title), styles["section_title"]))
    story.append(HRFlowable(width="100%", thickness=1.0, color=colors.black))
    story.append(Spacer(1, 0.1 * inch))


def build_school_export_table(data, col_widths, styles, span_empty_row=False):
    table = LongTable(data, repeatRows=1, colWidths=col_widths)
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9e9e9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.black),
        ("LINEBELOW", (0, 0), (-1, 0), 1.1, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for index in range(1, len(data)):
        table_style.append((
            "BACKGROUND",
            (0, index),
            (-1, index),
            colors.HexColor("#f7f7f7") if index % 2 == 0 else colors.white,
        ))
    if span_empty_row and len(data) == 2:
        table_style.append(("SPAN", (0, 1), (-1, 1)))
        table_style.append(("ALIGN", (0, 1), (-1, 1), "CENTER"))
    table.setStyle(TableStyle(table_style))
    return table


def build_report_signature_markup(label, name, title=""):
    safe_label = xml_escape(str(label or "").strip() or "Signature")
    raw_name = str(name or "").strip()
    raw_title = str(title or "").strip()
    display_lines = []
    normalized_name = raw_name.casefold()
    normalized_title = raw_title.casefold()

    if raw_name and normalized_name != "admin" and normalized_name != normalized_title:
        display_lines.append(xml_escape(raw_name))
    if raw_title and normalized_title != normalized_name:
        display_lines.append(xml_escape(raw_title))
    if not display_lines:
        display_lines.append(xml_escape(raw_title or raw_name or "Pending"))

    return (
        f"<b>{safe_label}</b><br/><br/><br/>"
        "______________________________<br/>"
        + "<br/>".join(display_lines)
    )


def build_report_signature_footer_block(styles, signatories=None):
    signatories = signatories or build_report_signatories()
    prepared_markup = build_report_signature_markup(
        signatories.get("prepared_by_label", "Prepared by"),
        signatories.get("prepared_by_name", "Admin"),
        signatories.get("prepared_by_title", REPORT_PREPARED_BY_TITLE),
    )
    approved_markup = build_report_signature_markup(
        signatories.get("approved_by_label", "Approved by"),
        signatories.get("approved_by_name", resolve_report_principal_name()),
        signatories.get("approved_by_title", REPORT_PRINCIPAL_TITLE),
    )
    return KeepTogether([
        Spacer(1, 0.24 * inch),
        HRFlowable(width="100%", thickness=1.0, color=colors.black),
        Spacer(1, 0.14 * inch),
        Table(
            [[
                Paragraph(prepared_markup, styles["signature_block"]),
                Paragraph(approved_markup, styles["signature_block"]),
            ]],
            colWidths=[4.35 * inch, 4.35 * inch],
            style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]),
        ),
        Spacer(1, 0.14 * inch),
        Paragraph("This document was generated by the CHS Gate Access system.", styles["note"]),
    ])


def build_school_export_footer_block(styles, signatories=None):
    return build_report_signature_footer_block(styles, signatories=signatories)


def sanitize_students_export_filename_part(value, fallback="records"):
    raw = normalize_text_value(value)
    if not raw:
        return fallback
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", raw).strip("_").lower()
    return cleaned or fallback


def sort_students_export_rows(rows):
    return sorted(
        rows,
        key=lambda row: (
            grade_sort_key(extract_grade_number(row.get("grade_level")) or row.get("grade_level") or ""),
            normalize_section_value(row.get("section")).lower(),
            normalize_student_name_value(row.get("name")).lower(),
            normalize_lrn_value(row.get("student_id") or row.get("lrn")).lower(),
        ),
    )


def build_students_plain_export_styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "StudentsPlainExportTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "subtitle": ParagraphStyle(
            "StudentsPlainExportSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=12,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "section_title": ParagraphStyle(
            "StudentsPlainExportSectionTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            alignment=TA_LEFT,
            textColor=colors.black,
            spaceAfter=4,
        ),
        "metadata_label": ParagraphStyle(
            "StudentsPlainExportMetadataLabel",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=10.5,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
        "metadata_value": ParagraphStyle(
            "StudentsPlainExportMetadataValue",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=10.5,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
        "table_header": ParagraphStyle(
            "StudentsPlainExportTableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "table_cell": ParagraphStyle(
            "StudentsPlainExportTableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
        "table_cell_center": ParagraphStyle(
            "StudentsPlainExportTableCellCenter",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "signature_block": ParagraphStyle(
            "StudentsPlainExportSignatureBlock",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            alignment=TA_CENTER,
            textColor=colors.black,
        ),
        "note": ParagraphStyle(
            "StudentsPlainExportNote",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            alignment=TA_LEFT,
            textColor=colors.black,
        ),
    }


def build_students_plain_export_footer(canvas_obj, doc):
    canvas_obj.saveState()
    if canvas_obj.getPageNumber() == 1:
        canvas_obj.setTitle(getattr(doc, "export_title", "Official Student Records Report"))
        canvas_obj.setAuthor(getattr(doc, "export_author", "Cawitan High School"))
        canvas_obj.setSubject(getattr(doc, "export_subject", "Student records export"))
    canvas_obj.setFont("Helvetica", 8)
    canvas_obj.setFillColor(colors.black)
    canvas_obj.drawString(doc.leftMargin, 28, "Cawitan High School")
    canvas_obj.drawRightString(doc.pagesize[0] - doc.rightMargin, 28, f"Page {canvas_obj.getPageNumber()}")
    canvas_obj.restoreState()


def build_students_plain_metadata_table(metadata_items, styles):
    rows = []
    items = list(metadata_items or [])
    if len(items) % 2 != 0:
        items.append(("", ""))
    for index in range(0, len(items), 2):
        left_label, left_value = items[index]
        right_label, right_value = items[index + 1]
        rows.append([
            Paragraph(f"<b>{xml_escape(str(left_label or ''))}</b>", styles["metadata_label"]),
            Paragraph(xml_escape(str(left_value or "")), styles["metadata_value"]),
            Paragraph(f"<b>{xml_escape(str(right_label or ''))}</b>", styles["metadata_label"]),
            Paragraph(xml_escape(str(right_value or "")), styles["metadata_value"]),
        ])

    table = Table(rows, colWidths=[1.2 * inch, 3.1 * inch, 1.2 * inch, 3.1 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f7f7f7")),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def build_students_plain_key_value_table(details, styles):
    rows = []
    for label, value in details:
        rows.append([
            Paragraph(f"<b>{xml_escape(str(label or ''))}</b>", styles["metadata_label"]),
            Paragraph(xml_escape(str(value or "")), styles["metadata_value"]),
        ])
    table = Table(rows, colWidths=[1.9 * inch, 6.9 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f7f7f7")),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def build_students_plain_records_table(rows, styles):
    data = [[
        Paragraph("LRN", styles["table_header"]),
        Paragraph("Name", styles["table_header"]),
        Paragraph("Grade", styles["table_header"]),
        Paragraph("Section", styles["table_header"]),
        Paragraph("Gender", styles["table_header"]),
        Paragraph("Status", styles["table_header"]),
        Paragraph("Face Registered", styles["table_header"]),
        Paragraph("Parent Contact", styles["table_header"]),
    ]]

    if rows:
        for row in rows:
            data.append([
                Paragraph(xml_escape(row.get("lrn") or row.get("student_id") or "N/A"), styles["table_cell"]),
                Paragraph(xml_escape(row.get("name") or "N/A"), styles["table_cell"]),
                Paragraph(xml_escape(row.get("grade_level") or "N/A"), styles["table_cell_center"]),
                Paragraph(xml_escape(row.get("section") or "N/A"), styles["table_cell"]),
                Paragraph(xml_escape(row.get("gender") or "N/A"), styles["table_cell_center"]),
                Paragraph(xml_escape(row.get("status") or "Active"), styles["table_cell_center"]),
                Paragraph("Yes" if row.get("face_registered") else "No", styles["table_cell_center"]),
                Paragraph(xml_escape(row.get("parent_contact") or "N/A"), styles["table_cell"]),
            ])
    else:
        data.append([
            Paragraph("No student records matched the selected export filter.", styles["table_cell_center"]),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    table = LongTable(
        data,
        repeatRows=1,
        colWidths=[1.15 * inch, 2.3 * inch, 0.8 * inch, 1.1 * inch, 0.9 * inch, 0.95 * inch, 1.1 * inch, 1.5 * inch],
    )
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9e9e9")),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for index in range(1, len(data)):
        table_style.append((
            "BACKGROUND",
            (0, index),
            (-1, index),
            colors.HexColor("#f7f7f7") if index % 2 == 0 else colors.white,
        ))
    if len(data) == 2 and not rows:
        table_style.append(("SPAN", (0, 1), (-1, 1)))
        table_style.append(("ALIGN", (0, 1), (-1, 1), "CENTER"))
    table.setStyle(TableStyle(table_style))
    return table


def resolve_students_export_scope(scope_value="", grade_level="", section_value="", student_record_id="", student_id="", q_value=""):
    scope = str(scope_value or "").strip().lower()
    if scope == "student" and (str(student_record_id or "").strip() or str(student_id or "").strip() or str(q_value or "").strip()):
        return "student"
    if scope == "section" and str(section_value or "").strip():
        return "section"
    if scope == "grade" and str(grade_level or "").strip():
        return "grade"
    if scope in {"all", "filtered"}:
        return "filtered"
    if str(student_record_id or "").strip() or str(student_id or "").strip():
        return "student"
    if str(grade_level or "").strip():
        return "section" if str(section_value or "").strip() else "grade"
    if str(section_value or "").strip() or str(q_value or "").strip():
        return "filtered"
    return "all"


def resolve_students_export_student_row(collection, school_year_label, student_record_id="", student_id="", q_value=""):
    record_id_text = str(student_record_id or "").strip()
    if record_id_text:
        try:
            raw_row = collection.find_one({"school_year": school_year_label, "_id": ObjectId(record_id_text)})
            if raw_row:
                return normalize_enrollment_doc(raw_row)
        except Exception:
            pass

    normalized_student_id = normalize_lrn_value(student_id)
    if normalized_student_id:
        raw_row = collection.find_one({
            "school_year": school_year_label,
            "$or": [{"student_id": normalized_student_id}, {"lrn": normalized_student_id}],
        })
        if raw_row:
            return normalize_enrollment_doc(raw_row)

    q_text = str(q_value or "").strip()
    if q_text:
        exact_student_id = normalize_lrn_value(q_text)
        if exact_student_id:
            raw_row = collection.find_one({
                "school_year": school_year_label,
                "$or": [{"student_id": exact_student_id}, {"lrn": exact_student_id}],
            })
            if raw_row:
                return normalize_enrollment_doc(raw_row)

        query, _q, _grade, _section, _face_status, _school_year = build_students_query(q_text, "", "", "", school_year_label)
        raw_row = collection.find_one(query, sort=[("name", 1), ("created_at", -1)])
        if raw_row:
            return normalize_enrollment_doc(raw_row)

    return None


def build_students_export_payload(args, school_year_label):
    collection = get_school_year_enrollment_collection(school_year_label)
    raw_scope = str(args.get("scope", "")).strip().lower()
    q_value = str(args.get("q", "")).strip()
    grade_value = str(args.get("grade", "") or args.get("grade_level", "")).strip()
    section_value = str(args.get("section", "")).strip()
    face_status_value = str(args.get("face_status", "")).strip()
    student_record_id = str(args.get("student_record_id", "")).strip()
    student_id = str(args.get("student_id", "")).strip()

    effective_scope = resolve_students_export_scope(
        raw_scope,
        grade_level=grade_value,
        section_value=section_value,
        student_record_id=student_record_id,
        student_id=student_id,
        q_value=q_value,
    )

    rows = []
    selected_student = None
    scope_label = "All Students"
    scope_detail = "All student records"
    filename_prefix = "students_export"
    grade_level = ""
    normalized_section = normalize_section_value(section_value)
    normalized_face_status = normalize_face_status_filter(face_status_value)

    if effective_scope == "student":
        selected_student = resolve_students_export_student_row(
            collection,
            school_year_label,
            student_record_id=student_record_id,
            student_id=student_id,
            q_value=q_value,
        )
        rows = [selected_student] if selected_student else []
        scope_label = "Individual Student"
        if selected_student:
            scope_detail = selected_student.get("name") or selected_student.get("student_id") or "Selected student"
            filename_prefix = f"student_{sanitize_students_export_filename_part(selected_student.get('student_id') or selected_student.get('name'), 'record')}"
        else:
            scope_detail = "Selected student record"
            filename_prefix = "student_record"
    elif effective_scope == "grade":
        query, _q, grade_level, _section, _face_status, _school_year = build_students_query("", grade_value, "", "", school_year_label)
        rows = [normalize_enrollment_doc(row) for row in collection.find(query)]
        rows = sort_students_export_rows(rows)
        scope_label = "Grade Level"
        scope_detail = grade_level or grade_value or "Selected grade level"
        filename_prefix = f"students_grade_{sanitize_students_export_filename_part(scope_detail, 'grade')}"
    else:
        if effective_scope == "section":
            query, _q, grade_level, normalized_section, _face_status, _school_year = build_students_query("", grade_value, section_value, "", school_year_label)
            scope_label = "Section"
            scope_detail = f"{grade_level} - {normalized_section}" if grade_level and normalized_section else (normalized_section or section_value or "Selected section")
            filename_prefix = f"students_section_{sanitize_students_export_filename_part(scope_detail, 'section')}"
        elif effective_scope == "filtered":
            query, q_value, grade_level, normalized_section, normalized_face_status, _school_year = build_students_query(
                q_value,
                grade_value,
                section_value,
                face_status_value,
                school_year_label,
            )
            scope_label = "Filtered Records"
            detail_parts = []
            if q_value:
                detail_parts.append(f"Search: {q_value}")
            if grade_level:
                detail_parts.append(grade_level)
            if normalized_section:
                detail_parts.append(normalized_section)
            if normalized_face_status:
                detail_parts.append(format_students_export_face_status(normalized_face_status))
            scope_detail = ", ".join(detail_parts) if detail_parts else "Filtered student records"
            filename_prefix = "students_filtered"
        else:
            query, _q, grade_level, normalized_section, normalized_face_status, _school_year = build_students_query("", "", "", "", school_year_label)

        rows = [normalize_enrollment_doc(row) for row in collection.find(query)]
        rows = sort_students_export_rows(rows)

    return {
        "scope": effective_scope,
        "scope_label": scope_label,
        "scope_detail": scope_detail,
        "rows": rows,
        "selected_student": selected_student,
        "school_year": school_year_label,
        "query": q_value,
        "grade_level": grade_level,
        "section": normalized_section,
        "face_status": normalized_face_status,
        "filename_prefix": filename_prefix,
    }


def collect_export_scope_student_ids(collection, query, cap=5000):
    student_ids = []
    seen = set()
    for row in collection.find(query, {"student_id": 1}).limit(max(1, int(cap or 5000))):
        student_id = normalize_lrn_value(row.get("student_id"))
        if not student_id or student_id in seen:
            continue
        seen.add(student_id)
        student_ids.append(student_id)
    return student_ids


def build_logs_export_scope_payload(args, school_year_label):
    collection = get_school_year_enrollment_collection(school_year_label)
    raw_scope = str(args.get("scope", "")).strip().lower()
    q_value = str(args.get("q", "")).strip()
    grade_value = str(args.get("grade", "") or args.get("grade_level", "")).strip()
    section_value = str(args.get("section", "")).strip()
    student_record_id = str(args.get("student_record_id", "")).strip()
    student_id = str(args.get("student_id", "")).strip()

    effective_scope = resolve_students_export_scope(
        raw_scope,
        grade_level=grade_value,
        section_value=section_value,
        student_record_id=student_record_id,
        student_id=student_id,
        q_value=q_value,
    )

    selected_student = None
    scope_label = "All Matching Records"
    scope_detail = "All students in the selected log filters"
    filename_prefix = "logs_export"
    grade_level = ""
    normalized_section = normalize_section_value(section_value)
    scoped_student_ids = None

    if effective_scope == "student":
        selected_student = resolve_students_export_student_row(
            collection,
            school_year_label,
            student_record_id=student_record_id,
            student_id=student_id,
            q_value=q_value,
        )
        selected_student_id = normalize_lrn_value(
            (selected_student or {}).get("student_id") or student_id or q_value
        )
        scoped_student_ids = [selected_student_id] if selected_student_id else []
        scope_label = "Individual Student"
        if selected_student:
            scope_detail = selected_student.get("name") or selected_student.get("student_id") or "Selected student"
            filename_prefix = f"log_student_{sanitize_students_export_filename_part(selected_student.get('student_id') or selected_student.get('name'), 'record')}"
        else:
            scope_detail = selected_student_id or q_value or "Selected student"
            filename_prefix = "log_student_record"
    elif effective_scope == "grade":
        query, _q, grade_level, _section, _face_status, _school_year = build_students_query("", grade_value, "", "", school_year_label)
        scoped_student_ids = collect_export_scope_student_ids(collection, query)
        scope_label = "Grade Level"
        scope_detail = grade_level or grade_value or "Selected grade level"
        filename_prefix = f"logs_grade_{sanitize_students_export_filename_part(scope_detail, 'grade')}"
    elif effective_scope == "section":
        query, _q, grade_level, normalized_section, _face_status, _school_year = build_students_query("", grade_value, section_value, "", school_year_label)
        scoped_student_ids = collect_export_scope_student_ids(collection, query)
        scope_label = "Section"
        scope_detail = f"{grade_level} - {normalized_section}" if grade_level and normalized_section else (normalized_section or section_value or "Selected section")
        filename_prefix = f"logs_section_{sanitize_students_export_filename_part(scope_detail, 'section')}"

    return {
        "scope": effective_scope,
        "scope_label": scope_label,
        "scope_detail": scope_detail,
        "student_ids": scoped_student_ids,
        "selected_student": selected_student,
        "grade_level": grade_level,
        "section": normalized_section,
        "school_year": school_year_label,
        "filename_prefix": filename_prefix,
    }


def build_students_plain_export_pdf(export_payload, signatories):
    generated_at_label = now_local().strftime("%B %d, %Y %I:%M:%S %p")
    rows = export_payload["rows"]
    scope = export_payload["scope"]
    school_year_label = export_payload["school_year"] or get_current_school_year_label()
    styles = build_students_plain_export_styles()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=42,
        leftMargin=42,
        topMargin=42,
        bottomMargin=42,
        pageCompression=0,
    )
    doc.export_author = signatories.get("prepared_by_name", "Admin")
    doc.export_title = "Official Student Records Report"
    doc.export_subject = f"Student records export - {export_payload['scope_label']}"

    story = [build_pdf_export_header_block("Official Student Records Report")]

    metadata_items = [
        ("Generated At", generated_at_label),
        ("School Year", school_year_label),
        ("Export Scope", export_payload["scope_label"]),
        ("Total Records", str(len(rows))),
        ("Selection", export_payload["scope_detail"]),
        ("Face Status", format_students_export_face_status(export_payload["face_status"])),
    ]

    if scope == "grade":
        metadata_items.extend([("Grade Level", export_payload["grade_level"] or "N/A"), ("Section", "All sections")])
    elif scope == "section":
        metadata_items.extend([("Grade Level", export_payload["grade_level"] or "All grades"), ("Section", export_payload["section"] or "N/A")])
    elif scope == "student":
        student = export_payload["selected_student"] or {}
        metadata_items.extend([("Student Name", student.get("name") or "N/A"), ("Student ID", student.get("student_id") or student.get("lrn") or "N/A")])
    else:
        metadata_items.extend([("Search Query", export_payload["query"] or "All records"), ("Section", export_payload["section"] or "All sections")])

    story.append(build_students_plain_metadata_table(metadata_items, styles))
    story.append(Spacer(1, 0.18 * inch))

    if scope == "student":
        story.append(Paragraph("Student Details", styles["section_title"]))
        if export_payload["selected_student"]:
            student = export_payload["selected_student"]
            story.append(build_students_plain_key_value_table([
                ("LRN", student.get("lrn") or student.get("student_id") or "N/A"),
                ("Student ID", student.get("student_id") or student.get("lrn") or "N/A"),
                ("Name", student.get("name") or "N/A"),
                ("Grade Level", student.get("grade_level") or "N/A"),
                ("Section", student.get("section") or "N/A"),
                ("Gender", student.get("gender") or "N/A"),
                ("Status", student.get("status") or "Active"),
                ("Face Registered", "Yes" if student.get("face_registered") else "No"),
                ("Parent Contact", student.get("parent_contact") or "N/A"),
                ("Created At", format_students_export_timestamp(student.get("created_at"))),
                ("Updated At", format_students_export_timestamp(student.get("updated_at"))),
            ], styles))
        else:
            story.append(Paragraph("No student record matched the selected export filter.", styles["note"]))
    else:
        story.append(Paragraph("Student Records", styles["section_title"]))
        story.append(build_students_plain_records_table(rows, styles))

    story.append(build_report_signature_footer_block(styles, signatories=signatories))
    doc.build(story, onFirstPage=build_students_plain_export_footer, onLaterPages=build_students_plain_export_footer)
    buffer.seek(0)
    return buffer


def format_export_date_range_label(start_date, end_date):
    start_text = str(start_date or "").strip()
    end_text = str(end_date or "").strip()
    if start_text and end_text:
        return f"{start_text} to {end_text}"
    if start_text:
        return f"From {start_text}"
    if end_text:
        return f"Until {end_text}"
    return "All dates"


def format_export_sort_label(sort_by):
    return "Oldest First" if str(sort_by or "").strip().lower() == "oldest" else "Newest First"


def wants_inline_pdf_response():
    requested = str(
        request.args.get("disposition")
        or request.args.get("pdf_disposition")
        or request.args.get("view")
        or ""
    ).strip().lower()
    return requested in {"inline", "preview", "print"}


def send_generated_pdf(buffer, filename):
    response = send_file(
        buffer,
        as_attachment=not wants_inline_pdf_response(),
        download_name=filename,
        mimetype="application/pdf",
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/students/export_pdf", methods=["GET"])
@require_permission("students_write")
def students_export_pdf():
    selected_school_year = resolve_selected_school_year(request.args.get("school_year", ""))
    school_year_label = selected_school_year or get_current_school_year_label()
    signatories = build_report_signatories(resolve_current_admin_export_name())

    export_payload = build_students_export_payload(request.args, school_year_label)
    buffer = build_students_plain_export_pdf(export_payload, signatories)
    timestamp_suffix = now_local().strftime("%Y%m%d_%H%M%S")
    filename = f"{export_payload['filename_prefix']}_{timestamp_suffix}.pdf"
    return send_generated_pdf(buffer, filename)

    query, q_value, grade_level, section_value, face_status, school_year_label = build_students_query(
        request.args.get("q", ""),
        request.args.get("grade", "") or request.args.get("grade_level", ""),
        request.args.get("section", ""),
        request.args.get("face_status", ""),
        selected_school_year,
    )

    collection = get_school_year_enrollment_collection(school_year_label)
    raw_rows = list(collection.find(query).sort([("created_at", -1), ("name", 1)]))
    rows = [normalize_enrollment_doc(row) for row in raw_rows]

    admin_username = session.get("admin", "Admin")
    admin_name = admin_username
    try:
        user_doc, profile = current_user_profile()
        if profile and profile.get("fullName"):
            admin_name = profile.get("fullName")
    except Exception:
        pass

    generated_at = now_local()
    generated_at_label = generated_at.strftime("%B %d, %Y %I:%M:%S %p")
    filters = {
        "School Year": school_year_label or get_current_school_year_label(),
        "Search Query": q_value or "All records",
        "Grade Filter": grade_level or "All grades",
        "Section Filter": section_value or "All sections",
        "Face Status": format_students_export_face_status(face_status),
    }
    section_breakdown = build_students_export_section_breakdown(rows)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72,
    )
    doc.export_author = admin_name
    doc.export_title = "Official Student Records Report"
    doc.export_subject = "Student records export"
    doc.export_footer_title = "Cawitan High School Student Records Export"
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "StudentsExportTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=2,
    )
    doc_title_style = ParagraphStyle(
        "StudentsExportDocTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#14532d"),
        spaceBefore=4,
    )
    metadata_label_style = ParagraphStyle(
        "StudentsExportMetaLabel",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=11,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
    )
    metadata_value_style = ParagraphStyle(
        "StudentsExportMetaValue",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#334155"),
    )
    section_title_style = ParagraphStyle(
        "StudentsExportSectionTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=2,
    )
    table_header_style = ParagraphStyle(
        "StudentsExportTableHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.2,
        leading=8.6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )
    table_cell_style = ParagraphStyle(
        "StudentsExportTableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=8.4,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#111827"),
    )
    centered_cell_style = ParagraphStyle(
        "StudentsExportCenteredCell",
        parent=table_cell_style,
        alignment=TA_CENTER,
    )
    note_style = ParagraphStyle(
        "StudentsExportNote",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#475569"),
    )
    footer_value_style = ParagraphStyle(
        "StudentsExportFooterValue",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
    )
    signature_style = ParagraphStyle(
        "StudentsExportSignature",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )

    logo_path = resolve_students_export_logo_path()
    logo_flowable = Spacer(1, 0.85 * inch)
    if logo_path:
        logo_width = 0.8 * inch
        logo_height = 0.8 * inch
        try:
            with Image.open(logo_path) as logo_image:
                src_width, src_height = logo_image.size
                if src_width and src_height:
                    ratio = min(float(logo_width) / float(src_width), float(logo_height) / float(src_height))
                    logo_width = max(src_width * ratio, 1)
                    logo_height = max(src_height * ratio, 1)
        except Exception:
            logo_width = 0.8 * inch
            logo_height = 0.8 * inch
        logo_flowable = RLImage(logo_path, width=logo_width, height=logo_height)

    school_header = Paragraph(
        "<b>Cawitan High School</b><br/>"
        "<font size='9'>Cawitan, Sta. Catalina, Negros Oriental</font><br/>"
        "<font size='8'>Student Records Export</font>",
        ParagraphStyle(
            "StudentsExportHeaderBlock",
            parent=title_style,
            fontSize=17,
            leading=20,
        ),
    )
    header_table = Table(
        [[logo_flowable, school_header, Spacer(1, 0.85 * inch)]],
        colWidths=[1.0 * inch, 7.0 * inch, 1.0 * inch],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "CENTER"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("ALIGN", (2, 0), (2, 0), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph("Official Student Records Report", doc_title_style))
    story.append(Spacer(1, 0.08 * inch))
    story.append(HRFlowable(width="100%", thickness=1.6, color=colors.HexColor("#0f172a")))
    story.append(Spacer(1, 0.16 * inch))

    metadata_rows = [
        [
            Paragraph("<b>Generated At</b>", metadata_label_style),
            Paragraph(xml_escape(generated_at_label), metadata_value_style),
            Paragraph("<b>School Year</b>", metadata_label_style),
            Paragraph(xml_escape(filters["School Year"]), metadata_value_style),
        ],
        [
            Paragraph("<b>Total Records</b>", metadata_label_style),
            Paragraph(xml_escape(str(len(rows))), metadata_value_style),
            Paragraph("", metadata_label_style),
            Paragraph("", metadata_value_style),
        ],
        [
            Paragraph("<b>Search Query</b>", metadata_label_style),
            Paragraph(xml_escape(filters["Search Query"]), metadata_value_style),
            Paragraph("<b>Grade Filter</b>", metadata_label_style),
            Paragraph(xml_escape(filters["Grade Filter"]), metadata_value_style),
        ],
        [
            Paragraph("<b>Section Filter</b>", metadata_label_style),
            Paragraph(xml_escape(filters["Section Filter"]), metadata_value_style),
            Paragraph("<b>Face Status</b>", metadata_label_style),
            Paragraph(xml_escape(filters["Face Status"]), metadata_value_style),
        ],
    ]
    metadata_table = Table(
        metadata_rows,
        colWidths=[1.15 * inch, 3.35 * inch, 1.15 * inch, 3.35 * inch],
    )
    metadata_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#94a3b8")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(metadata_table)
    story.append(Spacer(1, 0.18 * inch))

    story.append(Paragraph("Section Breakdown", section_title_style))
    story.append(HRFlowable(width="100%", thickness=1.4, color=colors.HexColor("#0f172a")))
    story.append(Spacer(1, 0.1 * inch))

    if section_breakdown:
        breakdown_data = [[
            Paragraph("Grade", table_header_style),
            Paragraph("Section", table_header_style),
            Paragraph("Students", table_header_style),
        ]]
        for grade_label, section_label, count in section_breakdown:
            breakdown_data.append([
                Paragraph(xml_escape(grade_label or "Unassigned"), table_cell_style),
                Paragraph(xml_escape(section_label or "Unassigned"), table_cell_style),
                Paragraph(xml_escape(str(count)), centered_cell_style),
            ])

        breakdown_table = LongTable(
            breakdown_data,
            repeatRows=1,
            colWidths=[2.2 * inch, 4.7 * inch, 2.1 * inch],
        )
        breakdown_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#94a3b8")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("LINEBELOW", (0, 0), (-1, 0), 1.1, colors.HexColor("#0f172a")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for index in range(1, len(breakdown_data)):
            breakdown_style.append((
                "BACKGROUND",
                (0, index),
                (-1, index),
                colors.HexColor("#f8fafc") if index % 2 == 0 else colors.white,
            ))
        breakdown_table.setStyle(TableStyle(breakdown_style))
        story.append(breakdown_table)
    else:
        story.append(Paragraph("No records available for the selected filters.", metadata_value_style))

    story.append(Spacer(1, 0.22 * inch))
    story.append(Paragraph("Student Records", section_title_style))
    story.append(HRFlowable(width="100%", thickness=1.4, color=colors.HexColor("#0f172a")))
    story.append(Spacer(1, 0.1 * inch))

    student_table_data = [[
        Paragraph("LRN", table_header_style),
        Paragraph("Student ID", table_header_style),
        Paragraph("Name", table_header_style),
        Paragraph("Grade", table_header_style),
        Paragraph("Section", table_header_style),
        Paragraph("Parent Contact", table_header_style),
        Paragraph("Gender", table_header_style),
        Paragraph("Status", table_header_style),
        Paragraph("Face Registered", table_header_style),
        Paragraph("Created At", table_header_style),
        Paragraph("Updated At", table_header_style),
    ]]
    for row in rows:
        student_table_data.append([
            Paragraph(xml_escape(row.get("lrn") or "N/A"), table_cell_style),
            Paragraph(xml_escape(row.get("student_id") or "N/A"), table_cell_style),
            Paragraph(xml_escape(row.get("name") or "N/A"), table_cell_style),
            Paragraph(xml_escape(row.get("grade_level") or "N/A"), centered_cell_style),
            Paragraph(xml_escape(row.get("section") or "N/A"), table_cell_style),
            Paragraph(xml_escape(row.get("parent_contact") or "N/A"), table_cell_style),
            Paragraph(xml_escape(row.get("gender") or "N/A"), centered_cell_style),
            Paragraph(xml_escape(row.get("status") or "Active"), centered_cell_style),
            Paragraph("Yes" if row.get("face_registered") else "No", centered_cell_style),
            Paragraph(format_students_export_timestamp(row.get("created_at"), multiline=True), centered_cell_style),
            Paragraph(format_students_export_timestamp(row.get("updated_at"), multiline=True), centered_cell_style),
        ])

    if len(student_table_data) == 1:
        student_table_data.append([
            Paragraph("No student records matched the selected filters.", table_cell_style),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    student_table = LongTable(
        student_table_data,
        repeatRows=1,
        colWidths=[
            0.82 * inch,
            0.82 * inch,
            1.55 * inch,
            0.62 * inch,
            0.75 * inch,
            0.95 * inch,
            0.58 * inch,
            0.62 * inch,
            0.67 * inch,
            0.81 * inch,
            0.81 * inch,
        ],
    )
    student_table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#94a3b8")),
        ("INNERGRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#cbd5e1")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.1, colors.HexColor("#0f172a")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for index in range(1, len(student_table_data)):
        student_table_style.append((
            "BACKGROUND",
            (0, index),
            (-1, index),
            colors.HexColor("#f8fafc") if index % 2 == 0 else colors.white,
        ))
    if len(student_table_data) == 2 and not rows:
        student_table_style.append(("SPAN", (0, 1), (-1, 1)))
        student_table_style.append(("ALIGN", (0, 1), (-1, 1), "CENTER"))
    student_table.setStyle(TableStyle(student_table_style))
    story.append(student_table)

    footer_block = KeepTogether([
        Spacer(1, 0.24 * inch),
        HRFlowable(width="100%", thickness=1.2, color=colors.HexColor("#0f172a")),
        Spacer(1, 0.14 * inch),
        Table(
            [[
                Paragraph(
                    "<b>Exported By:</b><br/><br/>______________________________",
                    footer_value_style,
                ),
                Paragraph(
                    "______________________________<br/>Principal",
                    signature_style,
                ),
            ]],
            colWidths=[4.5 * inch, 4.5 * inch],
            style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]),
        ),
        Spacer(1, 0.14 * inch),
        Paragraph("This document is system-generated and valid without signature.", note_style),
    ])
    story.append(footer_block)

    doc.build(story, onFirstPage=build_school_export_footer, onLaterPages=build_school_export_footer)
    buffer.seek(0)

    filename = f"students_export_{now_local().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_generated_pdf(buffer, filename)


@app.route("/api/students/stats", methods=["GET"])
@require_permission("students_read", api=True)
def api_students_stats():
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    return api_success({"stats": build_students_stats_payload(school_year_label), "school_year": school_year_label})


@app.route("/api/students", methods=["GET", "POST"])
@require_permission("students_read", api=True)
def api_students_collection():
    if request.method == "GET":
        selected_school_year = resolve_selected_school_year(request.args.get("school_year", ""))
        collection = get_school_year_enrollment_collection(selected_school_year)
        query, q_value, grade_level, section_value, face_status, school_year_label = build_students_query(
            request.args.get("q", ""),
            request.args.get("grade", "") or request.args.get("grade_level", ""),
            request.args.get("section", ""),
            request.args.get("face_status", ""),
            selected_school_year,
        )
        try:
            limit = int(request.args.get("limit", "5"))
        except (TypeError, ValueError):
            limit = 5
        try:
            page = int(request.args.get("page", "1"))
        except (TypeError, ValueError):
            page = 1

        limit = min(max(limit, 1), 100)
        page = max(page, 1)
        skip = (page - 1) * limit

        total = collection.count_documents(query)
        rows = collection.find(query).sort([("created_at", -1), ("name", 1)]).skip(skip).limit(limit)
        payload = {
            "students": [normalize_enrollment_doc(row) for row in rows],
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit if total else 1,
            "school_year": school_year_label,
            "archived_view": school_year_label != get_current_school_year_label(),
            "filters": {
                "q": q_value,
                "grade": grade_level,
                "section": section_value,
                "face_status": face_status,
            },
        }
        return api_success(payload)

    if not has_permission("students_write"):
        return api_error("Forbidden", 403)

    payload = request_payload()
    school_year_label = resolve_selected_school_year(payload.get("school_year", ""))
    if school_year_label != get_current_school_year_label():
        return api_error("Archived school years are read-only.", 403, "school_year")

    student_data, err_message, err_field = sanitize_personal_student_payload(payload, school_year=school_year_label)
    if err_message:
        return api_error(err_message, 400, err_field)

    collection = get_school_year_enrollment_collection(school_year_label)
    if collection.count_documents({"school_year": school_year_label, "student_id": student_data["lrn"]}) > 0:
        return api_error(f"Student is already enrolled for School Year {school_year_label}.", 400, "lrn")

    try:
        upsert_manual_section(
            student_data.get("grade_level", ""),
            student_data.get("section", ""),
            created_by=session.get("admin", ""),
            school_year=school_year_label,
        )
    except ValueError as exc:
        return api_error(str(exc), 400, "section")

    existing_student = students.find_one(build_lrn_duplicate_query(student_data["lrn"]))
    if existing_student:
        update_fields = {
            "name": student_data.get("name", existing_student.get("name", "")),
            "gender": student_data.get("gender", existing_student.get("gender", "")),
            "sex": student_data.get("gender", existing_student.get("sex", "")),
            "parent_contact": student_data.get("parent_contact", existing_student.get("parent_contact", "")),
            "grade_level": student_data.get("grade_level", existing_student.get("grade_level", "")),
            "grade": student_data.get("grade_level", existing_student.get("grade", "")),
            "section": student_data.get("section", existing_student.get("section", "")),
            "status": student_data.get("status", existing_student.get("status", "Active")),
            "updated_at": now_iso(),
        }
        students.update_one({"_id": existing_student["_id"]}, {"$set": update_fields})
        student_doc = students.find_one({"_id": existing_student["_id"]}) or existing_student
    else:
        student_doc = build_new_student_document(student_data)
        try:
            inserted = students.insert_one(student_doc)
        except DuplicateKeyError:
            return api_error("LRN already exists.", 400, "lrn")
        student_doc = students.find_one({"_id": inserted.inserted_id}) or student_doc

    try:
        saved_enrollment = upsert_student_enrollment(
            student_doc,
            school_year_label,
            grade_level=student_data.get("grade_level", ""),
            section=student_data.get("section", ""),
            status=student_data.get("status", "Active"),
            update_existing=False,
        )
    except ValueError as exc:
        return api_error(str(exc), 400, "lrn")

    sync_student_base_fields_to_enrollments(student_doc)
    signal_data_change("students")
    return api_success({
        "message": f"Student enrolled successfully for School Year {school_year_label}.",
        "student": normalize_enrollment_doc(saved_enrollment),
        "school_year": school_year_label,
    }, 201)


@app.route("/api/logs/export/students", methods=["GET"])
@require_permission("logs", api=True)
def api_logs_export_students():
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    collection = get_school_year_enrollment_collection(school_year_label)
    query, _q_value, _grade_level, _section_value, _face_status, school_year_label = build_students_query(
        request.args.get("q", ""),
        "",
        "",
        "",
        school_year_label,
    )
    try:
        limit = int(request.args.get("limit", "8"))
    except (TypeError, ValueError):
        limit = 8
    limit = min(max(limit, 1), 20)

    rows = collection.find(query).sort([("name", 1), ("created_at", -1)]).limit(limit)
    return api_success({
        "students": [normalize_enrollment_doc(row) for row in rows],
        "school_year": school_year_label,
        "limit": limit,
    })


@app.route("/api/students/import", methods=["POST"])
@require_permission("students_write", api=True)
def api_students_import():
    upload = request.files.get("file")
    if upload is None or not str(upload.filename or "").strip():
        return api_error("Excel file is required.", 400, "file")

    filename = secure_filename(upload.filename or "")
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in STUDENT_IMPORT_ALLOWED_EXTENSIONS:
        return api_error("Only .xlsx files are supported.", 400, "file")

    file_bytes = upload.read()
    try:
        parsed_rows = parse_student_import_workbook(file_bytes)
    except ValueError as exc:
        return api_error(str(exc), 400, "file")

    total_rows_read = len(parsed_rows)
    school_year_label = resolve_selected_school_year(request.form.get("school_year", ""))
    if school_year_label != get_current_school_year_label():
        return api_error("Archived school years are read-only.", 403, "school_year")
    default_grade_level = normalize_grade_level(request.form.get("default_grade_level", ""))
    default_section = normalize_section_value(request.form.get("default_section", ""))

    validation_errors = []
    pending_rows = []
    summary_skipped_count = 0
    for row in parsed_rows:
        if is_student_import_summary_row(row):
            summary_skipped_count += 1
            continue

        row_grade_level = row.get("grade_level", "") or default_grade_level
        row_section = row.get("section", "") or default_section
        student_data, err_message, _err_field = sanitize_personal_student_payload({
            "lrn": row.get("lrn", ""),
            "name": row.get("name", ""),
            "grade_level": row_grade_level,
            "gender": row.get("gender", ""),
            "section": row_section,
            "status": "Active",
        }, require_existing_section=True, school_year=school_year_label)
        if err_message:
            validation_errors.append(f"Row {row.get('row_number', '?')}: {err_message}")
            continue
        pending_rows.append({
            "row_number": row.get("row_number", ""),
            "student_data": student_data,
        })

    candidate_lrns = sorted({item["student_data"]["lrn"] for item in pending_rows if item.get("student_data")})
    existing_lrns = set()
    if candidate_lrns:
        collection = get_school_year_enrollment_collection(school_year_label)
        for row in collection.find(
            {"school_year": school_year_label, "student_id": {"$in": candidate_lrns}},
            {"student_id": 1, "lrn": 1},
        ):
            existing_lrn = normalize_lrn_value(row.get("lrn") or row.get("student_id"))
            if existing_lrn:
                existing_lrns.add(existing_lrn)

    imported_count = 0
    duplicate_count = 0
    in_file_seen_lrns = set()

    for item in pending_rows:
        row_number = item.get("row_number", "")
        student_data = dict(item.get("student_data") or {})
        lrn_value = student_data.get("lrn", "")
        if not lrn_value:
            validation_errors.append(f"Row {row_number}: LRN is required.")
            continue
        if lrn_value in in_file_seen_lrns:
            duplicate_count += 1
            validation_errors.append(f"Row {row_number}: Duplicate LRN in uploaded file ({lrn_value}).")
            continue
        if lrn_value in existing_lrns:
            duplicate_count += 1
            validation_errors.append(f"Row {row_number}: LRN already exists ({lrn_value}).")
            continue

        existing_student = students.find_one(build_lrn_duplicate_query(lrn_value))
        if existing_student:
            update_fields = {
                "name": student_data.get("name", existing_student.get("name", "")),
                "gender": student_data.get("gender", existing_student.get("gender", "")),
                "sex": student_data.get("gender", existing_student.get("sex", "")),
                "parent_contact": student_data.get("parent_contact", existing_student.get("parent_contact", "")),
                "grade_level": student_data.get("grade_level", existing_student.get("grade_level", "")),
                "grade": student_data.get("grade_level", existing_student.get("grade", "")),
                "section": student_data.get("section", existing_student.get("section", "")),
                "status": student_data.get("status", existing_student.get("status", "Active")),
                "updated_at": now_iso(),
            }
            students.update_one({"_id": existing_student["_id"]}, {"$set": update_fields})
            student_doc = students.find_one({"_id": existing_student["_id"]}) or existing_student
        else:
            student_doc = build_new_student_document(student_data)
            try:
                inserted = students.insert_one(student_doc)
            except DuplicateKeyError:
                duplicate_count += 1
                validation_errors.append(f"Row {row_number}: LRN already exists ({lrn_value}).")
                continue
            student_doc = students.find_one({"_id": inserted.inserted_id}) or student_doc

        try:
            upsert_student_enrollment(
                student_doc,
                school_year_label,
                grade_level=student_data.get("grade_level", ""),
                section=student_data.get("section", ""),
                status=student_data.get("status", "Active"),
                update_existing=False,
            )
            sync_student_base_fields_to_enrollments(student_doc)
        except ValueError as exc:
            duplicate_count += 1
            validation_errors.append(f"Row {row_number}: {exc}")
            continue

        imported_count += 1
        in_file_seen_lrns.add(lrn_value)

    if imported_count > 0:
        signal_data_change("students")

    error_count = len(validation_errors)
    invalid_count = max(error_count - duplicate_count, 0)
    skipped_count = max(total_rows_read - imported_count, 0)
    message = (
        f"Import completed. Rows read: {total_rows_read}. "
        f"Imported: {imported_count}. "
        f"Skipped/failed: {skipped_count}."
    )
    if skipped_count > 0:
        message = (
            f"{message} Duplicates: {duplicate_count}. "
            f"Invalid: {invalid_count}. "
            f"Summary rows skipped: {summary_skipped_count}."
        )

    response_payload = {
        "message": message,
        "school_year": school_year_label,
        "total_rows_read": total_rows_read,
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "duplicate_count": duplicate_count,
        "invalid_count": invalid_count,
        "summary_skipped_count": summary_skipped_count,
        "error_count": error_count,
        "errors": validation_errors[:100],
    }

    if imported_count == 0:
        return api_success(response_payload, 200)
    return api_success(response_payload, 201)


@app.route("/api/students/import/template", methods=["GET"])
@require_permission("students_write", api=True)
def api_students_import_template():
    try:
        template_bytes = build_student_import_template_bytes()
    except ValueError as exc:
        return api_error(str(exc), 500, "file")

    filename = f"student_import_template_{now_local().strftime('%Y%m%d')}.xlsx"
    return Response(
        template_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/api/students/<id>", methods=["GET", "PUT", "DELETE"])
@require_permission("students_read", api=True)
def api_students_item(id):
    enrollment_oid = parse_student_oid(id)
    if not enrollment_oid:
        return api_error("Invalid student record id.", 400, "id")

    enrollment_doc, enrollment_school_year = find_student_enrollment_record(enrollment_oid)
    if not enrollment_doc:
        return api_error("Student record not found.", 404)
    enrollment_school_year = normalize_school_year_value(enrollment_school_year or enrollment_doc.get("school_year"))
    collection = get_school_year_enrollment_collection(enrollment_school_year)

    if request.method == "GET":
        return api_success({
            "student": normalize_enrollment_doc(enrollment_doc),
            "archived_view": enrollment_school_year != get_current_school_year_label(),
        })

    if not has_permission("students_write"):
        return api_error("Forbidden", 403)

    if enrollment_school_year != get_current_school_year_label():
        return api_error("Archived school years are read-only.", 403, "school_year")

    if request.method == "DELETE":
        student_ref_id = parse_student_oid(enrollment_doc.get("student_ref_id"))
        base_student_doc = students.find_one({"_id": student_ref_id}) if student_ref_id else students.find_one(
            build_lrn_duplicate_query(normalize_lrn_value(enrollment_doc.get("student_id") or enrollment_doc.get("lrn")))
        )
        deletion_result = delete_student_profile_and_related_data(
            student_doc=base_student_doc,
            enrollment_doc=enrollment_doc,
        )
        if not deletion_result.get("ok"):
            return api_error(
                deletion_result.get("message") or "Student deletion could not be fully completed.",
                500,
                "student",
                {
                    "counts": deletion_result.get("counts", {}),
                    "validation": deletion_result.get("validation", {}),
                },
            )
        log_audit_event(
            action="student.delete",
            target_type="student",
            target_id=deletion_result.get("student_id") or str(enrollment_oid),
            details={
                "student_name": deletion_result.get("student_name", ""),
                "counts": deletion_result.get("counts", {}),
            },
        )
        signal_data_change("students", "sections")
        return api_success({
            "message": deletion_result.get("message") or f"Student removed from School Year {enrollment_school_year}.",
            "counts": deletion_result.get("counts", {}),
            "validation": deletion_result.get("validation", {}),
        })

    student_ref_id = parse_student_oid(enrollment_doc.get("student_ref_id"))
    existing_doc = students.find_one({"_id": student_ref_id}) if student_ref_id else None
    if not existing_doc:
        return api_error("Student profile not found.", 404)

    payload = request_payload()
    student_data, err_message, err_field = sanitize_personal_student_payload(
        payload,
        existing_doc=existing_doc,
        school_year=enrollment_school_year,
    )
    if err_message:
        return api_error(err_message, 400, err_field)

    if student_data["lrn"] != normalize_lrn_value(existing_doc.get("lrn") or existing_doc.get("student_id")):
        return api_error("LRN cannot be changed for an existing student profile.", 400, "lrn")

    try:
        upsert_manual_section(
            student_data.get("grade_level", ""),
            student_data.get("section", ""),
            created_by=session.get("admin", ""),
            school_year=enrollment_school_year,
        )
    except ValueError as exc:
        return api_error(str(exc), 400, "section")

    student_update = {
        "name": student_data.get("name", existing_doc.get("name", "")),
        "gender": student_data.get("gender", existing_doc.get("gender", "")),
        "sex": student_data.get("gender", existing_doc.get("sex", "")),
        "parent_contact": student_data.get("parent_contact", existing_doc.get("parent_contact", "")),
        "grade_level": student_data.get("grade_level", existing_doc.get("grade_level", "")),
        "grade": student_data.get("grade_level", existing_doc.get("grade", "")),
        "section": student_data.get("section", existing_doc.get("section", "")),
        "status": student_data.get("status", existing_doc.get("status", "Active")),
        "updated_at": now_iso(),
    }
    students.update_one({"_id": existing_doc["_id"]}, {"$set": student_update})
    sync_student_base_fields_to_enrollments(students.find_one({"_id": existing_doc["_id"]}) or existing_doc)
    updated_doc = upsert_student_enrollment(
        students.find_one({"_id": existing_doc["_id"]}) or existing_doc,
        enrollment_school_year,
        grade_level=student_data.get("grade_level", ""),
        section=student_data.get("section", ""),
        status=student_data.get("status", "Active"),
        update_existing=True,
    )
    signal_data_change("students")
    return api_success({
        "message": f"Student enrollment updated for School Year {enrollment_school_year}.",
        "student": normalize_enrollment_doc(updated_doc),
    })


@app.route("/api/sections", methods=["GET", "POST"])
@require_permission("students_read", api=True)
def api_sections():
    if request.method == "GET":
        school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
        grade_filter = request.args.get("grade", "").strip() or request.args.get("grade_level", "").strip()
        return api_success({
            "school_year": school_year_label,
            "sections_by_grade": build_sections_by_grade(grade_filter, school_year=school_year_label),
        })

    if not has_permission("students_write"):
        return api_error("Forbidden", 403)

    payload = request_payload()
    school_year_label = resolve_selected_school_year(payload.get("school_year", ""))
    if school_year_label != get_current_school_year_label():
        return api_error("Archived school years are read-only.", 403, "school_year")
    grade_value = str(payload.get("grade", payload.get("grade_level", "")) or "").strip()
    section_value = payload.get("section", "")
    try:
        section_doc = upsert_manual_section(
            grade_value,
            section_value,
            created_by=session.get("admin", ""),
            school_year=school_year_label,
        )
    except ValueError as exc:
        message = str(exc)
        field = "grade" if "grade" in message.lower() else "section"
        return api_error(message, 400, field)

    return api_success({
        "message": "Section saved successfully.",
        "school_year": school_year_label,
        "section": section_doc,
        "sections_by_grade": build_sections_by_grade(school_year=school_year_label),
    })


@app.route("/api/sections/stats", methods=["GET"])
@require_permission("students_read", api=True)
def api_sections_stats():
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    collection = get_school_year_enrollment_collection(school_year_label)
    grade_value = request.args.get("grade", "").strip() or request.args.get("grade_level", "").strip()
    section_value = normalize_section_value(request.args.get("section", ""))

    if not grade_value:
        return api_error("Grade level is required.", 400, "grade")
    if not section_value:
        return api_error("Section is required.", 400, "section")

    grade_clause, grade_level = build_grade_filter(grade_value)
    if not grade_clause:
        return api_error("Invalid grade level.", 400, "grade")

    section_clause = {"section": section_value}
    base_conditions = [{"school_year": school_year_label}, grade_clause, section_clause]
    base_query = {"$and": base_conditions}

    male_values = ["Male", "male", "M", "m"]
    female_values = ["Female", "female", "F", "f"]
    male_query = {
        "$and": base_conditions + [{
            "$or": [
                {"gender": {"$in": male_values}},
                {"sex": {"$in": male_values}},
            ]
        }]
    }
    female_query = {
        "$and": base_conditions + [{
            "$or": [
                {"gender": {"$in": female_values}},
                {"sex": {"$in": female_values}},
            ]
        }]
    }

    return api_success({
        "school_year": school_year_label,
        "grade_level": grade_level,
        "section": section_value,
        "stats": {
            "total": collection.count_documents(base_query),
            "male": collection.count_documents(male_query),
            "female": collection.count_documents(female_query),
        },
    })


@app.route("/api/sections/clear-students", methods=["POST"])
@require_permission("students_write", api=True)
def api_sections_clear_students():
    payload = request_payload()
    school_year_label = resolve_selected_school_year(payload.get("school_year", ""))
    collection = get_school_year_enrollment_collection(school_year_label)
    if school_year_label != get_current_school_year_label():
        return api_error("Archived school years are read-only.", 403, "school_year")
    grade_value = str(payload.get("grade", payload.get("grade_level", "")) or "").strip()
    section_value = normalize_section_value(payload.get("section", ""))

    if not grade_value:
        return api_error("Grade level is required.", 400, "grade")
    if not section_value:
        return api_error("Section is required.", 400, "section")

    grade_clause, grade_level = build_grade_filter(grade_value)
    if not grade_clause:
        return api_error("Invalid grade level.", 400, "grade")

    query = {"$and": [{"school_year": school_year_label}, grade_clause, {"section": section_value}]}
    update_result = collection.update_many(
        query,
        {
            "$set": {
                "section": "",
                "updated_at": now_iso(),
            }
        },
    )
    if update_result.modified_count > 0:
        signal_data_change("students", "sections")

    return api_success({
        "message": f"Removed {update_result.modified_count} student(s) from {grade_level} - {section_value} for School Year {school_year_label}.",
        "removed_count": int(update_result.modified_count),
        "school_year": school_year_label,
        "grade_level": grade_level,
        "section": section_value,
    })


def save_face_registration(student_id, is_update=False):
    student_oid = parse_student_oid(student_id)
    if not student_oid:
        return api_error("Invalid student id.", 400, "id")

    student_doc = students.find_one({"_id": student_oid})
    if not student_doc:
        return api_error("Student not found.", 404)

    normalized_student_id = normalize_lrn_value(student_doc.get("student_id") or student_doc.get("lrn"))
    existing_face_registration = bool(
        student_doc.get("face_registered")
        or student_doc.get("face_data")
        or student_doc.get("faces")
        or student_doc.get("face_encodings")
        or student_doc.get("face_embeddings")
    )
    if existing_face_registration and not is_update:
        return api_error(
            "This student already has a registered face profile. Use the update action instead.",
            409,
            "faces",
            {"requires_update": True},
        )

    payload = request_payload()
    faces_array, face_encodings, capture_details, err_message, err_field, err_extra = parse_faces_payload(payload)
    if err_message:
        return api_error(err_message, 400, err_field, err_extra)

    diversity_gate = validate_face_registration_diversity(face_encodings, capture_details)
    if not diversity_gate.get("ok"):
        return api_error(
            diversity_gate.get("message") or "The guided captures need more variation before saving.",
            400,
            "faces",
            diversity_gate.get("details") or {},
        )

    db_centroids, db_students = load_face_index_from_db(
        allow_legacy_fallback=False,
        prefer_current_roster=True,
    )
    face_conflict = detect_face_registration_conflict(
        face_encodings,
        {"students": db_students, "centroids": db_centroids},
        exclude_student_id=normalized_student_id,
        duplicate_distance_threshold=FACE_REGISTRATION_CONFLICT_DISTANCE,
        duplicate_mean_threshold=FACE_REGISTRATION_CONFLICT_MEAN_DISTANCE,
        support_distance_threshold=FACE_REGISTRATION_CONFLICT_SUPPORT_DISTANCE,
        min_support_count=FACE_REGISTRATION_CONFLICT_SUPPORT_COUNT,
    )
    if face_conflict:
        conflict_student = face_conflict.get("student") or {}
        conflict_name = str(conflict_student.get("name") or "another student").strip()
        conflict_student_id = str(conflict_student.get("student_id") or "").strip()
        return api_error(
            f"These face captures closely match the registered profile for {conflict_name}"
            f"{f' ({conflict_student_id})' if conflict_student_id else ''}.",
            409,
            "faces",
            {
                "conflict_student_id": conflict_student_id,
                "conflict_student_name": conflict_name,
                "best_distance": round(float(face_conflict.get("best_distance") or 0.0), 6),
                "support_count": int(face_conflict.get("support_count") or 0),
            },
        )

    face_metadata = diversity_gate.get("metadata") if isinstance(diversity_gate.get("metadata"), dict) else build_face_registration_metadata(face_encodings)
    pose_summary = diversity_gate.get("pose_summary") if isinstance(diversity_gate.get("pose_summary"), dict) else summarize_face_capture_pose_diversity(capture_details.get("capture_meta", []))

    update_doc = {
        "face_data": faces_array,
        "face_encodings": face_encodings,
        "face_capture_profile": capture_details.get("capture_profile", "standard") if isinstance(capture_details, dict) else "standard",
        "face_capture_count": int(face_metadata.get("face_encoding_count") or len(faces_array)),
        "face_capture_meta": capture_details.get("capture_meta", []) if isinstance(capture_details, dict) else [],
        "face_pose_bucket_count": int(pose_summary.get("unique_bucket_count") or 0),
        "face_pose_sample_count": int(pose_summary.get("sample_count") or 0),
        "face_pose_yaw_span": float(pose_summary.get("yaw_span") or 0.0),
        "face_pose_pitch_span": float(pose_summary.get("pitch_span") or 0.0),
        "profile_photo": faces_array[0] if faces_array else "",
        "face_registered": True,
        "face_updated_at": now_local(),
        "updated_at": now_iso(),
        **face_metadata,
    }
    students.update_one(
        {"_id": student_oid},
        {
            "$set": update_doc,
            "$unset": {
                "faces": "",
                "face_embeddings": "",
            },
        },
    )
    refresh_scan_face_index_if_active()
    saved_doc = students.find_one({"_id": student_oid})
    if saved_doc:
        sync_student_base_fields_to_enrollments(saved_doc)
    signal_data_change("students")
    message = "Face registration updated successfully." if is_update else "Face registered successfully."
    return api_success({"message": message, "student": normalize_student_doc(saved_doc)})


@app.route("/api/students/<id>/face/register", methods=["POST"])
@require_permission("face_register", api=True)
def api_student_face_register(id):
    return save_face_registration(id, is_update=False)


@app.route("/api/students/<id>/face/update", methods=["PUT"])
@require_permission("face_register", api=True)
def api_student_face_update(id):
    return save_face_registration(id, is_update=True)


@app.route("/students/delete/<id>", methods=["POST"])
@require_permission("students_write")
def delete_student(id):
    try:
        student_oid = parse_student_oid(id)
        base_student_doc = students.find_one({"_id": student_oid}) if student_oid else None
        if not base_student_doc:
            return redirect(url_for("students_page", message="Failed to delete student record.", message_type="error"))
        deletion_result = delete_student_profile_and_related_data(student_doc=base_student_doc)
        if deletion_result.get("ok"):
            log_audit_event(
                action="student.delete",
                target_type="student",
                target_id=deletion_result.get("student_id") or id,
                details={
                    "student_name": deletion_result.get("student_name", ""),
                    "counts": deletion_result.get("counts", {}),
                },
            )
            signal_data_change("students", "sections")
            return redirect(url_for("students_page", message=deletion_result.get("message") or "Student deleted successfully.", message_type="success"))
        return redirect(url_for("students_page", message=deletion_result.get("message") or "Failed to delete student record.", message_type="error"))
    except Exception:
        return redirect(url_for("students_page", message="Failed to delete student record.", message_type="error"))


# =====================================
# LOG ROUTES
# =====================================
def build_gate_logs_query(args, school_year="", include_export_scope=False):
    school_year_label = resolve_selected_school_year(school_year or args.get("school_year", ""))
    q = args.get("q", "").strip()
    start_date = args.get("start_date", "").strip()
    end_date = args.get("end_date", "").strip()
    status_filter = args.get("status", "").strip()
    session_filter = args.get("session", "").strip().upper()
    sort_by = args.get("sort", "newest").strip()

    query = {"school_year": school_year_label}
    q_regex = contains_regex_filter(q)
    if q_regex:
        query["$or"] = [
            {"student_name": q_regex},
            {"student_id": q_regex},
            {"date": q_regex},
        ]

    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        query["date"] = date_query

    if status_filter:
        query["status"] = status_filter

    if session_filter in ("IN", "OUT"):
        query["gate_action"] = session_filter

    sort_spec = [("timestamp", -1)] if sort_by != "oldest" else [("timestamp", 1)]

    filters_payload = {
        "school_year": school_year_label,
        "q": q,
        "start_date": start_date,
        "end_date": end_date,
        "status": status_filter,
        "session": session_filter,
        "sort": sort_by,
    }
    if include_export_scope:
        export_scope = build_logs_export_scope_payload(args, school_year_label)
        scoped_student_ids = export_scope.get("student_ids")
        if scoped_student_ids is not None:
            query["student_id"] = {"$in": scoped_student_ids}
        filters_payload["export_scope"] = export_scope
    return query, sort_spec, filters_payload, school_year_label


def build_student_photo_map(student_ids):
    normalized_ids = sorted({
        str(student_id or "").strip()
        for student_id in student_ids
        if str(student_id or "").strip()
    })
    if not normalized_ids:
        return {}

    photo_map = {}
    projection = {"student_id": 1, "profile_photo": 1, "face_data": 1, "faces": 1}
    for row in students.find({"student_id": {"$in": normalized_ids}}, projection):
        student_id = str(row.get("student_id", "")).strip()
        if not student_id or student_id in photo_map:
            continue
        normalized = normalize_student_doc(row)
        photo_map[student_id] = normalized.get("profile_photo", "")
    return photo_map


def build_sms_logs_query(args, school_year="", include_export_scope=False):
    school_year_label = resolve_selected_school_year(school_year or args.get("school_year", ""))
    q = args.get("q", "").strip()
    start_date = args.get("start_date", "").strip()
    end_date = args.get("end_date", "").strip()
    status_filter = args.get("status", "").strip()
    sort_by = args.get("sort", "newest").strip()

    query = {"school_year": school_year_label}
    q_regex = contains_regex_filter(q)
    if q_regex:
        query["$or"] = [
            {"name": q_regex},
            {"student_id": q_regex},
            {"parent_contact": q_regex},
        ]

    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        query["date"] = date_query

    if status_filter.strip().lower() in ("sent", "failed", "queued", "sending", "skipped"):
        query["status"] = sms_status_mongo_filter(status_filter.strip().lower())

    sort_spec = [("timestamp", -1)] if sort_by != "oldest" else [("timestamp", 1)]

    filters_payload = {
        "school_year": school_year_label,
        "q": q,
        "start_date": start_date,
        "end_date": end_date,
        "status": status_filter,
        "sort": sort_by,
    }
    if include_export_scope:
        export_scope = build_logs_export_scope_payload(args, school_year_label)
        scoped_student_ids = export_scope.get("student_ids")
        if scoped_student_ids is not None:
            query["student_id"] = {"$in": scoped_student_ids}
        filters_payload["export_scope"] = export_scope
    return query, sort_spec, filters_payload, school_year_label


def build_gate_logs_summary_rows(rows):
    summary = {
        "Total Entries": len(rows),
        "IN Logs": 0,
        "OUT Logs": 0,
        "Present": 0,
        "Late": 0,
        "Absent": 0,
    }
    extra_statuses = {}

    for row in rows:
        gate_action = str(row.get("gate_action") or "").strip().upper()
        if gate_action == "IN":
            summary["IN Logs"] += 1
        elif gate_action == "OUT":
            summary["OUT Logs"] += 1

        status = str(row.get("status") or "").strip().title()
        if status in {"Present", "Late", "Absent"}:
            summary[status] += 1
        elif status:
            extra_statuses[status] = extra_statuses.get(status, 0) + 1

    ordered = [(label, count) for label, count in summary.items()]
    for status_label in sorted(extra_statuses.keys()):
        ordered.append((status_label, extra_statuses[status_label]))
    return ordered


def build_sms_logs_summary_rows(rows):
    summary = {
        "Total Logs": len(rows),
        "Sent": 0,
        "Failed": 0,
        "Queued": 0,
        "Sending": 0,
        "Skipped": 0,
    }
    extra_statuses = {}

    for row in rows:
        status_raw = str(row.get("status") or "").strip()
        status_key = status_raw.lower()
        if status_key == "sent":
            summary["Sent"] += 1
        elif status_key == "failed":
            summary["Failed"] += 1
        elif status_key == "queued":
            summary["Queued"] += 1
        elif status_key == "sending":
            summary["Sending"] += 1
        elif status_key == "skipped":
            summary["Skipped"] += 1
        elif status_raw:
            extra_statuses[status_raw.title()] = extra_statuses.get(status_raw.title(), 0) + 1

    ordered = [(label, count) for label, count in summary.items()]
    for status_label in sorted(extra_statuses.keys()):
        ordered.append((status_label, extra_statuses[status_label]))
    return ordered


@app.route("/gate-logs")
@require_permission("logs")
def gate_logs_page():
    query, sort_spec, filters_payload, selected_school_year = build_gate_logs_query(request.args)
    logs_collection, selected_school_year, archived_view = get_attendance_logs_storage(selected_school_year)
    try:
        page = max(int(request.args.get("page", "1")), 1)
    except ValueError:
        page = 1

    per_page = 10
    total_filtered = logs_collection.count_documents(query)
    pagination = build_pagination_payload(page, per_page, total_filtered, filters_payload, "gate_logs_page")
    skip = (pagination["page"] - 1) * per_page

    rows = list(logs_collection.find(query).sort(sort_spec).skip(skip).limit(per_page))
    photo_map = build_student_photo_map([row.get("student_id", "") for row in rows])
    logs = [
        serialize_gate_log_display_row(
            row,
            profile_photo=photo_map.get(str(row.get("student_id", "")).strip(), ""),
        )
        for row in rows
    ]

    stats = {
        "total_entries": total_filtered,
        "total_in": logs_collection.count_documents({**query, "gate_action": "IN"}),
        "total_out": logs_collection.count_documents({**query, "gate_action": "OUT"}),
        "late_count": logs_collection.count_documents({**query, "status": "Late"}),
    }

    return render_template(
        "gate_logs.html",
        logs=logs,
        stats=stats,
        filters=filters_payload,
        pagination=pagination,
        export_query=urlencode({k: v for k, v in filters_payload.items() if v not in ("", None)}),
        grade_options=list(GRADE_LEVEL_OPTIONS),
        sections_by_grade=build_sections_by_grade(school_year=selected_school_year),
        archived_view=archived_view,
        **sidebar_context("gate_logs", selected_school_year),
    )


@app.route("/gate-logs/export")
@require_permission("logs")
def gate_logs_export():
    query, sort_spec, filters_payload, selected_school_year = build_gate_logs_query(request.args, include_export_scope=True)
    logs_collection, selected_school_year, _ = get_attendance_logs_storage(selected_school_year)
    export_scope = filters_payload.get("export_scope") or {}
    rows = list(logs_collection.find(query).sort(sort_spec).limit(5000))
    signatories = build_report_signatories(resolve_current_admin_export_name())
    generated_at_label = now_local().strftime("%B %d, %Y %I:%M:%S %p")
    metadata_items = [
        ("Generated At", generated_at_label),
        ("School Year", selected_school_year),
        ("Total Records", len(rows)),
        ("Search Query", filters_payload.get("q") or "All records"),
        ("Date Range", format_export_date_range_label(filters_payload.get("start_date"), filters_payload.get("end_date"))),
        ("Status", filters_payload.get("status") or "All statuses"),
        ("Gate Action", filters_payload.get("session") or "All actions"),
        ("Export Scope", export_scope.get("scope_label") or "All Matching Records"),
        ("Selection", export_scope.get("scope_detail") or "All students in the selected log filters"),
        ("Sort Order", format_export_sort_label(filters_payload.get("sort"))),
    ]
    buffer, doc, story, styles = build_school_export_document(
        document_title="Official Gate Logs Report",
        header_caption="Gate Logs Export",
        metadata_items=metadata_items,
        footer_title="Cawitan High School Gate Logs Export",
        export_subject="Gate logs export",
        export_author=signatories.get("prepared_by_name", "Admin"),
    )

    append_school_export_section_title(story, "Gate Log Summary", styles)
    summary_data = [[
        Paragraph("Metric", styles["table_header"]),
        Paragraph("Count", styles["table_header"]),
    ]]
    for label, count in build_gate_logs_summary_rows(rows):
        summary_data.append([
            Paragraph(xml_escape(str(label)), styles["table_cell"]),
            Paragraph(xml_escape(str(count)), styles["table_cell_center"]),
        ])
    if len(summary_data) == 1:
        summary_data.append([
            Paragraph("No gate log summary available.", styles["table_cell"]),
            "",
        ])
    story.append(build_school_export_table(
        summary_data,
        col_widths=[6.6 * inch, 2.4 * inch],
        styles=styles,
        span_empty_row=len(summary_data) == 2 and not rows,
    ))

    story.append(Spacer(1, 0.22 * inch))
    append_school_export_section_title(story, "Gate Log Records", styles)
    gate_table_data = [[
        Paragraph("Log ID", styles["table_header"]),
        Paragraph("Student ID", styles["table_header"]),
        Paragraph("Name", styles["table_header"]),
        Paragraph("Date", styles["table_header"]),
        Paragraph("Time", styles["table_header"]),
        Paragraph("Action", styles["table_header"]),
        Paragraph("Session", styles["table_header"]),
        Paragraph("Status", styles["table_header"]),
        Paragraph("Verification Label", styles["table_header"]),
        Paragraph("Source", styles["table_header"]),
        Paragraph("Timestamp", styles["table_header"]),
    ]]
    for row in rows:
        gate_table_data.append([
            Paragraph(xml_escape(str(row.get("_id") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("student_id") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("student_name") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("date") or "N/A")), styles["table_cell_center"]),
            Paragraph(xml_escape(format_time_for_display(row.get("time"), row.get("timestamp")) or "N/A"), styles["table_cell_center"]),
            Paragraph(xml_escape(str(row.get("gate_action") or "N/A")), styles["table_cell_center"]),
            Paragraph(xml_escape(str(row.get("session") or "N/A")), styles["table_cell_center"]),
            Paragraph(xml_escape(str(row.get("status") or "N/A")), styles["table_cell_center"]),
            Paragraph(xml_escape(str(row.get("verification_label") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("source") or "System")), styles["table_cell"]),
            Paragraph(format_students_export_timestamp(row.get("timestamp"), multiline=True), styles["table_cell_center"]),
        ])
    if len(gate_table_data) == 1:
        gate_table_data.append([
            Paragraph("No gate log records matched the selected filters.", styles["table_cell"]),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])
    story.append(build_school_export_table(
        gate_table_data,
        col_widths=[
            0.78 * inch,
            0.82 * inch,
            1.25 * inch,
            0.68 * inch,
            0.52 * inch,
            0.52 * inch,
            0.55 * inch,
            0.60 * inch,
            0.86 * inch,
            0.55 * inch,
            1.87 * inch,
        ],
        styles=styles,
        span_empty_row=len(gate_table_data) == 2 and not rows,
    ))
    story.append(build_school_export_footer_block(styles, signatories=signatories))

    doc.build(story, onFirstPage=build_school_export_footer, onLaterPages=build_school_export_footer)
    buffer.seek(0)

    filename = f"{export_scope.get('filename_prefix') or 'gate_logs'}_{now_local().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_generated_pdf(buffer, filename)


@app.route("/api/gate-logs/latest")
@require_permission("logs", api=True)
def gate_logs_latest():
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    logs_collection, school_year_label, archived_view = get_attendance_logs_storage(school_year_label)
    if archived_view:
        return jsonify({"status": "ok", "school_year": school_year_label, "logs": []})
    since_id = request.args.get("since_id", "").strip()
    query = {"school_year": school_year_label}
    if since_id:
        try:
            query["_id"] = {"$gt": ObjectId(since_id)}
        except Exception:
            pass

    rows = list(logs_collection.find(query).sort("_id", -1).limit(10))
    rows.reverse()
    photo_map = build_student_photo_map([row.get("student_id", "") for row in rows])

    payload = [
        serialize_gate_log_display_row(
            row,
            profile_photo=photo_map.get(str(row.get("student_id", "")).strip(), ""),
        )
        for row in rows
    ]

    return jsonify({"status": "ok", "school_year": school_year_label, "logs": payload})


@app.route("/api/gate-logs/corrections", methods=["GET", "POST"])
@require_permission("logs", api=True)
def gate_logs_corrections_api():
    school_year_label = resolve_selected_school_year(request.args.get("school_year", ""))
    corrections_collection, school_year_label, archived_view = get_attendance_corrections_storage(school_year_label)
    if request.method == "GET":
        status_filter = (request.args.get("status", "pending") or "").strip().lower()
        mine_only = (request.args.get("mine", "") or "").strip().lower() in {"1", "true", "yes"}
        try:
            limit_value = int(request.args.get("limit", "20"))
        except (TypeError, ValueError):
            limit_value = 20
        limit_value = max(1, min(limit_value, 100))

        query = {"school_year": school_year_label}
        if status_filter and status_filter != "all":
            query["status"] = status_filter

        if current_role() != "Full Admin":
            query["requested_by"] = session.get("admin", "")
        elif mine_only:
            query["requested_by"] = session.get("admin", "")

        docs = list(corrections_collection.find(query).sort("requestedAt", -1).limit(limit_value))
        rows = [serialize_attendance_correction(doc) for doc in docs]
        pending_count = corrections_collection.count_documents({"school_year": school_year_label, "status": "pending"})
        return jsonify({"status": "ok", "school_year": school_year_label, "rows": rows, "pending_count": pending_count})

    if archived_view:
        return jsonify({"status": "error", "message": "Archived school years are read-only."}), 403

    payload = request.get_json(silent=True) or {}
    log_id = str(payload.get("log_id") or "").strip()
    requested_status = str(payload.get("requested_status") or "").strip().title()
    reason = str(payload.get("reason") or "").strip()

    if not log_id:
        return jsonify({"status": "error", "message": "Gate log ID is required."}), 400
    if requested_status not in CORRECTION_ALLOWED_STATUSES:
        return jsonify({"status": "error", "message": "Invalid correction status requested."}), 400
    if len(reason) < 8:
        return jsonify({"status": "error", "message": "Please provide a clear reason (at least 8 characters)."}), 400

    try:
        log_oid = ObjectId(log_id)
    except Exception:
        return jsonify({"status": "error", "message": "Invalid gate log ID."}), 400

    row = attendance_logs.find_one({"_id": log_oid, "school_year": school_year_label})
    if not row:
        return jsonify({"status": "error", "message": "Gate log entry not found."}), 404

    if str(row.get("status") or "").title() == requested_status:
        return jsonify({"status": "error", "message": "Requested status matches current status."}), 400

    existing_pending = attendance_corrections.find_one({
        "attendance_log_id": str(log_oid),
        "school_year": school_year_label,
        "status": "pending",
    })
    if existing_pending:
        return jsonify({"status": "error", "message": "A pending correction already exists for this log."}), 409

    now_utc = _now_utc()
    correction_doc = {
        "attendance_log_id": str(log_oid),
        "school_year": normalize_school_year_value(row.get("school_year")) or derive_school_year_label_from_value(row.get("date")),
        "student_id": str(row.get("student_id") or ""),
        "student_name": str(row.get("student_name") or ""),
        "log_timestamp": str(row.get("timestamp") or ""),
        "current_status": str(row.get("status") or ""),
        "requested_status": requested_status,
        "reason": reason[:500],
        "status": "pending",
        "requested_by": session.get("admin", ""),
        "requested_by_role": current_role(),
        "requested_at": now_iso(),
        "requestedAt": now_utc,
        "reviewed_by": "",
        "reviewed_at": "",
        "review_note": "",
        "reviewedAt": None,
        "applied": False,
    }
    result = attendance_corrections.insert_one(correction_doc)
    signal_data_change("gate_logs")
    create_alert(
        "info",
        f"Correction requested for {correction_doc['student_name'] or correction_doc['student_id']}.",
        "attendance",
        {
            "attendance_log_id": str(log_oid),
            "requested_status": requested_status,
            "school_year": correction_doc.get("school_year", ""),
        },
    )
    log_audit_event(
        action="gate_log.correction_requested",
        outcome="success",
        severity="info",
        target_type="attendance_log",
        target_id=str(log_oid),
        details={"requested_status": requested_status},
    )
    return jsonify({
        "status": "ok",
        "message": "Correction request submitted.",
        "correction": serialize_attendance_correction({**correction_doc, "_id": result.inserted_id}),
    })


@app.route("/api/gate-logs/corrections/<correction_id>/review", methods=["POST"])
@require_permission("logs", api=True)
def gate_logs_correction_review_api(correction_id):
    if current_role() != "Full Admin":
        return jsonify({"status": "error", "message": "Only Full Admin can review corrections."}), 403

    try:
        correction_oid = ObjectId(correction_id)
    except Exception:
        return jsonify({"status": "error", "message": "Invalid correction ID."}), 400

    payload = request.get_json(silent=True) or {}
    decision = str(payload.get("decision") or "").strip().lower()
    review_note = str(payload.get("review_note") or "").strip()
    if decision not in {"approve", "reject"}:
        return jsonify({"status": "error", "message": "Decision must be approve or reject."}), 400

    correction, correction_collection, archived_correction = find_record_in_active_or_archive(
        attendance_corrections,
        attendance_corrections_archive,
        correction_oid,
    )
    if not correction:
        return jsonify({"status": "error", "message": "Correction request not found."}), 404
    if archived_correction or is_archived_school_year(correction.get("school_year")):
        return jsonify({"status": "error", "message": "Archived school years are read-only."}), 403
    if str(correction.get("status") or "").lower() != "pending":
        return jsonify({"status": "error", "message": "Correction request is already reviewed."}), 409

    now_utc = _now_utc()
    update_doc = {
        "status": "approved" if decision == "approve" else "rejected",
        "reviewed_by": session.get("admin", ""),
        "reviewed_at": now_iso(),
        "review_note": review_note[:500],
        "reviewedAt": now_utc,
    }

    applied = False
    attendance_log_id = str(correction.get("attendance_log_id") or "")
    if decision == "approve":
        try:
            log_oid = ObjectId(attendance_log_id)
        except Exception:
            return jsonify({"status": "error", "message": "Original gate log ID is invalid."}), 400

        update_result = attendance_logs.update_one(
            {"_id": log_oid},
            {
                "$set": {
                    "status": correction.get("requested_status"),
                    "corrected": True,
                    "corrected_at": now_iso(),
                    "corrected_by": session.get("admin", ""),
                    "correction_note": review_note[:500],
                }
            },
        )
        if update_result.matched_count == 0:
            return jsonify({"status": "error", "message": "Original gate log record was not found."}), 404
        applied = True
        signal_data_change("gate_logs")

    update_doc["applied"] = applied
    correction_collection.update_one({"_id": correction_oid}, {"$set": update_doc})
    create_alert(
        "info",
        f"Correction {decision}d for {correction.get('student_name') or correction.get('student_id')}.",
        "attendance",
        {
            "attendance_log_id": attendance_log_id,
            "decision": decision,
            "school_year": normalize_school_year_value(correction.get("school_year")),
        },
    )
    log_audit_event(
        action="gate_log.correction_reviewed",
        outcome="success",
        severity="info",
        target_type="attendance_correction",
        target_id=str(correction_oid),
        details={"decision": decision, "attendance_log_id": attendance_log_id},
    )
    return jsonify({"status": "ok", "message": f"Correction {decision}d successfully.", "applied": applied})


@app.route("/gate-logs/delete/<id>", methods=["POST", "DELETE"])
@require_permission("logs", api=True)
def gate_logs_delete(id):
    if current_role() != "Full Admin":
        log_audit_event(
            action="gate_log.delete",
            outcome="blocked",
            severity="warn",
            target_type="attendance_log",
            target_id=id,
            details={"reason": "insufficient_role"},
        )
        return jsonify({"status": "error", "message": "Only Full Admin can delete gate logs."}), 403

    try:
        log_oid = ObjectId(id)
        row, target_collection, archived_row = find_record_in_active_or_archive(
            attendance_logs,
            attendance_logs_archive,
            log_oid,
        )
        if not row:
            return jsonify({"status": "error", "message": "Gate log not found."}), 404
        if archived_row or is_archived_school_year(row.get("school_year")):
            return jsonify({"status": "error", "message": "Archived school years are read-only."}), 403
        result = target_collection.delete_one({"_id": log_oid})
        if result.deleted_count == 0:
            return jsonify({"status": "error", "message": "Gate log not found."}), 404
        signal_data_change("gate_logs")
        log_audit_event(
            action="gate_log.delete",
            outcome="success",
            severity="info",
            target_type="attendance_log",
            target_id=id,
        )
        return jsonify({"status": "ok", "message": "Gate log deleted."})
    except Exception as exc:
        log_audit_event(
            action="gate_log.delete",
            outcome="failed",
            severity="warn",
            target_type="attendance_log",
            target_id=id,
            details={"error": str(exc)},
        )
        return jsonify({"status": "error", "message": "Failed to delete gate log."}), 400


@app.route("/simulate-gate/<student_id>", methods=["POST"])
@require_permission("scan", api=True)
def simulate_gate(student_id):
    student = students.find_one({"student_id": student_id})
    if not student:
        return jsonify({"status": "FAILED", "error": "Student not found"}), 404

    result = log_attendance_and_sms(student, source="manual_simulation", send_notifications=False, mode="auto")
    if not result:
        return jsonify({"status": "FAILED", "error": "Unable to process gate simulation."}), 400

    if not result["duplicate"]:
        push_scan_event("verified", {
            "student_id": student.get("student_id", ""),
            "name": student.get("name", ""),
            "verified": True,
            "attendance_status": result["status"],
            "sms_status": result["sms_status"],
            "gate_action": result["gate_action"],
            "verification_label": result["verification_label"],
            "session": result["session"],
            "display_message": result["display_message"],
            "voice_message": result["voice_message"],
            "voice_key": f"{student.get('student_id', '')}:{result['gate_action']}:{result['timestamp']}",
            "duplicate": False,
            "duplicate_reason": "",
            "time": format_time_for_display(result.get("time"), result.get("timestamp")),
            "timestamp_display": format_timestamp_for_display(result.get("timestamp"), result.get("time")),
            "feed_update": bool(result.get("feed_update")),
            "activity_entry": result.get("activity_entry"),
            "tracking_mode": result.get("tracking_mode", get_scan_session_mode()),
        })
    return jsonify({
        "status": "SUCCESS",
        "name": student.get("name", ""),
        "action": result["gate_action"],
        "duplicate": result["duplicate"],
        "message": result["display_message"],
    })


@app.route("/sms-logs")
@require_permission("logs")
def sms_logs_page():
    query, sort_spec, filters_payload, selected_school_year = build_sms_logs_query(request.args)
    logs_collection, selected_school_year, archived_view = get_sms_logs_storage(selected_school_year)
    try:
        page = max(int(request.args.get("page", "1")), 1)
    except ValueError:
        page = 1

    per_page = 10
    total_filtered = logs_collection.count_documents(query)
    pagination = build_pagination_payload(page, per_page, total_filtered, filters_payload, "sms_logs_page")
    skip = (pagination["page"] - 1) * per_page

    rows = list(logs_collection.find(query).sort(sort_spec).skip(skip).limit(per_page))
    logs = [serialize_sms_log_display_row(row) for row in rows]

    stats = {
        "total_logs": total_filtered,
        "sent_count": logs_collection.count_documents({**query, "status": sms_status_mongo_filter("sent")}),
        "failed_count": logs_collection.count_documents({**query, "status": sms_status_mongo_filter("failed")}),
    }
    sms_template = get_attendance_sms_template_payload()
    sms_notification_settings = get_attendance_sms_notification_settings_payload()

    return render_template(
        "sms_logs.html",
        logs=logs,
        stats=stats,
        sms_template=sms_template,
        sms_notification_settings=sms_notification_settings,
        sms_template_variables=ATTENDANCE_SMS_TEMPLATE_VARIABLES,
        filters=filters_payload,
        pagination=pagination,
        export_query=urlencode({k: v for k, v in filters_payload.items() if v not in ("", None)}),
        grade_options=list(GRADE_LEVEL_OPTIONS),
        sections_by_grade=build_sections_by_grade(school_year=selected_school_year),
        archived_view=archived_view,
        **sidebar_context("sms_logs", selected_school_year),
    )


@app.route("/sms-logs/template", methods=["POST"])
@require_permission("logs", api=True)
def sms_logs_template_update():
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        raw_template = payload.get("template", "")
    else:
        raw_template = request.form.get("template", "")

    actor_username = session.get("admin", "system")
    actor_role = session.get("role", "System")

    try:
        saved_template = save_attendance_sms_template(
            raw_template,
            actor_username=actor_username,
            actor_role=actor_role,
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        log_audit_event(
            action="sms.template.update",
            outcome="failed",
            severity="warn",
            target_type="sms_template",
            target_id=ATTENDANCE_SMS_TEMPLATE_DOC_ID,
            details={"error": str(exc)},
        )
        return jsonify({"status": "error", "message": "Failed to save SMS template."}), 500

    log_audit_event(
        action="sms.template.update",
        outcome="success",
        severity="info",
        target_type="sms_template",
        target_id=ATTENDANCE_SMS_TEMPLATE_DOC_ID,
        details={
            "updated_by": actor_username,
            "template_preview": saved_template.get("template", "")[:120],
            "max_length": SMS_TEMPLATE_MAX_LENGTH,
        },
    )
    return jsonify(
        {
            "status": "ok",
            "message": "SMS template saved.",
            "template": saved_template.get("template", ""),
            "updated_at": saved_template.get("updated_at", ""),
            "updated_by": saved_template.get("updated_by", ""),
            "default_template": saved_template.get("default_template", ""),
            "max_length": saved_template.get("max_length", SMS_TEMPLATE_MAX_LENGTH),
            "variables": saved_template.get("variables", list(ATTENDANCE_SMS_TEMPLATE_VARIABLES)),
        }
    )


@app.route("/sms-logs/notifications", methods=["POST"])
@require_permission("logs", api=True)
def sms_logs_notification_settings_update():
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        raw_enabled = payload.get("enabled")
    else:
        raw_enabled = request.form.get("enabled")

    actor_username = session.get("admin", "system")
    actor_role = session.get("role", "System")

    try:
        saved_settings = save_attendance_sms_notification_settings(
            raw_enabled,
            actor_username=actor_username,
            actor_role=actor_role,
        )
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        log_audit_event(
            action="sms.notifications.update",
            outcome="failed",
            severity="warn",
            target_type="system_setting",
            target_id=ATTENDANCE_SMS_NOTIFICATION_SETTINGS_KEY,
            details={"error": str(exc)},
        )
        return jsonify({"status": "error", "message": "Failed to update SMS notification setting."}), 500

    enabled = bool(saved_settings.get("enabled"))
    status_message = (
        "Attendance SMS notifications enabled."
        if enabled
        else "Attendance SMS notifications disabled. Events will still be recorded in SMS Logs."
    )
    log_audit_event(
        action="sms.notifications.update",
        outcome="success",
        severity="info",
        target_type="system_setting",
        target_id=ATTENDANCE_SMS_NOTIFICATION_SETTINGS_KEY,
        details={
            "enabled": enabled,
            "updated_by": actor_username,
        },
    )
    return jsonify(
        {
            "status": "ok",
            "message": status_message,
            "enabled": enabled,
            "status_label": saved_settings.get("status_label", "Enabled" if enabled else "Disabled"),
            "updated_at": saved_settings.get("updated_at", ""),
            "updated_by": saved_settings.get("updated_by", ""),
        }
    )


@app.route("/sms-logs/export")
@require_permission("logs")
def sms_logs_export():
    query, sort_spec, filters_payload, selected_school_year = build_sms_logs_query(request.args, include_export_scope=True)
    logs_collection, selected_school_year, _ = get_sms_logs_storage(selected_school_year)
    export_scope = filters_payload.get("export_scope") or {}
    rows = list(logs_collection.find(query).sort(sort_spec).limit(5000))
    signatories = build_report_signatories(resolve_current_admin_export_name())
    generated_at_label = now_local().strftime("%B %d, %Y %I:%M:%S %p")
    metadata_items = [
        ("Generated At", generated_at_label),
        ("School Year", selected_school_year),
        ("Total Records", len(rows)),
        ("Search Query", filters_payload.get("q") or "All records"),
        ("Date Range", format_export_date_range_label(filters_payload.get("start_date"), filters_payload.get("end_date"))),
        ("Status", filters_payload.get("status") or "All statuses"),
        ("Export Scope", export_scope.get("scope_label") or "All Matching Records"),
        ("Selection", export_scope.get("scope_detail") or "All students in the selected log filters"),
        ("Sort Order", format_export_sort_label(filters_payload.get("sort"))),
    ]
    buffer, doc, story, styles = build_school_export_document(
        document_title="Official SMS Logs Report",
        header_caption="SMS Logs Export",
        metadata_items=metadata_items,
        footer_title="Cawitan High School SMS Logs Export",
        export_subject="SMS logs export",
        export_author=signatories.get("prepared_by_name", "Admin"),
    )

    append_school_export_section_title(story, "SMS Delivery Summary", styles)
    summary_data = [[
        Paragraph("Metric", styles["table_header"]),
        Paragraph("Count", styles["table_header"]),
    ]]
    for label, count in build_sms_logs_summary_rows(rows):
        summary_data.append([
            Paragraph(xml_escape(str(label)), styles["table_cell"]),
            Paragraph(xml_escape(str(count)), styles["table_cell_center"]),
        ])
    if len(summary_data) == 1:
        summary_data.append([
            Paragraph("No SMS summary available.", styles["table_cell"]),
            "",
        ])
    story.append(build_school_export_table(
        summary_data,
        col_widths=[6.6 * inch, 2.4 * inch],
        styles=styles,
        span_empty_row=len(summary_data) == 2 and not rows,
    ))

    story.append(Spacer(1, 0.22 * inch))
    append_school_export_section_title(story, "SMS Log Records", styles)
    sms_table_data = [[
        Paragraph("Student ID", styles["table_header"]),
        Paragraph("Name", styles["table_header"]),
        Paragraph("Parent Contact", styles["table_header"]),
        Paragraph("Date", styles["table_header"]),
        Paragraph("Time", styles["table_header"]),
        Paragraph("Status", styles["table_header"]),
        Paragraph("Message", styles["table_header"]),
        Paragraph("SID", styles["table_header"]),
        Paragraph("Error", styles["table_header"]),
        Paragraph("Timestamp", styles["table_header"]),
    ]]
    for row in rows:
        sms_table_data.append([
            Paragraph(xml_escape(str(row.get("student_id") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("name") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("parent_contact") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("date") or "N/A")), styles["table_cell_center"]),
            Paragraph(xml_escape(format_time_for_display(row.get("time"), row.get("timestamp")) or "N/A"), styles["table_cell_center"]),
            Paragraph(xml_escape(str(row.get("status") or "N/A").title()), styles["table_cell_center"]),
            Paragraph(xml_escape(str(row.get("message") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("sid") or "N/A")), styles["table_cell"]),
            Paragraph(xml_escape(str(row.get("error") or "N/A")), styles["table_cell"]),
            Paragraph(format_students_export_timestamp(row.get("timestamp"), multiline=True), styles["table_cell_center"]),
        ])
    if len(sms_table_data) == 1:
        sms_table_data.append([
            Paragraph("No SMS log records matched the selected filters.", styles["table_cell"]),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])
    story.append(build_school_export_table(
        sms_table_data,
        col_widths=[
            0.82 * inch,
            1.18 * inch,
            0.90 * inch,
            0.66 * inch,
            0.52 * inch,
            0.68 * inch,
            1.60 * inch,
            0.65 * inch,
            0.85 * inch,
            1.14 * inch,
        ],
        styles=styles,
        span_empty_row=len(sms_table_data) == 2 and not rows,
    ))
    story.append(build_school_export_footer_block(styles, signatories=signatories))

    doc.build(story, onFirstPage=build_school_export_footer, onLaterPages=build_school_export_footer)
    buffer.seek(0)

    filename = f"{export_scope.get('filename_prefix') or 'sms_logs'}_{now_local().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_generated_pdf(buffer, filename)


@app.route("/sms-logs/resend/<id>", methods=["POST"])
@require_permission("logs", api=True)
def sms_logs_resend(id):
    try:
        sms_log_id = ObjectId(id)
    except Exception:
        return jsonify({"status": "error", "message": "Invalid SMS log id."}), 400

    original, _, archived_log = find_record_in_active_or_archive(sms_logs, sms_logs_archive, sms_log_id)
    if not original:
        return jsonify({"status": "error", "message": "SMS log not found."}), 404
    if archived_log or is_archived_school_year(original.get("school_year")):
        return jsonify({"status": "error", "message": "Archived school years are read-only."}), 403

    parent_contact = original.get("parent_contact", "")
    message = original.get("message", "")
    if not parent_contact or not message:
        return jsonify({"status": "error", "message": "Missing recipient or message content."}), 400

    if not attendance_sms_notifications_enabled():
        log_skipped_sms(
            student_id=original.get("student_id", ""),
            student_name=original.get("name", ""),
            parent_contact=parent_contact,
            message=f"Attendance SMS notifications are disabled. Resend not sent for {original.get('name') or original.get('student_id') or 'Unknown student'}.",
            reason="notifications_disabled",
            sms_type=original.get("type", "transactional"),
            metadata={
                "context": "sms_resend",
                "resent_from": str(original.get("_id")),
                "school_year": normalize_school_year_value(original.get("school_year")),
                "notifications_enabled": False,
            },
        )
        log_audit_event(
            action="sms.resend",
            outcome="blocked",
            severity="info",
            target_type="sms_log",
            target_id=id,
            details={"reason": "notifications_disabled"},
        )
        return jsonify({
            "status": "ok",
            "message": "Attendance SMS notifications are disabled. Resend was recorded but not sent.",
            "sms_status": "DISABLED",
            "provider_message_id": "",
            "error": "",
            "timestamp": now_iso(),
        })

    sms_result = send_sms(
        parent_contact,
        message,
        sms_type=original.get("type", "transactional"),
        metadata={
            "context": "sms_resend",
            "resent_from": str(original.get("_id")),
            "school_year": normalize_school_year_value(original.get("school_year")),
        },
        student_id=original.get("student_id", ""),
        student_name=original.get("name", ""),
        parent_contact=parent_contact,
    )
    sms_status = "sent" if sms_result.get("status") == "sent" else "failed"
    sms_sid = sms_result.get("sid", "")
    sms_error = sms_result.get("error", "")
    now = now_local()

    if sms_status == "failed":
        create_alert(
            level="high",
            message=f"Failed SMS resend for {original.get('name', original.get('student_id', 'Unknown'))}.",
            category="sms",
            meta={
                "student_id": original.get("student_id", ""),
                "error": sms_error,
                "school_year": normalize_school_year_value(original.get("school_year")),
            },
        )
        log_audit_event(
            action="sms.resend",
            outcome="failed",
            severity="warn",
            target_type="sms_log",
            target_id=id,
            details={"error": sms_error},
        )
    else:
        log_audit_event(
            action="sms.resend",
            outcome="success",
            severity="info",
            target_type="sms_log",
            target_id=id,
            details={"provider_message_id": sms_sid},
        )

    return jsonify({
        "status": "ok",
        "message": "SMS resend queued." if sms_status == "sent" else "SMS resend failed.",
        "sms_status": sms_status.upper(),
        "provider_message_id": sms_sid,
        "error": sms_error,
        "timestamp": now_iso(),
    })


# =====================================
# ANALYTICS
# =====================================
def next_month_start(value):
    base = value.replace(day=28) + timedelta(days=4)
    return base.replace(day=1)


def build_daily_analytics_buckets(start_date, end_date):
    bucket_keys = []
    display_labels = []
    cursor = start_date
    while cursor <= end_date:
        bucket_keys.append(cursor.strftime("%Y-%m-%d"))
        display_labels.append(cursor.strftime("%b %d"))
        cursor += timedelta(days=1)
    return bucket_keys, display_labels


def build_monthly_analytics_buckets(start_date, end_date):
    bucket_keys = []
    display_labels = []
    cursor = start_date.replace(day=1)
    end_month = end_date.replace(day=1)
    while cursor <= end_month:
        bucket_keys.append(cursor.strftime("%Y-%m"))
        display_labels.append(cursor.strftime("%b %Y"))
        cursor = next_month_start(cursor)
    return bucket_keys, display_labels


def resolve_analytics_range(selected_school_year, range_type, requested_start="", requested_end=""):
    today = now_local().date()
    school_year_start, school_year_end = school_year_date_bounds(selected_school_year)
    range_anchor = today
    if school_year_start and school_year_end:
        if range_anchor < school_year_start:
            range_anchor = school_year_start
        elif range_anchor > school_year_end:
            range_anchor = school_year_end

    normalized_range = str(range_type or "month").strip().lower()
    if normalized_range not in {"week", "month", "year", "custom"}:
        normalized_range = "month"

    if normalized_range == "week":
        start_date = range_anchor - timedelta(days=range_anchor.weekday())
        end_date = range_anchor
        granularity = "daily"
        view_label = "Weekly"
        view_description = "Daily totals for the current week window."
    elif normalized_range == "year":
        start_date = school_year_start or range_anchor.replace(month=1, day=1)
        end_date = range_anchor
        granularity = "monthly"
        view_label = "Yearly"
        view_description = f"Monthly totals across school year {selected_school_year}."
    elif normalized_range == "custom":
        start_date = parse_date_or_none(requested_start) or (range_anchor - timedelta(days=29))
        end_date = parse_date_or_none(requested_end) or range_anchor
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        granularity = "daily"
        view_label = "Custom"
        view_description = "Daily totals for the custom date range."
    else:
        normalized_range = "month"
        start_date = range_anchor.replace(day=1)
        end_date = range_anchor
        granularity = "daily"
        view_label = "Monthly"
        view_description = "Daily totals for the current month window."

    if school_year_start and start_date < school_year_start:
        start_date = school_year_start
    if school_year_end and end_date > school_year_end:
        end_date = school_year_end
    if start_date > end_date:
        start_date = end_date

    return {
        "range_type": normalized_range,
        "start_date": start_date,
        "end_date": end_date,
        "granularity": granularity,
        "view_label": view_label,
        "view_description": view_description,
        "summary_label": f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
        "range_anchor": range_anchor,
    }


@app.route("/analytics")
@require_permission("analytics")
def analytics():
    selected_school_year = resolve_selected_school_year(request.args.get("school_year", ""))
    attendance_collection, selected_school_year, archived_view = get_attendance_logs_storage(selected_school_year)
    sms_collection, selected_school_year, _ = get_sms_logs_storage(selected_school_year)
    range_payload = resolve_analytics_range(
        selected_school_year,
        request.args.get("range", "month"),
        requested_start=request.args.get("start_date"),
        requested_end=request.args.get("end_date"),
    )
    range_type = range_payload["range_type"]
    start_date = range_payload["start_date"]
    end_date = range_payload["end_date"]
    granularity = range_payload["granularity"]
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")

    if granularity == "monthly":
        bucket_keys, chart_labels = build_monthly_analytics_buckets(start_date, end_date)
        attendance_pipeline = [
            {"$match": {"school_year": selected_school_year, "date": {"$gte": start_str, "$lte": end_str}}},
            {"$group": {"_id": {"$substrBytes": ["$date", 0, 7]}, "count": {"$sum": 1}}},
        ]
        sms_pipeline = [
            {
                "$match": {
                    "school_year": selected_school_year,
                    "date": {"$gte": start_str, "$lte": end_str},
                    "status": sms_status_mongo_filter("sent"),
                }
            },
            {"$group": {"_id": {"$substrBytes": ["$date", 0, 7]}, "count": {"$sum": 1}}},
        ]
    else:
        bucket_keys, chart_labels = build_daily_analytics_buckets(start_date, end_date)
        attendance_pipeline = [
            {"$match": {"school_year": selected_school_year, "date": {"$gte": start_str, "$lte": end_str}}},
            {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        ]
        sms_pipeline = [
            {
                "$match": {
                    "school_year": selected_school_year,
                    "date": {"$gte": start_str, "$lte": end_str},
                    "status": sms_status_mongo_filter("sent"),
                }
            },
            {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        ]

    attendance_map = {row["_id"]: row["count"] for row in attendance_collection.aggregate(attendance_pipeline)}
    sms_map = {row["_id"]: row["count"] for row in sms_collection.aggregate(sms_pipeline)}
    gate_series = [attendance_map.get(bucket_key, 0) for bucket_key in bucket_keys]
    sms_series = [sms_map.get(bucket_key, 0) for bucket_key in bucket_keys]

    total_gate_entries = sum(gate_series)
    total_sms_sent = sum(sms_series)

    enrollment_collection = get_school_year_enrollment_collection(selected_school_year)
    total_students = enrollment_collection.count_documents({"school_year": selected_school_year})
    late_ids = set(attendance_collection.distinct("student_id", {"school_year": selected_school_year, "date": end_str, "status": "Late"}))
    present_ids_all = set(attendance_collection.distinct("student_id", {"school_year": selected_school_year, "date": end_str}))
    present_ids = set([sid for sid in present_ids_all if sid]) - set([sid for sid in late_ids if sid])
    late_ids = set([sid for sid in late_ids if sid])
    absent_count = max(total_students - len(present_ids) - len(late_ids), 0)

    attendance_distribution = {
        "present": len(present_ids),
        "late": len(late_ids),
        "absent": absent_count,
    }

    chart_meta = {
        "view_label": range_payload["view_label"],
        "view_description": range_payload["view_description"],
        "summary_label": range_payload["summary_label"],
        "granularity_label": "Monthly" if granularity == "monthly" else "Daily",
        "gate_title": "Monthly Gate Entries" if granularity == "monthly" else "Daily Gate Entries",
        "gate_subtitle": "Attendance totals grouped by month for the selected school year."
        if granularity == "monthly"
        else "Attendance flow from gate scans within the selected range.",
        "sms_title": "SMS Logs Sent Per Month" if granularity == "monthly" else "SMS Logs Sent Per Day",
        "sms_subtitle": "Delivered guardian notification totals grouped by month."
        if granularity == "monthly"
        else "Delivered guardian notification totals within the selected range.",
        "attendance_subtitle": f"Present, absent, and late breakdown based on {end_date.strftime('%B %d, %Y')}.",
        "attendance_reference_label": end_date.strftime("%B %d, %Y"),
    }

    return render_template(
        "analytics.html",
        filters={
            "school_year": selected_school_year,
            "range": range_type,
            "start_date": start_date.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
        },
        stats={
            "total_gate_entries": total_gate_entries,
            "total_sms_sent": total_sms_sent,
            "present_on_reference": len(present_ids),
            "late_on_reference": len(late_ids),
        },
        chart_labels=chart_labels,
        gate_series=gate_series,
        sms_series=sms_series,
        attendance_distribution=attendance_distribution,
        chart_meta=chart_meta,
        archived_view=archived_view,
        grade_options=list(GRADE_LEVEL_OPTIONS),
        **sidebar_context("analytics", selected_school_year),
    )


@app.route("/api/analytics/scheduled-reports", methods=["GET", "POST"])
@require_permission("analytics", api=True)
def analytics_scheduled_reports_api():
    if request.method == "GET":
        try:
            limit = max(1, min(int(request.args.get("limit", "12")), SCHEDULED_REPORT_MAX_RESULTS))
        except Exception:
            limit = 12
        rows = list(scheduled_reports.find().sort("updated_at", -1).limit(limit))
        run_rows = list(scheduled_report_runs.find().sort("started_at", -1).limit(15))
        return api_success({
            "reports": [serialize_scheduled_report(row) for row in rows],
            "runs": [serialize_scheduled_report_run(row) for row in run_rows],
        })

    if current_role() != "Full Admin":
        return api_error("Only Full Admin can manage scheduled reports.", 403)

    payload = parse_json_payload()
    name = normalize_text_value(payload.get("name"))[:80]
    frequency = str(payload.get("frequency") or "weekly").strip().lower()
    send_time = parse_hhmm(payload.get("send_time") or SCHEDULED_REPORT_DEFAULT_SEND_TIME)
    recipients = parse_email_list(payload.get("recipients") or [])
    enabled = parse_bool_value(payload.get("enabled"), default=True)
    filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else {}
    filters_grade = normalize_grade_level(filters.get("grade") or "")
    filters_section = normalize_section_value(filters.get("section") or "")

    if not name:
        return api_error("Report name is required.", 400, "name")
    if frequency not in SCHEDULED_REPORT_ALLOWED_FREQUENCIES:
        return api_error("Invalid frequency. Allowed values: daily, weekly, monthly.", 400, "frequency")
    if not recipients:
        return api_error("At least one valid recipient email is required.", 400, "recipients")

    now_dt = now_local()
    doc = {
        "name": name,
        "frequency": frequency,
        "send_time": send_time,
        "recipients": recipients,
        "enabled": enabled,
        "filters": {"grade": filters_grade or "", "section": filters_section or ""},
        "next_run_at": compute_next_report_run_at(frequency, send_time, now_dt=now_dt),
        "last_run_at": None,
        "last_status": "",
        "last_error": "",
        "created_by": session.get("admin", "system"),
        "updated_by": session.get("admin", "system"),
        "created_at": now_dt,
        "updated_at": now_dt,
    }

    try:
        inserted = scheduled_reports.insert_one(doc)
    except DuplicateKeyError:
        return api_error("A scheduled report with this name already exists.", 409, "name")
    except Exception as exc:
        return api_error(f"Failed to create scheduled report: {exc}", 500)

    created = scheduled_reports.find_one({"_id": inserted.inserted_id})
    log_audit_event(
        action="analytics.scheduled_report_create",
        outcome="success",
        severity="info",
        target_type="scheduled_report",
        target_id=str(inserted.inserted_id),
        details={"name": name, "frequency": frequency, "enabled": enabled},
    )
    return api_success({"report": serialize_scheduled_report(created)}, status_code=201)


@app.route("/api/analytics/scheduled-reports/<report_id>", methods=["PUT", "DELETE"])
@require_permission("analytics", api=True)
def analytics_scheduled_report_detail_api(report_id):
    if current_role() != "Full Admin":
        return api_error("Only Full Admin can manage scheduled reports.", 403)
    try:
        report_oid = ObjectId(report_id)
    except Exception:
        return api_error("Invalid scheduled report ID.", 400)

    existing = scheduled_reports.find_one({"_id": report_oid})
    if not existing:
        return api_error("Scheduled report not found.", 404)

    if request.method == "DELETE":
        scheduled_reports.delete_one({"_id": report_oid})
        scheduled_report_runs.delete_many({"report_id": report_id})
        log_audit_event(
            action="analytics.scheduled_report_delete",
            outcome="success",
            severity="warn",
            target_type="scheduled_report",
            target_id=report_id,
            details={"name": existing.get("name", "")},
        )
        return api_success({"message": "Scheduled report deleted."})

    payload = parse_json_payload()
    update_doc = {}

    if "name" in payload:
        name = normalize_text_value(payload.get("name"))[:80]
        if not name:
            return api_error("Report name is required.", 400, "name")
        update_doc["name"] = name

    if "frequency" in payload:
        frequency = str(payload.get("frequency") or "").strip().lower()
        if frequency not in SCHEDULED_REPORT_ALLOWED_FREQUENCIES:
            return api_error("Invalid frequency. Allowed values: daily, weekly, monthly.", 400, "frequency")
        update_doc["frequency"] = frequency

    if "send_time" in payload:
        update_doc["send_time"] = parse_hhmm(payload.get("send_time") or SCHEDULED_REPORT_DEFAULT_SEND_TIME)

    if "recipients" in payload:
        recipients = parse_email_list(payload.get("recipients") or [])
        if not recipients:
            return api_error("At least one valid recipient email is required.", 400, "recipients")
        update_doc["recipients"] = recipients

    if "enabled" in payload:
        update_doc["enabled"] = parse_bool_value(payload.get("enabled"), default=True)

    if "filters" in payload and isinstance(payload.get("filters"), dict):
        filters = payload.get("filters") or {}
        update_doc["filters"] = {
            "grade": normalize_grade_level(filters.get("grade") or "") or "",
            "section": normalize_section_value(filters.get("section") or "") or "",
        }

    merged_frequency = update_doc.get("frequency", existing.get("frequency") or "weekly")
    merged_send_time = update_doc.get("send_time", existing.get("send_time") or SCHEDULED_REPORT_DEFAULT_SEND_TIME)
    update_doc["next_run_at"] = compute_next_report_run_at(merged_frequency, merged_send_time, now_dt=now_local())
    update_doc["updated_at"] = now_local()
    update_doc["updated_by"] = session.get("admin", "system")

    try:
        scheduled_reports.update_one({"_id": report_oid}, {"$set": update_doc})
    except DuplicateKeyError:
        return api_error("A scheduled report with this name already exists.", 409, "name")
    except Exception as exc:
        return api_error(f"Failed to update scheduled report: {exc}", 500)

    updated = scheduled_reports.find_one({"_id": report_oid})
    log_audit_event(
        action="analytics.scheduled_report_update",
        outcome="success",
        severity="info",
        target_type="scheduled_report",
        target_id=report_id,
        details={"fields": sorted(list(update_doc.keys()))},
    )
    return api_success({"report": serialize_scheduled_report(updated)})


@app.route("/api/analytics/scheduled-reports/<report_id>/run-now", methods=["POST"])
@require_permission("analytics", api=True)
def analytics_scheduled_report_run_now_api(report_id):
    if current_role() != "Full Admin":
        return api_error("Only Full Admin can run scheduled reports.", 403)
    try:
        report_oid = ObjectId(report_id)
    except Exception:
        return api_error("Invalid scheduled report ID.", 400)

    report_doc = scheduled_reports.find_one({"_id": report_oid})
    if not report_doc:
        return api_error("Scheduled report not found.", 404)
    result = run_single_scheduled_report(report_doc, trigger="manual")
    if result.get("status") != "ok":
        return api_error(result.get("message") or "Failed to run report.", 500)
    return api_success(result)


@app.route("/api/analytics/anomaly-rules", methods=["GET", "POST"])
@require_permission("analytics", api=True)
def analytics_anomaly_rules_api():
    if request.method == "GET":
        try:
            limit = max(1, min(int(request.args.get("limit", "20")), SCHEDULED_REPORT_MAX_RESULTS))
        except Exception:
            limit = 20
        rows = list(anomaly_rules.find().sort("updated_at", -1).limit(limit))
        events = list(anomaly_events.find().sort("triggered_at", -1).limit(20))
        return api_success({
            "rules": [serialize_anomaly_rule(row) for row in rows],
            "events": [serialize_anomaly_event(row) for row in events],
        })

    if current_role() != "Full Admin":
        return api_error("Only Full Admin can manage anomaly rules.", 403)

    payload = parse_json_payload()
    name = normalize_text_value(payload.get("name"))[:80]
    metric = str(payload.get("metric") or "").strip().lower()
    operator_value = str(payload.get("operator") or "").strip().lower()
    try:
        threshold = float(payload.get("threshold", 0))
    except Exception:
        return api_error("Threshold must be a numeric value.", 400, "threshold")
    try:
        window_days = int(payload.get("window_days", 1))
    except Exception:
        window_days = 1
    window_days = max(1, min(window_days, 90))
    severity = str(payload.get("severity") or "warn").strip().lower()
    cooldown_minutes = max(5, min(int(payload.get("cooldown_minutes") or ANOMALY_DEFAULT_COOLDOWN_MINUTES), 1440))
    enabled = parse_bool_value(payload.get("enabled"), default=True)
    notify_emails = parse_email_list(payload.get("notify_emails") or [])
    filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else {}
    filters_grade = normalize_grade_level(filters.get("grade") or "")
    filters_section = normalize_section_value(filters.get("section") or "")

    if not name:
        return api_error("Rule name is required.", 400, "name")
    if metric not in ANOMALY_ALLOWED_METRICS:
        return api_error("Invalid metric.", 400, "metric")
    if operator_value not in ANOMALY_ALLOWED_OPERATORS:
        return api_error("Invalid operator.", 400, "operator")
    if severity not in ANOMALY_ALLOWED_SEVERITIES:
        return api_error("Invalid severity. Allowed values: info, warn, high.", 400, "severity")

    now_dt = now_local()
    doc = {
        "name": name,
        "metric": metric,
        "operator": operator_value,
        "threshold": threshold,
        "window_days": window_days,
        "severity": severity,
        "cooldown_minutes": cooldown_minutes,
        "enabled": enabled,
        "notify_emails": notify_emails,
        "filters": {"grade": filters_grade or "", "section": filters_section or ""},
        "last_evaluated_at": None,
        "last_triggered_at": None,
        "last_value": 0,
        "last_result": "",
        "created_by": session.get("admin", "system"),
        "updated_by": session.get("admin", "system"),
        "created_at": now_dt,
        "updated_at": now_dt,
    }
    try:
        inserted = anomaly_rules.insert_one(doc)
    except DuplicateKeyError:
        return api_error("An anomaly rule with this name already exists.", 409, "name")
    except Exception as exc:
        return api_error(f"Failed to create anomaly rule: {exc}", 500)

    created = anomaly_rules.find_one({"_id": inserted.inserted_id})
    log_audit_event(
        action="analytics.anomaly_rule_create",
        outcome="success",
        severity="info",
        target_type="anomaly_rule",
        target_id=str(inserted.inserted_id),
        details={"metric": metric, "operator": operator_value, "threshold": threshold},
    )
    return api_success({"rule": serialize_anomaly_rule(created)}, status_code=201)


@app.route("/api/analytics/anomaly-rules/<rule_id>", methods=["PUT", "DELETE"])
@require_permission("analytics", api=True)
def analytics_anomaly_rule_detail_api(rule_id):
    if current_role() != "Full Admin":
        return api_error("Only Full Admin can manage anomaly rules.", 403)
    try:
        rule_oid = ObjectId(rule_id)
    except Exception:
        return api_error("Invalid anomaly rule ID.", 400)

    existing = anomaly_rules.find_one({"_id": rule_oid})
    if not existing:
        return api_error("Anomaly rule not found.", 404)

    if request.method == "DELETE":
        anomaly_rules.delete_one({"_id": rule_oid})
        anomaly_events.delete_many({"rule_id": rule_id})
        log_audit_event(
            action="analytics.anomaly_rule_delete",
            outcome="success",
            severity="warn",
            target_type="anomaly_rule",
            target_id=rule_id,
            details={"name": existing.get("name", "")},
        )
        return api_success({"message": "Anomaly rule deleted."})

    payload = parse_json_payload()
    update_doc = {}

    if "name" in payload:
        name = normalize_text_value(payload.get("name"))[:80]
        if not name:
            return api_error("Rule name is required.", 400, "name")
        update_doc["name"] = name
    if "metric" in payload:
        metric = str(payload.get("metric") or "").strip().lower()
        if metric not in ANOMALY_ALLOWED_METRICS:
            return api_error("Invalid metric.", 400, "metric")
        update_doc["metric"] = metric
    if "operator" in payload:
        operator_value = str(payload.get("operator") or "").strip().lower()
        if operator_value not in ANOMALY_ALLOWED_OPERATORS:
            return api_error("Invalid operator.", 400, "operator")
        update_doc["operator"] = operator_value
    if "threshold" in payload:
        try:
            update_doc["threshold"] = float(payload.get("threshold"))
        except Exception:
            return api_error("Threshold must be a numeric value.", 400, "threshold")
    if "window_days" in payload:
        try:
            update_doc["window_days"] = max(1, min(int(payload.get("window_days")), 90))
        except Exception:
            return api_error("Window days must be an integer value.", 400, "window_days")
    if "severity" in payload:
        severity = str(payload.get("severity") or "").strip().lower()
        if severity not in ANOMALY_ALLOWED_SEVERITIES:
            return api_error("Invalid severity. Allowed values: info, warn, high.", 400, "severity")
        update_doc["severity"] = severity
    if "cooldown_minutes" in payload:
        try:
            update_doc["cooldown_minutes"] = max(5, min(int(payload.get("cooldown_minutes")), 1440))
        except Exception:
            return api_error("Cooldown minutes must be an integer value.", 400, "cooldown_minutes")
    if "enabled" in payload:
        update_doc["enabled"] = parse_bool_value(payload.get("enabled"), default=True)
    if "notify_emails" in payload:
        update_doc["notify_emails"] = parse_email_list(payload.get("notify_emails") or [])
    if "filters" in payload and isinstance(payload.get("filters"), dict):
        filters = payload.get("filters") or {}
        update_doc["filters"] = {
            "grade": normalize_grade_level(filters.get("grade") or "") or "",
            "section": normalize_section_value(filters.get("section") or "") or "",
        }

    update_doc["updated_at"] = now_local()
    update_doc["updated_by"] = session.get("admin", "system")
    try:
        anomaly_rules.update_one({"_id": rule_oid}, {"$set": update_doc})
    except DuplicateKeyError:
        return api_error("An anomaly rule with this name already exists.", 409, "name")
    except Exception as exc:
        return api_error(f"Failed to update anomaly rule: {exc}", 500)

    updated = anomaly_rules.find_one({"_id": rule_oid})
    log_audit_event(
        action="analytics.anomaly_rule_update",
        outcome="success",
        severity="info",
        target_type="anomaly_rule",
        target_id=rule_id,
        details={"fields": sorted(list(update_doc.keys()))},
    )
    return api_success({"rule": serialize_anomaly_rule(updated)})


@app.route("/api/analytics/anomaly-rules/evaluate", methods=["POST"])
@require_permission("analytics", api=True)
def analytics_anomaly_rules_evaluate_api():
    if current_role() != "Full Admin":
        return api_error("Only Full Admin can evaluate anomaly rules.", 403)
    result = evaluate_all_anomaly_rules(trigger="manual", max_rules=100)
    return api_success(result)


@app.route("/api/analytics/anomaly-rules/<rule_id>/evaluate", methods=["POST"])
@require_permission("analytics", api=True)
def analytics_anomaly_rule_evaluate_api(rule_id):
    if current_role() != "Full Admin":
        return api_error("Only Full Admin can evaluate anomaly rules.", 403)
    try:
        rule_oid = ObjectId(rule_id)
    except Exception:
        return api_error("Invalid anomaly rule ID.", 400)
    rule_doc = anomaly_rules.find_one({"_id": rule_oid})
    if not rule_doc:
        return api_error("Anomaly rule not found.", 404)
    result = evaluate_anomaly_rule(rule_doc, trigger="manual")
    if result.get("status") != "ok":
        return api_error(result.get("message") or "Failed to evaluate rule.", 500)
    return api_success(result)


# =====================================
# TEST SMS
# =====================================
@app.route("/test_sms", methods=["POST"])
@require_permission("users_manage", api=True)
def test_sms():
    verified_recipient = os.getenv("TEST_SMS_RECIPIENT") or os.getenv("VERIFIED_RECIPIENT")
    if not verified_recipient:
        return jsonify({"status": "failed", "error": "No test recipient number set in TEST_SMS_RECIPIENT."}), 400

    message = "This is a PHILSMS test message from CHS Gate Access."
    result = send_sms(verified_recipient, message, sms_type="transactional", metadata={"context": "test_sms"})
    if result.get("status") == "sent":
        return jsonify({"status": "sent", "provider": "PHILSMS", "sid": result.get("sid")})
    return jsonify({"status": "failed", "error": result.get("error", "Unknown error")})


@app.route("/api/debug/sms/test", methods=["POST"])
@require_permission("users_manage", api=True)
def debug_sms_test():
    if current_role() != "Full Admin":
        return jsonify({"status": "error", "message": "Only Full Admin can run SMS debug test."}), 403

    payload = parse_json_payload()
    to_value = (payload.get("to") or os.getenv("TEST_SMS_RECIPIENT") or "").strip()
    message = (payload.get("message") or "PHILSMS debug test from CHS Gate Access.").strip()

    if not to_value:
        return jsonify({"status": "error", "message": "Recipient is required (payload.to or TEST_SMS_RECIPIENT)."}), 400

    health = sms_provider.health_check()
    if health.get("status") != "ok":
        return jsonify({"status": "error", "message": "SMS provider health check failed.", "health": health}), 503

    result = send_sms(
        to_value,
        message,
        sms_type="transactional",
        metadata={"context": "debug_sms_test"},
        parent_contact=to_value,
    )
    if not result.get("log_id"):
        return jsonify({
            "status": "error",
            "message": "SMS attempted but no log entry id was returned.",
            "health": health,
            "result": result,
        }), 500

    log_doc = sms_logs.find_one({"_id": ObjectId(result["log_id"])})
    if log_doc:
        log_doc["_id"] = str(log_doc["_id"])

    http_code = 200 if result.get("status") == "sent" else 502
    return jsonify({
        "status": "ok" if result.get("status") == "sent" else "error",
        "health": health,
        "result": result,
        "log": log_doc,
    }), http_code


# =====================================
# CALENDAR / SETTINGS API
# =====================================

# =====================================
# EARLY TIME-OUT (ETO) FEATURE
# =====================================

@app.route("/early-timeout")
@require_permission("logs")
def early_timeout_page():
    """Admin page for managing early time-out requests."""
    sy = resolve_selected_school_year(request.args.get("school_year", ""))
    pending_count = 0
    try:
        eto_collection, _, _ = get_early_timeout_requests_storage(sy)
        pending_count = int(eto_collection.count_documents({"status": "pending"}))
    except Exception:
        pass
    return render_template(
        "early_timeout.html",
        pending_count=pending_count,
        **sidebar_context("early_timeout", sy)
    )


@app.route("/api/early-timeout/requests/count", methods=["GET"])
def api_eto_count():
    """Return pending ETO request count for badge display."""
    if not login_required():
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    try:
        school_year = resolve_selected_school_year(request.args.get("school_year", ""))
        eto_collection, _, _ = get_early_timeout_requests_storage(school_year)
        count = int(eto_collection.count_documents({"status": "pending"}))
        return jsonify({"status": "ok", "pending_count": count})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/early-timeout/request", methods=["POST"])
@require_permission("logs")
def api_eto_submit():
    """Submit a new early time-out request for a student (by staff/admin)."""
    data = request.get_json(silent=True) or {}
    student_id = str(data.get("student_id") or "").strip()
    reason = str(data.get("reason") or "").strip()
    urgency = str(data.get("urgency") or "normal").strip().lower()
    notes = str(data.get("notes") or "").strip()
    requested_by = str(data.get("requested_by") or "").strip()
    if not student_id:
        return jsonify({"status": "error", "message": "Student ID is required."}), 400
    if len(reason) < 5:
        return jsonify({"status": "error", "message": "Please provide a clear reason (at least 5 characters)."}), 400

    # Look up student across current school year
    student = None
    sy = resolve_selected_school_year("")
    enroll_col = get_student_enrollment_collection(sy)
    student = enroll_col.find_one({"student_id": student_id})
    if not student:
        student = students.find_one({"student_id": student_id})
    student_name = (
        str(student.get("name") or student.get("fullName") or "").strip()
        if student else student_id
    )
    section = str(student.get("section") or "").strip() if student else ""
    grade_level = str(student.get("grade_level") or student.get("grade") or "").strip() if student else ""

    now = now_local()
    now_str = now.isoformat()
    date_str = now.strftime("%Y-%m-%d")
    submitted_by = requested_by[:120] or str(session.get("admin", "staff") or "staff").strip()

    doc = {
        "student_id": student_id,
        "student_name": student_name,
        "section": section,
        "grade_level": grade_level,
        "reason": reason,
        "urgency": urgency if urgency in ("normal", "urgent") else "normal",
        "notes": notes,
        "status": "pending",
        "requested_by": submitted_by,
        "requested_at": now_str,
        "date": date_str,
        "school_year": sy,
        "reviewed_by": None,
        "reviewed_at": None,
        "review_note": None,
        "attendance_log_id": None,
    }
    
    # Get the appropriate storage for the school year
    eto_collection, _, _ = get_early_timeout_requests_storage(sy)
    result = eto_collection.insert_one(doc)
    return jsonify({"status": "ok", "message": f"Early time-out request submitted for {student_name}.", "request_id": str(result.inserted_id)})


@app.route("/api/early-timeout/requests", methods=["GET"])
@require_permission("logs")
def api_eto_list():
    """List early time-out requests with optional filters."""
    status_filter = request.args.get("status", "").strip()
    date_filter = request.args.get("date", "").strip()
    student_q = request.args.get("q", "").strip()
    limit = min(int(request.args.get("limit", 50)), 200)
    school_year = resolve_selected_school_year(request.args.get("school_year", ""))

    query = {}
    if status_filter in ("pending", "approved", "denied"):
        query["status"] = status_filter
    if date_filter:
        query["date"] = date_filter
    if student_q:
        import re as _re
        pat = {"$regex": _re.escape(student_q), "$options": "i"}
        query["$or"] = [{"student_name": pat}, {"student_id": pat}]

    # Get the appropriate storage collection based on school year
    eto_collection, _, _ = get_early_timeout_requests_storage(school_year)

    rows = []
    try:
        for doc in eto_collection.find(query).sort("requested_at", -1).limit(limit):
            doc["_id"] = str(doc["_id"])
            rows.append(doc)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

    pending_count = int(eto_collection.count_documents({"status": "pending"}))
    return jsonify({"status": "ok", "rows": rows, "pending_count": pending_count})


@app.route("/api/early-timeout/requests/<request_id>/approve", methods=["POST"])
@require_permission("logs")
def api_eto_approve(request_id):
    """Approve an ETO request and create an attendance log exit entry."""
    from bson import ObjectId
    try:
        oid = ObjectId(request_id)
    except Exception:
        return jsonify({"status": "error", "message": "Invalid request ID."}), 400

    # Get school year from request or use current
    school_year = resolve_selected_school_year(request.args.get("school_year", "") or request.form.get("school_year", ""))
    eto_collection, _, _ = get_early_timeout_requests_storage(school_year)
    
    eto = eto_collection.find_one({"_id": oid})
    if not eto:
        return jsonify({"status": "error", "message": "Request not found."}), 404
    if eto.get("status") != "pending":
        return jsonify({"status": "error", "message": f"Request is already {eto.get('status')}."}), 409

    data = request.get_json(silent=True) or {}
    review_note = str(data.get("review_note") or "").strip()
    now = now_local()
    now_str = now.isoformat()
    date_str = eto.get("date") or now.strftime("%Y-%m-%d")
    time_str = now.strftime("%I:%M %p")
    reviewer = session.get("admin", "admin")
    
    # Use the school year from the ETO request or derive it from date
    eto_school_year = eto.get("school_year") or resolve_selected_school_year("")
    if not eto_school_year:
        # Derive school year from the ETO date
        from datetime import datetime
        eto_date = datetime.strptime(date_str, "%Y-%m-%d")
        eto_school_year = get_school_year_for_date(eto_date)
    
    # Write attendance log entry for the early exit
    log_entry = {
        "student_id": eto["student_id"],
        "name": eto.get("student_name", eto["student_id"]),
        "student_name": eto.get("student_name", eto["student_id"]),
        "section": eto.get("section", ""),
        "grade_level": eto.get("grade_level", ""),
        "grade": eto.get("grade_level", ""),
        "action": "OUT",
        "status": "Present",
        "verification_label": "Early Timeout",
        "early_timeout": True,
        "early_timeout_reason": eto.get("reason", ""),
        "early_timeout_request_id": str(oid),
        "session": "Afternoon",
        "date": date_str,
        "time": time_str,
        "timestamp": now_str,
        "school_year": eto_school_year,
        "gate_action": "OUT",
    }
    
    # Get the appropriate attendance storage for the school year
    attendance_collection, _, _ = get_attendance_logs_storage(eto_school_year)
    log_result = attendance_collection.insert_one(log_entry)
    log_id = str(log_result.inserted_id)

    # Update ETO request status
    eto_collection.update_one(
        {"_id": oid},
        {"$set": {
            "status": "approved",
            "reviewed_by": reviewer,
            "reviewed_at": now_str,
            "review_note": review_note,
            "attendance_log_id": log_id,
            "school_year": eto_school_year,  # Ensure school_year is set
        }}
    )
    return jsonify({
        "status": "ok",
        "message": f"Early time-out approved for {eto.get('student_name', eto['student_id'])}. Exit recorded.",
        "attendance_log_id": log_id,
    })


@app.route("/api/early-timeout/requests/<request_id>/deny", methods=["POST"])
@require_permission("logs")
def api_eto_deny(request_id):
    """Deny an ETO request."""
    from bson import ObjectId
    try:
        oid = ObjectId(request_id)
    except Exception:
        return jsonify({"status": "error", "message": "Invalid request ID."}), 400

    # Get school year from request or use current
    school_year = resolve_selected_school_year(request.args.get("school_year", "") or request.form.get("school_year", ""))
    eto_collection, _, _ = get_early_timeout_requests_storage(school_year)
    
    eto = eto_collection.find_one({"_id": oid})
    if not eto:
        return jsonify({"status": "error", "message": "Request not found."}), 404
    if eto.get("status") != "pending":
        return jsonify({"status": "error", "message": f"Request is already {eto.get('status')}."}), 409

    data = request.get_json(silent=True) or {}
    review_note = str(data.get("review_note") or "").strip()
    now = now_local()
    reviewer = session.get("admin", "admin")
    
    # Ensure school year is set
    eto_school_year = eto.get("school_year") or resolve_selected_school_year(school_year)

    eto_collection.update_one(
        {"_id": oid},
        {"$set": {
            "status": "denied",
            "reviewed_by": reviewer,
            "reviewed_at": now.isoformat(),
            "review_note": review_note,
            "school_year": eto_school_year,  # Ensure school_year is set
        }}
    )
    return jsonify({
        "status": "ok",
        "message": f"Early time-out request denied for {eto.get('student_name', eto['student_id'])}.",
    })


@app.route("/api/early-timeout/requests/<request_id>", methods=["DELETE"])
@require_permission("full_admin")
def api_eto_delete(request_id):
    """Cancel/delete an ETO request (Full Admin only)."""
    from bson import ObjectId
    try:
        oid = ObjectId(request_id)
    except Exception:
        return jsonify({"status": "error", "message": "Invalid request ID."}), 400
    school_year = resolve_selected_school_year(request.args.get("school_year", ""))
    eto_collection, _, _ = get_early_timeout_requests_storage(school_year)
    result = eto_collection.delete_one({"_id": oid})
    if result.deleted_count == 0:
        return jsonify({"status": "error", "message": "Request not found."}), 404
    return jsonify({"status": "ok", "message": "Request deleted."})


# =====================================
# CALENDAR / SETTINGS API
# =====================================

@app.route("/calendar")
@require_permission("analytics")  # Reuse analytics permission - calendar is for full-admin only
def admin_calendar():
    return render_template(
        "admin_calendar.html",
        school_year_start_month=SCHOOL_YEAR_START_MONTH,
        **sidebar_context("calendar", resolve_selected_school_year(request.args.get("school_year", "")))
    )

@app.route("/api/schedule/default", methods=["GET"])
def api_get_default_schedule():
    if not login_required():
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    return jsonify({"status": "ok", "schedule": get_default_schedule()})


@app.route("/api/schedule/default", methods=["POST"])
def api_update_default_schedule():
    if not login_required() or session.get("role") != ROLE_FULL_ADMIN:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401

    payload = request.json or {}
    raw_schedule = {
        "morning_start": payload.get("morning_start", "05:00"),
        "morning_late": payload.get("morning_late", ""),
        "noon_start": payload.get("noon_start", "12:00"),
        "afternoon_start": payload.get("afternoon_start", "13:00"),
        "afternoon_late": payload.get("afternoon_late", ""),
        "afternoon_end": payload.get("afternoon_end", "17:00"),
        "late_threshold_minutes": payload.get("late_threshold_minutes", 15),
        "scan_cooldown_minutes": payload.get("scan_cooldown_minutes", 30),
        "scan_in_cooldown_seconds": payload.get("scan_in_cooldown_seconds", SCAN_LIVE_IN_COOLDOWN_SECONDS),
    }
    schedule_data = normalize_attendance_schedule(raw_schedule)

    system_settings.update_one(
        {"key": "default_schedule"},
        {"$set": {"schedule": schedule_data}},
        upsert=True
    )
    return jsonify({"status": "ok", "message": "Default schedule updated", "schedule": schedule_data})


@app.route("/api/calendar/events", methods=["GET"])
def api_get_calendar_events():
    if not login_required():
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    
    start_date = request.args.get("start")
    end_date = request.args.get("end")
    school_year = resolve_selected_school_year(request.args.get("school_year", ""))
    
    # Get the appropriate storage collection based on school year
    calendar_collection, _, _ = get_calendar_events_storage(school_year)
    
    query = {"school_year": school_year}
    if start_date or end_date:
        date_q = {}
        if start_date: date_q["$gte"] = start_date
        if end_date: date_q["$lte"] = end_date
        query["date"] = date_q
        
    eventsCursor = calendar_collection.find(query).sort("date", ASCENDING)
    events = []
    for ev in eventsCursor:
        ev["_id"] = str(ev["_id"])
        events.append(ev)
        
    return jsonify({"status": "ok", "events": events})


@app.route("/api/calendar/events", methods=["POST"])
def api_create_calendar_event():
    if not login_required() or session.get("role") != ROLE_FULL_ADMIN:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
        
    payload = request.json or {}
    date_str = str(payload.get("date") or "").strip()
    type_str = str(payload.get("type", "holiday") or "holiday").strip()
    title = str(payload.get("title", "") or "").strip()
    special_cond = str(payload.get("special_condition", "") or "").strip()
    custom_schedule = payload.get("custom_schedule")
    school_year = normalize_school_year_value(payload.get("school_year", ""))
    
    if not date_str:
        return jsonify({"status": "error", "message": "Date is required"}), 400
    
    # Derive school year from date if not provided
    if not school_year:
        from datetime import datetime
        try:
            event_date = datetime.strptime(date_str, "%Y-%m-%d")
            school_year = get_school_year_for_date(event_date)
        except ValueError:
            school_year = get_current_school_year_label()
    
    # Get the appropriate storage collection based on school year
    calendar_collection, _, _ = get_calendar_events_storage(school_year)
        
    doc = {
        "date": date_str,
        "type": type_str,
        "title": title,
        "special_condition": special_cond,
        "school_year": school_year,
    }
    if custom_schedule:
        doc["custom_schedule"] = custom_schedule
        
    try:
        res = calendar_collection.insert_one(doc)
        doc["_id"] = str(res.inserted_id)
        return jsonify({"status": "ok", "event": doc})
    except DuplicateKeyError:
        return jsonify({"status": "error", "message": "An event already exists for this date."}), 400


@app.route("/api/calendar/events/<event_id>", methods=["PUT"])
def api_update_calendar_event(event_id):
    if not login_required() or session.get("role") != ROLE_FULL_ADMIN:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
        
    payload = request.json or {}
    date_str = str(payload.get("date") or "").strip()
    type_str = str(payload.get("type", "holiday") or "holiday").strip()
    title = str(payload.get("title", "") or "").strip()
    special_cond = str(payload.get("special_condition", "") or "").strip()
    custom_schedule = payload.get("custom_schedule")
    school_year = normalize_school_year_value(payload.get("school_year", ""))
    
    # First, try to find the event in active collection, then archive collections
    from bson import ObjectId
    try:
        oid = ObjectId(event_id)
    except Exception:
        return jsonify({"status": "error", "message": "Invalid event ID"}), 400
    
    event_doc, source_collection, _is_archived = find_record_in_active_or_archive(calendar_events, calendar_events_archive, oid)
    if not event_doc:
        return jsonify({"status": "error", "message": "Event not found"}), 404

    event_date_str = date_str or str(event_doc.get("date") or "").strip()
    if not event_date_str:
        return jsonify({"status": "error", "message": "Date is required"}), 400
    
    # Use the event's school year or derive from date
    event_school_year = school_year or event_doc.get("school_year")
    if not event_school_year and event_date_str:
        from datetime import datetime
        try:
            event_date = datetime.strptime(event_date_str, "%Y-%m-%d")
            event_school_year = get_school_year_for_date(event_date)
        except ValueError:
            event_school_year = get_current_school_year_label()
    
    # Get the appropriate storage collection
    calendar_collection, _, _ = get_calendar_events_storage(event_school_year)
    
    update_data = {
        "date": event_date_str,
        "type": type_str,
        "title": title,
        "special_condition": special_cond,
        "school_year": event_school_year,
    }
    if custom_schedule:
        update_data["custom_schedule"] = custom_schedule

    updated_doc = dict(event_doc)
    updated_doc.update(update_data)
    if custom_schedule:
        updated_doc["custom_schedule"] = custom_schedule
    else:
        updated_doc.pop("custom_schedule", None)

    try:
        if calendar_collection == source_collection:
            update_ops = {"$set": update_data}
            if not custom_schedule:
                update_ops["$unset"] = {"custom_schedule": ""}
            res = calendar_collection.update_one({"_id": oid}, update_ops)
            if res.matched_count == 0:
                return jsonify({"status": "error", "message": "Event not found"}), 404
        else:
            calendar_collection.insert_one(updated_doc)
            source_collection.delete_one({"_id": oid})
    except DuplicateKeyError:
        return jsonify({"status": "error", "message": "An event already exists for this date."}), 400

    return jsonify({"status": "ok", "message": "Event updated"})


@app.route("/api/calendar/events/<event_id>", methods=["DELETE"])
def api_delete_calendar_event(event_id):
    if not login_required() or session.get("role") != ROLE_FULL_ADMIN:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
        
    from bson import ObjectId
    try:
        oid = ObjectId(event_id)
    except Exception:
        return jsonify({"status": "error", "message": "Invalid event ID"}), 400
    
    event_doc, source_collection, _is_archived = find_record_in_active_or_archive(calendar_events, calendar_events_archive, oid)
    if event_doc:
        res = source_collection.delete_one({"_id": oid})
        if res.deleted_count == 0:
            return jsonify({"status": "error", "message": "Event not found"}), 404
        return jsonify({"status": "ok", "message": "Event deleted"})
    
    return jsonify({"status": "error", "message": "Event not found"}), 404


# =====================================
# RUN APP
# =====================================
def resolve_ssl_context():
    if not HTTPS_ENABLED:
        return None

    cert_path = SSL_CERT_FILE
    key_path = SSL_KEY_FILE

    if cert_path:
        if not os.path.exists(cert_path):
            print(f"[WARNING] SSL_CERT_FILE '{cert_path}' not found. Falling back to adhoc certificate.")
        elif key_path and not os.path.exists(key_path):
            print(f"[WARNING] SSL_KEY_FILE '{key_path}' not found. Falling back to adhoc certificate.")
        else:
            return (cert_path, key_path) if key_path else cert_path

    return "adhoc"


if __name__ == "__main__":
    debug_mode = FLASK_DEBUG_MODE
    use_reloader = bool(debug_mode or DEV_AUTO_RELOAD)
    ssl_context = resolve_ssl_context()

    if not HTTPS_ENABLED:
        print("[WARNING] HTTPS_ENABLED=0. Browsers will block camera access on non-HTTPS origins (except localhost).")

    scheme = "https" if ssl_context else "http"
    print(f"[INFO] Starting Flask server on {scheme}://{FLASK_HOST}:{FLASK_PORT} (debug={debug_mode}, auto_reload={use_reloader})")

    if should_bootstrap_live_monitoring_on_startup():
        live_monitoring_ready, live_monitoring_message = ensure_live_monitoring_service_ready(
            reason="app_start",
            fail_hard=False,
        )
        if live_monitoring_ready:
            print(f"[INFO] {live_monitoring_message}")
        else:
            print(f"[WARNING] {live_monitoring_message}")

    app.run(
        host=FLASK_HOST,
        port=FLASK_PORT,
        debug=debug_mode,
        use_reloader=use_reloader,
        ssl_context=ssl_context,
    )

